import cgi
import csv
import hashlib
import hmac
import json
import os
import re
import secrets
import smtplib
import sqlite3
import subprocess
import sys
import tempfile
import threading
import time
import traceback
from datetime import datetime, timedelta
from email.message import EmailMessage
from http import cookies
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qs, quote, unquote, urlparse

try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    import pdfplumber
except Exception:
    pdfplumber = None

try:
    import pypdfium2 as pdfium
except Exception:
    pdfium = None

try:
    from rapidocr_onnxruntime import RapidOCR
except Exception:
    RapidOCR = None

OCR_ENGINE = None


ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
UPLOAD_DIR = DATA_DIR / "uploads"
BRAND_DIR = DATA_DIR / "brand"
DB_PATH = DATA_DIR / "pm_tracker.sqlite3"
HOST = os.environ.get("PM_TRACKER_HOST", "127.0.0.1")
PORT = int(os.environ.get("PM_TRACKER_PORT", "8765"))
SMTP_HOST = os.environ.get("PM_TRACKER_SMTP_HOST", "")
SMTP_PORT = int(os.environ.get("PM_TRACKER_SMTP_PORT", "587"))
SMTP_USERNAME = os.environ.get("PM_TRACKER_SMTP_USERNAME", "")
SMTP_PASSWORD = os.environ.get("PM_TRACKER_SMTP_PASSWORD", "")
SMTP_FROM = os.environ.get("PM_TRACKER_SMTP_FROM", SMTP_USERNAME)
SMTP_USE_TLS = os.environ.get("PM_TRACKER_SMTP_TLS", "1").lower() not in ("0", "false", "no")
SESSION_IDLE_TIMEOUT_MINUTES = 30
DEFAULT_NEW_USER_PASSWORD = "TPE1776"
CO_MATERIAL_MARGIN = 0.35
CO_MATERIAL_COST_FACTOR = 1 - CO_MATERIAL_MARGIN
COG_CUSTOMER_OPTIONAL_NAMES = {"expenses", "truck & auto expense", "truck and auto expense"}
ROLE_NAMES = ("Admin", "User", "Read Only", "TX/Read Only", "Field PO")
PERMISSION_DEFINITIONS = [
    {"key": "projects", "label": "Project Dashboard", "group": "Projects"},
    {"key": "project_setup", "label": "Project Setup", "group": "Projects"},
    {"key": "fieldwise", "label": "Field Wise Import", "group": "Projects"},
    {"key": "review_exceptions", "label": "Review Exceptions", "group": "Projects"},
    {"key": "vendor_invoices", "label": "Vendor Invoices", "group": "Projects"},
    {"key": "customer_billing", "label": "Customer Billing", "group": "Projects"},
    {"key": "fieldwise_audit", "label": "Field Wise Audit", "group": "Company"},
    {"key": "bids", "label": "Bid Tracking", "group": "Company"},
    {"key": "job_order_report", "label": "Job Order Quick Reference", "group": "Company"},
    {"key": "archived_projects", "label": "Archived Projects", "group": "Projects"},
    {"key": "texas_ops", "label": "Texas Operations", "group": "Texas"},
    {"key": "texas_reminders", "label": "Texas Reminders Setup", "group": "Texas"},
    {"key": "po_requests", "label": "Create / My POs", "group": "Purchase Orders"},
    {"key": "po_review", "label": "Office PO Review", "group": "Purchase Orders"},
    {"key": "cog_setup", "label": "COG Setup", "group": "Purchase Orders"},
    {"key": "activity", "label": "Activity Log", "group": "Admin"},
    {"key": "admin", "label": "Admin / Users / Backups", "group": "Admin"},
]
DEFAULT_ROLE_PERMISSIONS = {
    "Admin": {p["key"]: (1, 1) for p in PERMISSION_DEFINITIONS},
    "User": {p["key"]: (1, 1) for p in PERMISSION_DEFINITIONS if p["key"] not in ("admin", "activity")},
    "Read Only": {p["key"]: (1, 0) for p in PERMISSION_DEFINITIONS if p["key"] not in ("admin", "activity")},
    "TX/Read Only": {"texas_ops": (1, 0)},
    "Field PO": {"po_requests": (1, 1), "job_order_report": (1, 0)},
}
BID_TRACKER_SOURCE = Path(r"C:\Users\rossc\Twin Peaks Electrical\Project Manager WIP - General\Bid Request Management\_Bid Tracker\Bid Tracker.xlsx")
SERVER_STARTED_AT = datetime.now().isoformat(timespec="seconds")
REVISION_LABEL = "server-health-version"
SOURCE_PATH_AT_STARTUP = Path(__file__).resolve()
SOURCE_STAT_AT_STARTUP = SOURCE_PATH_AT_STARTUP.stat()
LOADED_SOURCE_SHA256 = hashlib.sha256(SOURCE_PATH_AT_STARTUP.read_bytes()).hexdigest()
LOADED_SOURCE_MODIFIED_AT = datetime.fromtimestamp(SOURCE_STAT_AT_STARTUP.st_mtime).isoformat(timespec="seconds")
LOADED_SOURCE_SIZE_BYTES = SOURCE_STAT_AT_STARTUP.st_size


def git_output(*args):
    try:
        return subprocess.check_output(
            ["git", *args],
            cwd=ROOT,
            stderr=subprocess.DEVNULL,
            text=True,
            timeout=3,
        ).strip()
    except Exception:
        return ""


LOADED_GIT_COMMIT = git_output("rev-parse", "HEAD")
LOADED_GIT_BRANCH = git_output("branch", "--show-current")


def html_escape(value):
    return str(value or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;").replace("'", "&#39;")


def normalize_cog_name(value):
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def po_customer_required_for_ref(job_ref):
    if not job_ref or not job_ref.get("cog_category_id"):
        return False
    return normalize_cog_name(job_ref.get("job_number")) not in COG_CUSTOMER_OPTIONAL_NAMES


def app_revision_info():
    source = Path(__file__).resolve()
    stat = source.stat()
    disk_digest = hashlib.sha256(source.read_bytes()).hexdigest()
    db_stat = DB_PATH.stat() if DB_PATH.exists() else None
    current_git_commit = git_output("rev-parse", "HEAD")
    current_git_branch = git_output("branch", "--show-current")
    upstream_git_commit = git_output("rev-parse", "@{u}")
    git_dirty = bool(git_output("status", "--porcelain", "--untracked-files=no"))
    source_matches_disk = disk_digest == LOADED_SOURCE_SHA256 and stat.st_size == LOADED_SOURCE_SIZE_BYTES
    loaded_commit_matches_current = not LOADED_GIT_COMMIT or not current_git_commit or LOADED_GIT_COMMIT == current_git_commit
    upstream_matches_current = not upstream_git_commit or not current_git_commit or upstream_git_commit == current_git_commit
    running_latest_deployed_code = source_matches_disk and loaded_commit_matches_current and upstream_matches_current and not git_dirty
    return {
        "revision_label": REVISION_LABEL,
        "server_started_at": SERVER_STARTED_AT,
        "loaded_git_branch": LOADED_GIT_BRANCH,
        "loaded_git_commit": LOADED_GIT_COMMIT,
        "loaded_git_commit_short": LOADED_GIT_COMMIT[:12],
        "current_git_branch": current_git_branch,
        "current_git_commit": current_git_commit,
        "current_git_commit_short": current_git_commit[:12],
        "upstream_git_commit": upstream_git_commit,
        "upstream_git_commit_short": upstream_git_commit[:12],
        "git_dirty": git_dirty,
        "loaded_commit_matches_current": loaded_commit_matches_current,
        "upstream_matches_current": upstream_matches_current,
        "running_latest_deployed_code": running_latest_deployed_code,
        "loaded_source_modified_at": LOADED_SOURCE_MODIFIED_AT,
        "loaded_source_size_bytes": LOADED_SOURCE_SIZE_BYTES,
        "loaded_source_sha256": LOADED_SOURCE_SHA256,
        "loaded_source_sha256_short": LOADED_SOURCE_SHA256[:12],
        "disk_source_modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
        "disk_source_size_bytes": stat.st_size,
        "disk_source_sha256": disk_digest,
        "disk_source_sha256_short": disk_digest[:12],
        "source_matches_disk": source_matches_disk,
        "source_path": str(source),
        "database_path": str(DB_PATH),
        "database_exists": DB_PATH.exists(),
        "database_modified_at": datetime.fromtimestamp(db_stat.st_mtime).isoformat(timespec="seconds") if db_stat else "Missing",
        "database_size_bytes": db_stat.st_size if db_stat else 0,
        "app_bind": f"http://{HOST}:{PORT}",
        "python_version": sys.version.split()[0],
    }


def developer_revision_html(user):
    info = app_revision_info()
    health_ok = info["running_latest_deployed_code"]
    restart_ok = info["source_matches_disk"] and info["loaded_commit_matches_current"]
    git_status = "Clean" if not info["git_dirty"] else "Uncommitted changes"
    deployed_status = "Running latest deployed code" if health_ok else "Needs attention"
    restart_status = "Running current app file" if restart_ok else "Restart needed after code update"
    rows_html = "".join(
        f"<tr><th>{html_escape(label)}</th><td>{html_escape(value)}</td></tr>"
        for label, value in [
            ("Revision Label", info["revision_label"]),
            ("Deployment Status", deployed_status),
            ("Running App Status", restart_status),
            ("Server Started", info["server_started_at"]),
            ("Loaded Git Branch", info["loaded_git_branch"] or "Not available"),
            ("Loaded Git Commit", info["loaded_git_commit_short"] or "Not available"),
            ("Current Git Branch", info["current_git_branch"] or "Not available"),
            ("Current Git Commit", info["current_git_commit_short"] or "Not available"),
            ("Upstream Commit", info["upstream_git_commit_short"] or "No upstream found"),
            ("Git Working Copy", git_status),
            ("Loaded Source Hash", info["loaded_source_sha256_short"]),
            ("Current Disk Hash", info["disk_source_sha256_short"]),
            ("Server Matches Disk", "Yes" if info["source_matches_disk"] else "No - restart the app"),
            ("Database Modified", info["database_modified_at"]),
            ("Database Size", f"{info['database_size_bytes']:,} bytes"),
            ("App Bind", info["app_bind"]),
            ("Loaded Source Modified", info["loaded_source_modified_at"]),
            ("Current Disk Modified", info["disk_source_modified_at"]),
            ("Loaded Source Size", f"{info['loaded_source_size_bytes']:,} bytes"),
            ("Current Disk Size", f"{info['disk_source_size_bytes']:,} bytes"),
            ("Source Path", info["source_path"]),
            ("Database Path", info["database_path"]),
            ("Python Version", info["python_version"]),
            ("Signed In As", f"{user.get('display_name') or user.get('username')} / {user.get('role')}"),
            ("Loaded Full SHA-256", info["loaded_source_sha256"]),
            ("Current Disk Full SHA-256", info["disk_source_sha256"]),
        ]
    )
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <meta http-equiv="Cache-Control" content="no-store">
  <title>Server Health / Version</title>
  <style>
    :root {{ --ink:#17202a; --muted:#607080; --line:#d8dee5; --bg:#f6f8fa; --blue:#2266aa; --green:#137a45; --red:#b42318; --amber:#a15c07; }}
    body {{ margin:0; font-family:"Segoe UI", Arial, sans-serif; color:var(--ink); background:var(--bg); }}
    main {{ max-width:1120px; margin:0 auto; padding:28px 24px 48px; }}
    h1 {{ margin:0 0 6px; font-size:28px; }}
    p {{ color:var(--muted); margin:0 0 18px; }}
    .panel {{ background:white; border:1px solid var(--line); border-radius:8px; padding:16px; }}
    .cards {{ display:grid; grid-template-columns:repeat(4, minmax(0, 1fr)); gap:12px; margin:16px 0; }}
    .card {{ background:white; border:1px solid var(--line); border-radius:8px; padding:14px; min-height:96px; }}
    .card-label {{ color:#526376; font-size:12px; font-weight:800; text-transform:uppercase; }}
    .card-value {{ font-size:22px; font-weight:850; margin-top:10px; overflow-wrap:anywhere; }}
    .card-note {{ color:var(--muted); font-size:13px; margin-top:8px; }}
    .ok {{ color:var(--green); }}
    .bad {{ color:var(--red); }}
    .warn {{ color:var(--amber); }}
    table {{ width:100%; border-collapse:collapse; font-size:14px; }}
    th, td {{ text-align:left; vertical-align:top; padding:11px 10px; border-bottom:1px solid var(--line); }}
    th {{ width:210px; color:#34495e; background:#eef2f6; }}
    td {{ word-break:break-word; }}
    .status {{ display:inline-flex; align-items:center; border-radius:999px; padding:6px 10px; font-size:13px; font-weight:800; background:#e8f5ee; color:var(--green); }}
    .status.bad {{ background:#fdecea; color:var(--red); }}
    .actions {{ display:flex; gap:8px; flex-wrap:wrap; margin-top:16px; }}
    a, button {{ border:1px solid var(--line); background:white; color:var(--ink); padding:9px 12px; border-radius:6px; cursor:pointer; font-weight:650; text-decoration:none; font:inherit; }}
    .primary {{ background:var(--blue); color:white; border-color:var(--blue); }}
    @media (max-width: 860px) {{ .cards {{ grid-template-columns:1fr 1fr; }} }}
    @media (max-width: 560px) {{ .cards {{ grid-template-columns:1fr; }} }}
  </style>
</head>
<body>
  <main>
    <h1>Server Health / Version</h1>
    <p>Use this page after a deploy to confirm the app restarted on the expected code and to see when the database file last changed.</p>
    <span class="status {'ok' if health_ok else 'bad'}">{html_escape(deployed_status)}</span>
    <div class="cards">
      <div class="card">
        <div class="card-label">Current Git Commit</div>
        <div class="card-value">{html_escape(info['current_git_commit_short'] or 'N/A')}</div>
        <div class="card-note">{html_escape(info['current_git_branch'] or 'No branch found')}</div>
      </div>
      <div class="card">
        <div class="card-label">Running App File</div>
        <div class="card-value {'ok' if restart_ok else 'bad'}">{html_escape('Current' if restart_ok else 'Restart Needed')}</div>
        <div class="card-note">Loaded {html_escape(info['loaded_source_sha256_short'])}</div>
      </div>
      <div class="card">
        <div class="card-label">Database File</div>
        <div class="card-value">{html_escape(info['database_modified_at'])}</div>
        <div class="card-note">{info['database_size_bytes']:,} bytes</div>
      </div>
      <div class="card">
        <div class="card-label">Git Working Copy</div>
        <div class="card-value {'ok' if not info['git_dirty'] else 'warn'}">{html_escape(git_status)}</div>
        <div class="card-note">Uncommitted app files can mean the server differs from GitHub.</div>
      </div>
    </div>
    <div class="panel">
      <table>{rows_html}</table>
      <div class="actions">
        <button class="primary" type="button" onclick="window.location.reload()">Refresh Health</button>
        <a href="/">Back to App</a>
      </div>
    </div>
  </main>
</body>
</html>"""


def pwa_manifest_json():
    return json.dumps(
        {
            "name": "Twin Peaks Company Dashboard",
            "short_name": "TPE Field PO",
            "description": "Twin Peaks Electrical project tracking and field PO requests.",
            "start_url": "/",
            "scope": "/",
            "display": "standalone",
            "background_color": "#f6f8fa",
            "theme_color": "#152332",
            "orientation": "portrait-primary",
            "icons": [
                {
                    "src": "/brand/twin-peaks-logo.png",
                    "sizes": "192x192 512x512",
                    "type": "image/png",
                    "purpose": "any maskable",
                }
            ],
        }
    )


def service_worker_js():
    return """
self.addEventListener('install', event => {
  self.skipWaiting();
});

self.addEventListener('activate', event => {
  event.waitUntil(
    caches.keys()
      .then(keys => Promise.all(keys.map(key => caches.delete(key))))
      .then(() => self.registration.unregister())
      .then(() => self.clients.matchAll({ type: 'window' }))
      .then(clients => Promise.all(clients.map(client => client.navigate(client.url))))
  );
});
""".strip()


def offline_html():
    return """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <meta name="theme-color" content="#152332">
  <title>TPE Field PO Offline</title>
  <style>
    body { margin:0; min-height:100vh; display:grid; place-items:center; font-family:"Segoe UI", Arial, sans-serif; background:#f6f8fa; color:#17202a; padding:22px; }
    .card { max-width:460px; background:white; border:1px solid #d8dee5; border-radius:8px; padding:24px; box-shadow:0 14px 34px rgba(15,35,55,.12); text-align:center; }
    img { width:110px; max-width:70%; margin-bottom:14px; }
    h1 { margin:0 0 8px; font-size:24px; }
    p { color:#607080; line-height:1.45; }
  </style>
</head>
<body>
  <div class="card">
    <img src="/brand/twin-peaks-logo.png" alt="Twin Peaks Electrical">
    <h1>Connection needed</h1>
    <p>This app needs a connection before creating or updating POs so PO numbers stay accurate. Reconnect and try again.</p>
  </div>
</body>
</html>"""


def purchase_order_html(po):
    attachment = ""
    if po["attachment_file"]:
        safe_name = quote(po["attachment_file"])
        attachment = f'<p><strong>Attachment:</strong> <a href="/uploads/{safe_name}" target="_blank" rel="noopener">{html_escape(po["attachment_file"])}</a></p>'
    pickup = ""
    if po["pickup_file"]:
        safe_name = quote(po["pickup_file"])
        pickup = f'<p><strong>Pickup Ticket:</strong> <a href="/uploads/{safe_name}" target="_blank" rel="noopener">{html_escape(po["pickup_file"])}</a></p>'
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html_escape(po["po_number"])} Purchase Order</title>
  <style>
    body {{ margin:0; font-family:"Segoe UI", Arial, sans-serif; color:#17202a; background:#f6f8fa; }}
    main {{ max-width:860px; margin:0 auto; padding:28px 22px 44px; }}
    .po {{ background:white; border:1px solid #d8dee5; border-radius:8px; padding:28px; }}
    h1 {{ margin:0; font-size:28px; }}
    .meta {{ color:#607080; margin-top:5px; }}
    .grid {{ display:grid; grid-template-columns:1fr 1fr; gap:14px; margin-top:22px; }}
    .box {{ border:1px solid #d8dee5; border-radius:8px; padding:14px; }}
    .label {{ color:#607080; font-size:12px; font-weight:700; text-transform:uppercase; margin-bottom:5px; }}
    .value {{ font-size:18px; font-weight:700; }}
    .desc {{ white-space:pre-wrap; line-height:1.45; }}
    .actions {{ margin:0 0 16px; display:flex; gap:8px; }}
    button, a.btn {{ border:1px solid #d8dee5; background:white; color:#17202a; padding:9px 12px; border-radius:6px; cursor:pointer; font-weight:700; text-decoration:none; }}
    .primary {{ background:#2266aa; color:white; border-color:#2266aa; }}
    @media print {{ body {{ background:white; }} main {{ padding:0; }} .actions {{ display:none; }} .po {{ border:0; }} }}
    @media (max-width:720px) {{ .grid {{ grid-template-columns:1fr; }} .po {{ padding:18px; }} }}
  </style>
</head>
<body>
  <main>
    <div class="actions"><button class="primary" onclick="window.print()">Print / Save PDF</button><a class="btn" href="/">Back to App</a></div>
    <div class="po">
      <h1>Purchase Order {html_escape(po["po_number"])}</h1>
      <div class="meta">Status: {html_escape(po["status"])} / Created {html_escape((po["created_at"] or "").replace("T", " "))}</div>
      <div class="grid">
        <div class="box"><div class="label">Vendor</div><div class="value">{html_escape(po["vendor"])}</div></div>
        <div class="box"><div class="label">Estimated Amount</div><div class="value">${money(po["estimated_amount"]):,.2f}</div></div>
        <div class="box"><div class="label">Job / Order #</div><div class="value">{html_escape(po["job_number"])}</div><div class="meta">{html_escape(po["job_label"])}</div></div>
        <div class="box"><div class="label">Created By</div><div class="value">{html_escape(po["requested_by_username"])}</div></div>
        <div class="box"><div class="label">Customer</div><div class="value">{html_escape(po["customer_name"] or "")}</div></div>
      </div>
      <div class="box" style="margin-top:14px"><div class="label">PO Details</div><div class="desc">{html_escape(po["description"])}</div></div>
      {attachment}
      {pickup}
    </div>
  </main>
</body>
</html>"""


def db():
    DATA_DIR.mkdir(exist_ok=True)
    UPLOAD_DIR.mkdir(exist_ok=True)
    BRAND_DIR.mkdir(exist_ok=True)
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys = ON")
    return con


def init_db():
    with db() as con:
        con.executescript(
            """
            CREATE TABLE IF NOT EXISTS projects (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              project_code TEXT NOT NULL UNIQUE,
              name TEXT NOT NULL,
              customer TEXT,
              location TEXT,
              customer_po TEXT,
              description TEXT,
              rate_set_id INTEGER,
              contract_value REAL DEFAULT 0,
              status TEXT DEFAULT 'Active',
              closed_at TEXT,
              archived_at TEXT,
              created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS subprojects (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
              job_number TEXT,
              code TEXT NOT NULL,
              name TEXT NOT NULL,
              pricing_type TEXT DEFAULT 'Fixed',
              contract_value REAL DEFAULT 0,
              budget_labor_hours REAL DEFAULT 0,
              budget_labor REAL DEFAULT 0,
              budget_material REAL DEFAULT 0,
              budget_equipment REAL DEFAULT 0,
              budget_vendor REAL DEFAULT 0,
              budget_other REAL DEFAULT 0,
              UNIQUE(project_id, code)
            );

            CREATE TABLE IF NOT EXISTS change_orders (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
              subproject_id INTEGER REFERENCES subprojects(id) ON DELETE SET NULL,
              co_number TEXT NOT NULL,
              job_number TEXT,
              order_type TEXT DEFAULT 'Change Order',
              pricing_type TEXT DEFAULT 'Fixed',
              title TEXT,
              status TEXT DEFAULT 'Pending',
              quoted_value REAL DEFAULT 0,
              approved_value REAL DEFAULT 0,
              UNIQUE(project_id, subproject_id, co_number)
            );

            CREATE TABLE IF NOT EXISTS cost_records (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              project_id INTEGER REFERENCES projects(id) ON DELETE SET NULL,
              subproject_id INTEGER REFERENCES subprojects(id) ON DELETE SET NULL,
              change_order_id INTEGER REFERENCES change_orders(id) ON DELETE SET NULL,
              source TEXT NOT NULL,
              source_file TEXT,
              ticket_or_invoice TEXT,
              record_date TEXT,
              status TEXT,
              cost_type TEXT,
              item TEXT,
              description TEXT,
              qty REAL DEFAULT 0,
              rate REAL DEFAULT 0,
              amount REAL DEFAULT 0,
              sales_rate REAL DEFAULT 0,
              sales_amount REAL DEFAULT 0,
              raw_rate REAL DEFAULT 0,
              raw_cost_source TEXT,
              vendor TEXT,
              notes TEXT,
              created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS vendor_invoice_allocations (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
              source_file TEXT,
              ticket_or_invoice TEXT,
              vendor TEXT,
              original_total REAL DEFAULT 0,
              allocation_count INTEGER DEFAULT 0,
              allocated_by_user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
              allocated_by_username TEXT,
              allocated_at TEXT NOT NULL,
              notes TEXT
            );

            CREATE TABLE IF NOT EXISTS vendor_invoice_allocation_lines (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              allocation_id INTEGER NOT NULL REFERENCES vendor_invoice_allocations(id) ON DELETE CASCADE,
              subproject_id INTEGER REFERENCES subprojects(id) ON DELETE SET NULL,
              change_order_id INTEGER REFERENCES change_orders(id) ON DELETE SET NULL,
              amount REAL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS customer_invoices (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
              subproject_id INTEGER REFERENCES subprojects(id) ON DELETE SET NULL,
              change_order_id INTEGER REFERENCES change_orders(id) ON DELETE SET NULL,
              invoice_number TEXT,
              billing_type TEXT DEFAULT 'Progress',
              invoice_date TEXT,
              due_date TEXT,
              status TEXT DEFAULT 'Draft',
              amount REAL DEFAULT 0,
              paid_amount REAL DEFAULT 0,
              invoice_file TEXT,
              notes TEXT,
              created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS customer_invoice_allocations (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              customer_invoice_id INTEGER NOT NULL REFERENCES customer_invoices(id) ON DELETE CASCADE,
              project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
              subproject_id INTEGER REFERENCES subprojects(id) ON DELETE SET NULL,
              change_order_id INTEGER REFERENCES change_orders(id) ON DELETE SET NULL,
              amount REAL DEFAULT 0,
              notes TEXT,
              created_by_user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
              created_by_username TEXT,
              created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS project_invoice_alert_acknowledgements (
              project_id INTEGER PRIMARY KEY REFERENCES projects(id) ON DELETE CASCADE,
              acknowledged_at TEXT NOT NULL,
              acknowledged_by_user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
              acknowledged_by_username TEXT,
              notes TEXT
            );

            CREATE TABLE IF NOT EXISTS job_invoice_alert_acknowledgements (
              target_type TEXT NOT NULL,
              target_id INTEGER NOT NULL,
              acknowledged_at TEXT NOT NULL,
              acknowledged_by_user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
              acknowledged_by_username TEXT,
              notes TEXT,
              PRIMARY KEY (target_type, target_id)
            );

            CREATE TABLE IF NOT EXISTS internal_rates (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              rate_set_id INTEGER,
              category_type TEXT NOT NULL,
              category TEXT NOT NULL,
              raw_rate REAL DEFAULT 0,
              active INTEGER DEFAULT 1,
              UNIQUE(rate_set_id, category_type, category)
            );

            CREATE TABLE IF NOT EXISTS rate_sets (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              name TEXT NOT NULL UNIQUE,
              effective_date TEXT,
              active INTEGER DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS bid_requests (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              rfq_no TEXT NOT NULL UNIQUE,
              date_received TEXT,
              customer TEXT,
              project_name TEXT,
              estimator TEXT,
              stage TEXT,
              bid_due_date TEXT,
              go_no_go TEXT,
              estimated_cost REAL DEFAULT 0,
              target_margin REAL DEFAULT 0,
              bid_price REAL DEFAULT 0,
              probability REAL DEFAULT 0,
              weighted_value REAL DEFAULT 0,
              submission_status TEXT,
              outcome TEXT DEFAULT 'Pending',
              notes TEXT,
              created_at TEXT NOT NULL,
              updated_at TEXT
            );

            CREATE TABLE IF NOT EXISTS bid_risks (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              risk_id TEXT,
              rfq_no TEXT,
              risk_description TEXT,
              probability INTEGER DEFAULT 0,
              impact INTEGER DEFAULT 0,
              risk_rating INTEGER DEFAULT 0,
              pricing_action TEXT,
              mitigation TEXT,
              owner TEXT,
              status TEXT,
              created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS bid_win_loss (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              rfq_no TEXT,
              customer TEXT,
              outcome TEXT,
              bid_value REAL DEFAULT 0,
              known_competitor TEXT,
              primary_reason TEXT,
              pricing_position TEXT,
              lesson_learned TEXT,
              next_action TEXT,
              created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS users (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              username TEXT NOT NULL UNIQUE,
              display_name TEXT,
              password_hash TEXT NOT NULL,
              role TEXT DEFAULT 'User',
              active INTEGER DEFAULT 1,
              po_auto_issue INTEGER DEFAULT 0,
              must_change_password INTEGER DEFAULT 0,
              created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS user_sessions (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
              session_token TEXT NOT NULL UNIQUE,
              created_at TEXT NOT NULL,
              expires_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS role_permissions (
              role TEXT NOT NULL,
              permission_key TEXT NOT NULL,
              can_view INTEGER DEFAULT 0,
              can_edit INTEGER DEFAULT 0,
              updated_at TEXT,
              PRIMARY KEY (role, permission_key)
            );

            CREATE TABLE IF NOT EXISTS financial_reports (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              report_date TEXT NOT NULL,
              report_type TEXT NOT NULL,
              source_file TEXT,
              uploaded_at TEXT NOT NULL,
              notes TEXT
            );

            CREATE TABLE IF NOT EXISTS financial_metrics (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              report_id INTEGER NOT NULL REFERENCES financial_reports(id) ON DELETE CASCADE,
              metric_key TEXT NOT NULL,
              label TEXT NOT NULL,
              amount REAL DEFAULT 0,
              UNIQUE(report_id, metric_key)
            );

            CREATE TABLE IF NOT EXISTS texas_ap_snapshots (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              report_date TEXT NOT NULL UNIQUE,
              ap_total REAL DEFAULT 0,
              notes TEXT,
              created_by_user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
              created_by_username TEXT,
              created_at TEXT NOT NULL,
              updated_at TEXT
            );

            CREATE TABLE IF NOT EXISTS texas_upload_reminder_recipients (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              user_id INTEGER NOT NULL UNIQUE REFERENCES users(id) ON DELETE CASCADE,
              active INTEGER DEFAULT 1,
              created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS texas_upload_reminder_settings (
              id INTEGER PRIMARY KEY CHECK (id = 1),
              enabled INTEGER DEFAULT 1,
              weekday INTEGER DEFAULT 4,
              reminder_time TEXT DEFAULT '15:00',
              updated_at TEXT
            );

            CREATE TABLE IF NOT EXISTS texas_upload_reminder_log (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              reminder_date TEXT NOT NULL,
              recipient TEXT NOT NULL,
              subject TEXT,
              status TEXT NOT NULL,
              error TEXT,
              sent_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS billing_invoice_reminder_recipients (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              user_id INTEGER NOT NULL UNIQUE REFERENCES users(id) ON DELETE CASCADE,
              active INTEGER DEFAULT 1,
              created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS billing_invoice_reminder_settings (
              id INTEGER PRIMARY KEY CHECK (id = 1),
              enabled INTEGER DEFAULT 1,
              weekday INTEGER DEFAULT 4,
              reminder_time TEXT DEFAULT '15:00',
              lookback_days INTEGER DEFAULT 7,
              updated_at TEXT
            );

            CREATE TABLE IF NOT EXISTS billing_invoice_reminder_log (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              reminder_date TEXT NOT NULL,
              recipient TEXT NOT NULL,
              subject TEXT,
              status TEXT NOT NULL,
              error TEXT,
              job_count INTEGER DEFAULT 0,
              sent_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS po_untouched_reminder_recipients (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              user_id INTEGER NOT NULL UNIQUE REFERENCES users(id) ON DELETE CASCADE,
              active INTEGER DEFAULT 1,
              created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS po_untouched_reminder_settings (
              id INTEGER PRIMARY KEY CHECK (id = 1),
              enabled INTEGER DEFAULT 1,
              weekday INTEGER DEFAULT 4,
              reminder_time TEXT DEFAULT '15:00',
              untouched_days INTEGER DEFAULT 2,
              updated_at TEXT
            );

            CREATE TABLE IF NOT EXISTS po_untouched_reminder_log (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              reminder_date TEXT NOT NULL,
              recipient TEXT NOT NULL,
              subject TEXT,
              status TEXT NOT NULL,
              error TEXT,
              po_count INTEGER DEFAULT 0,
              sent_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS activity_log (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              actor_user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
              actor_username TEXT,
              action TEXT NOT NULL,
              entity_type TEXT,
              entity_label TEXT,
              details TEXT,
              created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS purchase_orders (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              po_number TEXT NOT NULL UNIQUE,
              project_id INTEGER REFERENCES projects(id) ON DELETE SET NULL,
              subproject_id INTEGER REFERENCES subprojects(id) ON DELETE SET NULL,
              change_order_id INTEGER REFERENCES change_orders(id) ON DELETE SET NULL,
              cog_category_id INTEGER REFERENCES cog_categories(id) ON DELETE SET NULL,
              job_number TEXT NOT NULL,
              job_label TEXT,
              customer_name TEXT,
              vendor TEXT NOT NULL,
              description TEXT NOT NULL,
              estimated_amount REAL DEFAULT 0,
              attachment_file TEXT,
              pickup_file TEXT,
              pickup_uploaded_at TEXT,
              pickup_uploaded_by_user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
              pickup_uploaded_by_username TEXT,
              status TEXT DEFAULT 'Pending Approval',
              closed_at TEXT,
              closed_by_user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
              closed_by_username TEXT,
              close_reason TEXT,
              voided_at TEXT,
              voided_by_user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
              voided_by_username TEXT,
              void_reason TEXT,
              requested_by_user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
              requested_by_username TEXT,
              created_at TEXT NOT NULL,
              updated_at TEXT
            );

            CREATE TABLE IF NOT EXISTS cog_categories (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              code TEXT NOT NULL,
              name TEXT NOT NULL UNIQUE,
              description TEXT,
              active INTEGER DEFAULT 1,
              created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS fieldwise_audit_omissions (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              ticket_number TEXT NOT NULL,
              order_number TEXT NOT NULL,
              customer TEXT,
              project_name TEXT,
              reason TEXT,
              omitted_by_user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
              omitted_by_username TEXT,
              created_at TEXT NOT NULL,
              UNIQUE(ticket_number, order_number)
            );
            """
        )
        existing_cols = [r["name"] for r in con.execute("PRAGMA table_info(subprojects)").fetchall()]
        if "job_number" not in existing_cols:
            con.execute("ALTER TABLE subprojects ADD COLUMN job_number TEXT")
        if "budget_labor_hours" not in existing_cols:
            con.execute("ALTER TABLE subprojects ADD COLUMN budget_labor_hours REAL DEFAULT 0")
        if "budget_equipment" not in existing_cols:
            con.execute("ALTER TABLE subprojects ADD COLUMN budget_equipment REAL DEFAULT 0")
        if "pricing_type" not in existing_cols:
            con.execute("ALTER TABLE subprojects ADD COLUMN pricing_type TEXT DEFAULT 'Fixed'")
        if "contract_value" not in existing_cols:
            con.execute("ALTER TABLE subprojects ADD COLUMN contract_value REAL DEFAULT 0")
        bid_cols = [r["name"] for r in con.execute("PRAGMA table_info(bid_requests)").fetchall()]
        if "updated_at" not in bid_cols:
            con.execute("ALTER TABLE bid_requests ADD COLUMN updated_at TEXT")
        con.execute("UPDATE bid_requests SET updated_at = COALESCE(updated_at, created_at)")
        project_cols = [r["name"] for r in con.execute("PRAGMA table_info(projects)").fetchall()]
        if "customer_po" not in project_cols:
            con.execute("ALTER TABLE projects ADD COLUMN customer_po TEXT")
        if "description" not in project_cols:
            con.execute("ALTER TABLE projects ADD COLUMN description TEXT")
        if "status" not in project_cols:
            con.execute("ALTER TABLE projects ADD COLUMN status TEXT DEFAULT 'Active'")
        if "closed_at" not in project_cols:
            con.execute("ALTER TABLE projects ADD COLUMN closed_at TEXT")
        if "archived_at" not in project_cols:
            con.execute("ALTER TABLE projects ADD COLUMN archived_at TEXT")
        con.execute("UPDATE projects SET status = COALESCE(status, 'Active')")
        co_cols = [r["name"] for r in con.execute("PRAGMA table_info(change_orders)").fetchall()]
        if "job_number" not in co_cols:
            con.execute("ALTER TABLE change_orders ADD COLUMN job_number TEXT")
        if "pricing_type" not in co_cols:
            con.execute("ALTER TABLE change_orders ADD COLUMN pricing_type TEXT DEFAULT 'Fixed'")
        if "order_type" not in co_cols:
            con.execute("ALTER TABLE change_orders ADD COLUMN order_type TEXT DEFAULT 'Change Order'")
        con.execute("UPDATE change_orders SET order_type = COALESCE(order_type, 'Change Order')")
        co_schema = con.execute("SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'change_orders'").fetchone()
        if co_schema and "UNIQUE(project_id, co_number)" in (co_schema["sql"] or ""):
            con.execute("PRAGMA foreign_keys = OFF")
            con.executescript(
                """
                CREATE TABLE change_orders_new (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                  subproject_id INTEGER REFERENCES subprojects(id) ON DELETE SET NULL,
                  co_number TEXT NOT NULL,
                  job_number TEXT,
                  order_type TEXT DEFAULT 'Change Order',
                  pricing_type TEXT DEFAULT 'Fixed',
                  title TEXT,
                  status TEXT DEFAULT 'Pending',
                  quoted_value REAL DEFAULT 0,
                  approved_value REAL DEFAULT 0,
                  UNIQUE(project_id, subproject_id, co_number)
                );
                INSERT INTO change_orders_new (
                  id, project_id, subproject_id, co_number, job_number, order_type, pricing_type, title, status, quoted_value, approved_value
                )
                SELECT
                  id, project_id, subproject_id, co_number, job_number, COALESCE(order_type, 'Change Order'), COALESCE(pricing_type, 'Fixed'), title, status, quoted_value, approved_value
                FROM change_orders;
                DROP TABLE change_orders;
                ALTER TABLE change_orders_new RENAME TO change_orders;
                """
            )
            con.execute("PRAGMA foreign_keys = ON")
        if "rate_set_id" not in project_cols:
            con.execute("ALTER TABLE projects ADD COLUMN rate_set_id INTEGER")
        customer_invoice_cols = [r["name"] for r in con.execute("PRAGMA table_info(customer_invoices)").fetchall()]
        if "invoice_file" not in customer_invoice_cols:
            con.execute("ALTER TABLE customer_invoices ADD COLUMN invoice_file TEXT")
        con.execute(
            """
            INSERT INTO customer_invoice_allocations (
              customer_invoice_id, project_id, subproject_id, change_order_id, amount,
              notes, created_by_user_id, created_by_username, created_at
            )
            SELECT
              ci.id,
              ci.project_id,
              ci.subproject_id,
              ci.change_order_id,
              COALESCE(ci.amount, 0),
              'Original invoice allocation',
              NULL,
              NULL,
              ci.created_at
            FROM customer_invoices ci
            WHERE NOT EXISTS (
              SELECT 1
              FROM customer_invoice_allocations cia
              WHERE cia.customer_invoice_id = ci.id
            )
            """
        )
        cost_cols = [r["name"] for r in con.execute("PRAGMA table_info(cost_records)").fetchall()]
        if "sales_rate" not in cost_cols:
            con.execute("ALTER TABLE cost_records ADD COLUMN sales_rate REAL DEFAULT 0")
        if "sales_amount" not in cost_cols:
            con.execute("ALTER TABLE cost_records ADD COLUMN sales_amount REAL DEFAULT 0")
        if "raw_rate" not in cost_cols:
            con.execute("ALTER TABLE cost_records ADD COLUMN raw_rate REAL DEFAULT 0")
        if "raw_cost_source" not in cost_cols:
            con.execute("ALTER TABLE cost_records ADD COLUMN raw_cost_source TEXT")
        con.execute(
            """
            UPDATE cost_records
            SET amount = CASE
                  WHEN change_order_id IS NOT NULL OR COALESCE((SELECT pricing_type FROM subprojects WHERE id = cost_records.subproject_id), 'Fixed') = 'T&M'
                    THEN COALESCE(sales_amount, 0) * ?
                  ELSE 0
                END,
                raw_rate = CASE
                  WHEN change_order_id IS NOT NULL OR COALESCE((SELECT pricing_type FROM subprojects WHERE id = cost_records.subproject_id), 'Fixed') = 'T&M'
                    THEN COALESCE(sales_rate, rate, 0) * ?
                  ELSE 0
                END,
                raw_cost_source = CASE
                  WHEN change_order_id IS NOT NULL THEN 'CO T&M material estimate at 35% margin'
                  WHEN COALESCE((SELECT pricing_type FROM subprojects WHERE id = cost_records.subproject_id), 'Fixed') = 'T&M' THEN 'Subproject T&M material estimate at 35% margin'
                  ELSE 'Usage only - not budget cost'
                END
            WHERE cost_type = 'Field Ticket Material'
            """,
            (CO_MATERIAL_COST_FACTOR, CO_MATERIAL_COST_FACTOR),
        )
        rate_cols = [r["name"] for r in con.execute("PRAGMA table_info(internal_rates)").fetchall()]
        if "rate_set_id" not in rate_cols:
            con.execute("ALTER TABLE internal_rates ADD COLUMN rate_set_id INTEGER")
        user_cols = [r["name"] for r in con.execute("PRAGMA table_info(users)").fetchall()]
        if "po_auto_issue" not in user_cols:
            con.execute("ALTER TABLE users ADD COLUMN po_auto_issue INTEGER DEFAULT 0")
            con.execute("UPDATE users SET po_auto_issue = 1 WHERE role IN ('Admin', 'User')")
        if "must_change_password" not in user_cols:
            con.execute("ALTER TABLE users ADD COLUMN must_change_password INTEGER DEFAULT 0")
        po_cols = [r["name"] for r in con.execute("PRAGMA table_info(purchase_orders)").fetchall()]
        if "pickup_file" not in po_cols:
            con.execute("ALTER TABLE purchase_orders ADD COLUMN pickup_file TEXT")
        if "cog_category_id" not in po_cols:
            con.execute("ALTER TABLE purchase_orders ADD COLUMN cog_category_id INTEGER REFERENCES cog_categories(id) ON DELETE SET NULL")
        if "customer_name" not in po_cols:
            con.execute("ALTER TABLE purchase_orders ADD COLUMN customer_name TEXT")
        po_extra_columns = {
            "pickup_uploaded_at": "TEXT",
            "pickup_uploaded_by_user_id": "INTEGER",
            "pickup_uploaded_by_username": "TEXT",
            "closed_at": "TEXT",
            "closed_by_user_id": "INTEGER",
            "closed_by_username": "TEXT",
            "close_reason": "TEXT",
            "voided_at": "TEXT",
            "voided_by_user_id": "INTEGER",
            "voided_by_username": "TEXT",
            "void_reason": "TEXT",
        }
        for column_name, column_type in po_extra_columns.items():
            if column_name not in po_cols:
                con.execute(f"ALTER TABLE purchase_orders ADD COLUMN {column_name} {column_type}")
        con.execute(
            """
            UPDATE purchase_orders
            SET status = CASE
              WHEN status IN ('Open', 'Issued') AND COALESCE(pickup_file, '') <> '' THEN 'Picked Up'
              WHEN status = 'Open' THEN 'Issued'
              WHEN status = 'Pending Review' THEN 'Pending Approval'
              WHEN status = 'Received' THEN 'Picked Up'
              ELSE status
            END
            WHERE status IN ('Open', 'Pending Review', 'Received')
            """
        )
        cog_schema = con.execute("SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'cog_categories'").fetchone()
        if cog_schema and "code TEXT NOT NULL UNIQUE" in (cog_schema["sql"] or ""):
            con.execute("PRAGMA foreign_keys = OFF")
            con.executescript(
                """
                CREATE TABLE cog_categories_new (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  code TEXT NOT NULL,
                  name TEXT NOT NULL UNIQUE,
                  description TEXT,
                  active INTEGER DEFAULT 1,
                  created_at TEXT NOT NULL
                );
                INSERT INTO cog_categories_new (id, code, name, description, active, created_at)
                SELECT id, code, name, description, COALESCE(active, 1), created_at
                FROM cog_categories;
                DROP TABLE cog_categories;
                ALTER TABLE cog_categories_new RENAME TO cog_categories;
                """
            )
            con.execute("PRAGMA foreign_keys = ON")
        default_rate_set = con.execute("SELECT id FROM rate_sets WHERE name = 'Current'").fetchone()
        if not default_rate_set:
            cur = con.execute("INSERT INTO rate_sets (name, effective_date, active) VALUES ('Current', '', 1)")
            default_rate_set_id = cur.lastrowid
        else:
            default_rate_set_id = default_rate_set["id"]
        con.execute("UPDATE internal_rates SET rate_set_id = ? WHERE rate_set_id IS NULL", (default_rate_set_id,))
        con.execute("UPDATE projects SET rate_set_id = ? WHERE rate_set_id IS NULL", (default_rate_set_id,))
        user_count = con.execute("SELECT COUNT(*) count FROM users").fetchone()["count"]
        if not user_count:
            con.execute(
                "INSERT INTO users (username, display_name, password_hash, role, active, created_at) VALUES (?, ?, ?, 'Admin', 1, ?)",
                ("admin", "Administrator", hash_password("ChangeMe123!"), datetime.now().isoformat(timespec="seconds")),
            )
        now = datetime.now().isoformat(timespec="seconds")
        for role, defaults in DEFAULT_ROLE_PERMISSIONS.items():
            for permission in PERMISSION_DEFINITIONS:
                can_view, can_edit = defaults.get(permission["key"], (0, 0))
                con.execute(
                    """
                    INSERT OR IGNORE INTO role_permissions (role, permission_key, can_view, can_edit, updated_at)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (role, permission["key"], can_view, can_edit, now),
                )
    seed_bid_tracker_from_workbook()


def rows(sql, params=()):
    with db() as con:
        return [dict(r) for r in con.execute(sql, params).fetchall()]


def one(sql, params=()):
    with db() as con:
        r = con.execute(sql, params).fetchone()
        return dict(r) if r else None


def execute(sql, params=()):
    with db() as con:
        cur = con.execute(sql, params)
        return cur.lastrowid


def date_text(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value or "").strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def bid_price_value(estimated_cost, target_margin, explicit_bid_price=0):
    if explicit_bid_price:
        return money(explicit_bid_price)
    cost = money(estimated_cost)
    margin = money(target_margin)
    return cost / (1 - margin) if cost and margin < 1 else 0


def weighted_bid_value(outcome, bid_price, probability):
    outcome = str(outcome or "Pending")
    if outcome == "Won":
        return money(bid_price)
    if outcome == "Lost":
        return 0
    return money(bid_price) * money(probability)


def hash_password(password, salt=None):
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", str(password).encode("utf-8"), salt.encode("utf-8"), 200000)
    return f"pbkdf2_sha256$200000${salt}${digest.hex()}"


def verify_password(password, password_hash):
    try:
        algorithm, rounds, salt, digest = str(password_hash).split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        test = hashlib.pbkdf2_hmac("sha256", str(password).encode("utf-8"), salt.encode("utf-8"), int(rounds)).hex()
        return hmac.compare_digest(test, digest)
    except Exception:
        return False


def create_session(user_id):
    token = secrets.token_urlsafe(32)
    now = datetime.now()
    expires = now + timedelta(minutes=SESSION_IDLE_TIMEOUT_MINUTES)
    execute("DELETE FROM user_sessions WHERE expires_at <= ?", (now.isoformat(timespec="seconds"),))
    execute(
        "INSERT INTO user_sessions (user_id, session_token, created_at, expires_at) VALUES (?, ?, ?, ?)",
        (user_id, token, now.isoformat(timespec="seconds"), expires.isoformat(timespec="seconds")),
    )
    return token


def parse_cookie_header(header):
    jar = cookies.SimpleCookie()
    if header:
        try:
            jar.load(header)
        except cookies.CookieError:
            return {}
    return {key: morsel.value for key, morsel in jar.items()}


def current_user(handler):
    token = parse_cookie_header(handler.headers.get("Cookie")).get("pm_session")
    if not token:
        return None
    now = datetime.now()
    user = one(
        """
        SELECT users.id, users.username, users.display_name, users.role, users.active,
               COALESCE(users.po_auto_issue, 0) AS po_auto_issue,
               COALESCE(users.must_change_password, 0) AS must_change_password
        FROM user_sessions
        JOIN users ON users.id = user_sessions.user_id
        WHERE user_sessions.session_token = ?
          AND users.active = 1
          AND user_sessions.expires_at > ?
        """,
        (token, now.isoformat(timespec="seconds")),
    )
    if user:
        new_expires = now + timedelta(minutes=SESSION_IDLE_TIMEOUT_MINUTES)
        execute(
            "UPDATE user_sessions SET expires_at = ? WHERE session_token = ?",
            (new_expires.isoformat(timespec="seconds"), token),
        )
    return user


def role_permissions(role):
    role = clean_role(role)
    if role == "Admin":
        return {p["key"]: {"can_view": 1, "can_edit": 1} for p in PERMISSION_DEFINITIONS}
    records = rows("SELECT permission_key, can_view, can_edit FROM role_permissions WHERE role = ?", (role,))
    permissions = {r["permission_key"]: {"can_view": int(r["can_view"] or 0), "can_edit": int(r["can_edit"] or 0)} for r in records}
    defaults = DEFAULT_ROLE_PERMISSIONS.get(role, {})
    for permission in PERMISSION_DEFINITIONS:
        if permission["key"] not in permissions:
            can_view, can_edit = defaults.get(permission["key"], (0, 0))
            permissions[permission["key"]] = {"can_view": can_view, "can_edit": can_edit}
    return permissions


def current_user_with_permissions(handler):
    user = current_user(handler)
    if user:
        user["permissions"] = role_permissions(user.get("role"))
    return user


def can_view_permission(user, permission_key):
    if not user:
        return False
    if user.get("role") == "Admin":
        return True
    return bool(role_permissions(user.get("role")).get(permission_key, {}).get("can_view"))


def can_edit_permission(user, permission_key):
    if not user:
        return False
    if user.get("role") == "Admin":
        return True
    return bool(role_permissions(user.get("role")).get(permission_key, {}).get("can_edit"))


def require_admin(handler):
    user = current_user(handler)
    return user if user and user.get("role") == "Admin" else None


def require_editor(handler):
    user = current_user(handler)
    return user if user and user.get("role") in ("Admin", "User") else None


def can_use_field_po(user):
    return bool(user and user.get("role") in ("Admin", "User", "Field PO"))


def can_auto_issue_po(user):
    return bool(user and (user.get("role") == "Admin" or int(user.get("po_auto_issue") or 0) == 1))


def is_texas_read_only(user):
    return bool(user and user.get("role") == "TX/Read Only")


def is_field_po_only(user):
    return bool(user and user.get("role") == "Field PO")


def clean_role(role):
    role = str(role or "User").strip()
    return role if role in ROLE_NAMES else "User"


def clean_order_type(order_type):
    order_type = str(order_type or "Change Order").strip()
    return order_type if order_type in ("Change Order", "Child Project") else "Change Order"


def duplicate_job_order_message(con, project_id, job_number, exclude_change_order_id=None):
    job_number = str(job_number or "").strip()
    if not job_number:
        return None
    subproject = con.execute(
        """
        SELECT job_number, code, name
        FROM subprojects
        WHERE project_id = ?
          AND LOWER(TRIM(COALESCE(job_number, ''))) = LOWER(TRIM(?))
        LIMIT 1
        """,
        (project_id, job_number),
    ).fetchone()
    if subproject:
        label = " ".join(str(part or "").strip() for part in (subproject["job_number"], subproject["code"], subproject["name"]) if str(part or "").strip())
        return f"Job / Order # {job_number} is already used by subproject {label}."
    params = [project_id, job_number]
    exclude_clause = ""
    if exclude_change_order_id:
        exclude_clause = " AND id != ?"
        params.append(exclude_change_order_id)
    change_order = con.execute(
        f"""
        SELECT order_type, co_number, job_number, title
        FROM change_orders
        WHERE project_id = ?
          AND LOWER(TRIM(COALESCE(job_number, ''))) = LOWER(TRIM(?))
          {exclude_clause}
        LIMIT 1
        """,
        params,
    ).fetchone()
    if change_order:
        label = " / ".join(str(part or "").strip() for part in (change_order["co_number"], change_order["job_number"]) if str(part or "").strip())
        title = str(change_order["title"] or "").strip()
        if title:
            label = f"{label} - {title}" if label else title
        return f"Job / Order # {job_number} is already used by {clean_order_type(change_order['order_type']).lower()} {label}."
    return None


def parse_fieldwise_audit_export(path):
    if openpyxl is None:
        raise RuntimeError("Excel import needs openpyxl, but it is not available.")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["LineItems"] if "LineItems" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(v or "").strip() for v in next(rows_iter)]
    except StopIteration:
        return {}
    header_map = {h.lower(): i for i, h in enumerate(headers)}
    required = ["customer name", "ticket date", "ticket number", "order number", "status", "sub total"]
    missing = [h for h in required if h not in header_map]
    if missing:
        raise RuntimeError(f"Missing required Field Wise export columns: {', '.join(missing)}")

    def cell(row, name):
        idx = header_map.get(name.lower())
        if idx is None or idx >= len(row):
            return ""
        return row[idx]

    tickets = {}
    line_count = 0
    for row in rows_iter:
        if not row or not any(row):
            continue
        ticket_number = str(cell(row, "ticket number") or "").strip()
        if not ticket_number:
            continue
        order_number = str(cell(row, "order number") or "").strip()
        key = f"{ticket_number}|{order_number}"
        ticket = tickets.setdefault(
            key,
            {
                "ticket_number": ticket_number,
                "order_number": order_number,
                "customer": str(cell(row, "customer name") or "").strip(),
                "ticket_date": date_text(cell(row, "ticket date")),
                "status": str(cell(row, "status") or "").strip(),
                "line_count": 0,
                "export_total": 0,
            },
        )
        ticket["line_count"] += 1
        ticket["export_total"] += money(cell(row, "sub total"))
        if not ticket["customer"]:
            ticket["customer"] = str(cell(row, "customer name") or "").strip()
        if not ticket["ticket_date"]:
            ticket["ticket_date"] = date_text(cell(row, "ticket date"))
        if not ticket["status"]:
            ticket["status"] = str(cell(row, "status") or "").strip()
        line_count += 1
    return {"line_count": line_count, "tickets": tickets}


def fieldwise_audit_result(path):
    export = parse_fieldwise_audit_export(path)
    export_tickets = export.get("tickets", {})
    with db() as con:
        omission_rows = con.execute(
            """
            SELECT *
            FROM fieldwise_audit_omissions
            ORDER BY created_at DESC, ticket_number
            """
        ).fetchall()
        tracked_rows = con.execute(
            """
            SELECT
              'Subproject' AS item_type,
              sp.job_number,
              p.name AS project_name,
              p.customer,
              sp.code AS reference_code
            FROM subprojects sp
            JOIN projects p ON p.id = sp.project_id
            WHERE COALESCE(p.status, 'Active') <> 'Archived'
              AND TRIM(COALESCE(sp.job_number, '')) <> ''
            UNION ALL
            SELECT
              COALESCE(co.order_type, 'Change Order') AS item_type,
              co.job_number,
              p.name AS project_name,
              p.customer,
              co.co_number AS reference_code
            FROM change_orders co
            JOIN projects p ON p.id = co.project_id
            WHERE COALESCE(p.status, 'Active') <> 'Archived'
              AND TRIM(COALESCE(co.job_number, '')) <> ''
            """
        ).fetchall()
        tracked_jobs = {str(r["job_number"] or "").strip(): dict(r) for r in tracked_rows}
        imported_rows = con.execute(
            """
            SELECT
              cr.ticket_or_invoice,
              COALESCE(NULLIF(TRIM(co.job_number), ''), NULLIF(TRIM(sp.job_number), '')) AS job_number,
              COUNT(*) AS line_count,
              COALESCE(SUM(cr.sales_amount), 0) AS imported_sales_total,
              COALESCE(SUM(cr.amount), 0) AS imported_raw_total,
              MAX(cr.record_date) AS last_record_date
            FROM cost_records cr
            LEFT JOIN subprojects sp ON sp.id = cr.subproject_id
            LEFT JOIN change_orders co ON co.id = cr.change_order_id
            WHERE cr.source IN ('Field Wise', 'Field Wise PDF')
              AND TRIM(COALESCE(cr.ticket_or_invoice, '')) <> ''
            GROUP BY cr.ticket_or_invoice, COALESCE(NULLIF(TRIM(co.job_number), ''), NULLIF(TRIM(sp.job_number), ''))
            """
        ).fetchall()

    omissions_by_key = {f"{str(r['ticket_number'] or '').strip()}|{str(r['order_number'] or '').strip()}": dict(r) for r in omission_rows}
    imported_by_key = {}
    for row in imported_rows:
        ticket_number = str(row["ticket_or_invoice"] or "").strip()
        job_number = str(row["job_number"] or "").strip()
        if not ticket_number or not job_number:
            continue
        imported_by_key[f"{ticket_number}|{job_number}"] = dict(row)

    missing = []
    matched = []
    untracked = []
    no_order = []
    omitted = []
    for key, ticket in export_tickets.items():
        order_number = str(ticket["order_number"] or "").strip()
        if not order_number:
            no_order.append(ticket)
            continue
        job = tracked_jobs.get(order_number)
        if not job:
            untracked.append(ticket)
            continue
        imported = imported_by_key.get(key)
        row = {**ticket, **job, "imported_total": money(imported["imported_sales_total"]) if imported else 0, "imported_line_count": imported["line_count"] if imported else 0}
        if not imported:
            omission = omissions_by_key.get(key)
            if omission:
                omitted.append({**row, "omission_id": omission["id"], "omission_reason": omission["reason"] or "", "omitted_by_username": omission["omitted_by_username"] or "", "omitted_at": omission["created_at"] or ""})
            else:
                missing.append(row)
        else:
            matched.append(row)

    extra = []
    for key, imported in imported_by_key.items():
        job_number = str(imported["job_number"] or "").strip()
        if job_number not in tracked_jobs:
            continue
        if key in export_tickets:
            continue
        job = tracked_jobs[job_number]
        ticket_number = str(imported["ticket_or_invoice"] or "").strip()
        extra.append({
            "ticket_number": ticket_number,
            "order_number": job_number,
            "customer": job.get("customer") or "",
            "ticket_date": imported.get("last_record_date") or "",
            "status": "",
            "line_count": 0,
            "export_total": 0,
            "imported_total": money(imported["imported_sales_total"]),
            "imported_line_count": imported["line_count"],
            **job,
        })

    sort_key = lambda r: (str(r.get("project_name") or ""), str(r.get("order_number") or ""), str(r.get("ticket_number") or ""))
    missing.sort(key=sort_key)
    matched.sort(key=sort_key)
    extra.sort(key=sort_key)
    omitted.sort(key=sort_key)
    untracked.sort(key=lambda r: (str(r.get("order_number") or ""), str(r.get("ticket_number") or "")))
    no_order.sort(key=lambda r: str(r.get("ticket_number") or ""))
    return {
        "summary": {
            "export_line_count": export.get("line_count", 0),
            "export_ticket_count": len(export_tickets),
            "tracked_job_count": len(tracked_jobs),
            "matched_count": len(matched),
            "missing_count": len(missing),
            "omitted_count": len(omitted),
            "extra_imported_count": len(extra),
            "untracked_count": len(untracked),
            "no_order_count": len(no_order),
        },
        "missing": missing,
        "omitted": omitted,
        "extra_imported": extra,
        "untracked": untracked,
        "no_order": no_order,
    }


def next_po_number(con):
    year = datetime.now().year
    prefix = f"PO-{year}-"
    row = con.execute(
        "SELECT po_number FROM purchase_orders WHERE po_number LIKE ? ORDER BY po_number DESC LIMIT 1",
        (f"{prefix}%",),
    ).fetchone()
    seq = 1
    if row:
        try:
            seq = int(str(row["po_number"]).rsplit("-", 1)[-1]) + 1
        except Exception:
            seq = 1
    return f"{prefix}{seq:04d}"


def job_reference_for_po(con, job_key):
    parts = str(job_key or "").split(":", 1)
    if len(parts) != 2:
        return None
    kind, raw_id = parts
    try:
        item_id = int(raw_id)
    except Exception:
        return None
    if kind == "subproject":
        row = con.execute(
            """
            SELECT
              sp.id AS subproject_id,
              NULL AS change_order_id,
              sp.project_id,
              sp.job_number,
              'Subproject' AS item_type,
              p.customer,
              p.name AS project_name,
              p.project_code,
              sp.code AS reference_code,
              sp.name AS description
            FROM subprojects sp
            JOIN projects p ON p.id = sp.project_id
            WHERE sp.id = ?
              AND COALESCE(p.status, 'Active') <> 'Archived'
            """,
            (item_id,),
        ).fetchone()
    elif kind == "change_order":
        row = con.execute(
            """
            SELECT
              co.subproject_id,
              co.id AS change_order_id,
              co.project_id,
              co.job_number,
              COALESCE(co.order_type, 'Change Order') AS item_type,
              p.customer,
              p.name AS project_name,
              p.project_code,
              CASE
                WHEN sp.code IS NOT NULL THEN co.co_number || ' / ' || sp.code
                ELSE co.co_number
              END AS reference_code,
              co.title AS description
            FROM change_orders co
            JOIN projects p ON p.id = co.project_id
            LEFT JOIN subprojects sp ON sp.id = co.subproject_id
            WHERE co.id = ?
              AND COALESCE(p.status, 'Active') <> 'Archived'
            """,
            (item_id,),
        ).fetchone()
    elif kind == "cog":
        row = con.execute(
            """
            SELECT
              NULL AS subproject_id,
              NULL AS change_order_id,
              id AS cog_category_id,
              NULL AS project_id,
              name AS job_number,
              'COG' AS item_type,
              '' AS customer,
              'COG Category' AS project_name,
              code AS project_code,
              code AS reference_code,
              description
            FROM cog_categories
            WHERE id = ?
              AND COALESCE(active, 1) = 1
            """,
            (item_id,),
        ).fetchone()
    else:
        return None
    if not row or not str(row["job_number"] or "").strip():
        return None
    label_parts = [
        row["job_number"],
        row["item_type"],
        row["customer"],
        row["project_name"],
        row["reference_code"],
    ]
    result = dict(row)
    result.setdefault("cog_category_id", None)
    result["job_label"] = " - ".join(str(part or "").strip() for part in label_parts if str(part or "").strip())
    return result


def seed_bid_tracker_from_workbook():
    if openpyxl is None or not BID_TRACKER_SOURCE.exists():
        return
    existing = one("SELECT COUNT(*) count FROM bid_requests")
    if existing and existing["count"]:
        return
    wb = openpyxl.load_workbook(BID_TRACKER_SOURCE, data_only=True, read_only=True)
    now = datetime.now().isoformat(timespec="seconds")
    with db() as con:
        if "Bid Tracker" in wb.sheetnames:
            ws = wb["Bid Tracker"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                rfq_no = str(row[0] or "").strip()
                estimated_cost = money(row[8] if len(row) > 8 else 0)
                target_margin = money(row[9] if len(row) > 9 else 0)
                bid_price = bid_price_value(estimated_cost, target_margin, money(row[10] if len(row) > 10 else 0))
                probability = money(row[11] if len(row) > 11 else 0)
                outcome = str(row[14] if len(row) > 14 and row[14] else "Pending")
                weighted = weighted_bid_value(outcome, bid_price, probability)
                con.execute(
                    """
                    INSERT OR IGNORE INTO bid_requests (
                      rfq_no, date_received, customer, project_name, estimator, stage, bid_due_date,
                      go_no_go, estimated_cost, target_margin, bid_price, probability, weighted_value,
                      submission_status, outcome, notes, created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        rfq_no,
                        date_text(row[1] if len(row) > 1 else ""),
                        str(row[2] or "") if len(row) > 2 else "",
                        str(row[3] or "") if len(row) > 3 else "",
                        str(row[4] or "") if len(row) > 4 and row[4] else "",
                        str(row[5] or "") if len(row) > 5 else "",
                        date_text(row[6] if len(row) > 6 else ""),
                        str(row[7] or "") if len(row) > 7 else "",
                        estimated_cost,
                        target_margin,
                        bid_price,
                        probability,
                        weighted,
                        str(row[13] or "") if len(row) > 13 else "",
                        outcome,
                        str(row[15] or "") if len(row) > 15 else "",
                        now,
                        now,
                    ),
                )
        if "Risk_Register" in wb.sheetnames:
            ws = wb["Risk_Register"]
            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row or not row[0]:
                    continue
                con.execute(
                    """
                    INSERT INTO bid_risks (risk_id, rfq_no, risk_description, probability, impact, risk_rating, pricing_action, mitigation, owner, status, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        str(row[0] or ""),
                        str(row[1] or ""),
                        str(row[2] or ""),
                        int(money(row[3])),
                        int(money(row[4])),
                        int(money(row[5])),
                        str(row[6] or ""),
                        str(row[7] or ""),
                        str(row[8] or ""),
                        str(row[9] or ""),
                        now,
                    ),
                )
        if "Win_Loss_Analysis" in wb.sheetnames:
            ws = wb["Win_Loss_Analysis"]
            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row or not row[0]:
                    continue
                con.execute(
                    """
                    INSERT INTO bid_win_loss (rfq_no, customer, outcome, bid_value, known_competitor, primary_reason, pricing_position, lesson_learned, next_action, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        str(row[0] or ""),
                        str(row[1] or ""),
                        str(row[2] or ""),
                        money(row[3]),
                        str(row[4] or ""),
                        str(row[5] or ""),
                        str(row[6] or ""),
                        str(row[7] or ""),
                        str(row[8] or ""),
                        now,
                    ),
                )
    wb.close()


def money(value):
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def parse_money_text(value):
    if value is None:
        return 0.0
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    return money(cleaned)


def first_regex(patterns, text, flags=re.IGNORECASE):
    for pattern in patterns:
        match = re.search(pattern, text, flags)
        if match:
            return match.group(1).strip()
    return ""


def pdf_text_is_garbled(text):
    if not text:
        return False
    cid_count = text.count("(cid:")
    if cid_count >= 10:
        return True
    compact = re.sub(r"\s+", "", text)
    if not compact:
        return False
    printable_letters = sum(1 for ch in compact if ch.isalpha())
    return cid_count >= 3 and printable_letters / max(len(compact), 1) < 0.15


def extract_pdf_text(path):
    pages = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            pages.append(page.extract_text() or "")
    text = "\n".join(pages).strip()
    if text and not pdf_text_is_garbled(text):
        return text
    if pdfium is None or RapidOCR is None:
        return "" if pdf_text_is_garbled(text) else text

    global OCR_ENGINE
    if OCR_ENGINE is None:
        OCR_ENGINE = RapidOCR()

    ocr_pages = []
    pdf = pdfium.PdfDocument(str(path))
    for i in range(len(pdf)):
        page = pdf[i]
        image = page.render(scale=2).to_pil()
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                tmp_path = tmp.name
            image.save(tmp_path)
            result, _ = OCR_ENGINE(tmp_path)
            if result:
                ocr_pages.append("\n".join(line[1] for line in result))
        finally:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)
    return "\n".join(ocr_pages).strip()


def extract_pdf_ocr_pages(path):
    if pdfium is None or RapidOCR is None:
        return []

    global OCR_ENGINE
    if OCR_ENGINE is None:
        OCR_ENGINE = RapidOCR()

    pages = []
    pdf = pdfium.PdfDocument(str(path))
    for i in range(len(pdf)):
        page = pdf[i]
        image = page.render(scale=2).to_pil()
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                tmp_path = tmp.name
            image.save(tmp_path)
            result, _ = OCR_ENGINE(tmp_path)
            page_lines = []
            for box, text, score in result or []:
                xs = [p[0] for p in box]
                ys = [p[1] for p in box]
                page_lines.append({
                    "text": str(text).strip(),
                    "x1": min(xs),
                    "y1": min(ys),
                    "x2": max(xs),
                    "y2": max(ys),
                })
            pages.append(sorted(page_lines, key=lambda r: (r["y1"], r["x1"])))
        finally:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)
    return pages


def parse_dsg_online_invoice_ocr(path):
    item_lines = []
    def pick_product_line(info):
        meaningful = [txt for txt in info if not re.search(r"^(MFG|DSG|UPC)", txt, flags=re.IGNORECASE)]
        if not meaningful:
            return info[0] if info else ""
        if len(meaningful) > 1 and len(meaningful[0]) < 35:
            return meaningful[1]
        return meaningful[0]

    for page_lines in extract_pdf_ocr_pages(path):
        qty_labels = [
            line for line in page_lines
            if line["x1"] > 900 and re.sub(r"\s+", "", line["text"].lower()) == "qtyinvoiced"
        ]
        qty_labels.sort(key=lambda r: r["y1"])
        used_subtotal_ys = []
        for idx, label in enumerate(qty_labels):
            top = max(0, label["y1"] - 45)
            bottom = qty_labels[idx + 1]["y1"] - 20 if idx + 1 < len(qty_labels) else label["y1"] + 180
            block = [line for line in page_lines if top <= line["y1"] < bottom]
            left = [line for line in block if line["x1"] < 850]
            right = [line for line in block if line["x1"] > 900]

            qty = 0
            for line in sorted(right, key=lambda r: r["y1"]):
                txt = line["text"].replace(",", "")
                if line["y1"] > label["y1"] and re.fullmatch(r"[0-9]+(?:\.[0-9]+)?", txt):
                    qty = money(txt)
                    break

            amount = 0
            subtotal_seen = False
            subtotal_y = None
            for line in sorted(right, key=lambda r: r["y1"]):
                if "subtotal" in line["text"].lower():
                    subtotal_seen = True
                    subtotal_y = line["y1"]
                    continue
                if subtotal_seen and "$" in line["text"]:
                    amount = parse_money_text(line["text"])
                    break

            rate = 0
            info = []
            for line in sorted(left, key=lambda r: (r["y1"], r["x1"])):
                txt = line["text"].strip()
                if not txt:
                    continue
                if txt.lower().startswith(("invoice", "billing", "delivery", "order summary")):
                    continue
                if "$" in txt and "/" in txt and not rate:
                    rate = parse_money_text(txt)
                    continue
                info.append(txt)

            if not info or not amount:
                continue
            if subtotal_y is not None:
                used_subtotal_ys.append(subtotal_y)
            if not qty and rate:
                qty = round(amount / rate, 4)
            product_line = pick_product_line(info)
            product_code = product_line.split()[0] if product_line.split() else info[0]
            item_lines.append({
                "product_code": product_code,
                "description": " ".join(info),
                "qty": qty,
                "unit_price": rate if rate else (amount / qty if qty else 0),
                "amount": amount,
            })
        subtotal_lines = [
            line for line in page_lines
            if line["x1"] > 900 and "subtotal" in line["text"].lower()
        ]
        for subtotal in sorted(subtotal_lines, key=lambda r: r["y1"]):
            if any(abs(subtotal["y1"] - used_y) < 25 for used_y in used_subtotal_ys):
                continue
            top = max(0, subtotal["y1"] - 160)
            bottom = subtotal["y1"] + 70
            block = [line for line in page_lines if top <= line["y1"] < bottom]
            left = [line for line in block if line["x1"] < 850]
            right = [line for line in block if line["x1"] > 900]

            amount = 0
            for line in sorted(right, key=lambda r: r["y1"]):
                if line["y1"] > subtotal["y1"] and "$" in line["text"]:
                    amount = parse_money_text(line["text"])
                    break
            if not amount:
                continue

            qty = 0
            for line in sorted(right, key=lambda r: r["y1"], reverse=True):
                txt = line["text"].replace(",", "")
                if line["y1"] < subtotal["y1"] and re.fullmatch(r"[0-9]+(?:\.[0-9]+)?", txt):
                    qty = money(txt)
                    break

            rate = 0
            info = []
            for line in sorted(left, key=lambda r: (r["y1"], r["x1"])):
                txt = line["text"].strip()
                if not txt:
                    continue
                if txt.lower().startswith(("invoice", "billing", "delivery", "order summary")):
                    continue
                if "$" in txt and "/" in txt and not rate:
                    rate = parse_money_text(txt)
                    continue
                info.append(txt)
            if not info:
                continue
            if not qty and rate:
                qty = round(amount / rate, 4)
            product_line = pick_product_line(info)
            product_code = product_line.split()[0] if product_line.split() else info[0]
            item_lines.append({
                "product_code": product_code,
                "description": " ".join(info),
                "qty": qty,
                "unit_price": rate if rate else (amount / qty if qty else 0),
                "amount": amount,
            })
    return item_lines


def labor_category(item):
    parts = [p.strip() for p in str(item or "").split(" - ") if p.strip()]
    if len(parts) >= 2:
        category = " ".join(parts[1:])
    else:
        category = str(item or "").strip()
    category = category.replace(" Reg", " ST").replace(" - ", " ")
    return re.sub(r"\s+", " ", category).strip()


def equipment_category(item):
    category = re.sub(r"\s+", " ", str(item or "").strip())
    aliases = {
        "Service Trucks": "Service Truck",
        "Work Truck": "Service Truck",
        "Work Truck 01": "Service Truck",
        "Bobcat": "Skid Steer/Bobcat",
        "Skid Steer": "Skid Steer/Bobcat",
        "Trailer (Daily Rate)": "Trailer",
        "Transport Trailer": "Transport Truck & Trailer",
        "Compactor": "Trench Compactor",
    }
    return aliases.get(category, category)


def is_equipment_item(item):
    category = equipment_category(item)
    known = {
        "Service Truck",
        "Trencher Summer",
        "Trencher Winter",
        "Transport Truck & Trailer",
        "Trench Compactor",
        "Reel Trailer",
        "Mini Excavator",
        "Skid Steer/Bobcat",
        "Trailer",
    }
    return category in known


def raw_rate_for(cost_type, item, project_id=None):
    if cost_type == "Labor":
        category = labor_category(item)
    elif cost_type == "Equipment":
        category = equipment_category(item)
    else:
        return {"category": "", "raw_rate": 0.0, "source": "No internal rate needed"}
    project_rate_set_id = None
    if project_id:
        project = one("SELECT rate_set_id FROM projects WHERE id = ?", (project_id,))
        project_rate_set_id = project["rate_set_id"] if project else None
    rate = one(
        """
        SELECT raw_rate FROM internal_rates
        WHERE category_type = ?
          AND category = ?
          AND active = 1
          AND (rate_set_id = ? OR (? IS NULL AND rate_set_id IS NULL))
        """,
        (cost_type, category, project_rate_set_id, project_rate_set_id),
    )
    if rate:
        return {"category": category, "raw_rate": money(rate["raw_rate"]), "source": "Project rate set"}
    return {"category": category, "raw_rate": 0.0, "source": "Missing project rate"}


def apply_internal_rate(category_type, category, raw_rate, rate_set_id=None):
    if category_type not in ("Labor", "Equipment"):
        return 0
    updated = 0
    candidates = rows(
        """
        SELECT cost_records.id, cost_records.item, cost_records.qty
        FROM cost_records
        JOIN projects p ON p.id = cost_records.project_id
        WHERE cost_records.cost_type = ?
          AND source IN ('Field Wise', 'Field Wise PDF')
          AND p.rate_set_id = ?
        """,
        (category_type, rate_set_id),
    )
    with db() as con:
        for record in candidates:
            record_category = labor_category(record["item"]) if category_type == "Labor" else equipment_category(record["item"])
            if record_category != category:
                continue
            con.execute(
                "UPDATE cost_records SET raw_rate = ?, rate = ?, amount = ?, raw_cost_source = 'Project rate set' WHERE id = ?",
                (money(raw_rate), money(raw_rate), money(record["qty"]) * money(raw_rate), record["id"]),
            )
            updated += 1
    return updated


def json_response(handler, payload, status=200):
    body = json.dumps(payload, default=str).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json")
    handler.send_header("Cache-Control", "no-store")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def text_response(handler, body, content_type="text/html; charset=utf-8", status=200):
    data = body.encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", content_type)
    handler.send_header("Cache-Control", "no-store")
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def file_response(handler, path, content_type="application/octet-stream", status=200):
    data = path.read_bytes()
    handler.send_response(status)
    handler.send_header("Content-Type", content_type)
    handler.send_header("Cache-Control", "no-store")
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def redirect_response(handler, location):
    handler.send_response(302)
    handler.send_header("Location", location)
    handler.end_headers()


def login_success_response(handler, token):
    body = b'{"ok": true}'
    handler.send_response(200)
    handler.send_header("Content-Type", "application/json")
    handler.send_header("Set-Cookie", f"pm_session={token}; Path=/; HttpOnly; SameSite=Lax")
    handler.send_header("Cache-Control", "no-store")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def logout_response(handler):
    body = b'{"ok": true}'
    handler.send_response(200)
    handler.send_header("Content-Type", "application/json")
    handler.send_header("Set-Cookie", "pm_session=; Path=/; HttpOnly; SameSite=Lax; Max-Age=0")
    handler.send_header("Cache-Control", "no-store")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def bytes_response(handler, data, content_type="application/octet-stream", status=200):
    handler.send_response(status)
    handler.send_header("Content-Type", content_type)
    handler.send_header("Cache-Control", "no-store")
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def download_response(handler, data, filename, content_type="application/octet-stream", status=200):
    handler.send_response(status)
    handler.send_header("Content-Type", content_type)
    handler.send_header("Cache-Control", "no-store")
    handler.send_header("Content-Disposition", f'attachment; filename="{filename}"')
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def sqlite_backup_bytes():
    backup_path = Path(tempfile.gettempdir()) / f"pm-tracker-backup-{secrets.token_hex(8)}.sqlite3"
    try:
        with sqlite3.connect(DB_PATH) as source, sqlite3.connect(backup_path) as target:
            source.backup(target)
        return backup_path.read_bytes()
    finally:
        try:
            backup_path.unlink()
        except Exception:
            pass


def log_activity(actor=None, action="", entity_type="", entity_label="", details=None):
    try:
        actor = actor or {}
        detail_text = details if isinstance(details, str) else json.dumps(details or {}, default=str)
        execute(
            """
            INSERT INTO activity_log (actor_user_id, actor_username, action, entity_type, entity_label, details, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                actor.get("id"),
                actor.get("username") or actor.get("display_name") or "",
                str(action or "").strip(),
                str(entity_type or "").strip(),
                str(entity_label or "").strip(),
                detail_text,
                datetime.now().isoformat(timespec="seconds"),
            ),
        )
    except Exception:
        traceback.print_exc()


def upload_pdf_path(file_name):
    safe_name = Path(unquote(file_name)).name
    path = (UPLOAD_DIR / safe_name).resolve()
    upload_root = UPLOAD_DIR.resolve()
    if upload_root not in path.parents or not path.exists() or path.suffix.lower() != ".pdf":
        return None
    return path


def upload_attachment_path(file_name):
    safe_name = Path(unquote(file_name)).name
    path = (UPLOAD_DIR / safe_name).resolve()
    upload_root = UPLOAD_DIR.resolve()
    if upload_root not in path.parents or not path.exists():
        return None
    return path


def pdf_viewer_html(file_name):
    path = upload_pdf_path(file_name)
    if not path:
        return None
    if pdfium is None:
        direct = f"/uploads/{path.name}"
        return f"""<!doctype html><html><head><title>{path.name}</title></head><body>
        <p>PDF preview is unavailable in this runtime.</p>
        <p><a href="{direct}" target="_blank" rel="noopener">Open PDF directly</a></p>
        </body></html>"""
    pdf = pdfium.PdfDocument(str(path))
    encoded = quote(path.name)
    pages = "\n".join(
        f'<img class="pdf-page" src="/pdf-page/{encoded}/{idx}.png" alt="Page {idx + 1}">'
        for idx in range(len(pdf))
    )
    direct = f"/uploads/{encoded}"
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{path.name}</title>
  <style>
    body {{ margin: 0; font-family: "Segoe UI", Arial, sans-serif; background: #eef2f6; color: #17202a; }}
    header {{ position: sticky; top: 0; z-index: 2; display: flex; justify-content: space-between; align-items: center; gap: 12px; padding: 12px 18px; background: #152332; color: white; border-bottom: 4px solid #ffc20e; }}
    h1 {{ margin: 0; font-size: 16px; }}
    a {{ color: #ffffff; font-weight: 700; }}
    main {{ max-width: 980px; margin: 18px auto; padding: 0 16px 30px; }}
    .pdf-page {{ display: block; width: 100%; height: auto; background: white; margin: 0 0 16px; border: 1px solid #d8dee5; box-shadow: 0 8px 24px rgba(0,0,0,.12); }}
  </style>
</head>
<body>
  <header><h1>{path.name}</h1><a href="{direct}" target="_blank" rel="noopener">Open Original PDF</a></header>
  <main>{pages}</main>
</body>
</html>"""


def render_pdf_page_png(file_name, page_index):
    path = upload_pdf_path(file_name)
    if not path or pdfium is None:
        return None
    pdf = pdfium.PdfDocument(str(path))
    idx = int(page_index)
    if idx < 0 or idx >= len(pdf):
        return None
    page = pdf[idx]
    bitmap = page.render(scale=1.8)
    image = bitmap.to_pil()
    out = BytesIO()
    image.save(out, format="PNG")
    return out.getvalue()


def parse_json(handler):
    length = int(handler.headers.get("Content-Length", "0"))
    if not length:
        return {}
    return json.loads(handler.rfile.read(length).decode("utf-8"))


def project_summary(project_id):
    project = one("SELECT * FROM projects WHERE id = ?", (project_id,))
    if not project:
        return None

    totals = one(
        """
        SELECT
          COALESCE(SUM(
            CASE
              WHEN cr.change_order_id IS NOT NULL AND COALESCE(co.pricing_type, 'Fixed') = 'T&M' THEN
                CASE WHEN cr.source IN ('Field Wise', 'Field Wise PDF') THEN cr.amount ELSE 0 END
              WHEN cr.change_order_id IS NULL AND COALESCE(sp.pricing_type, 'Fixed') = 'T&M' THEN
                CASE WHEN cr.source IN ('Field Wise', 'Field Wise PDF') THEN cr.amount ELSE 0 END
              ELSE cr.amount
            END
          ), 0) actual_cost,
          COALESCE(SUM(CASE WHEN cr.subproject_id IS NULL OR cr.cost_type IS NULL OR cr.cost_type = 'Uncoded' OR cr.raw_cost_source IN ('Missing project rate', 'Missing internal rate') THEN cr.amount ELSE 0 END), 0) uncoded_cost,
          COUNT(*) record_count,
          SUM(CASE WHEN cr.subproject_id IS NULL OR cr.cost_type IS NULL OR cr.cost_type = 'Uncoded' OR cr.raw_cost_source IN ('Missing project rate', 'Missing internal rate') THEN 1 ELSE 0 END) uncoded_count
        FROM cost_records cr
        LEFT JOIN subprojects sp ON sp.id = cr.subproject_id
        LEFT JOIN change_orders co ON co.id = cr.change_order_id
        WHERE cr.project_id = ?
        """,
        (project_id,),
    )
    cos = one(
        """
        SELECT
          COALESCE(SUM(
            CASE WHEN status IN ('Approved', 'Billed') THEN
              CASE WHEN pricing_type = 'T&M' THEN
                COALESCE((SELECT SUM(cr.sales_amount) FROM cost_records cr WHERE cr.change_order_id = change_orders.id AND cr.source IN ('Field Wise', 'Field Wise PDF')), 0)
              ELSE approved_value END
            ELSE 0 END
          ), 0) approved_co_value,
          COALESCE(SUM(
            CASE WHEN status NOT IN ('Approved', 'Billed') THEN
              CASE WHEN pricing_type = 'T&M' THEN
                COALESCE((SELECT SUM(cr.sales_amount) FROM cost_records cr WHERE cr.change_order_id = change_orders.id AND cr.source IN ('Field Wise', 'Field Wise PDF')), 0)
              ELSE quoted_value END
            ELSE 0 END
          ), 0) pending_co_value
        FROM change_orders
        WHERE project_id = ?
        """,
        (project_id,),
    )
    subproject_contract = one(
        """
        SELECT COALESCE(SUM(
          CASE WHEN COALESCE(sp.pricing_type, 'Fixed') = 'T&M' THEN COALESCE(fw.fieldwise_sales, 0)
          ELSE sp.contract_value END
        ), 0) base_contract_value
        FROM subprojects sp
        LEFT JOIN (
          SELECT subproject_id, COALESCE(SUM(sales_amount), 0) fieldwise_sales
          FROM cost_records
          WHERE project_id = ?
            AND change_order_id IS NULL
            AND source IN ('Field Wise', 'Field Wise PDF')
          GROUP BY subproject_id
        ) fw ON fw.subproject_id = sp.id
        WHERE sp.project_id = ?
        """,
        (project_id, project_id),
    )
    base_contract_value = money(subproject_contract["base_contract_value"])
    contract = base_contract_value + money(cos["approved_co_value"])
    actual = money(totals["actual_cost"])
    profit = contract - actual
    margin = profit / contract if contract else 0

    subprojects = rows(
        """
        SELECT
          sp.id,
          sp.job_number,
          sp.code,
          sp.name,
          COALESCE(sp.pricing_type, 'Fixed') pricing_type,
          sp.contract_value,
          sp.budget_labor_hours,
          sp.budget_labor + sp.budget_material + COALESCE(sp.budget_equipment, 0) AS budget_total,
          COALESCE(SUM(
            CASE
              WHEN cr.change_order_id IS NULL AND COALESCE(sp.pricing_type, 'Fixed') = 'T&M' AND cr.source IN ('Field Wise', 'Field Wise PDF') THEN cr.amount
              WHEN cr.change_order_id IS NULL AND COALESCE(sp.pricing_type, 'Fixed') <> 'T&M' THEN cr.amount
              ELSE 0
            END
          ), 0) actual_cost,
          COALESCE(SUM(CASE WHEN cr.cost_type = 'Labor' AND cr.change_order_id IS NULL THEN cr.qty ELSE 0 END), 0) labor_hours_used,
          COALESCE(SUM(CASE WHEN cr.change_order_id IS NULL AND cr.source IN ('Field Wise', 'Field Wise PDF') THEN cr.sales_amount ELSE 0 END), 0) fieldwise_sales,
          CASE WHEN COALESCE(sp.pricing_type, 'Fixed') = 'T&M' THEN
            COALESCE(SUM(CASE WHEN cr.change_order_id IS NULL AND cr.source IN ('Field Wise', 'Field Wise PDF') THEN cr.sales_amount ELSE 0 END), 0)
          ELSE sp.contract_value END sales_value,
          COUNT(CASE WHEN cr.change_order_id IS NULL THEN 1 END) record_count
        FROM subprojects sp
        LEFT JOIN cost_records cr ON cr.subproject_id = sp.id
        WHERE sp.project_id = ?
        GROUP BY sp.id
        ORDER BY sp.job_number, sp.code
        """,
        (project_id,),
    )
    for subproject in subprojects:
        sales_value = money(subproject.get("sales_value"))
        actual_cost = money(subproject.get("actual_cost"))
        subproject["profit"] = sales_value - actual_cost
        subproject["margin"] = subproject["profit"] / sales_value if sales_value else 0

    change_orders = rows(
        """
        SELECT
          co.id,
          co.co_number,
          co.job_number,
          COALESCE(co.order_type, 'Change Order') order_type,
          co.title,
          co.status,
          COALESCE(co.pricing_type, 'Fixed') pricing_type,
          co.approved_value,
          co.quoted_value,
          co.subproject_id,
          sp.code AS subproject_code,
          COALESCE(SUM(cr.amount), 0) actual_cost,
          COALESCE(SUM(CASE WHEN cr.cost_type = 'Labor' THEN cr.qty ELSE 0 END), 0) labor_hours_used,
          COALESCE(SUM(CASE WHEN cr.source IN ('Field Wise', 'Field Wise PDF') THEN cr.sales_amount ELSE 0 END), 0) fieldwise_sales,
          CASE WHEN COALESCE(co.pricing_type, 'Fixed') = 'T&M' THEN
            COALESCE(SUM(CASE WHEN cr.source IN ('Field Wise', 'Field Wise PDF') THEN cr.sales_amount ELSE 0 END), 0)
          ELSE
            CASE WHEN co.status IN ('Approved', 'Billed') THEN co.approved_value ELSE co.quoted_value END
          END sales_value
        FROM change_orders co
        LEFT JOIN subprojects sp ON sp.id = co.subproject_id
        LEFT JOIN cost_records cr ON cr.change_order_id = co.id
        WHERE co.project_id = ?
        GROUP BY co.id
        ORDER BY co.co_number
        """,
        (project_id,),
    )

    by_type = rows(
        """
        SELECT
          COALESCE(cost_type, 'Uncoded') label,
          COALESCE(SUM(CASE WHEN cost_type = 'Field Ticket Material' THEN sales_amount ELSE amount END), 0) amount
        FROM cost_records
        WHERE project_id = ?
        GROUP BY COALESCE(cost_type, 'Uncoded')
        ORDER BY amount DESC
        """,
        (project_id,),
    )
    material_compare = one(
        """
        SELECT
          COALESCE(SUM(CASE WHEN cost_type = 'Field Ticket Material' THEN sales_amount ELSE 0 END), 0) field_ticket_material,
          COALESCE(SUM(CASE WHEN source = 'Vendor Invoice' AND cost_type = 'Material' THEN amount ELSE 0 END), 0) vendor_material
        FROM cost_records
        WHERE project_id = ?
        """,
        (project_id,),
    )
    billing = one(
        """
        SELECT
          COALESCE(SUM(CASE WHEN status != 'Void' THEN amount ELSE 0 END), 0) billed_amount,
          COALESCE(SUM(CASE WHEN status != 'Void' THEN paid_amount ELSE 0 END), 0) paid_amount,
          COALESCE(SUM(CASE WHEN status = 'Draft' THEN amount ELSE 0 END), 0) draft_amount,
          COALESCE(SUM(CASE WHEN status NOT IN ('Draft', 'Paid', 'Void') THEN amount - paid_amount ELSE 0 END), 0) open_amount,
          COUNT(CASE WHEN status != 'Void' THEN 1 END) invoice_count
        FROM customer_invoices
        WHERE project_id = ?
        """,
        (project_id,),
    )
    customer_invoices = rows(
        """
        SELECT ci.*, COALESCE(SUM(cia.amount), ci.amount, 0) AS allocated_amount
        FROM customer_invoices ci
        LEFT JOIN customer_invoice_allocations cia ON cia.customer_invoice_id = ci.id
        WHERE ci.project_id = ?
        GROUP BY ci.id
        ORDER BY invoice_date DESC, id DESC
        """,
        (project_id,),
    )
    customer_invoice_allocations = rows(
        """
        SELECT
          cia.*,
          sp.job_number,
          sp.code AS subproject_code,
          sp.name AS subproject_name,
          co.co_number,
          co.job_number AS co_job_number,
          co.title AS co_title,
          COALESCE(co.order_type, 'Change Order') AS order_type
        FROM customer_invoice_allocations cia
        LEFT JOIN subprojects sp ON sp.id = cia.subproject_id
        LEFT JOIN change_orders co ON co.id = cia.change_order_id
        WHERE cia.project_id = ?
        ORDER BY cia.id
        """,
        (project_id,),
    )
    invoices_by_id = {invoice["id"]: invoice for invoice in customer_invoices}
    for invoice in customer_invoices:
        invoice["allocations"] = []
    for allocation in customer_invoice_allocations:
        if allocation.get("change_order_id"):
            pieces = [allocation.get("co_number"), allocation.get("co_job_number")]
            label = " / ".join(str(x) for x in pieces if x) or allocation.get("order_type") or "Change Order"
            if allocation.get("co_title"):
                label = f"{label} - {allocation.get('co_title')}"
        else:
            pieces = [allocation.get("job_number"), allocation.get("subproject_code")]
            label = " ".join(str(x) for x in pieces if x) or "Unassigned"
            if allocation.get("subproject_name"):
                label = f"{label} - {allocation.get('subproject_name')}"
        allocation["target_label"] = label
        invoice = invoices_by_id.get(allocation.get("customer_invoice_id"))
        if invoice is not None:
            invoice["allocations"].append(allocation)
    billed_amount = money(billing["billed_amount"])
    paid_amount = money(billing["paid_amount"])
    open_amount = money(billing["open_amount"])
    remaining_to_bill = contract - billed_amount
    if remaining_to_bill < 0:
        remaining_to_bill = 0
    billing_stage = "Not billed"
    if contract and paid_amount >= contract:
        billing_stage = "Paid in full"
    elif contract and billed_amount >= contract:
        billing_stage = "Fully billed"
    elif billed_amount:
        billing_stage = "Partially billed"

    return {
        "project": project,
        "contract_value": contract,
        "base_contract_value": base_contract_value,
        "actual_cost": actual,
        "profit": profit,
        "margin": margin,
        "record_count": totals["record_count"] or 0,
        "uncoded_count": totals["uncoded_count"] or 0,
        "uncoded_cost": money(totals["uncoded_cost"]),
        "approved_co_value": money(cos["approved_co_value"]),
        "pending_co_value": money(cos["pending_co_value"]),
        "subprojects": subprojects,
        "change_orders": change_orders,
        "by_type": by_type,
        "material_compare": material_compare,
        "billing": {
            "billed_amount": billed_amount,
            "paid_amount": paid_amount,
            "open_amount": open_amount,
            "draft_amount": money(billing["draft_amount"]),
            "remaining_to_bill": remaining_to_bill,
            "invoice_count": billing["invoice_count"] or 0,
            "stage": billing_stage,
        },
        "customer_invoices": customer_invoices,
    }


def subproject_detail(subproject_id, change_order_id=None):
    subproject = one(
        """
        SELECT sp.*, p.name AS project_name
        FROM subprojects sp
        JOIN projects p ON p.id = sp.project_id
        WHERE sp.id = ?
        """,
        (subproject_id,),
    )
    if not subproject:
        return None
    selected_co = None
    scope_label = "Base Contract"
    cost_scope_sql = "subproject_id = ? AND change_order_id IS NULL"
    cost_scope_params = [subproject_id]
    if change_order_id:
        selected_co = one(
            """
            SELECT
              co.*,
              CASE WHEN COALESCE(co.pricing_type, 'Fixed') = 'T&M' THEN
                COALESCE((SELECT SUM(cr.sales_amount) FROM cost_records cr WHERE cr.change_order_id = co.id AND cr.source IN ('Field Wise', 'Field Wise PDF')), 0)
              ELSE
                CASE WHEN co.status IN ('Approved', 'Billed') THEN co.approved_value ELSE co.quoted_value END
              END sales_value
            FROM change_orders co
            WHERE co.id = ? AND co.subproject_id = ?
            """,
            (change_order_id, subproject_id),
        )
        if selected_co:
            scope_label = f"CO {selected_co.get('co_number') or ''}"
            if selected_co.get("job_number"):
                scope_label += f" / {selected_co.get('job_number')}"
            cost_scope_sql = "subproject_id = ? AND change_order_id = ?"
            cost_scope_params = [subproject_id, change_order_id]
    is_tm_scope = (
        (selected_co and (selected_co.get("pricing_type") or "Fixed") == "T&M")
        or (not selected_co and (subproject.get("pricing_type") or "Fixed") == "T&M")
    )
    totals = one(
        f"""
        SELECT
          COALESCE(SUM(
            CASE
              WHEN ? = 1 THEN CASE WHEN source IN ('Field Wise', 'Field Wise PDF') THEN amount ELSE 0 END
              ELSE amount
            END
          ), 0) raw_actual,
          COALESCE(SUM(CASE WHEN cost_type = 'Labor' THEN qty ELSE 0 END), 0) labor_hours_used,
          COALESCE(SUM(CASE WHEN cost_type = 'Field Ticket Material' THEN sales_amount ELSE 0 END), 0) field_ticket_material,
          COALESCE(SUM(CASE WHEN source = 'Vendor Invoice' AND cost_type = 'Material' THEN amount ELSE 0 END), 0) vendor_material,
          COALESCE(SUM(sales_amount), 0) fieldwise_sales
        FROM cost_records
        WHERE {cost_scope_sql}
        """,
        tuple([1 if is_tm_scope else 0] + cost_scope_params),
    )
    by_type = rows(
        f"""
        SELECT
          COALESCE(cost_type, 'Uncoded') label,
          COALESCE(SUM(
            CASE
              WHEN ? = 1 AND source NOT IN ('Field Wise', 'Field Wise PDF') THEN 0
              WHEN cost_type = 'Field Ticket Material' THEN sales_amount
              ELSE amount
            END
          ), 0) amount
        FROM cost_records
        WHERE {cost_scope_sql}
        GROUP BY COALESCE(cost_type, 'Uncoded')
        ORDER BY amount DESC
        """,
        tuple([1 if is_tm_scope else 0] + cost_scope_params),
    )
    records = rows(
        f"""
        SELECT *
        FROM cost_records
        WHERE {cost_scope_sql}
        ORDER BY record_date DESC, ticket_or_invoice, id
        """,
        tuple(cost_scope_params),
    )
    budget_total = money(subproject["budget_labor"]) + money(subproject["budget_material"]) + money(subproject.get("budget_equipment"))
    contract = money(selected_co["sales_value"]) if selected_co else (
        money(totals["fieldwise_sales"]) if (subproject.get("pricing_type") or "Fixed") == "T&M" else money(subproject["contract_value"])
    )
    raw_actual = money(totals["raw_actual"])
    profit = contract - raw_actual
    margin = profit / contract if contract else 0
    budget_used = raw_actual / budget_total if budget_total else 0
    labor_budget = money(subproject["budget_labor_hours"])
    labor_used = money(totals["labor_hours_used"])
    labor_used_pct = labor_used / labor_budget if labor_budget else 0
    return {
        "subproject": subproject,
        "scope_label": scope_label,
        "selected_change_order": selected_co,
        "contract_value": contract,
        "budget_total": budget_total,
        "raw_actual": raw_actual,
        "profit": profit,
        "margin": margin,
        "budget_used": budget_used,
        "labor_hours_budget": labor_budget,
        "labor_hours_used": labor_used,
        "labor_hours_used_pct": labor_used_pct,
        "field_ticket_material": money(totals["field_ticket_material"]),
        "vendor_material": money(totals["vendor_material"]),
        "fieldwise_sales": money(totals["fieldwise_sales"]),
        "by_type": by_type,
        "records": records,
    }


def master_project_detail(project_id):
    summary = project_summary(project_id)
    if not summary:
        return None
    project = summary["project"]
    labor = one(
        """
        SELECT
          COALESCE((SELECT SUM(budget_labor_hours) FROM subprojects WHERE project_id = ?), 0) labor_hours_budget,
          COALESCE((SELECT SUM(qty) FROM cost_records WHERE project_id = ? AND cost_type = 'Labor' AND change_order_id IS NULL), 0) labor_hours_used
        """,
        (project_id, project_id),
    )
    budget = one(
        """
        SELECT COALESCE(SUM(budget_labor + budget_material + COALESCE(budget_equipment, 0)), 0) budget_total
        FROM subprojects
        WHERE project_id = ?
        """,
        (project_id,),
    )
    records = rows(
        """
        SELECT cr.*, sp.job_number, sp.code AS subproject_code
        FROM cost_records cr
        LEFT JOIN subprojects sp ON sp.id = cr.subproject_id
        WHERE cr.project_id = ?
        ORDER BY cr.record_date DESC, cr.ticket_or_invoice, cr.id
        """,
        (project_id,),
    )
    labor_budget = money(labor["labor_hours_budget"])
    labor_used = money(labor["labor_hours_used"])
    budget_total = money(budget["budget_total"])
    return {
        "project": project,
        "contract_value": summary["contract_value"],
        "base_contract_value": summary["base_contract_value"],
        "approved_co_value": summary["approved_co_value"],
        "pending_co_value": summary["pending_co_value"],
        "budget_total": budget_total,
        "raw_actual": summary["actual_cost"],
        "profit": summary["profit"],
        "margin": summary["margin"],
        "budget_used": summary["actual_cost"] / budget_total if budget_total else 0,
        "labor_hours_budget": labor_budget,
        "labor_hours_used": labor_used,
        "labor_hours_used_pct": labor_used / labor_budget if labor_budget else 0,
        "field_ticket_material": money(summary["material_compare"]["field_ticket_material"]),
        "vendor_material": money(summary["material_compare"]["vendor_material"]),
        "by_type": summary["by_type"],
        "subprojects": summary["subprojects"],
        "records": records,
    }


def bid_summary():
    bids = rows("SELECT * FROM bid_requests ORDER BY bid_due_date, rfq_no")
    open_bids = [b for b in bids if b.get("outcome") not in ("Won", "Lost") and b.get("go_no_go") != "No Go"]
    won = [b for b in bids if b.get("outcome") == "Won"]
    lost = [b for b in bids if b.get("outcome") == "Lost"]
    open_pipeline = sum(money(b.get("bid_price")) for b in open_bids)
    weighted_forecast = sum(money(b.get("weighted_value")) for b in bids)
    margins = [money(b.get("target_margin")) for b in bids if money(b.get("target_margin")) > 0]
    stage = {}
    estimator = {}
    for b in bids:
        stage_name = b.get("stage") or "Unstaged"
        stage.setdefault(stage_name, {"stage": stage_name, "count": 0, "value": 0})
        stage[stage_name]["count"] += 1
        stage[stage_name]["value"] += money(b.get("bid_price"))
        name = b.get("estimator") or "Unassigned"
        estimator.setdefault(name, {"estimator": name, "open_rfqs": 0, "open_value": 0})
        if b in open_bids:
            estimator[name]["open_rfqs"] += 1
            estimator[name]["open_value"] += money(b.get("bid_price"))
    return {
        "bids": bids,
        "open_pipeline": open_pipeline,
        "weighted_forecast": weighted_forecast,
        "win_rate": len(won) / (len(won) + len(lost)) if (won or lost) else 0,
        "avg_target_margin": sum(margins) / len(margins) if margins else 0,
        "open_count": len(open_bids),
        "stage": list(stage.values()),
        "estimator": list(estimator.values()),
        "risks": rows("SELECT * FROM bid_risks ORDER BY id DESC"),
        "win_loss": rows("SELECT * FROM bid_win_loss ORDER BY id DESC"),
    }


def jobs_missing_recent_customer_invoice_attention(days=5):
    cutoff = (datetime.now() - timedelta(days=int(days or 5))).isoformat(timespec="seconds")
    subprojects = rows(
        """
        SELECT *
        FROM (
          SELECT
            'subproject' AS target_type,
            sp.id AS target_id,
            p.id AS project_id,
            p.project_code,
            p.name AS project_name,
            p.customer,
            sp.job_number,
            sp.code AS reference_code,
            sp.name AS job_name,
            'Subproject' AS job_type,
            ci.last_invoice_added_at,
            ci.last_invoice_date,
            ack.acknowledged_at,
            ack.acknowledged_by_username,
            MAX(
              COALESCE(ci.last_invoice_added_at, ''),
              COALESCE(ack.acknowledged_at, ''),
              COALESCE(p.created_at, '')
            ) AS last_activity_at
          FROM subprojects sp
          JOIN projects p ON p.id = sp.project_id
          LEFT JOIN (
            SELECT cia.subproject_id, MAX(ci.created_at) AS last_invoice_added_at, MAX(ci.invoice_date) AS last_invoice_date
            FROM customer_invoice_allocations cia
            JOIN customer_invoices ci ON ci.id = cia.customer_invoice_id
            WHERE COALESCE(ci.status, '') <> 'Void'
              AND cia.change_order_id IS NULL
            GROUP BY cia.subproject_id
          ) ci ON ci.subproject_id = sp.id
          LEFT JOIN job_invoice_alert_acknowledgements ack ON ack.target_type = 'subproject' AND ack.target_id = sp.id
          WHERE COALESCE(p.status, 'Active') <> 'Archived'
            AND TRIM(COALESCE(sp.job_number, '')) <> ''
        )
        WHERE COALESCE(last_activity_at, '') <> ''
          AND last_activity_at <= ?
        ORDER BY last_activity_at ASC, customer, project_name, job_number
        """,
        (cutoff,),
    )
    change_orders = rows(
        """
        SELECT *
        FROM (
          SELECT
            'change_order' AS target_type,
            co.id AS target_id,
            p.id AS project_id,
            p.project_code,
            p.name AS project_name,
            p.customer,
            co.job_number,
            co.co_number AS reference_code,
            co.title AS job_name,
            COALESCE(co.order_type, 'Change Order') AS job_type,
            ci.last_invoice_added_at,
            ci.last_invoice_date,
            ack.acknowledged_at,
            ack.acknowledged_by_username,
            MAX(
              COALESCE(ci.last_invoice_added_at, ''),
              COALESCE(ack.acknowledged_at, ''),
              COALESCE(p.created_at, '')
            ) AS last_activity_at
          FROM change_orders co
          JOIN projects p ON p.id = co.project_id
          LEFT JOIN (
            SELECT cia.change_order_id, MAX(ci.created_at) AS last_invoice_added_at, MAX(ci.invoice_date) AS last_invoice_date
            FROM customer_invoice_allocations cia
            JOIN customer_invoices ci ON ci.id = cia.customer_invoice_id
            WHERE COALESCE(ci.status, '') <> 'Void'
            GROUP BY cia.change_order_id
          ) ci ON ci.change_order_id = co.id
          LEFT JOIN job_invoice_alert_acknowledgements ack ON ack.target_type = 'change_order' AND ack.target_id = co.id
          WHERE COALESCE(p.status, 'Active') <> 'Archived'
            AND TRIM(COALESCE(co.job_number, '')) <> ''
            AND COALESCE(co.status, '') NOT IN ('Rejected', 'Void')
        )
        WHERE COALESCE(last_activity_at, '') <> ''
          AND last_activity_at <= ?
        ORDER BY last_activity_at ASC, customer, project_name, job_number
        """,
        (cutoff,),
    )
    return sorted(subprojects + change_orders, key=lambda item: (item.get("last_activity_at") or "", item.get("customer") or "", item.get("project_name") or "", item.get("job_number") or ""))


def home_alerts():
    today = datetime.now().date()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    stale_cutoff = (datetime.now() - timedelta(days=5)).isoformat(timespec="seconds")
    today_text = today.isoformat()
    uncoded = one(
        """
        SELECT
          COUNT(*) AS line_count,
          COUNT(DISTINCT COALESCE(NULLIF(TRIM(ticket_or_invoice), ''), source_file, id)) AS ticket_count,
          COALESCE(SUM(COALESCE(sales_amount, amount, 0)), 0) AS total_amount
        FROM cost_records
        WHERE source IN ('Field Wise', 'Field Wise PDF')
          AND (
            subproject_id IS NULL
            OR cost_type IS NULL
            OR cost_type = 'Uncoded'
            OR raw_cost_source IN ('Missing project rate', 'Missing internal rate')
          )
        """
    )
    pnl = one(
        """
        SELECT COUNT(*) AS count
        FROM financial_reports
        WHERE report_date BETWEEN ? AND ?
          AND report_type IN ('pnl', 'combined')
        """,
        (week_start.isoformat(), week_end.isoformat()),
    )
    balance_sheet = one(
        """
        SELECT COUNT(*) AS count
        FROM financial_reports
        WHERE report_date BETWEEN ? AND ?
          AND report_type IN ('balance_sheet', 'combined')
        """,
        (week_start.isoformat(), week_end.isoformat()),
    )
    overdue_bids = one(
        """
        SELECT COUNT(*) AS count, COALESCE(SUM(COALESCE(bid_price, 0)), 0) AS value
        FROM bid_requests
        WHERE COALESCE(outcome, 'Pending') NOT IN ('Won', 'Lost')
          AND COALESCE(go_no_go, '') <> 'No Go'
          AND TRIM(COALESCE(bid_due_date, '')) <> ''
          AND bid_due_date < ?
        """,
        (today_text,),
    )
    stale_bids = one(
        """
        SELECT COUNT(*) AS count, COALESCE(SUM(COALESCE(bid_price, 0)), 0) AS value
        FROM bid_requests
        WHERE COALESCE(outcome, 'Pending') NOT IN ('Won', 'Lost')
          AND COALESCE(go_no_go, '') <> 'No Go'
          AND COALESCE(updated_at, created_at, '') <> ''
          AND COALESCE(updated_at, created_at) <= ?
        """,
        (stale_cutoff,),
    )
    open_pos = one(
        """
        SELECT COUNT(*) AS count, COALESCE(SUM(COALESCE(estimated_amount, 0)), 0) AS value
        FROM purchase_orders
        WHERE COALESCE(status, 'Pending Approval') NOT IN ('Closed', 'Void')
        """
    )
    stale_invoice_jobs = jobs_missing_recent_customer_invoice_attention(5)
    missing_pnl = not bool(pnl and pnl["count"])
    missing_balance_sheet = not bool(balance_sheet and balance_sheet["count"])
    alerts = [
        {
            "key": "uncoded_fieldwise",
            "label": "Field Wise Review / Coding",
            "count": int(uncoded["ticket_count"] or 0),
            "detail": f"{int(uncoded['line_count'] or 0)} line(s) need review, coding, or rate setup",
            "amount": money(uncoded["total_amount"] if uncoded else 0),
            "severity": "warn" if uncoded and uncoded["ticket_count"] else "ok",
            "tab": "review",
        },
        {
            "key": "missing_texas_pnl",
            "label": "Missing Texas P&L",
            "count": 1 if missing_pnl else 0,
            "detail": f"Current week {week_start.isoformat()} to {week_end.isoformat()}",
            "severity": "bad" if missing_pnl else "ok",
            "tab": "texasOps",
        },
        {
            "key": "missing_balance_sheet",
            "label": "Missing Balance Sheet",
            "count": 1 if missing_balance_sheet else 0,
            "detail": f"Current week {week_start.isoformat()} to {week_end.isoformat()}",
            "severity": "bad" if missing_balance_sheet else "ok",
            "tab": "texasOps",
        },
        {
            "key": "overdue_bids",
            "label": "Overdue Bids",
            "count": int(overdue_bids["count"] or 0),
            "detail": "Open bids past due date",
            "amount": money(overdue_bids["value"] if overdue_bids else 0),
            "severity": "bad" if overdue_bids and overdue_bids["count"] else "ok",
            "tab": "bids",
        },
        {
            "key": "open_pos",
            "label": "Open POs",
            "count": int(open_pos["count"] or 0),
            "detail": "Not closed or void",
            "amount": money(open_pos["value"] if open_pos else 0),
            "severity": "warn" if open_pos and open_pos["count"] else "ok",
            "tab": "officePo",
        },
        {
            "key": "stale_job_invoices",
            "label": "Jobs / Subprojects Not Invoiced",
            "count": len(stale_invoice_jobs),
            "detail": "No customer invoice added or acknowledgement in 5+ days",
            "severity": "warn" if stale_invoice_jobs else "ok",
            "tab": "billing",
            "items": [
                {
                    "target_type": item["target_type"],
                    "target_id": item["target_id"],
                    "project_id": item["project_id"],
                    "label": " / ".join(str(x or "").strip() for x in (item["job_number"], item["reference_code"]) if str(x or "").strip()) or str(item["job_name"] or "Job"),
                    "customer": item["customer"] or "",
                    "project_name": item["project_name"] or "",
                    "job_name": item["job_name"] or "",
                    "job_type": item["job_type"] or "",
                    "last_invoice_added_at": item["last_invoice_added_at"] or "",
                    "last_invoice_date": item["last_invoice_date"] or "",
                    "acknowledged_at": item["acknowledged_at"] or "",
                    "acknowledged_by_username": item["acknowledged_by_username"] or "",
                    "last_activity_at": item["last_activity_at"] or "",
                }
                for item in stale_invoice_jobs[:10]
            ],
        },
        {
            "key": "stale_pending_bids",
            "label": "Stale Pending Bid Rows",
            "count": int(stale_bids["count"] or 0),
            "detail": "No update in 5+ days",
            "amount": money(stale_bids["value"] if stale_bids else 0),
            "severity": "warn" if stale_bids and stale_bids["count"] else "ok",
            "tab": "bids",
        },
    ]
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "week_start": week_start.isoformat(),
        "week_end": week_end.isoformat(),
        "attention_count": sum(1 for alert in alerts if alert["severity"] != "ok" and alert["count"]),
        "alerts": alerts,
    }


FINANCIAL_METRICS = {
    "revenue": ("Revenue", ["total for income", "total income", "total revenue", "revenue", "sales"], "pnl"),
    "gross_profit": ("Gross Profit", ["gross profit", "gross margin"], "pnl"),
    "operating_expenses": ("Operating Expenses", ["total for expenses", "total operating expenses", "operating expenses", "total expenses"], "pnl"),
    "net_income": ("Net Income", ["net income", "net profit", "net earnings"], "pnl"),
    "cash": ("Cash", ["cash and cash equivalents", "cash in bank", "bank accounts", "cash"], "balance_sheet"),
    "accounts_receivable": ("Accounts Receivable", ["accounts receivable", "a/r", "ar"], "balance_sheet"),
    "accounts_payable": ("Accounts Payable", ["accounts payable", "a/p", "ap"], "balance_sheet"),
    "current_assets": ("Current Assets", ["total for current assets", "total current assets", "current assets"], "balance_sheet"),
    "total_assets": ("Total Assets", ["total for assets", "total assets"], "balance_sheet"),
    "current_liabilities": ("Current Liabilities", ["total for current liabilities", "total current liabilities", "current liabilities"], "balance_sheet"),
    "total_liabilities": ("Total Liabilities", ["total for liabilities", "total liabilities"], "balance_sheet"),
    "equity": ("Equity", ["total for equity", "total equity", "owner equity", "shareholder equity", "members equity", "equity"], "balance_sheet"),
}


def financial_amount(value):
    text = str(value or "").strip()
    if not text:
        return 0.0
    negative = "(" in text and ")" in text
    amount = parse_money_text(text)
    return -abs(amount) if negative else amount


def normalize_financial_label(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def match_financial_metric(label, report_type):
    normalized = normalize_financial_label(label)
    if normalized in ("total for liabilities and equity", "liabilities and equity"):
        return None
    matches = []
    for key, (display, aliases, metric_type) in FINANCIAL_METRICS.items():
        if report_type != "combined" and metric_type != report_type:
            continue
        for alias in aliases:
            alias_norm = normalize_financial_label(alias)
            if normalized == alias_norm:
                matches.append((3, len(alias_norm), key))
            elif normalized == f"total for {alias_norm}":
                matches.append((1, len(alias_norm), key))
    return sorted(matches, reverse=True)[0][2] if matches else None


def parse_report_date_text(value):
    text = str(value or "")
    match = re.search(r"(?:as of|through|january\s+\d+\s*-\s*)?\s*([A-Za-z]{3,9}\s+\d{1,2},\s+\d{4})", text, re.IGNORECASE)
    if not match:
        return ""
    for fmt in ("%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(match.group(1), fmt).date().isoformat()
        except ValueError:
            pass
    return ""


def infer_financial_report_date(path):
    suffix = Path(path).suffix.lower()
    candidates = []
    if suffix in (".xlsx", ".xlsm") and openpyxl is not None:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            for ws in wb.worksheets:
                for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
                    candidates.extend(str(v) for v in (row or []) if v)
        finally:
            wb.close()
    elif suffix == ".pdf":
        candidates = extract_pdf_text(path).splitlines()[:12]
    for candidate in candidates:
        parsed = parse_report_date_text(candidate)
        if parsed:
            return parsed
    return ""


def extract_financial_pairs_from_xlsx(path):
    if openpyxl is None:
        raise RuntimeError("Excel import needs openpyxl, but it is not available.")
    wb = openpyxl.load_workbook(path, data_only=True)
    pairs = []
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                text_cells = []
                amount_cells = []
                for idx, value in enumerate(row or []):
                    if value is None:
                        continue
                    if isinstance(value, (int, float)) and not isinstance(value, bool):
                        amount_cells.append((idx, float(value)))
                    else:
                        amount = financial_amount(value)
                        if amount:
                            amount_cells.append((idx, amount))
                        elif str(value).strip():
                            text_cells.append((idx, str(value).strip()))
                if not text_cells or not amount_cells:
                    continue
                value_idx, amount = amount_cells[-1]
                label = " ".join(text for idx, text in text_cells if idx < value_idx).strip()
                if label:
                    pairs.append((label, amount))
    finally:
        wb.close()
    return pairs


def extract_financial_pairs_from_csv(path):
    pairs = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            text_cells = []
            amount_cells = []
            for idx, value in enumerate(row):
                amount = financial_amount(value)
                if amount:
                    amount_cells.append((idx, amount))
                elif str(value or "").strip():
                    text_cells.append((idx, str(value).strip()))
            if not text_cells or not amount_cells:
                continue
            value_idx, amount = amount_cells[-1]
            label = " ".join(text for idx, text in text_cells if idx < value_idx).strip()
            if label:
                pairs.append((label, amount))
    return pairs


def extract_financial_pairs_from_pdf(path):
    pairs = []
    text = extract_pdf_text(path)
    for line in text.splitlines():
        cleaned = re.sub(r"\s+", " ", line).strip()
        match = re.search(r"(.+?)\s+(\(?\$?\s*-?[0-9][0-9,]*\.?[0-9]*\)?)$", cleaned)
        if match:
            label = re.sub(r"\s+\(?\$?\s*-?[0-9][0-9,]*\.?[0-9]*\)?$", "", match.group(1).strip()).strip()
            pairs.append((label, financial_amount(match.group(2))))
    return pairs


def extract_financial_metrics(path, report_type):
    suffix = Path(path).suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        pairs = extract_financial_pairs_from_xlsx(path)
    elif suffix in (".csv", ".tsv"):
        pairs = extract_financial_pairs_from_csv(path)
    elif suffix == ".pdf":
        pairs = extract_financial_pairs_from_pdf(path)
    else:
        raise RuntimeError("Upload an Excel, CSV, or PDF financial report.")
    metrics = {}
    for label, amount in pairs:
        key = match_financial_metric(label, report_type)
        if key:
            display = FINANCIAL_METRICS[key][0]
            metrics[key] = {"metric_key": key, "label": display, "amount": amount, "source_label": label}
    return list(metrics.values())


def import_financial_report(path, report_date, report_type):
    requested_report_type = report_type or "combined"
    metrics = extract_financial_metrics(path, requested_report_type)
    if not metrics:
        raise RuntimeError("No recognizable Balance Sheet or P&L metrics were found.")
    report_date = report_date or infer_financial_report_date(path) or datetime.now().date().isoformat()
    metric_types = {FINANCIAL_METRICS[m["metric_key"]][2] for m in metrics if m.get("metric_key") in FINANCIAL_METRICS}
    report_type = next(iter(metric_types)) if requested_report_type == "combined" and len(metric_types) == 1 else requested_report_type
    source_name = Path(path).name
    original_source_name = re.sub(r"^\d{14}-", "", source_name)
    existing_reports = rows(
        "SELECT id, source_file FROM financial_reports WHERE report_date = ? AND report_type = ?",
        (report_date, report_type),
    )
    for existing in existing_reports:
        existing_source_name = re.sub(r"^\d{14}-", "", existing["source_file"] or "")
        if existing_source_name == original_source_name:
            return {
                "report_id": existing["id"],
                "count": 0,
                "metrics": [],
                "duplicate": True,
                "report_date": report_date,
                "report_type": report_type,
                "source_file": source_name,
                "original_source_file": original_source_name,
            }
    now = datetime.now().isoformat(timespec="seconds")
    report_id = execute(
        "INSERT INTO financial_reports (report_date, report_type, source_file, uploaded_at, notes) VALUES (?, ?, ?, ?, ?)",
        (report_date, report_type, source_name, now, json.dumps({"matched": [m["source_label"] for m in metrics], "original_source_file": original_source_name}, default=str)),
    )
    for metric in metrics:
        execute(
            "INSERT OR REPLACE INTO financial_metrics (report_id, metric_key, label, amount) VALUES (?, ?, ?, ?)",
            (report_id, metric["metric_key"], metric["label"], metric["amount"]),
        )
    return {"report_id": report_id, "count": len(metrics), "metrics": metrics, "duplicate": False, "report_date": report_date, "report_type": report_type, "source_file": source_name, "original_source_file": original_source_name}


def texas_financial_summary():
    reports = rows("SELECT * FROM financial_reports ORDER BY report_date DESC, id DESC")
    ap_snapshots = rows("SELECT * FROM texas_ap_snapshots ORDER BY report_date DESC, id DESC")
    metrics = rows(
        """
        SELECT fr.report_date, fr.report_type, fr.source_file, fm.metric_key, fm.label, fm.amount
        FROM financial_reports fr
        JOIN financial_metrics fm ON fm.report_id = fr.id
        ORDER BY fr.report_date, fr.id, fm.metric_key
        """
    )
    latest = {}
    for metric in metrics:
        latest[metric["metric_key"]] = metric
    current_assets = money(latest.get("current_assets", {}).get("amount"))
    current_liabilities = money(latest.get("current_liabilities", {}).get("amount"))
    same_current_date = latest.get("current_assets", {}).get("report_date") == latest.get("current_liabilities", {}).get("report_date")
    latest["working_capital"] = {
        "label": "Working Capital",
        "amount": current_assets - current_liabilities if same_current_date else 0,
        "report_date": latest.get("current_assets", {}).get("report_date") if same_current_date else "",
    }
    latest["current_ratio"] = {
        "label": "Current Ratio",
        "amount": current_assets / current_liabilities if same_current_date and current_liabilities else 0,
        "report_date": latest.get("current_assets", {}).get("report_date") if same_current_date else "",
    }
    latest_ap = ap_snapshots[0] if ap_snapshots else None
    manual_ap_total = money(latest_ap["ap_total"] if latest_ap else 0)
    future_current_liabilities = current_liabilities + manual_ap_total
    latest["manual_accounts_payable"] = {
        "label": "Provided AP",
        "amount": manual_ap_total,
        "report_date": latest_ap["report_date"] if latest_ap else "",
    }
    latest["future_current_liabilities"] = {
        "label": "Future Current Liabilities",
        "amount": future_current_liabilities,
        "report_date": latest_ap["report_date"] if latest_ap else latest.get("current_liabilities", {}).get("report_date", ""),
    }
    latest["future_working_capital"] = {
        "label": "Future Working Capital",
        "amount": current_assets - future_current_liabilities if current_assets or future_current_liabilities else 0,
        "report_date": latest_ap["report_date"] if latest_ap else latest.get("current_assets", {}).get("report_date", ""),
    }
    latest["future_current_ratio"] = {
        "label": "Future Current Ratio",
        "amount": current_assets / future_current_liabilities if current_assets and future_current_liabilities else 0,
        "report_date": latest_ap["report_date"] if latest_ap else latest.get("current_assets", {}).get("report_date", ""),
    }
    history = {}
    for metric in metrics:
        history.setdefault(metric["metric_key"], []).append({"report_date": metric["report_date"], "amount": metric["amount"]})
    if ap_snapshots:
        history["manual_accounts_payable"] = [{"report_date": ap["report_date"], "amount": money(ap["ap_total"])} for ap in sorted(ap_snapshots, key=lambda p: p["report_date"])]
    by_date = {}
    for metric in metrics:
        bucket = by_date.setdefault(metric["report_date"], {})
        bucket[metric["metric_key"]] = money(metric["amount"])
    for ap in ap_snapshots:
        bucket = by_date.setdefault(ap["report_date"], {})
        bucket["manual_accounts_payable"] = money(ap["ap_total"])
    working_capital_history = []
    current_ratio_history = []
    future_current_liabilities_history = []
    future_working_capital_history = []
    future_current_ratio_history = []
    for report_date, metric_set in by_date.items():
        if "current_assets" in metric_set and "current_liabilities" in metric_set:
            current_assets_for_date = metric_set["current_assets"]
            current_liabilities_for_date = metric_set["current_liabilities"]
            working_capital_history.append({"report_date": report_date, "amount": current_assets_for_date - current_liabilities_for_date})
            if current_liabilities_for_date:
                current_ratio_history.append({"report_date": report_date, "amount": current_assets_for_date / current_liabilities_for_date})
            if "manual_accounts_payable" in metric_set:
                future_liabilities_for_date = current_liabilities_for_date + metric_set["manual_accounts_payable"]
                future_current_liabilities_history.append({"report_date": report_date, "amount": future_liabilities_for_date})
                future_working_capital_history.append({"report_date": report_date, "amount": current_assets_for_date - future_liabilities_for_date})
                if future_liabilities_for_date:
                    future_current_ratio_history.append({"report_date": report_date, "amount": current_assets_for_date / future_liabilities_for_date})
    if working_capital_history:
        history["working_capital"] = sorted(working_capital_history, key=lambda p: p["report_date"])
    if current_ratio_history:
        history["current_ratio"] = sorted(current_ratio_history, key=lambda p: p["report_date"])
    if future_current_liabilities_history:
        history["future_current_liabilities"] = sorted(future_current_liabilities_history, key=lambda p: p["report_date"])
    if future_working_capital_history:
        history["future_working_capital"] = sorted(future_working_capital_history, key=lambda p: p["report_date"])
    if future_current_ratio_history:
        history["future_current_ratio"] = sorted(future_current_ratio_history, key=lambda p: p["report_date"])
    reports_by_week = {}
    for report in reports:
        date = report["report_date"] or ""
        bucket = reports_by_week.setdefault(date, {"report_date": date, "pnl": [], "balance_sheet": [], "combined": [], "report_count": 0})
        report_type = report["report_type"] or "combined"
        if report_type not in bucket:
            report_type = "combined"
        bucket[report_type].append(report)
        bucket["report_count"] += 1
    week_status = []
    for bucket in reports_by_week.values():
        has_pnl = bool(bucket["pnl"] or bucket["combined"])
        has_balance = bool(bucket["balance_sheet"] or bucket["combined"])
        if has_pnl and has_balance:
            status = "Complete"
        elif has_pnl:
            status = "Missing Balance Sheet"
        elif has_balance:
            status = "Missing P&L"
        else:
            status = "Missing Reports"
        bucket["has_pnl"] = has_pnl
        bucket["has_balance_sheet"] = has_balance
        bucket["status"] = status
        week_status.append(bucket)
    week_status.sort(key=lambda x: x["report_date"], reverse=True)
    return {"reports": reports[:24], "report_weeks": week_status[:24], "ap_snapshots": ap_snapshots[:24], "latest_metrics": latest, "history": history}


def texas_pnl_uploaded_for_date(report_date):
    report_date = str(report_date or "").strip()
    if not report_date:
        return False
    found = one(
        """
        SELECT fr.id
        FROM financial_reports fr
        LEFT JOIN financial_metrics fm ON fm.report_id = fr.id
        WHERE fr.report_date = ?
          AND (
            fr.report_type IN ('pnl', 'combined')
            OR fm.metric_key IN ('revenue', 'gross_profit', 'operating_expenses', 'net_income')
          )
        LIMIT 1
        """,
        (report_date,),
    )
    return bool(found)


REMINDER_WEEKDAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


def texas_upload_reminder_settings():
    setting = one("SELECT * FROM texas_upload_reminder_settings WHERE id = 1")
    if not setting:
        execute(
            "INSERT INTO texas_upload_reminder_settings (id, enabled, weekday, reminder_time, updated_at) VALUES (1, 1, 4, '15:00', ?)",
            (datetime.now().isoformat(timespec="seconds"),),
        )
        setting = one("SELECT * FROM texas_upload_reminder_settings WHERE id = 1")
    result = dict(setting)
    try:
        result["weekday"] = max(0, min(6, int(result.get("weekday", 4))))
    except Exception:
        result["weekday"] = 4
    reminder_time = str(result.get("reminder_time") or "15:00").strip()
    if not re.match(r"^\d{2}:\d{2}$", reminder_time):
        reminder_time = "15:00"
    result["reminder_time"] = reminder_time
    result["weekday_label"] = REMINDER_WEEKDAYS[result["weekday"]]
    result["schedule_label"] = f"Every {result['weekday_label']} at {reminder_time} server time"
    return result


def texas_upload_reminder_status(reminder_date=None):
    reminder_date = reminder_date or datetime.now().date().isoformat()
    settings = texas_upload_reminder_settings()
    recipients = rows(
        """
        SELECT u.id, u.username, u.display_name, u.role, COALESCE(r.active, 0) AS selected
        FROM users u
        LEFT JOIN texas_upload_reminder_recipients r ON r.user_id = u.id AND COALESCE(r.active, 1) = 1
        WHERE u.active = 1
        ORDER BY u.display_name, u.username
        """
    )
    logs = rows(
        """
        SELECT *
        FROM texas_upload_reminder_log
        ORDER BY sent_at DESC, id DESC
        LIMIT 20
        """
    )
    return {
        "smtp_configured": bool(SMTP_HOST and SMTP_FROM),
        "reminder_date": reminder_date,
        "pnl_uploaded": texas_pnl_uploaded_for_date(reminder_date),
        "settings": settings,
        "schedule": settings["schedule_label"],
        "recipients": recipients,
        "logs": logs,
    }


def active_texas_upload_reminder_recipients():
    return rows(
        """
        SELECT u.id, u.username, u.display_name
        FROM texas_upload_reminder_recipients r
        JOIN users u ON u.id = r.user_id
        WHERE COALESCE(r.active, 1) = 1
          AND u.active = 1
        ORDER BY u.display_name, u.username
        """
    )


def billing_invoice_reminder_settings():
    setting = one("SELECT * FROM billing_invoice_reminder_settings WHERE id = 1")
    if not setting:
        execute(
            "INSERT INTO billing_invoice_reminder_settings (id, enabled, weekday, reminder_time, lookback_days, updated_at) VALUES (1, 1, 4, '15:00', 7, ?)",
            (datetime.now().isoformat(timespec="seconds"),),
        )
        setting = one("SELECT * FROM billing_invoice_reminder_settings WHERE id = 1")
    result = dict(setting)
    try:
        result["weekday"] = max(0, min(6, int(result.get("weekday", 4))))
    except Exception:
        result["weekday"] = 4
    reminder_time = str(result.get("reminder_time") or "15:00").strip()
    if not re.match(r"^\d{2}:\d{2}$", reminder_time):
        reminder_time = "15:00"
    result["reminder_time"] = reminder_time
    try:
        result["lookback_days"] = max(1, min(60, int(result.get("lookback_days", 7))))
    except Exception:
        result["lookback_days"] = 7
    result["weekday_label"] = REMINDER_WEEKDAYS[result["weekday"]]
    result["schedule_label"] = f"Every {result['weekday_label']} at {reminder_time} server time"
    return result


def jobs_missing_recent_customer_invoice(lookback_days=7):
    cutoff = (datetime.now() - timedelta(days=int(lookback_days or 7))).isoformat(timespec="seconds")
    subproject_jobs = rows(
        """
        SELECT
          'Subproject' AS job_type,
          p.customer,
          p.project_code,
          p.name AS project_name,
          sp.job_number,
          sp.code AS reference_code,
          sp.name AS description,
          MAX(ci.created_at) AS last_invoice_added_at,
          MAX(ci.invoice_date) AS last_invoice_date,
          SUM(CASE WHEN ci.created_at >= ? THEN 1 ELSE 0 END) AS recent_invoice_count
        FROM subprojects sp
        JOIN projects p ON p.id = sp.project_id
        LEFT JOIN customer_invoice_allocations cia ON cia.subproject_id = sp.id AND cia.change_order_id IS NULL
        LEFT JOIN customer_invoices ci ON ci.id = cia.customer_invoice_id AND COALESCE(ci.status, '') <> 'Void'
        WHERE COALESCE(p.status, 'Active') <> 'Archived'
          AND TRIM(COALESCE(sp.job_number, '')) <> ''
        GROUP BY sp.id
        HAVING COALESCE(recent_invoice_count, 0) = 0
        ORDER BY p.customer, p.name, sp.job_number
        """,
        (cutoff,),
    )
    change_order_jobs = rows(
        """
        SELECT
          COALESCE(co.order_type, 'Change Order') AS job_type,
          p.customer,
          p.project_code,
          p.name AS project_name,
          co.job_number,
          co.co_number AS reference_code,
          co.title AS description,
          MAX(ci.created_at) AS last_invoice_added_at,
          MAX(ci.invoice_date) AS last_invoice_date,
          SUM(CASE WHEN ci.created_at >= ? THEN 1 ELSE 0 END) AS recent_invoice_count
        FROM change_orders co
        JOIN projects p ON p.id = co.project_id
        LEFT JOIN customer_invoice_allocations cia ON cia.change_order_id = co.id
        LEFT JOIN customer_invoices ci ON ci.id = cia.customer_invoice_id AND COALESCE(ci.status, '') <> 'Void'
        WHERE COALESCE(p.status, 'Active') <> 'Archived'
          AND TRIM(COALESCE(co.job_number, '')) <> ''
          AND COALESCE(co.status, '') NOT IN ('Rejected', 'Void')
        GROUP BY co.id
        HAVING COALESCE(recent_invoice_count, 0) = 0
        ORDER BY p.customer, p.name, co.job_number
        """,
        (cutoff,),
    )
    return subproject_jobs + change_order_jobs


def active_billing_invoice_reminder_recipients():
    return rows(
        """
        SELECT u.id, u.username, u.display_name
        FROM billing_invoice_reminder_recipients r
        JOIN users u ON u.id = r.user_id
        WHERE COALESCE(r.active, 1) = 1
          AND u.active = 1
        ORDER BY u.display_name, u.username
        """
    )


def billing_invoice_reminder_status():
    settings = billing_invoice_reminder_settings()
    recipients = rows(
        """
        SELECT u.id, u.username, u.display_name, u.role, COALESCE(r.active, 0) AS selected
        FROM users u
        LEFT JOIN billing_invoice_reminder_recipients r ON r.user_id = u.id AND COALESCE(r.active, 1) = 1
        WHERE u.active = 1
        ORDER BY u.display_name, u.username
        """
    )
    logs = rows(
        """
        SELECT *
        FROM billing_invoice_reminder_log
        ORDER BY sent_at DESC, id DESC
        LIMIT 20
        """
    )
    missing_jobs = jobs_missing_recent_customer_invoice(settings["lookback_days"])
    return {
        "smtp_configured": bool(SMTP_HOST and SMTP_FROM),
        "settings": settings,
        "schedule": settings["schedule_label"],
        "recipients": recipients,
        "logs": logs,
        "missing_jobs": missing_jobs[:100],
        "missing_job_count": len(missing_jobs),
    }


def po_untouched_reminder_settings():
    setting = one("SELECT * FROM po_untouched_reminder_settings WHERE id = 1")
    if not setting:
        execute(
            "INSERT INTO po_untouched_reminder_settings (id, enabled, weekday, reminder_time, untouched_days, updated_at) VALUES (1, 1, 4, '15:00', 2, ?)",
            (datetime.now().isoformat(timespec="seconds"),),
        )
        setting = one("SELECT * FROM po_untouched_reminder_settings WHERE id = 1")
    result = dict(setting)
    try:
        result["weekday"] = max(0, min(6, int(result.get("weekday", 4))))
    except Exception:
        result["weekday"] = 4
    reminder_time = str(result.get("reminder_time") or "15:00").strip()
    if not re.match(r"^\d{2}:\d{2}$", reminder_time):
        reminder_time = "15:00"
    result["reminder_time"] = reminder_time
    try:
        result["untouched_days"] = max(1, min(30, int(result.get("untouched_days", 2))))
    except Exception:
        result["untouched_days"] = 2
    result["weekday_label"] = REMINDER_WEEKDAYS[result["weekday"]]
    result["schedule_label"] = f"Every {result['weekday_label']} at {reminder_time} server time"
    return result


def untouched_purchase_orders(untouched_days=2):
    cutoff = (datetime.now() - timedelta(days=int(untouched_days or 2))).isoformat(timespec="seconds")
    return rows(
        """
        SELECT
          po.*,
          COALESCE(po.updated_at, po.created_at) AS last_touched_at,
          p.customer AS project_customer,
          p.name AS project_name,
          p.project_code
        FROM purchase_orders po
        LEFT JOIN projects p ON p.id = po.project_id
        WHERE COALESCE(po.status, 'Pending Approval') IN ('Pending Approval', 'Issued')
          AND COALESCE(po.pickup_file, '') = ''
          AND COALESCE(po.updated_at, po.created_at) <= ?
        ORDER BY COALESCE(po.updated_at, po.created_at), po.created_at, po.po_number
        """,
        (cutoff,),
    )


def active_po_untouched_reminder_recipients():
    return rows(
        """
        SELECT u.id, u.username, u.display_name
        FROM po_untouched_reminder_recipients r
        JOIN users u ON u.id = r.user_id
        WHERE COALESCE(r.active, 1) = 1
          AND u.active = 1
        ORDER BY u.display_name, u.username
        """
    )


def po_untouched_reminder_status():
    settings = po_untouched_reminder_settings()
    recipients = rows(
        """
        SELECT u.id, u.username, u.display_name, u.role, COALESCE(r.active, 0) AS selected
        FROM users u
        LEFT JOIN po_untouched_reminder_recipients r ON r.user_id = u.id AND COALESCE(r.active, 1) = 1
        WHERE u.active = 1
        ORDER BY u.display_name, u.username
        """
    )
    logs = rows(
        """
        SELECT *
        FROM po_untouched_reminder_log
        ORDER BY sent_at DESC, id DESC
        LIMIT 20
        """
    )
    pos = untouched_purchase_orders(settings["untouched_days"])
    return {
        "smtp_configured": bool(SMTP_HOST and SMTP_FROM),
        "settings": settings,
        "schedule": settings["schedule_label"],
        "recipients": recipients,
        "logs": logs,
        "purchase_orders": pos[:100],
        "po_count": len(pos),
    }


def send_email(to_address, subject, body):
    if not SMTP_HOST or not SMTP_FROM:
        raise RuntimeError("Email is not configured. Set PM_TRACKER_SMTP_HOST and PM_TRACKER_SMTP_FROM on the server.")
    message = EmailMessage()
    message["From"] = SMTP_FROM
    message["To"] = to_address
    message["Subject"] = subject
    message.set_content(body)
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as smtp:
        if SMTP_USE_TLS:
            smtp.starttls()
        if SMTP_USERNAME:
            smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
        smtp.send_message(message)


def log_texas_upload_reminder(reminder_date, recipient, subject, status, error=""):
    execute(
        """
        INSERT INTO texas_upload_reminder_log (reminder_date, recipient, subject, status, error, sent_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (reminder_date, recipient, subject, status, str(error or ""), datetime.now().isoformat(timespec="seconds")),
    )


def texas_upload_reminder_already_attempted(reminder_date, recipient):
    return bool(
        one(
            """
            SELECT id
            FROM texas_upload_reminder_log
            WHERE reminder_date = ?
              AND recipient = ?
              AND status IN ('sent', 'failed')
            LIMIT 1
            """,
            (reminder_date, recipient),
        )
    )


def billing_invoice_reminder_already_attempted(reminder_date, recipient):
    return bool(
        one(
            """
            SELECT id
            FROM billing_invoice_reminder_log
            WHERE reminder_date = ?
              AND recipient = ?
              AND status IN ('sent', 'failed')
            LIMIT 1
            """,
            (reminder_date, recipient),
        )
    )


def po_untouched_reminder_already_attempted(reminder_date, recipient):
    return bool(
        one(
            """
            SELECT id
            FROM po_untouched_reminder_log
            WHERE reminder_date = ?
              AND recipient = ?
              AND status IN ('sent', 'failed')
            LIMIT 1
            """,
            (reminder_date, recipient),
        )
    )


def log_billing_invoice_reminder(reminder_date, recipient, subject, status, error="", job_count=0):
    execute(
        """
        INSERT INTO billing_invoice_reminder_log (reminder_date, recipient, subject, status, error, job_count, sent_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (reminder_date, recipient, subject, status, str(error or ""), int(job_count or 0), datetime.now().isoformat(timespec="seconds")),
    )


def log_po_untouched_reminder(reminder_date, recipient, subject, status, error="", po_count=0):
    execute(
        """
        INSERT INTO po_untouched_reminder_log (reminder_date, recipient, subject, status, error, po_count, sent_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (reminder_date, recipient, subject, status, str(error or ""), int(po_count or 0), datetime.now().isoformat(timespec="seconds")),
    )


def run_texas_upload_reminder(force=False):
    reminder_date = datetime.now().date().isoformat()
    if texas_pnl_uploaded_for_date(reminder_date):
        return {"sent": 0, "failed": 0, "skipped": 0, "pnl_uploaded": True, "message": "P&L is already uploaded for today."}
    recipients = active_texas_upload_reminder_recipients()
    subject = f"Texas Ops P&L upload needed for {reminder_date}"
    body = (
        f"Reminder: the Texas Operations weekly P&L has not been uploaded for {reminder_date}.\n\n"
        "Please upload the P&L report in the Company Dashboard under Texas Ops before end of day.\n"
    )
    result = {"sent": 0, "failed": 0, "skipped": 0, "pnl_uploaded": False, "message": ""}
    for recipient in recipients:
        email = str(recipient["username"] or "").strip()
        if "@" not in email:
            result["skipped"] += 1
            if force:
                log_texas_upload_reminder(reminder_date, email or f"user:{recipient['id']}", subject, "skipped", "Username is not an email address.")
            continue
        if not force and texas_upload_reminder_already_attempted(reminder_date, email):
            result["skipped"] += 1
            continue
        try:
            send_email(email, subject, body)
            log_texas_upload_reminder(reminder_date, email, subject, "sent")
            result["sent"] += 1
        except Exception as exc:
            log_texas_upload_reminder(reminder_date, email, subject, "failed", exc)
            result["failed"] += 1
    result["message"] = f"Sent {result['sent']}, failed {result['failed']}, skipped {result['skipped']}."
    return result


def run_po_untouched_reminder(force=False):
    settings = po_untouched_reminder_settings()
    reminder_date = datetime.now().date().isoformat()
    pos = untouched_purchase_orders(settings["untouched_days"])
    if not pos:
        return {"sent": 0, "failed": 0, "skipped": 0, "po_count": 0, "message": "No untouched POs are currently past the alert window."}
    recipients = active_po_untouched_reminder_recipients()
    subject = f"Untouched POs need follow-up: {len(pos)}"
    po_lines = []
    for po in pos[:200]:
        label = " / ".join(str(x or "").strip() for x in [po.get("po_number"), po.get("job_number"), po.get("job_label")] if str(x or "").strip())
        po_lines.append(
            f"- {label} | status: {po.get('status') or ''} | vendor: {po.get('vendor') or ''} | created by: {po.get('requested_by_username') or ''} | last touched: {po.get('last_touched_at') or po.get('created_at') or ''}"
        )
    body = (
        f"PO follow-up reminder for {reminder_date}\n\n"
        f"The following POs have not had a pickup ticket uploaded or been closed/voided after {settings['untouched_days']} day(s):\n\n"
        + "\n".join(po_lines)
        + "\n\nPlease review Office PO Review in the Company Dashboard."
    )
    result = {"sent": 0, "failed": 0, "skipped": 0, "po_count": len(pos), "message": ""}
    for recipient in recipients:
        email = str(recipient["username"] or "").strip()
        if "@" not in email:
            result["skipped"] += 1
            if force:
                log_po_untouched_reminder(reminder_date, email or f"user:{recipient['id']}", subject, "skipped", "Username is not an email address.", len(pos))
            continue
        if not force and po_untouched_reminder_already_attempted(reminder_date, email):
            result["skipped"] += 1
            continue
        try:
            send_email(email, subject, body)
            log_po_untouched_reminder(reminder_date, email, subject, "sent", "", len(pos))
            result["sent"] += 1
        except Exception as exc:
            log_po_untouched_reminder(reminder_date, email, subject, "failed", exc, len(pos))
            result["failed"] += 1
    result["message"] = f"Found {len(pos)} untouched PO(s). Sent {result['sent']}, failed {result['failed']}, skipped {result['skipped']}."
    return result


def run_billing_invoice_reminder(force=False):
    settings = billing_invoice_reminder_settings()
    reminder_date = datetime.now().date().isoformat()
    jobs = jobs_missing_recent_customer_invoice(settings["lookback_days"])
    if not jobs:
        return {"sent": 0, "failed": 0, "skipped": 0, "job_count": 0, "message": "All active jobs have a customer invoice added inside the lookback window."}
    recipients = active_billing_invoice_reminder_recipients()
    subject = f"Jobs missing recent customer invoices: {len(jobs)}"
    job_lines = []
    for job in jobs[:200]:
        label = " / ".join(str(x or "").strip() for x in [job.get("job_number"), job.get("reference_code")] if str(x or "").strip())
        last_added = job.get("last_invoice_added_at") or "Never"
        job_lines.append(f"- {label} | {job.get('customer') or ''} | {job.get('project_name') or ''} | {job.get('job_type') or ''} | last added: {last_added}")
    body = (
        f"Customer invoice reminder for {reminder_date}\n\n"
        f"The following active jobs have not had a customer invoice added in the last {settings['lookback_days']} day(s):\n\n"
        + "\n".join(job_lines)
        + "\n\nPlease review Customer Billing in the Company Dashboard."
    )
    result = {"sent": 0, "failed": 0, "skipped": 0, "job_count": len(jobs), "message": ""}
    for recipient in recipients:
        email = str(recipient["username"] or "").strip()
        if "@" not in email:
            result["skipped"] += 1
            if force:
                log_billing_invoice_reminder(reminder_date, email or f"user:{recipient['id']}", subject, "skipped", "Username is not an email address.", len(jobs))
            continue
        if not force and billing_invoice_reminder_already_attempted(reminder_date, email):
            result["skipped"] += 1
            continue
        try:
            send_email(email, subject, body)
            log_billing_invoice_reminder(reminder_date, email, subject, "sent", "", len(jobs))
            result["sent"] += 1
        except Exception as exc:
            log_billing_invoice_reminder(reminder_date, email, subject, "failed", exc, len(jobs))
            result["failed"] += 1
    result["message"] = f"Found {len(jobs)} job(s). Sent {result['sent']}, failed {result['failed']}, skipped {result['skipped']}."
    return result


def texas_upload_reminder_scheduler():
    while True:
        try:
            now = datetime.now()
            settings = texas_upload_reminder_settings()
            reminder_hour, reminder_minute = [int(x) for x in settings["reminder_time"].split(":", 1)]
            is_due_day = now.weekday() == int(settings["weekday"])
            is_due_time = now.hour > reminder_hour or (now.hour == reminder_hour and now.minute >= reminder_minute)
            if int(settings.get("enabled") or 0) and is_due_day and is_due_time:
                run_texas_upload_reminder(force=False)
            billing_settings = billing_invoice_reminder_settings()
            billing_hour, billing_minute = [int(x) for x in billing_settings["reminder_time"].split(":", 1)]
            billing_due_day = now.weekday() == int(billing_settings["weekday"])
            billing_due_time = now.hour > billing_hour or (now.hour == billing_hour and now.minute >= billing_minute)
            if int(billing_settings.get("enabled") or 0) and billing_due_day and billing_due_time:
                run_billing_invoice_reminder(force=False)
            po_settings = po_untouched_reminder_settings()
            po_hour, po_minute = [int(x) for x in po_settings["reminder_time"].split(":", 1)]
            po_due_day = now.weekday() == int(po_settings["weekday"])
            po_due_time = now.hour > po_hour or (now.hour == po_hour and now.minute >= po_minute)
            if int(po_settings.get("enabled") or 0) and po_due_day and po_due_time:
                run_po_untouched_reminder(force=False)
        except Exception:
            traceback.print_exc()
        time.sleep(60)


def start_background_jobs():
    thread = threading.Thread(target=texas_upload_reminder_scheduler, daemon=True)
    thread.start()


def import_fieldwise_xlsx(path, project_id):
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed. Run this app with the bundled Codex Python or install openpyxl.")

    wb = openpyxl.load_workbook(path, data_only=True)
    header = {}
    if "Header" in wb.sheetnames:
        ws = wb["Header"]
        for row in ws.iter_rows(values_only=True):
            if row and row[0]:
                header[str(row[0]).strip()] = row[1] if len(row) > 1 else None

    if "Line Items" not in wb.sheetnames:
        raise RuntimeError("Expected a 'Line Items' sheet in the Field Wise export.")

    order_number = str(header.get("Order #") or "").strip()
    matched_subproject = None
    matched_change_order = None
    if order_number:
        matched_change_order = one(
            "SELECT id, subproject_id FROM change_orders WHERE project_id = ? AND job_number = ? ORDER BY id LIMIT 1",
            (project_id, order_number),
        )
        matched_subproject = one(
            "SELECT id FROM subprojects WHERE project_id = ? AND job_number = ?",
            (project_id, order_number),
        )
        if not matched_subproject:
            matched_subproject = one(
                """
                SELECT id FROM subprojects
                WHERE project_id = ?
                  AND code <> ''
                  AND instr(upper(?), upper(code)) > 0
                ORDER BY length(code) DESC
                LIMIT 1
                """,
                (project_id, order_number),
            )
    matched_subproject_id = matched_subproject["id"] if matched_subproject else None
    matched_change_order_id = matched_change_order["id"] if matched_change_order else None
    if matched_change_order and matched_change_order["subproject_id"]:
        matched_subproject_id = matched_change_order["subproject_id"]
    matched_subproject_is_tm = False
    if matched_subproject_id:
        sp_rate_row = one("SELECT COALESCE(pricing_type, 'Fixed') pricing_type FROM subprojects WHERE id = ?", (matched_subproject_id,))
        matched_subproject_is_tm = bool(sp_rate_row and sp_rate_row["pricing_type"] == "T&M")

    ws = wb["Line Items"]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    index = {name: i for i, name in enumerate(headers)}
    required = ["Ticket", "Date", "Status", "Type", "Item", "Description", "Qty", "Rate", "Total"]
    missing = [h for h in required if h not in index]
    if missing:
        raise RuntimeError("Missing Field Wise columns: " + ", ".join(missing))

    source_file = Path(path).name
    count = 0
    skipped = 0
    with db() as con:
        if matched_subproject_id:
            sp_rate_row = con.execute("SELECT COALESCE(pricing_type, 'Fixed') pricing_type FROM subprojects WHERE id = ?", (matched_subproject_id,)).fetchone()
            matched_subproject_is_tm = bool(sp_rate_row and sp_rate_row["pricing_type"] == "T&M")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            date_value = row[index["Date"]]
            if isinstance(date_value, datetime):
                record_date = date_value.date().isoformat()
            else:
                record_date = str(date_value or "")
            type_value = str(row[index["Type"]] or "").strip()
            cost_type = "Labor" if "time" in type_value.lower() else "Field Ticket Material" if "material" in type_value.lower() else "Equipment" if "equipment" in type_value.lower() else type_value
            ticket = str(row[index["Ticket"]] or "")
            item = str(row[index["Item"]] or "")
            description = str(row[index["Description"]] or "")
            qty = money(row[index["Qty"]])
            sales_rate = money(row[index["Rate"]])
            sales_amount = money(row[index["Total"]])
            rate_info = raw_rate_for(cost_type, item, project_id)
            raw_rate = rate_info["raw_rate"]
            if cost_type == "Field Ticket Material":
                accrue_material = bool(matched_change_order_id or matched_subproject_is_tm)
                raw_rate = sales_rate * CO_MATERIAL_COST_FACTOR if accrue_material else 0
                amount = sales_amount * CO_MATERIAL_COST_FACTOR if accrue_material else 0
                rate_info = {
                    "source": "CO T&M material estimate at 35% margin" if matched_change_order_id else
                    "Subproject T&M material estimate at 35% margin" if matched_subproject_is_tm else
                    "Usage only - not budget cost"
                }
            else:
                amount = qty * raw_rate if raw_rate else sales_amount
            duplicate = con.execute(
                """
                SELECT id FROM cost_records
                WHERE project_id = ?
                  AND source = 'Field Wise'
                  AND ticket_or_invoice = ?
                  AND record_date = ?
                  AND cost_type = ?
                  AND item = ?
                  AND description = ?
                  AND sales_amount = ?
                LIMIT 1
                """,
                (project_id, ticket, record_date, cost_type, item, description, sales_amount),
            ).fetchone()
            if duplicate:
                skipped += 1
                continue
            con.execute(
                """
                INSERT INTO cost_records (
                  project_id, source, source_file, ticket_or_invoice, record_date, status,
                  subproject_id, change_order_id, cost_type, item, description, qty, rate, amount, sales_rate, sales_amount, raw_rate, raw_cost_source, notes, created_at
                )
                VALUES (?, 'Field Wise', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_id,
                    source_file,
                    ticket,
                    record_date,
                    str(row[index["Status"]] or ""),
                    matched_subproject_id,
                    matched_change_order_id,
                    cost_type,
                    item,
                    description,
                    qty,
                    raw_rate if raw_rate else sales_rate,
                    amount,
                    sales_rate,
                    sales_amount,
                    raw_rate,
                    rate_info["source"],
                    json.dumps(header, default=str),
                    datetime.now().isoformat(timespec="seconds"),
                ),
            )
            count += 1
    wb.close()
    return {"count": count, "skipped": skipped, "order_number": order_number, "matched_subproject_id": matched_subproject_id, "matched_change_order_id": matched_change_order_id}


def import_fieldwise_pdf(path, project_id):
    if pdfplumber is None:
        raise RuntimeError("PDF import needs pdfplumber, but it is not available in this Python runtime.")

    pages = []
    tables = []
    with pdfplumber.open(path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            pages.append(page.extract_text() or "")
            for table in page.extract_tables() or []:
                tables.append((page_number, table))
    text = "\n".join(pages).strip()
    if not text:
        raise RuntimeError("No readable text was found in this PDF. It may be a scanned image PDF.")

    def table_value(label):
        for _, table in tables:
            for row in table:
                if row and len(row) > 1 and str(row[0] or "").strip().lower() == label.lower():
                    return str(row[1] or "").strip()
        return ""

    job_text = table_value("Job #") or first_regex([r"Job\s*#\s+(.+)"], text)
    order_number = first_regex([r"^([A-Za-z0-9\-]+)"], job_text, flags=0)
    ticket_number = first_regex([r"Field\s*Ticket\s*#\s*([A-Za-z0-9\-]+)"], text)
    record_date = table_value("Ticket Date") or first_regex([r"Ticket\s*Date\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})"], text)
    status = "Imported"

    work_description = first_regex([r"Work\s*Description\s*\n(.+?)\nLabor"], text, flags=re.IGNORECASE | re.DOTALL)
    work_description = re.sub(r"\s+", " ", work_description).strip()
    if not work_description:
        work_description = "Imported Field Wise PDF ticket"

    matched_subproject = None
    matched_change_order = None
    if order_number:
        matched_change_order = one(
            "SELECT id, subproject_id FROM change_orders WHERE project_id = ? AND job_number = ? ORDER BY id LIMIT 1",
            (project_id, order_number),
        )
        matched_subproject = one(
            "SELECT id FROM subprojects WHERE project_id = ? AND job_number = ?",
            (project_id, order_number),
        )
        if not matched_subproject:
            matched_subproject = one(
                """
                SELECT id FROM subprojects
                WHERE project_id = ?
                  AND code <> ''
                  AND instr(upper(?), upper(code)) > 0
                ORDER BY length(code) DESC
                LIMIT 1
                """,
                (project_id, order_number),
            )
    matched_subproject_id = matched_subproject["id"] if matched_subproject else None
    matched_change_order_id = matched_change_order["id"] if matched_change_order else None
    if matched_change_order and matched_change_order["subproject_id"]:
        matched_subproject_id = matched_change_order["subproject_id"]
    matched_subproject_is_tm = False
    if matched_subproject_id:
        sp_rate_row = one("SELECT COALESCE(pricing_type, 'Fixed') pricing_type FROM subprojects WHERE id = ?", (matched_subproject_id,))
        matched_subproject_is_tm = bool(sp_rate_row and sp_rate_row["pricing_type"] == "T&M")

    source_file = Path(path).name
    records = []
    seen_extracted_records = set()

    def add_extracted_record(record):
        key = (
            record["cost_type"],
            record["item"],
            record["description"],
            record["qty"],
            record["rate"],
            record["amount"],
        )
        if key in seen_extracted_records:
            return
        seen_extracted_records.add(key)
        records.append(record)

    def looks_like_fieldwise_item_row(row):
        if len(row) < 5 or not row[0]:
            return False
        qty = money(row[2] if len(row) > 2 else 0)
        rate = parse_money_text(row[3] if len(row) > 3 else 0)
        amount = parse_money_text(row[4] if len(row) > 4 else 0)
        price_text = f"{row[3] if len(row) > 3 else ''} {row[4] if len(row) > 4 else ''}"
        if "$" not in price_text and not amount:
            return False
        return bool(qty or rate or amount)

    def add_fieldwise_item_row(row, page_number):
        if not looks_like_fieldwise_item_row(row):
            return
        item_name = str(row[0] or "").strip()
        is_equipment = is_equipment_item(item_name)
        add_extracted_record(
            {
                "cost_type": "Equipment" if is_equipment else "Field Ticket Material",
                "item": item_name,
                "description": str(row[1] or work_description).strip(),
                "qty": money(row[2] if len(row) > 2 else 0),
                "rate": parse_money_text(row[3] if len(row) > 3 else 0),
                "amount": parse_money_text(row[4] if len(row) > 4 else 0),
                "source_page": page_number,
            }
        )

    for page_number, table in tables:
        if not table:
            continue
        header = [str(c or "").strip().lower() for c in table[0]]
        if header == ["work type / comp item", "qty", "rate", "total"]:
            for row in table[1:]:
                if not row or not row[0]:
                    continue
                add_extracted_record(
                    {
                        "cost_type": "Labor",
                        "item": str(row[0] or "").strip(),
                        "description": work_description,
                        "qty": money(row[1] if len(row) > 1 else 0),
                        "rate": parse_money_text(row[2] if len(row) > 2 else 0),
                        "amount": parse_money_text(row[3] if len(row) > 3 else 0),
                        "source_page": page_number,
                    }
                )
        elif header[:5] == ["item", "item description", "qty", "rate", "amount"]:
            for row in table[1:]:
                add_fieldwise_item_row(row, page_number)
        elif looks_like_fieldwise_item_row(table[0]):
            for row in table:
                add_fieldwise_item_row(row, page_number)
        else:
            for row in table:
                add_fieldwise_item_row(row, page_number)

    if not records:
        amount_text = first_regex([r"TOTAL\s*\$?\s*([0-9,]+\.[0-9]{2})"], text)
        records.append(
            {
                "cost_type": "Uncoded",
                "item": "PDF Field Ticket",
                "description": work_description,
                "qty": 1,
                "rate": parse_money_text(amount_text),
                "amount": parse_money_text(amount_text),
                "source_page": None,
            }
        )

    count = 0
    skipped = 0
    seen_record_counts = {}
    for record in records:
        rate_info = raw_rate_for(record["cost_type"], record["item"], project_id)
        sales_rate = money(record["rate"])
        sales_amount = money(record["amount"])
        raw_rate = rate_info["raw_rate"]
        if record["cost_type"] == "Field Ticket Material":
            accrue_material = bool(matched_change_order_id or matched_subproject_is_tm)
            raw_rate = sales_rate * CO_MATERIAL_COST_FACTOR if accrue_material else 0
            raw_amount = sales_amount * CO_MATERIAL_COST_FACTOR if accrue_material else 0
            rate_info = {
                "category": "",
                "raw_rate": raw_rate,
                "source": "CO T&M material estimate at 35% margin" if matched_change_order_id else
                "Subproject T&M material estimate at 35% margin" if matched_subproject_is_tm else
                "Usage only - not budget cost",
            }
        else:
            raw_amount = record["qty"] * raw_rate if raw_rate else sales_amount
        duplicate_key = (
            ticket_number,
            record["cost_type"],
            record["item"],
            record["description"][:500],
            record["qty"],
            sales_rate,
            sales_amount,
        )
        seen_record_counts[duplicate_key] = seen_record_counts.get(duplicate_key, 0) + 1
        existing_count = one(
            """
            SELECT COUNT(*) AS existing_count FROM cost_records
            WHERE project_id = ?
              AND source = 'Field Wise PDF'
              AND ticket_or_invoice = ?
              AND cost_type = ?
              AND item = ?
              AND description = ?
              AND qty = ?
              AND sales_rate = ?
              AND sales_amount = ?
            """,
            (project_id, ticket_number, record["cost_type"], record["item"], record["description"][:500], record["qty"], sales_rate, sales_amount),
        )
        if existing_count and existing_count["existing_count"] >= seen_record_counts[duplicate_key]:
            skipped += 1
            continue
        execute(
            """
            INSERT INTO cost_records (
              project_id, subproject_id, change_order_id, source, source_file, ticket_or_invoice, record_date, status,
              cost_type, item, description, qty, rate, amount, sales_rate, sales_amount, raw_rate, raw_cost_source, notes, created_at
            )
            VALUES (?, ?, ?, 'Field Wise PDF', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                project_id,
                matched_subproject_id,
                matched_change_order_id,
                source_file,
                ticket_number,
                record_date,
                status,
                record["cost_type"],
                record["item"],
                record["description"][:500],
                record["qty"],
                raw_rate if raw_rate else sales_rate,
                raw_amount,
                sales_rate,
                sales_amount,
                raw_rate,
                rate_info["source"],
                json.dumps({"job_text": job_text, "rate_category": rate_info["category"], "source_page": record.get("source_page"), "extracted_text": text[:12000]}, default=str),
                datetime.now().isoformat(timespec="seconds"),
            ),
        )
        count += 1
    return {"count": count, "skipped": skipped, "order_number": order_number, "matched_subproject_id": matched_subproject_id, "matched_change_order_id": matched_change_order_id}


def import_vendor_invoice_pdf(path, project_id):
    if pdfplumber is None:
        raise RuntimeError("Vendor invoice PDF import needs pdfplumber, but it is not available.")
    text = extract_pdf_text(path)
    if not text:
        raise RuntimeError("No readable text was found in this vendor invoice PDF. It may need OCR or manual entry.")

    source_file = Path(path).name
    lower = text.lower()
    compact_lower = re.sub(r"\s+", "", lower)
    if "dsgsupply" in lower or "dakota supply group" in lower or "dsg truck delivery" in lower or "dsg#" in lower:
        vendor = "Dakota Supply Group"
        item_lines = []
        if ("invoicesummary" in compact_lower or "invoice summary" in lower) and "qty" in lower and "subtotal" in lower:
            invoice_number = first_regex([r"Invoice#\s*:?\s*([A-Za-z0-9.\-]+)"], text)
            invoice_date = first_regex([r"Invoice Date\s*:?\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})"], text)
            order_number = first_regex([r"PO number\s*:?\s*([A-Za-z0-9\-]+)"], text)
            total_due = parse_money_text(first_regex([r"Total\s*\n?\s*\$?([0-9,]+\.[0-9]{2})"], text))
            item_lines = parse_dsg_online_invoice_ocr(path)
            lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
            i = 0
            while not item_lines and i < len(lines):
                if i + 7 < len(lines) and re.search(r"QTY\s*Invoiced", lines[i + 2], flags=re.IGNORECASE):
                    manufacturer = lines[i]
                    description = lines[i + 1]
                    qty = first_regex([r"^([0-9.]+)$"], lines[i + 3], flags=0)
                    detail_parts = []
                    j = i + 4
                    while j < len(lines) and not re.match(r"^\$?[0-9,.]+/[A-Za-z]+$", lines[j]) and lines[j].lower() != "subtotal":
                        detail_parts.append(lines[j])
                        j += 1
                    if j + 2 < len(lines):
                        unit_price = parse_money_text(lines[j])
                        amount = parse_money_text(lines[j + 2])
                        item_lines.append({
                            "product_code": description.split()[0] if description.split() else manufacturer,
                            "description": " ".join([description] + detail_parts).strip(),
                            "qty": qty,
                            "unit_price": unit_price,
                            "amount": amount,
                        })
                        i = j + 3
                        continue
                i += 1
        else:
            header = re.search(r"INVOICE DATE\s+INVOICE NUMBER\s*\n([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})\s+([A-Za-z0-9.\-]+)", text)
            invoice_date = header.group(1).strip() if header else ""
            invoice_number = header.group(2).strip() if header else ""
            order_number = first_regex([r"CUSTOMER NUMBER\s+CUSTOMER PO NUMBER.*?\n[0-9]+\s+([0-9]{3,})\s+"], text, flags=re.IGNORECASE | re.DOTALL)
            total_due = parse_money_text(first_regex([r"Amount Due\s+\$?([0-9,]+\.[0-9]{2})"], text))
            block_match = re.search(r"ORDER QTY\s+SHIP QTY\s+DESCRIPTION\s+UNIT PRICE\s+EXT PRICE\s*\n(.+?)\nONLINE BILLPAY", text, flags=re.IGNORECASE | re.DOTALL)
            if block_match:
                current = None
                for raw_line in block_match.group(1).splitlines():
                    line = raw_line.strip()
                    if not line:
                        continue
                    match = re.match(r"^([0-9.]+)\s*([A-Za-z]+)\s+([0-9.]+)\s*([A-Za-z]+)\s+(.+?)\s+([0-9,]+\.[0-9]+/[A-Za-z]+)\s+([0-9,]+\.[0-9]{2})$", line)
                    if match:
                        if current:
                            item_lines.append(current)
                        ordered_qty, ordered_unit, shipped_qty, shipped_unit, description, unit_text, extension = match.groups()
                        current = {
                            "ordered_qty": ordered_qty,
                            "ordered_unit": ordered_unit,
                            "shipped_qty": shipped_qty,
                            "shipped_unit": shipped_unit,
                            "product_code": description.split()[0] if description.split() else description,
                            "description": description,
                            "unit_text": unit_text,
                            "extension": extension,
                        }
                    elif current:
                        current["description"] += " " + line
                if current:
                    item_lines.append(current)
    elif "vega americas" in lower:
        vendor = "VEGA Americas"
        invoice_number = first_regex([r"Invoice\s+No\.\s*([A-Za-z0-9\-]+)"], text, flags=re.IGNORECASE)
        invoice_date = first_regex([r"\bDate:\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})"], text, flags=re.IGNORECASE)
        order_number = first_regex([r"Purchase Order:\s*(.+)"], text, flags=re.IGNORECASE)
        total_due = parse_money_text(first_regex([r"Sales Total USD\s+([0-9,]+\.[0-9]{2})"], text))
        item_lines = []
        current = None
        stop_prefixes = (
            "Unit Price", "Pos.", "Carried over", "Sales Total", "Shipping cost",
            "Tariff Surcharge", "Transaction Information", "Order No.", "Packing List",
            "Net Due Date", "Payment Terms", "Ship Date", "Incoterm", "Ship To Address",
            "Contact Information", "ORIGINAL", "- Page", "VEGA Americas"
        )
        for raw_line in text.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            match = re.match(r"^([0-9]+)\s+([0-9.]+)\s+(.+?)\s+([0-9,]+\.[0-9]{2})\s+([0-9,]+\.[0-9]{2})$", line)
            if match:
                if current:
                    item_lines.append(current)
                position, qty, description, unit_price, extension = match.groups()
                current = {
                    "product_code": description.split()[0] if description.split() else description,
                    "description": description,
                    "qty": qty,
                    "unit_price": unit_price,
                    "amount": extension,
                }
                continue
            if current:
                if line.startswith(stop_prefixes):
                    item_lines.append(current)
                    current = None
                elif not re.search(r"^(Serial Number|Order Position|Packing List Position)", line, flags=re.IGNORECASE):
                    current["description"] += " " + line
        if current:
            item_lines.append(current)
        shipping = parse_money_text(first_regex([r"Shipping cost USD:\s*([0-9,]+\.[0-9]{2})"], text, flags=re.IGNORECASE))
        tariff = parse_money_text(first_regex([r"Tariff Surcharge USD:\s*([0-9,]+\.[0-9]{2})"], text, flags=re.IGNORECASE))
        if shipping:
            item_lines.append({"product_code": "Shipping", "description": "Shipping cost", "qty": 1, "unit_price": shipping, "amount": shipping})
        if tariff:
            item_lines.append({"product_code": "Tariff", "description": "Tariff surcharge", "qty": 1, "unit_price": tariff, "amount": tariff})
    elif "primec controls" in lower:
        vendor = "Prime Controls LLC"
        invoice_number = first_regex([r"INVOICE\s+PAGE:\s*[0-9]+\s*\n\s*([A-Za-z0-9\-]+)\s*\n\s*Invoice#"], text, flags=re.IGNORECASE)
        invoice_date = first_regex([r"Invoice#\s*\n\s*([0-9]{1,2}-[0-9]{1,2}-[0-9]{2,4})\s*\n\s*InvoiceDate"], text, flags=re.IGNORECASE)
        order_number = first_regex([r"SALESPERSON\s+.+?\s+PO\s+(.+)", r"\bFACILITY\s+(.+)"], text, flags=re.IGNORECASE)
        total_due = parse_money_text(first_regex([r"SUBTOTAL\s+([0-9,]+\.[0-9]{2})", r"INVOICETOTAL\s*\n\s*([0-9,]+\.[0-9]{2})"], text, flags=re.IGNORECASE))
        item_lines = []
        current = None
        block_match = re.search(r"Item\s+Description\s+Ordered\s+Shipped\s+UM\s+UnitPrice\s+Extension\s*\n(.+?)\nLast Page", text, flags=re.IGNORECASE | re.DOTALL)
        if block_match:
            for raw_line in block_match.group(1).splitlines():
                line = raw_line.strip()
                if not line:
                    continue
                match = re.match(r"^([A-Za-z0-9\-]+)\s+(.+?)\s+([0-9.]+)\s+([0-9.]+)\s+([A-Za-z]+)\s+([0-9,]+\.[0-9]{2})\s+([0-9,]+\.[0-9]{2})$", line)
                if match:
                    if current:
                        item_lines.append(current)
                    product_code, description, ordered_qty, shipped_qty, unit, unit_price, extension = match.groups()
                    current = {
                        "product_code": product_code,
                        "description": f"{product_code} {description}",
                        "qty": shipped_qty or ordered_qty,
                        "unit_price": unit_price,
                        "amount": extension,
                    }
                elif current:
                    current["description"] += " " + line
            if current:
                item_lines.append(current)
    elif "ced williston" in lower or "cedwilliston" in compact_lower:
        vendor = "CED Williston"
        invoice_header = re.search(r"([0-9]{3,}\s*-\s*[0-9]{3,})\s+([0-9]{2}/[0-9]{2}/[0-9]{2,4})", text)
        invoice_number = invoice_header.group(1).strip() if invoice_header else first_regex([r"\b([0-9]{3,}\s*-\s*[0-9]{3,})\b"], text)
        invoice_date = invoice_header.group(2).strip() if invoice_header else first_regex(
            [
                r"INVOICE\s+DATE\s*\n\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})",
                r"INVOICE\s+NO\.\s*\n\s*[0-9]{3,}\s*-\s*[0-9]{3,}\s*\n\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})",
                r"SHIP\s+DATE\s*\n\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})",
            ],
            text,
            flags=re.IGNORECASE,
        )
        if not invoice_date and invoice_number:
            invoice_date = first_regex([re.escape(invoice_number) + r".*?([0-9]{2}/[0-9]{2}/[0-9]{4})", re.escape(invoice_number) + r".*?([0-9]{2}/[0-9]{2}/[0-9]{2})"], text, flags=re.IGNORECASE | re.DOTALL)
        invoice_date = date_text(invoice_date)
        order_number = first_regex(
            [
                r"\b(HUNT-[0-9]+)\b",
                r"CUSTOMER\s+ORDER\s+NO\.\s*\n\s*([A-Za-z0-9\-]+)",
                r"CUSTOMERORDERNO\.\s*\n\s*([A-Za-z0-9\-]+)",
                r"JOB\s+NAME\s+CUSTOMER\s+ORDER\s+NO\.\s*\n.*?\b([A-Za-z0-9]+-[A-Za-z0-9]+)\b",
                r"CUSTOMER\s*\nORDER\s*\n.*?([A-Za-z0-9\-]+)\s*\nSALES PERSON",
                r"\bORDER\s*\n.*?([0-9]{3,})",
            ],
            text,
            flags=re.IGNORECASE | re.DOTALL,
        )
        total_due = parse_money_text(first_regex([r"TOTAL\s*DUE\s*[-=]*>?\s*([0-9,]+\.[0-9]{2})", r"TOTALDUE\s*\n?\s*([0-9,]+\.[0-9]{2})"], text, flags=re.IGNORECASE))
        if not total_due:
            money_values = [parse_money_text(value) for value in re.findall(r"\b[0-9,]+\.[0-9]{2}\b", text)]
            if money_values:
                total_due = money_values[-1]
        item_lines = []
        current = None
        stop_re = re.compile(r"^(TITLE TO|AT POINT|MERCHANDISE|A SERVICE|THIS SALE|CODE:|B -|C -|SALES TAX|SHIPPING CHARGE|TOTAL DUE)\b", re.IGNORECASE)
        for raw_line in text.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            match = re.match(
                r"^T\s+([0-9.]+)\s+(.+?)\s+([0-9.]+)\s+([0-9,]+\.[0-9]{2})\s+([A-Za-z])\s+([0-9,]+\.[0-9]{2})$",
                line,
            )
            if match:
                if current:
                    item_lines.append(current)
                ordered_qty, product_description, shipped_qty, quoted_price, price_code, extension = match.groups()
                price_code = price_code.upper()
                parts = product_description.split()
                product_code = " ".join(parts[:2]) if len(parts) > 1 else product_description
                current = {
                    "product_code": product_code,
                    "description": product_description,
                    "qty": shipped_qty or ordered_qty,
                    "unit_price": (parse_money_text(extension) / money(shipped_qty or ordered_qty)) if money(shipped_qty or ordered_qty) else 0,
                    "amount": extension,
                    "ced_price_code": price_code,
                    "ced_quoted_price": quoted_price,
                }
                continue
            match = re.match(
                r"^T?\s*([0-9.]+)\s+(.+?)\s+([0-9.]+)\s+([0-9,]+\.[0-9]{2})\s+([A-Za-z])\s+([0-9,]+\.[0-9]{2})$",
                line,
            )
            if match:
                if current:
                    item_lines.append(current)
                ordered_qty, product_description, shipped_qty, quoted_price, price_code, extension = match.groups()
                parts = product_description.split()
                product_code = " ".join(parts[:2]) if len(parts) > 1 else product_description
                qty = money(shipped_qty or ordered_qty)
                amount = parse_money_text(extension)
                current = {
                    "product_code": product_code,
                    "description": product_description,
                    "qty": qty,
                    "unit_price": amount / qty if qty else parse_money_text(quoted_price),
                    "amount": amount,
                    "ced_price_code": price_code.upper(),
                    "ced_quoted_price": quoted_price,
                }
                continue
            if current and stop_re.search(line):
                item_lines.append(current)
                current = None
                break
            if current and not stop_re.search(line) and not re.search(r"^(INVOICE|QUANTITY|PRODUCT CODE|ORDERED|SOLD TO|SHIP TO|ACCOUNT NO\.|SALES PERSON)\b", line, flags=re.IGNORECASE):
                current["description"] += " " + line
        if current:
            item_lines.append(current)
        if not item_lines:
            lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
            for i, line in enumerate(lines):
                start = re.match(r"^([0-9]+)([A-Z][A-Z0-9\-]*)$", line)
                if not start or i + 3 >= len(lines):
                    continue
                ordered_qty, maker = start.groups()
                description = lines[i + 1]
                shipped_qty = money(lines[i + 2])
                unit_price = parse_money_text(lines[i + 3])
                if not shipped_qty or not unit_price:
                    continue
                product_code_tail = ""
                for later in lines[i + 4:i + 8]:
                    if re.fullmatch(r"[A-Z0-9][A-Z0-9\-]{2,}", later) and not re.fullmatch(r"[A-Z]", later):
                        product_code_tail = later
                        break
                item_lines.append({
                    "product_code": " ".join([maker, product_code_tail]).strip(),
                    "description": description,
                    "qty": shipped_qty or money(ordered_qty),
                    "unit_price": unit_price,
                    "amount": (shipped_qty or money(ordered_qty)) * unit_price,
                })
                break
    elif "border states" in lower:
        vendor = "Border States"
        invoice_number = first_regex([r"Invoice:\s*([A-Za-z0-9\-]+)"], text, flags=re.IGNORECASE)
        invoice_date = first_regex([r"Invoice:\s*[A-Za-z0-9\-]+\s+Date:\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})"], text, flags=re.IGNORECASE)
        order_number = first_regex([r"P\.O\.#:\s*(.+)"], text, flags=re.IGNORECASE)
        total_due = parse_money_text(first_regex([r"Net Invoice Amount\s+\$\s*([0-9,]+\.[0-9]{2})", r"Total\s+\$\s*([0-9,]+\.[0-9]{2})"], text, flags=re.IGNORECASE))
        item_lines = []
        current = None
        for raw_line in text.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            start = re.match(r"^([0-9]{6})\s+([A-Za-z0-9\-]+)\s+([0-9,]+(?:\.[0-9]+)?)$", line)
            if start:
                if current:
                    item_lines.append(current)
                line_no, material, ordered_qty = start.groups()
                current = {
                    "product_code": material,
                    "description": material,
                    "ordered_qty": ordered_qty,
                    "qty": ordered_qty,
                    "unit_price": 0,
                    "amount": 0,
                }
                continue
            if not current:
                continue
            batch = re.match(
                r"^Batch Total:\s*([0-9,]+(?:\.[0-9]+)?)\s+([A-Za-z]+)\s+([0-9,]+(?:\.[0-9]+)?)\s+([0-9,]+\.[0-9]{2})\s*/\s*([0-9,]+(?:\.[0-9]+)?)\s+([A-Za-z]+)\s+([0-9,]+\.[0-9]{2})$",
                line,
                flags=re.IGNORECASE,
            )
            if batch:
                batch_qty, batch_uom, ship_qty, quoted_price, price_per_qty, price_per_uom, extension = batch.groups()
                qty = parse_money_text(ship_qty or batch_qty or current.get("ordered_qty"))
                amount = parse_money_text(extension)
                current["qty"] = qty
                current["unit_price"] = amount / qty if qty else parse_money_text(quoted_price)
                current["amount"] = amount
                current["description"] = re.sub(r"\s+", " ", current["description"]).strip()
                item_lines.append(current)
                current = None
                continue
            if re.search(r"^(Cust Material #|Batch:|OD |Williston stocked)\b", line, flags=re.IGNORECASE):
                current["description"] += " " + line
            elif not re.search(r"^(INVOICE|Page |Invoice:|Cust Acct|Cash discount|Total due|Shipping and Handling|State Tax|County Tax|Local Tax|Other Tax|Tax Subtotal|Net Invoice Amount|ORIGINAL REPRINT)\b", line, flags=re.IGNORECASE):
                current["description"] += " " + line
        if current and parse_money_text(current.get("amount")):
            item_lines.append(current)
    elif "mccody" in lower or "fromthedirtup" in lower:
        vendor = "McCody Concrete Products"
        header = re.search(
            r"DATE\s+INVOICE\s*#\s*\n(?:.*\n){0,3}?\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})\s+([A-Za-z0-9\-]+)",
            text,
            flags=re.IGNORECASE,
        )
        invoice_date = date_text(header.group(1).strip()) if header else first_regex([r"\b([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})\b"], text)
        invoice_number = header.group(2).strip() if header else ""
        if not invoice_number:
            invoice_number = first_regex([r"Inv_([A-Za-z0-9\-]+)_from_MCCODY"], source_file, flags=re.IGNORECASE)
        order_number = first_regex([r"JOB\s*/\s*PO\s*#:\s*([A-Za-z0-9\-]+)"], text, flags=re.IGNORECASE)
        total_due = parse_money_text(first_regex([r"BALANCE\s+DUE\s+\$?([0-9,]+\.[0-9]{2})", r"TOTAL\s+\$?([0-9,]+\.[0-9]{2})"], text, flags=re.IGNORECASE))
        item_lines = []
        for raw_line in text.splitlines():
            line = raw_line.strip()
            match = re.match(r"^(.+?)\s+([0-9]+(?:\.[0-9]+)?)\s+([A-Za-z]+)\s+(.+?)\s+([0-9,]+\.[0-9]{2})\s+([0-9,]+\.[0-9]{2})T?$", line)
            if match and "SUBTOTAL" not in line.upper():
                item_name, qty, unit, item_id, unit_price, extension = match.groups()
                item_lines.append({
                    "product_code": item_id.split()[0] if item_id.split() else item_name[:40],
                    "description": item_name,
                    "qty": qty,
                    "unit_price": unit_price,
                    "amount": extension,
                })
                break
    else:
        vendor = first_regex([r"^([A-Z][A-Za-z0-9 &.\-]+)\s*\n"], text, flags=re.MULTILINE) or "Vendor"
        invoice_number = first_regex([r"Invoice\s*(?:No\.?|#)\s*:?\s*([A-Za-z0-9\-]+)"], text)
        invoice_date = first_regex([r"Invoice\s*Date\s*:?\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})"], text)
        order_number = first_regex([r"(?:PO|Purchase Order|Customer PO)(?:\s*number|\s*#)?\s*:?\s*([A-Za-z0-9\-]+)"], text)
        total_due = parse_money_text(first_regex([r"(?:Total|Amount Due)\s*:?\s*\$?\s*([0-9,]+\.[0-9]{2})"], text))
        item_lines = []

    matched_subproject = None
    if order_number:
        matched_subproject = one(
            "SELECT id FROM subprojects WHERE project_id = ? AND job_number = ?",
            (project_id, order_number),
        )
        if not matched_subproject:
            matched_subproject = one(
                """
                SELECT id FROM subprojects
                WHERE project_id = ?
                  AND code <> ''
                  AND instr(upper(?), upper(code)) > 0
                ORDER BY length(code) DESC
                LIMIT 1
                """,
                (project_id, order_number),
            )
    matched_subproject_id = matched_subproject["id"] if matched_subproject else None

    duplicate_invoice = None
    if invoice_number:
        duplicate_invoice = one(
            """
            SELECT COUNT(*) AS line_count, COALESCE(SUM(amount), 0) AS total_amount, MAX(source_file) AS source_file
            FROM cost_records
            WHERE project_id = ?
              AND source = 'Vendor Invoice'
              AND ticket_or_invoice = ?
              AND (? = '' OR vendor = ?)
            """,
            (project_id, invoice_number, vendor or "", vendor or ""),
        )
        if duplicate_invoice and duplicate_invoice["line_count"]:
            return {
                "count": 0,
                "skipped": duplicate_invoice["line_count"],
                "order_number": order_number,
                "matched_subproject_id": matched_subproject_id,
                "vendor": vendor,
                "invoice_number": invoice_number,
                "duplicate": True,
                "existing_line_count": duplicate_invoice["line_count"],
                "existing_total": duplicate_invoice["total_amount"],
                "existing_source_file": duplicate_invoice["source_file"],
            }

    count = 0
    skipped = 0
    if item_lines:
        for line in item_lines:
            if isinstance(line, dict):
                product_code = line["product_code"]
                description = line["description"]
                if "qty" in line:
                    qty = money(line["qty"])
                    amount = parse_money_text(line["amount"])
                    unit_price = parse_money_text(line.get("unit_price")) if line.get("unit_price") is not None else (amount / qty if qty else 0)
                else:
                    qty = money(line["shipped_qty"] or line["ordered_qty"])
                    amount = parse_money_text(line["extension"])
                    unit_price = amount / qty if qty else 0
            elif len(line) == 6:
                ordered_qty, product_code, description, shipped_qty, per_hundred_price, extension = line
                qty = money(shipped_qty or ordered_qty)
                amount = parse_money_text(extension)
                unit_price = amount / qty if qty else 0
            else:
                ordered_qty, product_description, shipped_qty, per_hundred_price, extension = line
                parts = product_description.split()
                product_code = " ".join(parts[:2]) if len(parts) > 1 else product_description
                description = " ".join(parts[2:]) if len(parts) > 2 else product_description
                qty = money(shipped_qty or ordered_qty)
                amount = parse_money_text(extension)
                unit_price = amount / qty if qty else 0
            execute(
                """
                INSERT INTO cost_records (
                  project_id, subproject_id, source, source_file, ticket_or_invoice, record_date, status,
                  cost_type, item, description, qty, rate, amount, sales_rate, sales_amount, raw_rate, raw_cost_source,
                  vendor, notes, created_at
                )
                VALUES (?, ?, 'Vendor Invoice', ?, ?, ?, 'Imported', 'Material', ?, ?, ?, ?, ?, 0, 0, ?, 'Vendor invoice actual material cost', ?, ?, ?)
                """,
                (
                    project_id,
                    matched_subproject_id,
                    source_file,
                    invoice_number,
                    invoice_date,
                    product_code,
                    description.strip(),
                    qty,
                    unit_price,
                    amount,
                    unit_price,
                    vendor,
                    json.dumps({"order_number": order_number, "text": text[:12000]}, default=str),
                    datetime.now().isoformat(timespec="seconds"),
                ),
            )
            count += 1
    else:
        if not total_due:
            raise RuntimeError("Could not find invoice line items or total due in this vendor invoice.")
        duplicate = one(
            "SELECT id FROM cost_records WHERE project_id = ? AND source = 'Vendor Invoice' AND source_file = ? AND ticket_or_invoice = ? LIMIT 1",
            (project_id, source_file, invoice_number),
        )
        if duplicate:
            skipped += 1
        else:
            execute(
                """
                INSERT INTO cost_records (
                  project_id, subproject_id, source, source_file, ticket_or_invoice, record_date, status,
                  cost_type, item, description, qty, rate, amount, raw_rate, raw_cost_source, vendor, notes, created_at
                )
                VALUES (?, ?, 'Vendor Invoice', ?, ?, ?, 'Imported', 'Material', 'Vendor Invoice Total', ?, 1, ?, ?, ?, 'Vendor invoice actual material cost', ?, ?, ?)
                """,
                (
                    project_id,
                    matched_subproject_id,
                    source_file,
                    invoice_number,
                    invoice_date,
                    f"{vendor} invoice total",
                    total_due,
                    total_due,
                    total_due,
                    vendor,
                    json.dumps({"order_number": order_number, "text": text[:12000]}, default=str),
                    datetime.now().isoformat(timespec="seconds"),
                ),
            )
            count += 1
    return {"count": count, "skipped": skipped, "order_number": order_number, "matched_subproject_id": matched_subproject_id, "vendor": vendor, "invoice_number": invoice_number, "duplicate": False}


LOGIN_HTML = r"""
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <meta name="theme-color" content="#152332">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-title" content="TPE Field PO">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <link rel="manifest" href="/manifest.json">
  <link rel="apple-touch-icon" href="/brand/twin-peaks-logo.png">
  <title>Twin Peaks Login</title>
  <style>
    :root { --blue:#2f69b1; --ink:#17202a; --muted:#5b6b7f; --line:#d5dde6; --yellow:#ffc20e; }
    * { box-sizing: border-box; }
    body { margin:0; min-height:100vh; display:grid; place-items:center; background:#f3f6f9; color:var(--ink); font-family:"Segoe UI", Arial, sans-serif; }
    .login { width:min(420px, calc(100vw - 32px)); background:white; border:1px solid var(--line); border-radius:8px; padding:26px; box-shadow:0 16px 40px rgba(23,32,42,.12); }
    .brand { display:flex; align-items:center; gap:14px; margin-bottom:22px; }
    .brand img { width:82px; max-height:52px; object-fit:contain; }
    h1 { margin:0; font-size:24px; }
    .subtitle { color:var(--muted); font-weight:650; margin-top:3px; }
    label { display:block; color:#31445a; font-size:13px; font-weight:750; margin:14px 0 6px; }
    input { width:100%; border:1px solid var(--line); border-radius:7px; padding:12px; font-size:15px; }
    button { width:100%; margin-top:18px; border:0; border-radius:7px; padding:12px 14px; font-weight:800; color:white; background:var(--blue); cursor:pointer; }
    .error { min-height:20px; color:#b42318; font-weight:700; margin-top:12px; }
    .hint { color:var(--muted); font-size:13px; margin-top:14px; line-height:1.4; }
  </style>
</head>
<body>
  <form class="login" id="loginForm" method="post" action="/api/login" autocomplete="on">
    <div class="brand">
      <img src="/brand/twin-peaks-logo.png" alt="Twin Peaks Electrical">
      <div><h1>Company Dashboard</h1><div class="subtitle">Twin Peaks Electrical</div></div>
    </div>
    <label for="username">Username</label>
    <input id="username" name="username" type="text" autocomplete="username" autocapitalize="none" autocorrect="off" spellcheck="false" enterkeyhint="next" required autofocus>
    <label for="password">Password</label>
    <input id="password" name="password" type="password" autocomplete="current-password" autocapitalize="none" autocorrect="off" spellcheck="false" enterkeyhint="done" required>
    <button type="submit">Log In</button>
    <div class="error" id="loginError"></div>
  </form>
  <script>
    if ('serviceWorker' in navigator) {
      window.addEventListener('load', async () => {
        const registrations = await navigator.serviceWorker.getRegistrations();
        await Promise.all(registrations.map(registration => registration.unregister()));
        if ('caches' in window) {
          const keys = await caches.keys();
          await Promise.all(keys.map(key => caches.delete(key)));
        }
      });
    }
    document.getElementById('loginForm').onsubmit = async event => {
      event.preventDefault();
      const data = Object.fromEntries(new FormData(event.target).entries());
      const error = document.getElementById('loginError');
      error.textContent = '';
      const res = await fetch('/api/login', { method:'POST', body: JSON.stringify(data) });
      if (res.ok) window.location.href = '/';
      else error.textContent = 'Login failed. Check username and password.';
    };
  </script>
</body>
</html>
"""


HTML = r"""
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <meta name="theme-color" content="#152332">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-title" content="TPE Field PO">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <link rel="manifest" href="/manifest.json">
  <link rel="apple-touch-icon" href="/brand/twin-peaks-logo.png">
  <title>Twin Peaks Company Dashboard</title>
  <style>
    :root {
      --ink: #17202a;
      --muted: #607080;
      --line: #d8dee5;
      --bg: #f6f8fa;
      --panel: #ffffff;
      --blue: #2266aa;
      --green: #138a5b;
      --red: #b42318;
      --gold: #9a6700;
    }
    * { box-sizing: border-box; }
    body { margin: 0; font-family: "Segoe UI", Arial, sans-serif; color: var(--ink); background: var(--bg); }
    header { background: #152332; color: white; padding: 14px 24px; display: flex; align-items: center; justify-content: space-between; gap: 20px; border-bottom: 4px solid #ffc20e; }
    .brand-lockup { display: flex; align-items: center; gap: 14px; min-width: 280px; }
    .brand-logo { display: flex; align-items: center; justify-content: center; background: white; border-radius: 8px; padding: 8px 10px; border: 1px solid rgba(255,255,255,.3); box-shadow: 0 6px 18px rgba(0,0,0,.18); cursor: pointer; }
    .brand-logo img { display: block; height: 46px; width: auto; }
    .system-menu-wrap { position: relative; }
    .system-menu-btn { width: 34px; height: 34px; border-radius: 6px; border: 1px solid rgba(255,255,255,.35); background: rgba(255,255,255,.12); color: white; cursor: pointer; font-size: 20px; font-weight: 800; line-height: 1; }
    .system-menu { position: absolute; top: 40px; left: 0; z-index: 40; width: 160px; padding: 6px; background: white; border: 1px solid var(--line); border-radius: 8px; box-shadow: 0 12px 32px rgba(0,0,0,.2); }
    .system-menu button { display: block; width: 100%; border: 0; background: white; color: var(--ink); text-align: left; padding: 9px 10px; border-radius: 6px; cursor: pointer; font-weight: 650; }
    .system-menu button:hover { background: #eef2f6; }
    .brand-title h1 { margin: 0; font-size: 21px; font-weight: 750; }
    .brand-title .subtitle { color: #d8e6f3; font-size: 12px; font-weight: 650; margin-top: 2px; text-transform: uppercase; letter-spacing: .04em; }
    header select { min-width: 360px; padding: 10px 12px; border-radius: 6px; border: 1px solid #8aa0b5; font-weight: 650; }
    .project-switcher { display: flex; align-items: center; gap: 10px; }
    .project-switcher label { color: #d8e6f3; margin: 0; font-size: 12px; text-transform: uppercase; letter-spacing: .04em; }
    main { padding: 20px 24px 36px; max-width: 1500px; margin: 0 auto; }
    .app-shell { display: grid; grid-template-columns: 230px minmax(0, 1fr); gap: 18px; align-items: start; }
    nav { position: sticky; top: 16px; display: flex; flex-direction: column; gap: 8px; background: white; border: 1px solid var(--line); border-radius: 8px; padding: 12px; }
    nav::before { content: "Navigation"; color: var(--muted); font-size: 12px; font-weight: 750; text-transform: uppercase; letter-spacing: .04em; margin: 2px 2px 4px; }
    nav button { width: 100%; text-align: left; }
    .nav-subgroup { display: flex; flex-direction: column; gap: 6px; margin: -3px 0 6px 12px; padding: 0 0 0 10px; border-left: 2px solid #e6edf4; }
    .nav-subgroup.hidden { display: none; }
    .nav-subgroup button { font-size: 13px; padding: 8px 10px; }
    nav button, .btn { border: 1px solid var(--line); background: white; color: var(--ink); padding: 9px 12px; border-radius: 6px; cursor: pointer; font-weight: 600; transition: border-color .15s ease, box-shadow .15s ease, transform .15s ease, background-color .15s ease; }
    nav button:hover, nav button:focus-visible { border-color: var(--blue); box-shadow: 0 8px 24px rgba(25, 99, 176, .12); outline: none; transform: translateY(-1px); }
    .btn:hover, .btn:focus-visible { border-color: var(--blue); box-shadow: 0 8px 24px rgba(25, 99, 176, .12); outline: none; transform: translateY(-1px); }
    nav button.active, .btn.primary { background: var(--blue); color: white; border-color: var(--blue); }
    .btn.danger { color: var(--red); border-color: #f0b8b2; }
    .btn.danger:hover, .btn.danger:focus-visible { border-color: var(--red); box-shadow: 0 8px 24px rgba(180, 35, 24, .12); }
    .grid { display: grid; gap: 14px; }
    .grid.cols-4 { grid-template-columns: repeat(4, minmax(180px, 1fr)); }
    .grid.cols-2 { grid-template-columns: repeat(2, minmax(260px, 1fr)); }
    .panel { background: var(--panel); border: 1px solid var(--line); border-radius: 8px; padding: 16px; min-width: 0; }
    .home-card { cursor: pointer; transition: border-color .15s ease, box-shadow .15s ease; }
    .home-card:hover { border-color: var(--blue); box-shadow: 0 8px 24px rgba(25, 99, 176, .12); }
    .attention-panel { margin-bottom: 14px; }
    .section-head { display: flex; align-items: flex-start; justify-content: space-between; gap: 12px; margin-bottom: 12px; }
    .section-head h2 { margin: 0 0 4px; }
    .attention-grid { display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 10px; }
    .attention-card { display: grid; grid-template-columns: 1fr auto; gap: 8px 12px; align-items: center; border: 1px solid var(--line); border-radius: 8px; padding: 12px; background: #fbfcfd; cursor: pointer; transition: border-color .15s ease, box-shadow .15s ease, transform .15s ease; }
    .attention-card:hover { border-color: var(--blue); box-shadow: 0 8px 22px rgba(25, 99, 176, .11); transform: translateY(-1px); }
    .attention-card.ok { border-color: #cfe3d7; background: #f7fbf8; }
    .attention-card.warn { border-color: #f0d08b; background: #fffbef; }
    .attention-card.bad { border-color: #e5aaa5; background: #fff6f5; }
    .attention-card .label { color: #34495e; font-size: 12px; font-weight: 800; text-transform: uppercase; }
    .attention-card .count { font-size: 28px; font-weight: 800; line-height: 1; }
    .attention-card.ok .count { color: var(--green); }
    .attention-card.warn .count { color: var(--amber); }
    .attention-card.bad .count { color: var(--red); }
    .attention-card .detail { color: var(--muted); font-size: 13px; }
    .attention-card .amount { color: #34495e; font-weight: 700; text-align: right; }
    .attention-card .open-link { grid-column: 2; color: var(--blue); font-weight: 800; font-size: 13px; text-align: right; }
    .attention-list { grid-column: 1 / -1; display: grid; gap: 6px; margin-top: 4px; }
    .attention-item { display: grid; grid-template-columns: 1fr auto; align-items: center; gap: 8px; padding: 8px; border: 1px solid rgba(0,0,0,.08); border-radius: 8px; background: rgba(255,255,255,.65); }
    .attention-item strong { display: block; font-size: 13px; }
    .attention-item span { color: var(--muted); font-size: 12px; }
    .attention-item .btn { padding: 6px 8px; font-size: 12px; }
    .permission-role { border: 1px solid var(--line); border-radius: 8px; overflow: hidden; margin-top: 12px; }
    .permission-role-header { display: flex; justify-content: space-between; align-items: center; gap: 12px; padding: 12px 14px; background: #eef2f6; font-weight: 800; }
    .permission-role table { margin: 0; }
    .permission-role td:last-child, .permission-role th:last-child,
    .permission-role td:nth-last-child(2), .permission-role th:nth-last-child(2) { text-align: center; width: 90px; }
    .permission-role input[type="checkbox"] { width: auto; }
    .kpi { min-height: 98px; }
    .help-card { position: relative; overflow: visible; }
    .help-marker { position: absolute; top: 10px; right: 10px; z-index: 4; display: inline-flex; align-items: center; justify-content: center; width: 22px; height: 22px; border: 1px solid var(--line); border-radius: 50%; background: #fbfcfd; color: var(--blue); font-size: 13px; font-weight: 800; cursor: help; }
    .help-marker:focus { outline: 2px solid rgba(34, 102, 170, .3); outline-offset: 2px; }
    .inline-help-cell { position: relative; display: inline-flex; align-items: center; gap: 8px; padding-right: 30px; }
    .inline-help-cell .help-marker { position: relative; top: auto; right: auto; flex: 0 0 auto; }
    .help-popover { position: absolute; top: 28px; right: 0; z-index: 30; display: none; width: min(310px, calc(100vw - 52px)); padding: 10px 12px; border: 1px solid #bfd0e0; border-radius: 8px; background: white; color: var(--ink); box-shadow: 0 14px 34px rgba(15, 35, 55, .18); font-size: 13px; font-weight: 500; line-height: 1.38; text-transform: none; }
    .help-marker:hover .help-popover, .help-marker:focus .help-popover { display: block; }
    .kpi.clickable { cursor: pointer; transition: border-color .15s ease, box-shadow .15s ease, transform .15s ease; }
    .kpi.clickable:hover, .kpi.clickable:focus { border-color: var(--blue); box-shadow: 0 8px 24px rgba(25, 99, 176, .12); outline: none; transform: translateY(-1px); }
    .kpi .label { color: var(--muted); font-size: 13px; font-weight: 650; text-transform: uppercase; }
    .kpi .value { font-size: 28px; font-weight: 750; margin-top: 10px; }
    .kpi .hint { color: var(--muted); font-size: 13px; margin-top: 6px; }
    h2 { font-size: 18px; margin: 0 0 12px; }
    h3 { font-size: 15px; margin: 16px 0 8px; }
    label { display: block; font-size: 13px; color: var(--muted); font-weight: 650; margin: 10px 0 5px; }
    input, select, textarea { width: 100%; border: 1px solid var(--line); border-radius: 6px; padding: 9px 10px; font: inherit; background: white; }
    textarea { min-height: 76px; }
    table { width: 100%; border-collapse: collapse; font-size: 14px; background: white; }
    th, td { padding: 9px 10px; border-bottom: 1px solid var(--line); text-align: left; vertical-align: top; }
    th { color: #34495e; font-size: 12px; text-transform: uppercase; background: #eef2f6; position: sticky; top: 0; }
    .sort-header { width: 100%; border: 0; background: transparent; color: inherit; cursor: pointer; font: inherit; font-weight: 800; text-align: left; text-transform: uppercase; padding: 0; }
    .sort-header:hover, .sort-header:focus-visible { color: var(--blue); outline: none; }
    .sort-indicator { color: var(--blue); font-size: 11px; margin-left: 4px; }
    .table-wrap { max-height: 520px; overflow: auto; border: 1px solid var(--line); border-radius: 8px; }
    #bidTable { min-width: 1780px; table-layout: fixed; }
    #bidTable th, #bidTable td { white-space: nowrap; }
    #bidTable input, #bidTable select { min-width: 0; padding: 8px 9px; }
    #bidTable th:nth-child(1), #bidTable td:nth-child(1) { width: 120px; }
    #bidTable th:nth-child(2), #bidTable td:nth-child(2) { width: 160px; }
    #bidTable th:nth-child(3), #bidTable td:nth-child(3) { width: 135px; }
    #bidTable th:nth-child(4), #bidTable td:nth-child(4) { width: 160px; }
    #bidTable th:nth-child(5), #bidTable td:nth-child(5) { width: 120px; }
    #bidTable th:nth-child(6), #bidTable td:nth-child(6) { width: 155px; }
    #bidTable th:nth-child(7), #bidTable td:nth-child(7) { width: 160px; }
    #bidTable th:nth-child(8), #bidTable td:nth-child(8) { width: 110px; }
    #bidTable th:nth-child(9), #bidTable td:nth-child(9) { width: 120px; }
    #bidTable th:nth-child(10), #bidTable td:nth-child(10) { width: 105px; }
    #bidTable th:nth-child(11), #bidTable td:nth-child(11) { width: 125px; }
    #bidTable th:nth-child(12), #bidTable td:nth-child(12) { width: 100px; }
    #bidTable th:nth-child(13), #bidTable td:nth-child(13) { width: 130px; }
    #bidTable th:nth-child(14), #bidTable td:nth-child(14) { width: 125px; }
    #bidTable th:nth-child(15), #bidTable td:nth-child(15) { width: 190px; }
    #bidTable th:nth-child(16), #bidTable td:nth-child(16) { width: 80px; }
    #bidTable tr.bid-stale td { background: #fff1f0; }
    #bidTable tr.bid-stale td:first-child { box-shadow: inset 4px 0 0 var(--red); }
    #bidTable tr.bid-stale input, #bidTable tr.bid-stale select { background: #fff8f7; border-color: #f0b8b5; }
    #bidTable tr.bid-dirty td { background: #fff8e1; }
    #bidTable tr.bid-dirty td:first-child { box-shadow: inset 4px 0 0 var(--gold); }
    #bidTable tr.bid-dirty input, #bidTable tr.bid-dirty select { background: #fffdf3; border-color: #d6a700; }
    #bidTable tr.bid-dirty [data-save-bid] { background: var(--gold); color: white; border-color: var(--gold); }
    #subprojectEditTable tr.setup-dirty td,
    #changeOrderEditTable tr.setup-dirty td { background: #fff8e1; }
    #subprojectEditTable tr.setup-dirty td:first-child,
    #changeOrderEditTable tr.setup-dirty td:first-child { box-shadow: inset 4px 0 0 var(--gold); }
    #subprojectEditTable tr.setup-dirty input,
    #subprojectEditTable tr.setup-dirty select,
    #changeOrderEditTable tr.setup-dirty input,
    #changeOrderEditTable tr.setup-dirty select { background: #fffdf3; border-color: #d6a700; }
    #subprojectEditTable tr.setup-dirty [data-save-sp],
    #changeOrderEditTable tr.setup-dirty [data-save-co] { background: var(--gold); color: white; border-color: var(--gold); }
    .dashboard-split { display: grid; grid-template-columns: minmax(0, 1fr) minmax(0, 1fr); gap: 14px; margin-top: 14px; }
    .dashboard-table-wrap { overflow: auto; max-height: 360px; }
    .dashboard-table-wrap table { min-width: 760px; }
    .hierarchy-wrap { display: grid; gap: 10px; }
    .hierarchy-node { border: 1px solid var(--line); border-radius: 8px; background: #fbfcfd; padding: 12px; }
    .hierarchy-node.collapsible { cursor: pointer; transition: border-color .15s ease, box-shadow .15s ease; }
    .hierarchy-node.collapsible:hover, .hierarchy-node.collapsible:focus { border-color: var(--blue); box-shadow: 0 8px 24px rgba(25, 99, 176, .10); outline: none; }
    .hierarchy-node.master { border-left: 5px solid var(--blue); background: white; }
    .hierarchy-node.subproject { margin-left: 24px; border-left: 4px solid var(--green); }
    .hierarchy-node.change-order { margin-left: 52px; border-left: 4px solid var(--gold); }
    .hierarchy-node .node-title { display: flex; align-items: baseline; justify-content: space-between; gap: 12px; font-weight: 750; }
    .hierarchy-title-left { display: inline-flex; align-items: center; gap: 8px; min-width: 0; }
    .hierarchy-toggle { display: inline-flex; align-items: center; justify-content: center; width: 22px; height: 22px; border: 1px solid var(--line); border-radius: 50%; background: white; color: var(--blue); font-weight: 800; flex: 0 0 auto; }
    .hierarchy-children.collapsed { display: none; }
    .hierarchy-node .node-meta { color: var(--muted); font-size: 13px; margin-top: 4px; }
    .hierarchy-node .node-values { display: flex; gap: 14px; flex-wrap: wrap; color: var(--muted); font-size: 13px; margin-top: 8px; }
    .actions { display: flex; gap: 8px; flex-wrap: wrap; margin-top: 12px; }
    .muted { color: var(--muted); }
    .good { color: var(--green); }
    .bad { color: var(--red); }
    .warn { color: var(--gold); }
    .hidden { display: none; }
    body.read-only form input, body.read-only form select, body.read-only form textarea { pointer-events: none; background: #f6f8fa; color: #607080; }
    body.read-only #changePasswordForm input { pointer-events: auto; background: white; color: var(--ink); }
    body.read-only form .actions,
    body.read-only [data-save],
    body.read-only [data-save-sp],
    body.read-only [data-delete-sp],
    body.read-only [data-delete-co],
    body.read-only [data-save-rate],
    body.read-only [data-save-cost-group],
    body.read-only [data-save-bid],
    body.read-only [data-delete-import],
    body.read-only [data-reset-user],
    body.read-only [data-toggle-user],
    body.read-only [data-save-invoice-subproject],
    body.read-only [data-save-customer-invoice],
    body.read-only [data-allocate-invoice],
    body.read-only [data-delete-financial-report],
    body.read-only [data-omit-fieldwise],
    body.read-only [data-delete-audit-omission],
    body.read-only [data-save-office-po],
    body.read-only [data-save-cog],
    body.read-only [data-delete-cog],
    body.read-only #addRate,
    body.read-only #addCogCategory,
    body.read-only #saveTexasReminderRecipients,
    body.read-only #sendTexasReminderNow,
    body.read-only #saveBillingReminder,
    body.read-only #sendBillingReminderNow,
    body.read-only #savePoReminder,
    body.read-only #sendPoReminderNow,
    body.read-only #importVendorInvoice { display: none !important; }
    body.read-only .home-card[data-open-tab="officePo"] { display: none !important; }
    body.read-only #vendorAllocationForm .actions { display: none !important; }
    body.read-only #changePasswordForm .actions { display: flex !important; }
    body.tx-read-only .project-switcher,
    body.tx-read-only nav button:not([data-tab="texasOps"]),
    body.tx-read-only nav .nav-subgroup,
    body.tx-read-only .home-card:not([data-open-tab="texasOps"]),
    body.tx-read-only #financialUploadForm,
    body.tx-read-only [data-delete-financial-report] { display: none !important; }
    body.field-po-only .project-switcher,
    body.field-po-only nav,
    body.field-po-only .home-card:not([data-open-tab="fieldPo"]) { display: none !important; }
    .field-po-hero { max-width: 760px; margin: 0 auto 14px; }
    .field-po-hero h2 { font-size: 28px; margin-bottom: 4px; }
    .field-po-steps { display: grid; grid-template-columns: repeat(3, 1fr); gap: 8px; margin-top: 12px; }
    .field-po-step { border: 1px solid var(--line); border-radius: 8px; padding: 10px; background: #fbfcfd; font-weight: 750; }
    .field-po-step span { display: block; color: var(--muted); font-size: 12px; margin-top: 3px; font-weight: 650; }
    .field-po-panel { max-width: 760px; margin: 0 auto; }
    .field-po-panel label { font-size: 15px; color: var(--ink); margin-top: 16px; }
    .field-po-panel input,
    .field-po-panel select,
    .field-po-panel textarea { font-size: 18px; padding: 14px 13px; min-height: 52px; }
    .field-po-panel textarea { min-height: 132px; }
    .field-po-submit { width: 100%; min-height: 58px; font-size: 18px; }
    .field-po-list { display: grid; gap: 12px; }
    .field-po-card { border: 1px solid var(--line); border-radius: 8px; background: white; padding: 14px; display: grid; gap: 10px; }
    .field-po-card-header { display: flex; justify-content: space-between; gap: 10px; align-items: flex-start; }
    .field-po-card-title { font-size: 18px; font-weight: 800; }
    .field-po-card-meta { color: var(--muted); font-size: 13px; margin-top: 3px; }
    .field-po-status { border-radius: 999px; padding: 5px 9px; font-size: 12px; font-weight: 800; background: #eef2f6; white-space: nowrap; }
    .field-po-status-pending-approval { background: #fff7e6; color: #8a5a00; }
    .field-po-status-open, .field-po-status-issued { background: #eef5ff; color: var(--blue); }
    .field-po-status-picked-up { background: #e8f7ef; color: var(--green); }
    .field-po-status-closed { background: #e9eef5; color: var(--muted); }
    .field-po-status-void { background: #fdecea; color: var(--red); }
    .pickup-upload { border: 1px dashed #9eb4cb; background: #f8fbff; border-radius: 8px; padding: 12px; }
    .pickup-upload input { width: 100%; min-height: 48px; }
    .pickup-upload .btn { width: 100%; min-height: 48px; justify-content: center; margin-top: 8px; }
    .combo-wrap { position: relative; }
    .combo-list {
      position: absolute;
      z-index: 20;
      top: calc(100% + 4px);
      left: 0;
      right: 0;
      max-height: 260px;
      overflow: auto;
      background: white;
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: 0 14px 30px rgba(15, 23, 42, 0.16);
    }
    .combo-option {
      width: 100%;
      border: 0;
      background: white;
      text-align: left;
      padding: 10px 12px;
      border-bottom: 1px solid #eef2f6;
      cursor: pointer;
    }
    .combo-option:hover,
    .combo-option:focus { background: #eef5ff; outline: none; }
    .combo-option strong { display: block; color: var(--ink); }
    .combo-option span { display: block; color: var(--muted); font-size: 12px; margin-top: 2px; }
    .bar { height: 12px; background: #e8edf2; border-radius: 999px; overflow: hidden; }
    .bar span { display: block; height: 100%; background: var(--blue); }
    .invoice-summary { background: #fbfcfd; cursor: pointer; }
    .invoice-summary td { font-weight: 650; }
    .selectable-row { cursor: pointer; }
    .selectable-row.selected { background: #eef5ff; }
    .selectable-row.selected td { font-weight: 700; }
    .trend-metric-row { cursor: pointer; transition: background-color .15s ease, box-shadow .15s ease; }
    .trend-metric-row:hover,
    .trend-metric-row:focus-within { background: #eef5ff; box-shadow: inset 4px 0 0 var(--blue); }
    .trend-metric-row td:first-child::after { content: "View trend"; display: block; margin-top: 2px; color: var(--muted); font-size: 11px; font-weight: 650; }
    .invoice-detail { background: #ffffff; }
    .invoice-detail td:first-child { padding-left: 34px; }
    .invoice-toggle { display: inline-flex; align-items: center; justify-content: center; width: 22px; height: 22px; margin-right: 8px; border: 1px solid var(--line); border-radius: 6px; background: white; color: var(--blue); font-weight: 800; }
    .invoice-line-count { color: var(--muted); font-weight: 600; }
    .pdf-link { color: var(--blue); font-weight: 700; text-decoration: none; }
    .pdf-link:hover { text-decoration: underline; }
    .cost-filter-bar { display: grid; grid-template-columns: minmax(220px, 1.4fr) repeat(3, minmax(150px, .7fr)) auto; gap: 8px; align-items: end; margin-bottom: 10px; }
    .cost-filter-bar label { margin-top: 0; }
    .cost-filter-count { color: var(--muted); font-size: 13px; font-weight: 650; margin: 0 0 10px; }
    .filter-summary { display: flex; gap: 10px; flex-wrap: wrap; margin: 0 0 10px; }
    .filter-summary .summary-pill { border: 1px solid var(--line); background: #fbfcfd; border-radius: 8px; padding: 9px 12px; min-width: 160px; }
    .filter-summary .summary-label { color: var(--muted); font-size: 11px; font-weight: 750; text-transform: uppercase; letter-spacing: .04em; }
    .filter-summary .summary-value { color: var(--ink); font-size: 18px; font-weight: 800; margin-top: 3px; }
    .filter-summary .summary-value.amount { color: var(--green); }
    .segmented { display: inline-flex; gap: 0; border: 1px solid var(--line); border-radius: 8px; overflow: hidden; background: white; }
    .segmented button { border: 0; border-right: 1px solid var(--line); background: white; padding: 7px 10px; cursor: pointer; font-weight: 650; }
    .segmented button:last-child { border-right: 0; }
    .segmented button.active { background: var(--blue); color: white; }
    .trend-chart { width: 100%; height: 300px; display: block; }
    .trend-axis { stroke: #ccd6df; stroke-width: 1; }
    .trend-line-revenue { fill: none; stroke: var(--blue); stroke-width: 3; }
    .trend-line-profit { fill: none; stroke: var(--green); stroke-width: 3; }
    .trend-point { stroke: white; stroke-width: 2; }
    .trend-label { fill: var(--muted); font-size: 12px; }
    .trend-legend { display: flex; gap: 14px; flex-wrap: wrap; color: var(--muted); font-size: 13px; font-weight: 650; margin-top: 8px; }
    .trend-swatch { display: inline-block; width: 18px; height: 4px; border-radius: 999px; vertical-align: middle; margin-right: 6px; }
    .project-banner { display: grid; grid-template-columns: 1.4fr repeat(3, minmax(130px, .45fr)); gap: 12px; align-items: stretch; margin-bottom: 14px; }
    .project-title { background: #ffffff; border: 1px solid var(--line); border-left: 5px solid var(--blue); border-radius: 8px; padding: 14px 16px; }
    .project-title.clickable, .job-chip.clickable { cursor: pointer; transition: border-color .15s ease, box-shadow .15s ease, transform .15s ease; }
    .project-title.clickable:hover, .project-title.clickable:focus, .job-chip.clickable:hover, .job-chip.clickable:focus { border-color: var(--blue); box-shadow: 0 8px 24px rgba(25, 99, 176, .12); outline: none; transform: translateY(-1px); }
    .project-title .eyebrow { color: var(--muted); font-size: 12px; font-weight: 750; text-transform: uppercase; }
    .project-title .name { font-size: 24px; font-weight: 780; margin-top: 4px; }
    .job-chip { background: #ffffff; border: 1px solid var(--line); border-radius: 8px; padding: 12px; }
    .job-chip .job { font-size: 20px; font-weight: 780; }
    .job-chip .label { color: var(--muted); font-size: 13px; margin-top: 4px; }
    .modal-backdrop { position: fixed; inset: 0; background: rgba(15, 23, 42, .44); display: flex; align-items: center; justify-content: center; z-index: 50; padding: 20px; }
    .modal-backdrop.hidden { display: none; }
    .modal { background: white; border-radius: 8px; border: 1px solid var(--line); width: min(460px, 100%); padding: 18px; box-shadow: 0 18px 50px rgba(0,0,0,.22); }
    .modal.large { width: min(1180px, calc(100vw - 40px)); max-height: calc(100vh - 40px); overflow: auto; }
    .modal h2 { margin-bottom: 8px; }
    .modal p { color: var(--muted); margin: 0; line-height: 1.45; }
    .trend-modal-body .trend-chart { height: 560px; }
    details { border: 1px solid var(--line); border-radius: 8px; margin-top: 14px; overflow: hidden; background: white; }
    summary { cursor: pointer; font-weight: 750; padding: 12px 14px; background: #eef2f6; }
    details .detail-body { padding: 12px; }
    @media (max-width: 1250px) {
      .dashboard-split { grid-template-columns: 1fr; }
      .attention-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
    }
    @media (max-width: 900px) {
      header { align-items: flex-start; flex-direction: column; }
      .brand-logo img { height: 38px; }
      .brand-lockup { min-width: 0; }
      header select { min-width: 0; width: 100%; }
      .project-switcher { width: 100%; align-items: stretch; flex-direction: column; }
      .app-shell { grid-template-columns: 1fr; }
      nav { position: static; flex-direction: row; overflow-x: auto; }
      nav::before { display: none; }
      .nav-subgroup { flex-direction: row; margin: 0; padding: 0; border-left: 0; }
      nav button { width: auto; white-space: nowrap; }
      .field-po-steps { grid-template-columns: 1fr; }
      .field-po-hero h2 { font-size: 24px; }
      .field-po-card-header { flex-direction: column; }
      .field-po-status { align-self: flex-start; }
      .grid.cols-4, .grid.cols-2 { grid-template-columns: 1fr; }
      .attention-grid { grid-template-columns: 1fr; }
      .section-head { flex-direction: column; align-items: stretch; }
      .project-banner { grid-template-columns: 1fr; }
      .cost-filter-bar { grid-template-columns: 1fr; }
    }
  </style>
</head>
<body>
  <header>
    <div class="brand-lockup">
      <button class="brand-logo" id="homeLogo" type="button" title="Home"><img src="/brand/twin-peaks-logo.png" alt="Twin Peaks Electrical"></button>
      <div class="system-menu-wrap">
        <button class="system-menu-btn" id="systemMenuButton" type="button" title="System settings">...</button>
        <div class="system-menu hidden" id="systemMenu">
          <button id="systemAccountBtn" type="button">My Account</button>
          <button id="systemAdminBtn" type="button" class="hidden">Admin</button>
          <button id="systemRevisionBtn" type="button" class="hidden">Server Health</button>
          <button id="systemLogoutBtn" type="button">Logout</button>
        </div>
      </div>
      <div class="brand-title">
        <h1>Company Dashboard</h1>
        <div class="subtitle">Twin Peaks Electrical</div>
      </div>
    </div>
    <div class="project-switcher hidden">
      <label for="projectSelect">Master Project</label>
      <select id="projectSelect"></select>
      <div style="margin-top:8px;display:flex;gap:8px;align-items:center;justify-content:flex-end">
        <span class="muted" id="currentUser"></span>
      </div>
    </div>
  </header>
  <main>
    <div class="app-shell">
      <nav>
        <button data-tab="home" data-nav-area="main" data-nav-level="main" class="active">Home</button>
        <div class="nav-subgroup hidden" data-nav-group="home">
          <button data-tab="bids" data-nav-area="home" data-nav-level="sub">Bid Tracking</button>
          <button data-tab="jobOrderReport" data-nav-area="home" data-nav-level="sub">Job Order Quick Reference</button>
        </div>
        <button data-tab="dashboard" data-nav-area="main" data-nav-level="main">Project Dashboard</button>
        <div class="nav-subgroup hidden" data-nav-group="project">
          <button data-tab="newMasterProject" data-nav-area="project" data-nav-level="sub">New Master Project</button>
          <button data-tab="setup" data-nav-area="project" data-nav-level="sub">Setup</button>
          <button data-tab="import" data-nav-area="project" data-nav-level="sub">Field Wise</button>
          <button data-tab="review" data-nav-area="project" data-nav-level="sub">Review Exceptions</button>
          <button data-tab="invoices" data-nav-area="project" data-nav-level="sub">Vendor Invoices</button>
          <button data-tab="billing" data-nav-area="project" data-nav-level="sub">Customer Billing</button>
          <button data-tab="projectPo" data-nav-area="project" data-nav-level="sub" class="po-feature-disabled">POs</button>
          <button data-tab="archivedProjects" data-nav-area="project" data-nav-level="sub">Archived Projects</button>
        </div>
        <button data-tab="fieldwiseAudit" data-nav-area="main" data-nav-level="main">Field Wise Audit</button>
        <button data-tab="fieldPo" data-nav-area="main" data-nav-level="main">Purchase Orders</button>
        <div class="nav-subgroup hidden" data-nav-group="po">
          <button data-tab="fieldPo" data-nav-area="po" data-nav-level="sub">Create PO</button>
          <button data-tab="officePo" data-nav-area="po" data-nav-level="sub">Office PO Review</button>
          <button data-tab="cogSetup" data-nav-area="po" data-nav-level="sub">COG's Setup</button>
        </div>
        <button data-tab="texasOps" data-nav-area="main" data-nav-level="main">Texas Ops</button>
        <div class="nav-subgroup hidden" data-nav-group="texas">
          <button data-tab="texasReminders" data-nav-area="texas" data-nav-level="sub">Reminders Setup</button>
        </div>
      </nav>
      <div class="app-content">

    <section id="home" class="tab">
      <div class="panel attention-panel">
        <div class="section-head">
          <div>
            <h2>Needs Attention</h2>
            <p class="muted">Quick checks for items that may need action today.</p>
          </div>
          <button class="btn" id="refreshHomeAlerts" type="button">Refresh</button>
        </div>
        <div id="homeAlerts" class="attention-grid"></div>
      </div>
    </section>

    <section id="fieldPo" class="tab hidden">
      <div class="panel field-po-hero">
        <h2>Field Purchase Orders</h2>
        <p class="muted">Create the PO, then come back here to upload the pickup ticket after the vendor gives it to you.</p>
        <div class="field-po-steps">
          <div class="field-po-step">1. Pick job<span>Search job/order or COG</span></div>
          <div class="field-po-step">2. Create PO<span>Vendor and materials</span></div>
          <div class="field-po-step">3. Add ticket<span>Photo after pickup</span></div>
        </div>
      </div>
      <div class="panel field-po-panel">
        <h2>Create PO</h2>
        <form id="fieldPoForm" enctype="multipart/form-data">
          <label>Job / Order # / COG</label>
          <div class="combo-wrap">
            <input id="fieldPoJobSearch" placeholder="Type job, order, COG, customer, or project..." autocomplete="off" required>
            <input type="hidden" name="job_key" id="fieldPoJobKey" required>
            <div id="fieldPoJobList" class="combo-list hidden"></div>
          </div>
          <div id="fieldPoCustomerWrap" class="hidden">
            <label>Customer <span class="bad">*</span></label>
            <input name="customer_name" id="fieldPoCustomer" placeholder="Customer name" autocomplete="organization" disabled>
          </div>
          <label>Vendor <span class="bad">*</span></label>
          <input name="vendor" id="fieldPoVendor" placeholder="Vendor name" autocomplete="organization" required>
          <label>What are you buying?</label>
          <textarea name="description" id="fieldPoDescription" placeholder="Example: 2 boxes 3/4 EMT, 20 couplings, lift rental" required></textarea>
          <label>Estimated Amount</label>
          <input name="estimated_amount" id="fieldPoAmount" type="number" min="0" step="0.01" placeholder="Optional">
          <label>Photo / Quote</label>
          <input name="attachment" id="fieldPoAttachment" type="file" accept=".pdf,.png,.jpg,.jpeg,.webp">
          <div class="actions field-po-actions">
            <button class="btn primary field-po-submit" type="submit">Create PO</button>
          </div>
          <div id="fieldPoResult" class="muted"></div>
        </form>
      </div>
      <div class="panel" style="margin-top:14px">
        <div style="display:flex;align-items:center;justify-content:space-between;gap:12px">
          <h2>My POs</h2>
          <button class="btn" id="refreshFieldPos" type="button">Refresh</button>
        </div>
        <div id="fieldPoList" class="field-po-list"></div>
      </div>
    </section>

    <section id="officePo" class="tab hidden">
      <div class="panel">
        <div style="display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap">
          <div>
            <h2>Office PO Review</h2>
            <p class="muted">Use status to track pickup tickets and close or void completed POs with a reason.</p>
          </div>
          <button class="btn" id="refreshOfficePos" type="button">Refresh</button>
        </div>
        <div class="grid cols-4">
          <div><label>Status</label><select id="officePoStatusFilter"><option value="">All statuses</option><option>Pending Approval</option><option>Issued</option><option>Picked Up</option><option>Closed</option><option>Void</option></select></div>
          <div style="grid-column:span 3"><label>Search</label><input id="officePoSearch" placeholder="Search PO, job/order, vendor, requester, or description"></div>
        </div>
        <div class="muted" id="officePoCount" style="margin-top:8px"></div>
        <div class="table-wrap" style="margin-top:10px"><table id="officePoTable"></table></div>
      </div>
    </section>

    <section id="cogSetup" class="tab hidden">
      <div class="panel">
        <h2>COG Categories</h2>
        <p class="muted">Company-wide COG choices for POs that do not belong to a tracked job/order number.</p>
        <div class="grid cols-4">
          <div><label>Name</label><input id="cogName" placeholder="Truck 12"></div>
          <div><label>Internal COG Code</label><input id="cogCode" placeholder="TRUCK"></div>
          <div style="grid-column:span 2"><label>Description</label><input id="cogDescription" placeholder="Optional notes for office review"></div>
        </div>
        <div class="actions"><button class="btn primary" id="addCogCategory" type="button">Add COG Category</button></div>
        <div class="table-wrap" style="margin-top:12px"><table id="cogCategoryTable"></table></div>
      </div>
    </section>

    <section id="archivedProjects" class="tab hidden">
      <div class="panel">
        <div style="display:flex;align-items:center;justify-content:space-between;gap:12px">
          <h2>Archived Projects</h2>
          <button class="btn" id="refreshArchivedProjects" type="button">Refresh</button>
        </div>
        <div class="table-wrap" id="archivedProjectsTable"></div>
      </div>
    </section>

    <section id="jobOrderReport" class="tab hidden">
      <div class="panel">
        <div style="display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap">
          <div>
            <h2>Job Order Quick Reference</h2>
            <p class="muted">Active master projects only. Includes subprojects, change orders, and child projects.</p>
          </div>
          <button class="btn" id="refreshJobOrderReport" type="button">Refresh</button>
        </div>
        <div class="grid cols-4">
          <div><label>Search</label><input id="jobOrderSearch" placeholder="Search job/order, customer, project, description, or status"></div>
          <div><label>Customer</label><select id="jobOrderCustomerFilter"><option value="">All customers</option></select></div>
          <div><label>Master Project</label><select id="jobOrderProjectFilter"><option value="">All master projects</option></select></div>
          <div><label>Type</label><select id="jobOrderTypeFilter"><option value="">All types</option></select></div>
        </div>
        <div class="actions">
          <label style="min-width:220px;margin-top:0">Status<select id="jobOrderStatusFilter"><option value="">All statuses</option></select></label>
          <button class="btn" id="clearJobOrderFilters" type="button">Clear Filters</button>
        </div>
        <div class="muted" id="jobOrderReportCount" style="margin-top:8px"></div>
        <div class="table-wrap" style="margin-top:10px"><table id="jobOrderReportTable"></table></div>
      </div>
    </section>

    <section id="texasOps" class="tab hidden">
      <div class="panel">
        <h2>Texas Operations Financial Overview</h2>
        <form id="financialUploadForm">
          <div class="grid cols-4">
            <div><label>Week Ending Override</label><input name="report_date" type="date"></div>
            <div><label>Report Type</label><select name="report_type"><option value="combined">Combined / Auto</option><option value="balance_sheet">Balance Sheet</option><option value="pnl">P&L</option></select></div>
            <div style="grid-column:span 2"><label>Reports</label><input name="files" type="file" accept=".xlsx,.xlsm,.csv,.tsv,.pdf" multiple required></div>
          </div>
          <div class="muted" style="margin-top:8px">Leave week ending blank when uploading multiple weeks; the app will use the report date inside each file.</div>
          <div class="actions"><button class="btn primary" type="submit">Upload Reports</button></div>
          <div class="muted" id="financialUploadResult"></div>
        </form>
      </div>
      <div class="grid cols-4" id="financialKpis" style="margin-top:14px"></div>
      <div class="panel" style="margin-top:14px">
        <h2>Texas AP For Future View</h2>
        <p class="muted">Enter the weekly Texas AP total when accounting cannot provide AP by class. This does not change the uploaded balance sheet; it adds a projected view to the snapshot below.</p>
        <form id="texasApForm">
          <div class="grid cols-4">
            <div><label>Week Ending</label><input name="report_date" type="date" required></div>
            <div><label>AP Total</label><input name="ap_total" type="number" step="0.01" required></div>
            <div style="grid-column:span 2"><label>Notes</label><input name="notes" placeholder="Optional note about the AP report source"></div>
          </div>
          <div class="actions"><button class="btn primary" type="submit">Save AP Total</button></div>
          <div class="muted" id="texasApResult"></div>
        </form>
        <div id="texasApHistory" style="margin-top:12px"></div>
      </div>
      <div class="dashboard-split">
        <div class="panel"><h2>Profitability Trend</h2><div id="financialProfitTrend"></div></div>
        <div class="panel"><h2>Balance Sheet Snapshot</h2><div id="financialBalanceSnapshot"></div></div>
      </div>
      <div class="panel" style="margin-top:14px"><h2>Uploaded Reports</h2><div id="financialReports"></div></div>
    </section>

    <section id="texasReminders" class="tab hidden">
      <div class="panel">
        <h2>Reminders Setup</h2>
        <p class="muted">Set a recurring reminder for the Texas Ops P&amp;L upload. When the selected day and time arrives, the server checks whether that day&apos;s P&amp;L has been uploaded. If it is missing, selected users receive an email reminder.</p>
        <div class="grid cols-4">
          <label style="display:flex;align-items:center;gap:8px;margin-top:28px"><input id="texasReminderEnabled" type="checkbox" style="width:auto"> Enabled</label>
          <div><label>Day</label><select id="texasReminderWeekday">
            <option value="0">Monday</option>
            <option value="1">Tuesday</option>
            <option value="2">Wednesday</option>
            <option value="3">Thursday</option>
            <option value="4">Friday</option>
            <option value="5">Saturday</option>
            <option value="6">Sunday</option>
          </select></div>
          <div><label>Time</label><input id="texasReminderTime" type="time" value="15:00"></div>
        </div>
        <div id="texasReminderStatus" class="muted" style="margin-top:10px"></div>
        <h3>Recipients</h3>
        <div id="texasReminderRecipients" style="margin-top:10px"></div>
        <div class="actions">
          <button class="btn primary" id="saveTexasReminderRecipients" type="button">Save Reminder Setup</button>
          <button class="btn" id="sendTexasReminderNow" type="button">Check / Send Reminder Now</button>
        </div>
        <div id="texasReminderResult" class="muted"></div>
        <div id="texasReminderLog" style="margin-top:10px"></div>
      </div>
      <div class="panel" style="margin-top:14px">
        <h2>Customer Invoice Reminder</h2>
        <p class="muted">Send a weekly list of active jobs that have not had a customer invoice added within the selected lookback window.</p>
        <div class="grid cols-4">
          <label style="display:flex;align-items:center;gap:8px;margin-top:28px"><input id="billingReminderEnabled" type="checkbox" style="width:auto"> Enabled</label>
          <div><label>Day</label><select id="billingReminderWeekday">
            <option value="0">Monday</option>
            <option value="1">Tuesday</option>
            <option value="2">Wednesday</option>
            <option value="3">Thursday</option>
            <option value="4">Friday</option>
            <option value="5">Saturday</option>
            <option value="6">Sunday</option>
          </select></div>
          <div><label>Time</label><input id="billingReminderTime" type="time" value="15:00"></div>
          <div><label>Lookback Days</label><input id="billingReminderLookback" type="number" min="1" max="60" value="7"></div>
        </div>
        <div id="billingReminderStatus" class="muted" style="margin-top:10px"></div>
        <h3>Recipients</h3>
        <div id="billingReminderRecipients" style="margin-top:10px"></div>
        <div class="actions">
          <button class="btn primary" id="saveBillingReminder" type="button">Save Billing Reminder</button>
          <button class="btn" id="sendBillingReminderNow" type="button">Check / Send Billing Reminder Now</button>
        </div>
        <div id="billingReminderResult" class="muted"></div>
        <div id="billingReminderJobs" style="margin-top:10px"></div>
        <div id="billingReminderLog" style="margin-top:10px"></div>
      </div>
      <div class="panel" style="margin-top:14px">
        <h2>Untouched PO Reminder</h2>
        <p class="muted">Send a recurring list of POs that are still pending approval or issued, have no pickup ticket, and have not been updated inside the selected window.</p>
        <div class="grid cols-4">
          <label style="display:flex;align-items:center;gap:8px;margin-top:28px"><input id="poReminderEnabled" type="checkbox" style="width:auto"> Enabled</label>
          <div><label>Day</label><select id="poReminderWeekday">
            <option value="0">Monday</option>
            <option value="1">Tuesday</option>
            <option value="2">Wednesday</option>
            <option value="3">Thursday</option>
            <option value="4">Friday</option>
            <option value="5">Saturday</option>
            <option value="6">Sunday</option>
          </select></div>
          <div><label>Time</label><input id="poReminderTime" type="time" value="15:00"></div>
          <div><label>Untouched Days</label><input id="poReminderUntouchedDays" type="number" min="1" max="30" value="2"></div>
        </div>
        <div id="poReminderStatus" class="muted" style="margin-top:10px"></div>
        <h3>Recipients</h3>
        <div id="poReminderRecipients" style="margin-top:10px"></div>
        <div class="actions">
          <button class="btn primary" id="savePoReminder" type="button">Save PO Reminder</button>
          <button class="btn" id="sendPoReminderNow" type="button">Check / Send PO Reminder Now</button>
        </div>
        <div id="poReminderResult" class="muted"></div>
        <div id="poReminderList" style="margin-top:10px"></div>
        <div id="poReminderLog" style="margin-top:10px"></div>
      </div>
    </section>

    <section id="dashboard" class="tab hidden">
      <div id="projectBanner"></div>
      <div class="actions" style="margin-bottom:14px"><button class="btn primary" id="openMasterDetail" type="button">Open Master Project Detail</button></div>
      <div class="panel hidden" style="margin-bottom:14px" id="masterDetailPanel">
        <div style="display:flex;align-items:center;justify-content:space-between;gap:12px">
          <h2 id="masterDetailTitle">Master Project Detail</h2>
          <button class="btn" id="closeMasterDetail" type="button">Close</button>
        </div>
        <div id="masterDetail"></div>
      </div>
      <div class="grid cols-4" id="kpis"></div>
      <div class="panel" style="margin-top:14px">
        <h2>Project Hierarchy</h2>
        <div id="projectHierarchy"></div>
      </div>
      <div class="panel" style="margin-top:14px">
        <div style="display:flex;align-items:center;justify-content:space-between;gap:12px">
          <h2>Invoicing</h2>
          <button class="btn" id="openBillingFromDashboard" type="button">Open Customer Billing</button>
        </div>
        <div id="billingSummary"></div>
      </div>
      <div class="dashboard-split">
        <div class="panel"><h2>Subprojects</h2><div class="dashboard-table-wrap" id="subprojectSummary"></div></div>
        <div class="panel">
          <div style="display:flex;align-items:center;justify-content:space-between;gap:10px">
            <h2 id="coSummaryTitle">Change Orders / Child Projects</h2>
            <button class="btn hidden" id="showAllCos" type="button">Show All</button>
          </div>
          <div class="dashboard-table-wrap" id="coSummary"></div>
        </div>
      </div>
      <div class="panel hidden" style="margin-top:14px" id="subprojectDetailPanel">
        <div style="display:flex;align-items:center;justify-content:space-between;gap:12px">
          <h2 id="subprojectDetailTitle">Subproject Detail</h2>
          <button class="btn" id="closeSubprojectDetail" type="button">Close</button>
        </div>
        <div id="subprojectDetail"></div>
      </div>
      <div class="panel" style="margin-top:14px"><h2>Material Comparison</h2><div id="materialComparison"></div></div>
      <div class="panel" style="margin-top:14px"><h2>Cost By Type</h2><div id="typeSummary"></div></div>
    </section>

    <section id="newMasterProject" class="tab hidden">
      <form class="panel" id="newProjectForm">
        <h2>New Master Project</h2>
        <p class="muted">Create a company-level master project. Subprojects and change orders are added afterward in Setup.</p>
        <div class="grid cols-2">
          <div><label>Project Code</label><input name="project_code" placeholder="Alexandria 33-6" required></div>
          <div><label>Project Name</label><input name="name" placeholder="Alexandria 33-6" required></div>
          <div><label>Customer</label><input name="customer" placeholder="Hunt Oil"></div>
          <div><label>Location</label><input name="location" placeholder="Alexandria 33-6"></div>
          <div><label>Customer Provided PO #</label><input name="customer_po" placeholder="PO number"></div>
          <div><label>Project Rate Set</label><select name="rate_set_id" id="newProjectRateSet"></select></div>
        </div>
        <label>Description</label><textarea name="description" placeholder="Scope notes, project summary, or contract description"></textarea>
        <input type="hidden" name="contract_value" value="0">
        <div class="actions"><button class="btn primary" type="submit">Create Master Project</button></div>
      </form>
    </section>

    <section id="setup" class="tab hidden">
      <div class="grid cols-2">
        <form class="panel" id="projectForm">
          <h2>Master Project</h2>
          <label>Project Code</label><input name="project_code" placeholder="Oakland 17-20" required>
          <label>Project Name</label><input name="name" placeholder="Oakland 17-20" required>
          <label>Customer</label><input name="customer" placeholder="Hunt Oil">
          <label>Location</label><input name="location" placeholder="Oakland 17-20">
          <label>Customer Provided PO #</label><input name="customer_po" placeholder="PO number">
          <label>Base Contract Value</label><input name="contract_value" type="number" step="0.01" value="0" readonly>
          <label>Description</label><textarea name="description" placeholder="Scope notes, project summary, or contract description"></textarea>
          <label>Project Rate Set</label><select name="rate_set_id" id="projectRateSet"></select>
          <div class="actions">
            <button class="btn primary" id="saveProjectBtn" type="submit">Save Project</button>
            <button class="btn danger" id="archiveProjectBtn" type="button">Close & Archive Project</button>
          </div>
        </form>
        <form class="panel" id="subprojectForm">
          <h2>Subproject</h2>
          <label>Job / Order #</label><input name="job_number" placeholder="304">
          <label>Code</label><input name="code" placeholder="FC">
          <label>Name</label><input name="name" placeholder="Flow Computer">
          <label>Pricing Method</label><select name="pricing_type"><option>Fixed</option><option>T&M</option></select>
          <label>Contract Value</label><input name="contract_value" type="number" step="0.01" value="0">
          <label>Labor Hours Budget</label><input name="budget_labor_hours" type="number" step="0.01" value="0">
          <label>Budget Labor</label><input name="budget_labor" type="number" step="0.01" value="0">
          <label>Budget Material</label><input name="budget_material" type="number" step="0.01" value="0">
          <label>Budget Equipment</label><input name="budget_equipment" type="number" step="0.01" value="0">
          <div class="actions"><button class="btn primary" type="submit">Add Subproject</button></div>
        </form>
      </div>
      <form class="panel" id="coForm" style="margin-top:14px">
        <h2>Change Order / Child Project</h2>
        <div class="grid cols-4">
          <div><label>Subproject</label><select name="subproject_id" id="coSubproject"></select></div>
          <div><label>Type</label><select name="order_type"><option>Change Order</option><option>Child Project</option></select></div>
          <div><label>CO Number</label><input name="co_number" placeholder="CO-001"></div>
          <div><label>Job / Order #</label><input name="job_number" placeholder="304-CO1"></div>
          <div><label>Pricing Method</label><select name="pricing_type" id="coPricingType"><option>Fixed</option><option>T&M</option></select></div>
        </div>
        <div class="grid cols-4">
          <div><label>Status</label><select name="status"><option>Pending</option><option>Approved</option><option>Rejected</option><option>Billed</option></select></div>
          <div><label>Approved Value</label><input name="approved_value" id="coApprovedValue" type="number" step="0.01" value="0"></div>
        </div>
        <label>Title</label><input name="title" placeholder="Added instruments / wiring">
        <label>Quoted Value</label><input name="quoted_value" type="number" step="0.01" value="0">
        <div class="actions"><button class="btn primary" type="submit">Add Change Order / Child Project</button></div>
      </form>
      <div class="panel" style="margin-top:14px">
        <h2>Edit Subprojects</h2>
        <div class="table-wrap"><table id="subprojectEditTable"></table></div>
      </div>
      <div class="panel" style="margin-top:14px">
        <h2>Edit Change Orders / Child Projects</h2>
        <div class="table-wrap"><table id="changeOrderEditTable"></table></div>
      </div>
      <div class="panel" style="margin-top:14px">
        <h2>Internal Raw Rates</h2>
        <div class="grid cols-4">
          <div><label>Rate Set</label><select id="rateSetSelect"></select></div>
          <div><label>Type</label><select id="rateType"><option>Labor</option><option>Equipment</option></select></div>
          <div><label>Category</label><input id="rateCategory" placeholder="Foreman ST"></div>
          <div><label>Raw Rate</label><input id="rateRaw" type="number" step="0.01" value="0"></div>
        </div>
        <div class="actions"><button class="btn primary" id="addRate" type="button">Add Rate</button></div>
        <div class="table-wrap" style="margin-top:12px"><table id="rateEditTable"></table></div>
      </div>
    </section>

    <section id="import" class="tab hidden">
      <form class="panel" id="importForm">
        <h2>Import Field Wise</h2>
        <p class="muted">Upload one or more Field Wise job summary Excel files or Field Wise field ticket PDFs. Imported records go into Review Costs until coded.</p>
        <label>Field Wise Files</label><input type="file" name="file" accept=".xlsx,.pdf" multiple required>
        <div class="actions"><button class="btn primary" type="submit">Import</button></div>
        <p id="importResult" class="muted"></p>
      </form>
      <div class="panel" style="margin-top:14px">
        <h2>Imported Files</h2>
        <div id="importHistoryFilters" class="cost-filter" style="margin-bottom:10px"></div>
        <div class="muted" id="importHistoryCount" style="margin-bottom:8px"></div>
        <div class="table-wrap"><table id="importHistoryTable"></table></div>
      </div>
      <div class="panel" style="margin-top:14px">
        <h2>Field Ticket Lines</h2>
        <p class="muted">Expand a ticket to correct individual labor, equipment, or material lines when the ticket was written against the wrong job.</p>
        <div id="fieldTicketLineFilters"></div>
        <div class="table-wrap"><table id="fieldTicketLinesTable"></table></div>
      </div>
    </section>

    <section id="fieldwiseAudit" class="tab hidden">
      <form class="panel" id="fieldWiseAuditForm">
        <h2>Field Wise Ticket Audit</h2>
        <p class="muted">Upload the all-customer Field Wise line-item export to check tracked job/order numbers against tickets already imported here.</p>
        <label>Field Wise Ticket Export</label><input type="file" name="file" accept=".xlsx,.xlsm" required>
        <label style="display:flex;align-items:center;gap:8px"><input id="omitUntrackedAuditTickets" type="checkbox" style="width:auto" checked> Omit tickets that do not match a tracked job/order number</label>
        <div class="actions">
          <button class="btn primary" type="submit">Run Audit</button>
          <button class="btn" id="exportMissingTickets" type="button" disabled>Export Missing Tickets</button>
        </div>
        <p id="fieldWiseAuditResult" class="muted"></p>
        <div id="fieldWiseAuditSummary" class="grid cols-4" style="margin-top:12px"></div>
        <div id="fieldWiseAuditTables" style="margin-top:12px"></div>
        <div id="fieldWiseAuditOmissions" style="margin-top:12px"></div>
      </form>
    </section>

    <section id="review" class="tab hidden">
      <div class="panel">
        <h2>Review Exceptions</h2>
        <p class="muted">Only records that need attention show here: missing subproject, missing internal rate, or uncoded cost type.</p>
        <label style="display:flex;align-items:center;gap:8px;margin-bottom:10px"><input id="showAllCosts" type="checkbox" style="width:auto"> Show all cost records</label>
        <div class="table-wrap"><table id="costTable"></table></div>
      </div>
    </section>

    <section id="invoices" class="tab hidden">
      <form class="panel" id="invoiceForm">
        <h2>Vendor Invoice</h2>
        <p class="muted">Upload a readable vendor invoice PDF, or enter an invoice manually below.</p>
        <label>Vendor Invoice PDFs</label><input type="file" id="vendorInvoiceFile" accept=".pdf" multiple>
        <div class="actions"><button class="btn primary" id="importVendorInvoice" type="button">Import Vendor PDF(s)</button></div>
        <p id="vendorImportResult" class="muted"></p>
        <div class="grid cols-4">
          <div><label>Subproject</label><select name="subproject_id" id="invoiceSubproject"></select></div>
          <div><label>Change Order</label><select name="change_order_id" id="invoiceCo"></select></div>
          <div><label>Vendor</label><input name="vendor" placeholder="Graybar"></div>
          <div><label>Invoice #</label><input name="ticket_or_invoice"></div>
        </div>
        <div class="grid cols-4">
          <div><label>Date</label><input name="record_date" type="date"></div>
          <div><label>Cost Type</label><select name="cost_type"><option>Material</option><option>Rental</option><option>Equipment</option><option>Labor</option><option>Other</option></select></div>
          <div><label>Amount</label><input name="amount" type="number" step="0.01" value="0"></div>
          <div><label>Status</label><select name="status"><option>Pending</option><option>Approved</option><option>Paid</option><option>Disputed</option></select></div>
        </div>
        <label>Description</label><textarea name="description"></textarea>
        <div class="actions"><button class="btn primary" type="submit">Save Invoice</button></div>
      </form>
      <div class="panel" style="margin-top:14px">
        <h2>Vendor Invoice Lines</h2>
        <div id="vendorInvoiceLineFilters"></div>
        <div class="table-wrap"><table id="vendorInvoiceLinesTable"></table></div>
      </div>
      <div class="panel" style="margin-top:14px">
        <h2>Vendor Allocation History</h2>
        <div class="table-wrap"><table id="vendorAllocationHistoryTable"></table></div>
      </div>
    </section>

    <section id="billing" class="tab hidden">
      <form class="panel" id="customerInvoiceForm">
        <h2>Customer Invoice</h2>
        <div class="grid cols-4">
          <div><label>Subproject</label><select name="subproject_id" id="billingSubproject"></select></div>
          <div><label>Change Order</label><select name="change_order_id" id="billingCo"></select></div>
          <div><label>Invoice #</label><input name="invoice_number" placeholder="INV-1001"></div>
          <div><label>Billing Type</label><select name="billing_type"><option>Progress</option><option>Base Contract</option><option>Change Order</option><option>T&M</option><option>Retainage</option><option>Final</option></select></div>
        </div>
        <div class="grid cols-4">
          <div><label>Invoice Date</label><input name="invoice_date" type="date"></div>
          <div><label>Due Date</label><input name="due_date" type="date"></div>
          <div><label>Status</label><select name="status"><option>Draft</option><option>Sent</option><option>Partial</option><option>Paid</option><option>Overdue</option><option>Void</option></select></div>
          <div><label>Invoice Amount</label><input name="amount" type="number" step="0.01" value="0"></div>
        </div>
        <div class="grid cols-4">
          <div><label>Paid Amount</label><input name="paid_amount" type="number" step="0.01" value="0"></div>
        </div>
        <div class="panel soft-panel" style="margin-top:12px">
          <h3>Invoice Allocation</h3>
          <p class="muted">Use one line for a normal invoice, or add lines to split this customer invoice across multiple jobs.</p>
          <div id="customerInvoiceAllocations"></div>
          <div class="actions">
            <button class="btn" id="addCustomerInvoiceAllocation" type="button">Add Allocation Line</button>
            <button class="btn" id="splitCustomerInvoiceEvenly" type="button">Split Evenly</button>
          </div>
          <div class="muted" id="customerInvoiceAllocationTotal"></div>
        </div>
        <label>Our Invoice PDF</label><input name="invoice_file" type="file" accept=".pdf" required>
        <label>Notes</label><textarea name="notes" placeholder="Billing notes, customer comments, payment reference"></textarea>
        <div class="actions"><button class="btn primary" type="submit">Add Customer Invoice</button></div>
      </form>
      <div class="panel" style="margin-top:14px">
        <h2>Customer Invoice Tracking</h2>
        <div id="customerInvoiceSummary"></div>
        <div class="table-wrap" style="margin-top:12px"><table id="customerInvoiceTable"></table></div>
      </div>
    </section>

    <section id="projectPo" class="tab hidden">
      <div class="panel">
        <div style="display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap">
          <div>
            <h2>Project POs</h2>
            <p class="muted">Purchase orders tied to the selected master project.</p>
          </div>
          <button class="btn" id="refreshProjectPos" type="button">Refresh</button>
        </div>
        <div class="grid cols-4">
          <div><label>Status</label><select id="projectPoStatusFilter"><option value="">All statuses</option><option>Pending Approval</option><option>Issued</option><option>Picked Up</option><option>Closed</option><option>Void</option></select></div>
          <div style="grid-column:span 3"><label>Search</label><input id="projectPoSearch" placeholder="Search PO, job/order, vendor, requester, or description"></div>
        </div>
        <div class="muted" id="projectPoCount" style="margin-top:8px"></div>
        <div class="table-wrap" style="margin-top:10px"><table id="projectPoTable"></table></div>
      </div>
    </section>

    <section id="bids" class="tab hidden">
      <div class="panel">
        <h2>Bid Tracking</h2>
        <div class="grid cols-4" id="bidKpis"></div>
      </div>
      <div class="grid cols-2" style="margin-top:14px">
        <div class="panel"><h2>Pipeline By Stage</h2><div id="bidStageSummary"></div></div>
        <div class="panel"><h2>Estimator Workload</h2><div id="bidEstimatorSummary"></div></div>
      </div>
      <form class="panel" id="bidForm" style="margin-top:14px">
        <h2>Add Bid / RFQ</h2>
        <div class="grid cols-4">
          <div><label>RFQ No.</label><input name="rfq_no" placeholder="RFQ-1012" required></div>
          <div><label>Date Received</label><input name="date_received" type="date"></div>
          <div><label>Customer</label><input name="customer"></div>
          <div><label>Project Name</label><input name="project_name" required></div>
        </div>
        <div class="grid cols-4">
          <div><label>Estimator</label><input name="estimator" placeholder="Ross Stewart"></div>
          <div><label>Stage</label><select name="stage"><option>New RFQ</option><option>Go/No-Go</option><option>Estimating</option><option>Submitted</option><option>Award Pending</option><option>Closed</option></select></div>
          <div><label>Bid Due Date</label><input name="bid_due_date" type="date"></div>
          <div><label>Go / No-Go</label><select name="go_no_go"><option>Go</option><option>No Go</option><option>Review</option></select></div>
        </div>
        <div class="grid cols-4">
          <div><label>Estimated Cost</label><input name="estimated_cost" type="number" step="0.01" value="0"></div>
          <div><label>Target Margin</label><input name="target_margin" type="number" step="0.01" value="0.25"></div>
          <div><label>Probability</label><input name="probability" type="number" step="0.01" value="0.25"></div>
          <div><label>Outcome</label><select name="outcome"><option>Pending</option><option>Won</option><option>Lost</option></select></div>
        </div>
        <label>Notes</label><textarea name="notes"></textarea>
        <div class="actions"><button class="btn primary" type="submit">Add Bid</button></div>
      </form>
      <div class="panel" style="margin-top:14px">
        <h2>Bid List</h2>
        <div id="bidFilters"></div>
        <div class="table-wrap"><table id="bidTable"></table></div>
      </div>
    </section>

    <section id="admin" class="tab hidden">
      <div class="grid cols-2">
        <form class="panel" id="userForm">
          <h2>Add User</h2>
          <label>Username</label><input name="username" required>
          <label>Display Name</label><input name="display_name">
          <label>Role</label><select name="role"><option>User</option><option>Read Only</option><option>TX/Read Only</option><option>Field PO</option><option>Admin</option></select>
          <label><input name="po_auto_issue" type="checkbox" value="1"> Auto-issue this user's POs</label>
          <p class="muted">New users start with temporary password TPE1776 and must change it at first login.</p>
          <div class="actions"><button class="btn primary" type="submit">Add User</button></div>
        </form>
        <div class="panel">
          <h2>Users</h2>
          <div class="table-wrap"><table id="usersTable"></table></div>
        </div>
        <div class="panel" style="grid-column:1 / -1">
          <h2>Role Permissions</h2>
          <p class="muted">Control what each role can see or change. Admin is always full access so you cannot lock yourself out.</p>
          <div id="rolePermissions"></div>
        </div>
        <div class="panel">
          <h2>Manual Backup</h2>
          <p class="muted">Download a snapshot of the SQLite database before major imports, server updates, or cleanup work.</p>
          <div class="actions"><a class="btn primary" href="/api/admin/database-backup">Download Database Backup</a></div>
        </div>
        <div class="panel">
          <h2>Activity Log</h2>
          <p class="muted">Review sensitive changes such as deletes, user updates, reminder changes, and audit exceptions.</p>
          <div class="actions"><button class="btn primary" data-open-tab="activity" type="button">Open Activity Log</button></div>
        </div>
      </div>
    </section>

    <section id="activity" class="tab hidden">
      <div class="panel">
        <h2>Activity Log</h2>
        <div class="grid cols-4">
          <div><label>Search</label><input id="activitySearch" placeholder="User, action, target, details"></div>
          <div><label>Action</label><select id="activityAction"><option value="">All actions</option></select></div>
          <div><label>Limit</label><select id="activityLimit"><option>50</option><option selected>100</option><option>250</option><option>500</option></select></div>
        </div>
        <div class="actions"><button class="btn" id="refreshActivity" type="button">Refresh</button></div>
        <div class="muted" id="activityCount" style="margin-top:10px"></div>
        <div class="table-wrap" style="margin-top:10px"><table id="activityTable"></table></div>
      </div>
    </section>
      </div>
    </div>
  </main>
  <div class="modal-backdrop hidden" id="unsavedModal">
    <div class="modal">
      <h2>Unsaved changes</h2>
      <p>You have changes that have not been saved. Save them before leaving, or discard the changes and continue.</p>
      <div class="actions">
        <button class="btn primary" id="stayOnPage" type="button">Stay Here</button>
        <button class="btn" id="discardChanges" type="button">Discard Changes</button>
      </div>
    </div>
  </div>
  <div class="modal-backdrop hidden" id="trendModal">
    <div class="modal large">
      <div style="display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:10px">
        <div>
          <h2 style="margin:0">Profitability Trend</h2>
          <p id="trendModalSubtitle"></p>
        </div>
        <button class="btn" id="closeTrendModal" type="button">Close</button>
      </div>
      <div class="trend-modal-body" id="trendModalBody"></div>
    </div>
  </div>
  <div class="modal-backdrop hidden" id="financialDuplicateModal">
    <div class="modal">
      <h2>Duplicate Report Skipped</h2>
      <p>The following Texas financial report upload was skipped because it already exists.</p>
      <div id="financialDuplicateList"></div>
      <div class="actions">
        <button class="btn primary" id="closeFinancialDuplicateModal" type="button">OK</button>
      </div>
    </div>
  </div>
  <div class="modal-backdrop hidden" id="accountModal">
    <div class="modal">
      <h2>Change Password</h2>
      <p id="accountPasswordMessage" class="muted"></p>
      <form id="changePasswordForm">
        <label>Current Password</label><input name="current_password" type="password" autocomplete="current-password" required>
        <label>New Password</label><input name="new_password" type="password" autocomplete="new-password" required>
        <label>Confirm New Password</label><input name="confirm_password" type="password" autocomplete="new-password" required>
        <div class="error" id="changePasswordError"></div>
        <div class="actions">
          <button class="btn primary" type="submit">Save Password</button>
          <button class="btn" id="closeAccountModal" type="button">Cancel</button>
        </div>
      </form>
    </div>
  </div>
  <div class="modal-backdrop hidden" id="copySubprojectModal">
    <div class="modal">
      <h2>Copy Subproject</h2>
      <p id="copySubprojectMessage"></p>
      <form id="copySubprojectForm">
        <input type="hidden" name="subproject_id">
        <label>New Job / Order #<input name="job_number" required></label>
        <label>Code<input name="code"></label>
        <label>Name<input name="name" required></label>
        <label>Pricing
          <select name="pricing_type">
            <option>Fixed</option>
            <option>T&amp;M</option>
          </select>
        </label>
        <label>Contract Value<input name="contract_value" type="number" step="0.01"></label>
        <label>Labor Hours Budget<input name="budget_labor_hours" type="number" step="0.01"></label>
        <label>Labor $ Budget<input name="budget_labor" type="number" step="0.01"></label>
        <label>Material Budget<input name="budget_material" type="number" step="0.01"></label>
        <label>Equipment Budget<input name="budget_equipment" type="number" step="0.01"></label>
        <div id="copySubprojectError" class="bad"></div>
        <div class="actions">
          <button class="btn primary" type="submit">Create Copy</button>
          <button class="btn" id="closeCopySubprojectModal" type="button">Cancel</button>
        </div>
      </form>
    </div>
  </div>
  <div class="modal-backdrop hidden" id="copyChangeOrderModal">
    <div class="modal">
      <h2>Copy Change Order / Child Project</h2>
      <p id="copyChangeOrderMessage"></p>
      <form id="copyChangeOrderForm">
        <input type="hidden" name="change_order_id">
        <label>Subproject<select name="subproject_id"></select></label>
        <label>Type
          <select name="order_type">
            <option>Change Order</option>
            <option>Child Project</option>
          </select>
        </label>
        <label>New CO Number<input name="co_number" required></label>
        <label>New Job / Order #<input name="job_number" required></label>
        <label>Pricing
          <select name="pricing_type">
            <option>Fixed</option>
            <option>T&amp;M</option>
          </select>
        </label>
        <label>Status
          <select name="status">
            <option>Pending</option>
            <option>Approved</option>
            <option>Rejected</option>
            <option>Billed</option>
          </select>
        </label>
        <label>Title<input name="title"></label>
        <label>Quoted Value<input name="quoted_value" type="number" step="0.01"></label>
        <label>Approved Value<input name="approved_value" type="number" step="0.01"></label>
        <div id="copyChangeOrderError" class="bad"></div>
        <div class="actions">
          <button class="btn primary" type="submit">Create Copy</button>
          <button class="btn" id="closeCopyChangeOrderModal" type="button">Cancel</button>
        </div>
      </form>
    </div>
  </div>
  <div class="modal-backdrop hidden" id="vendorAllocationModal">
    <div class="modal large">
      <div style="display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:10px">
        <div>
          <h2 style="margin:0">Allocate Vendor Invoice</h2>
          <p id="vendorAllocationSubtitle" class="muted"></p>
        </div>
        <button class="btn" id="closeVendorAllocationModal" type="button">Close</button>
      </div>
      <form id="vendorAllocationForm">
        <div id="vendorAllocationTargets"></div>
        <div class="error" id="vendorAllocationError"></div>
        <div class="actions">
          <button class="btn primary" type="submit">Split Evenly</button>
        </div>
      </form>
    </div>
  </div>

  <script>
    const PO_FEATURE_ENABLED = true;
    const COG_CUSTOMER_OPTIONAL_NAMES = new Set(['expenses', 'truck & auto expense', 'truck and auto expense']);
    let state = { projects: [], projectId: null, subprojects: [], changeOrders: [], cogCategories: [], internalRates: [], rateSets: [], currentUser: null };
    let openSubprojectDetailId = null;
    let masterDetailIsOpen = false;
    let hasUnsavedChanges = false;
    let selectedDashboardSubprojectId = null;
    let selectedDashboardChangeOrderId = null;
    let invoiceGroupSeq = 0;
    let costFilterSeq = 0;
    let vendorAllocationGroups = {};
    let jobOrderReportRows = [];
    let jobOrderSort = { field: 'job_number', direction: 'asc' };
    let fieldPoJobChoices = [];
    let officePoRows = [];
    let projectPoRows = [];
    let rolePermissionsData = null;
    let fieldWiseAuditData = null;
    let fieldWiseAuditOmissions = [];
    const collapsedHierarchyNodes = new Set();
    const initializedHierarchyNodes = new Set();
    const money = v => Number(v || 0).toLocaleString(undefined, { style: 'currency', currency: 'USD' });
    const pct = v => `${(Number(v || 0) * 100).toFixed(1)}%`;
    const htmlEscape = v => String(v ?? '').replace(/[&<>"']/g, ch => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[ch]));
    const help = text => `<span class="help-marker" tabindex="0" aria-label="${htmlEscape(text)}">?<span class="help-popover">${htmlEscape(text)}</span></span>`;
    const TAB_PERMISSIONS = {
      dashboard: 'projects',
      newMasterProject: 'project_setup',
      setup: 'project_setup',
      import: 'fieldwise',
      review: 'review_exceptions',
      invoices: 'vendor_invoices',
      billing: 'customer_billing',
      projectPo: 'po_review',
      fieldwiseAudit: 'fieldwise_audit',
      bids: 'bids',
      jobOrderReport: 'job_order_report',
      archivedProjects: 'archived_projects',
      texasOps: 'texas_ops',
      texasReminders: 'texas_reminders',
      fieldPo: 'po_requests',
      officePo: 'po_review',
      cogSetup: 'cog_setup',
      activity: 'activity',
      admin: 'admin'
    };
    const plainTable = (headers, rows) => `<table><thead><tr>${headers.map(h => `<th>${h}</th>`).join('')}</tr></thead><tbody>${rows.map(r => `<tr>${r.map(c => `<td>${c}</td>`).join('')}</tr>`).join('')}</tbody></table>`;
    const api = async (url, opts={}) => {
      const res = await fetch(url, opts);
      if (res.status === 401) {
        window.location.href = '/login';
        throw new Error('Login required');
      }
      const data = await res.json();
      if (!res.ok) throw new Error(data.error || 'Request failed');
      return data;
    };
    const formDataObj = form => Object.fromEntries(new FormData(form).entries());
    const markSaved = () => { hasUnsavedChanges = false; };
    const isTexasReadOnly = () => state.currentUser?.role === 'TX/Read Only';
    const isFieldPoOnly = () => state.currentUser?.role === 'Field PO';
    const canViewPermission = key => {
      if (!key || state.currentUser?.role === 'Admin') return true;
      return !!Number(state.currentUser?.permissions?.[key]?.can_view || 0);
    };
    const canEditPermission = key => {
      if (!key || state.currentUser?.role === 'Admin') return true;
      return !!Number(state.currentUser?.permissions?.[key]?.can_edit || 0);
    };
    const canViewTab = tabName => canViewPermission(TAB_PERMISSIONS[tabName]);
    const isTemporaryViewControl = target => {
      return target.closest('.cost-filter') || target.closest('#financialProfitTrend') || target.id === 'showAllCosts';
    };
    const markUnsaved = event => {
      if (isTemporaryViewControl(event.target)) return;
      if (event.target.matches('input, textarea, select') && event.target.type !== 'file' && event.target.id !== 'projectSelect') {
        hasUnsavedChanges = true;
      }
    };
    const confirmDiscard = () => {
      if (!hasUnsavedChanges) return Promise.resolve(true);
      const modal = document.getElementById('unsavedModal');
      modal.classList.remove('hidden');
      return new Promise(resolve => {
        document.getElementById('stayOnPage').onclick = () => {
          modal.classList.add('hidden');
          resolve(false);
        };
        document.getElementById('discardChanges').onclick = async () => {
          modal.classList.add('hidden');
          markSaved();
          document.querySelectorAll('form').forEach(form => form.reset());
          await refresh();
          resolve(true);
        };
      });
    };
    window.addEventListener('beforeunload', event => {
      if (!hasUnsavedChanges) return;
      event.preventDefault();
      event.returnValue = '';
    });
    document.addEventListener('input', markUnsaved);
    document.addEventListener('change', markUnsaved);
    document.addEventListener('click', event => {
      if (event.target.closest('.help-marker')) event.stopPropagation();
    }, true);
    document.addEventListener('keydown', event => {
      if (!event.target.closest('.help-marker')) return;
      if (event.key === 'Enter' || event.key === ' ') {
        event.preventDefault();
        event.stopPropagation();
      }
    }, true);
    document.getElementById('closeTrendModal').onclick = () => {
      document.getElementById('trendModal').classList.add('hidden');
    };
    document.getElementById('trendModal').onclick = event => {
      if (event.target.id === 'trendModal') event.currentTarget.classList.add('hidden');
    };
    document.getElementById('closeFinancialDuplicateModal').onclick = () => {
      document.getElementById('financialDuplicateModal').classList.add('hidden');
    };
    document.getElementById('financialDuplicateModal').onclick = event => {
      if (event.target.id === 'financialDuplicateModal') event.currentTarget.classList.add('hidden');
    };
    function openAccountModal(force=false) {
      const modal = document.getElementById('accountModal');
      const message = document.getElementById('accountPasswordMessage');
      message.textContent = force ? 'You are using a temporary password. Change it before continuing.' : 'Update your own login password.';
      document.getElementById('closeAccountModal').classList.toggle('hidden', force);
      document.getElementById('changePasswordError').textContent = '';
      document.getElementById('changePasswordForm').reset();
      modal.dataset.force = force ? '1' : '0';
      modal.classList.remove('hidden');
      setTimeout(() => document.querySelector('#changePasswordForm input[name="current_password"]')?.focus(), 50);
    }
    function closeAccountModal() {
      if (document.getElementById('accountModal').dataset.force === '1') return;
      document.getElementById('accountModal').classList.add('hidden');
    }
    document.getElementById('closeAccountModal').onclick = closeAccountModal;
    document.getElementById('accountModal').onclick = event => {
      if (event.target.id === 'accountModal') closeAccountModal();
    };
    function closeVendorAllocationModal() {
      document.getElementById('vendorAllocationModal').classList.add('hidden');
    }
    document.getElementById('closeVendorAllocationModal').onclick = closeVendorAllocationModal;
    document.getElementById('vendorAllocationModal').onclick = event => {
      if (event.target.id === 'vendorAllocationModal') closeVendorAllocationModal();
    };
    function closeCopySubprojectModal() {
      document.getElementById('copySubprojectModal').classList.add('hidden');
    }
    document.getElementById('closeCopySubprojectModal').onclick = closeCopySubprojectModal;
    document.getElementById('copySubprojectModal').onclick = event => {
      if (event.target.id === 'copySubprojectModal') closeCopySubprojectModal();
    };
    function closeCopyChangeOrderModal() {
      document.getElementById('copyChangeOrderModal').classList.add('hidden');
    }
    document.getElementById('closeCopyChangeOrderModal').onclick = closeCopyChangeOrderModal;
    document.getElementById('copyChangeOrderModal').onclick = event => {
      if (event.target.id === 'copyChangeOrderModal') closeCopyChangeOrderModal();
    };
    document.addEventListener('keydown', event => {
      if (event.key === 'Escape') document.getElementById('trendModal').classList.add('hidden');
      if (event.key === 'Escape') document.getElementById('financialDuplicateModal').classList.add('hidden');
      if (event.key === 'Escape') closeAccountModal();
      if (event.key === 'Escape') closeVendorAllocationModal();
      if (event.key === 'Escape') closeCopySubprojectModal();
      if (event.key === 'Escape') closeCopyChangeOrderModal();
    });

    function navOptionsForTab(tabName) {
      const projectTabs = ['dashboard','newMasterProject','setup','import','review','invoices','billing','projectPo','archivedProjects'];
      const poTabs = ['fieldPo','officePo','cogSetup'];
      const homeTabs = ['home','bids','jobOrderReport','admin','activity'];
      if (projectTabs.includes(tabName)) return 'project';
      if (poTabs.includes(tabName)) return 'po';
      if (tabName === 'fieldwiseAudit') return 'audit';
      if (['texasOps','texasReminders'].includes(tabName)) return 'texas';
      if (homeTabs.includes(tabName)) return 'home';
      return 'home';
    }

    function attentionCountText(alert) {
      if (alert.key === 'missing_texas_pnl' || alert.key === 'missing_balance_sheet') {
        return Number(alert.count || 0) ? 'Missing' : 'OK';
      }
      return Number(alert.count || 0).toLocaleString();
    }

    function attentionItemsHtml(alert) {
      if (alert.key !== 'stale_job_invoices' || !Array.isArray(alert.items) || !alert.items.length) return '';
      return `<div class="attention-list">${alert.items.map(item => `
        <div class="attention-item">
          <div>
            <strong>${htmlEscape(item.label || 'Job')}</strong>
            <span>${htmlEscape([item.customer, item.project_name, item.job_type].filter(Boolean).join(' / '))}${item.last_invoice_added_at ? ` / last invoice added ${htmlEscape(item.last_invoice_added_at.replace('T', ' '))}` : ''}${item.acknowledged_at ? ` / acknowledged ${htmlEscape(item.acknowledged_at.replace('T', ' '))}` : ''}</span>
          </div>
          <button class="btn" data-ack-job-invoice="${htmlEscape(item.target_id)}" data-ack-job-type="${htmlEscape(item.target_type)}" type="button">Acknowledge</button>
        </div>
      `).join('')}</div>`;
    }

    async function openAttentionTab(tabName) {
      if (!(await confirmDiscard())) return;
      openTab(tabName);
    }

    async function loadHomeAlerts() {
      const target = document.getElementById('homeAlerts');
      if (!target) return;
      target.innerHTML = '<div class="muted">Checking alerts...</div>';
      try {
        const summary = await api('/api/home-alerts');
        const alerts = (summary.alerts || []).filter(alert => canViewTab(alert.tab || 'home'));
        target.innerHTML = alerts.length
          ? alerts.map(alert => `
              <div class="attention-card ${alert.severity || 'ok'}" data-alert-tab="${htmlEscape(alert.tab || 'home')}" role="button" tabindex="0">
                <div>
                  <div class="label">${htmlEscape(alert.label)}</div>
                  <div class="count">${htmlEscape(attentionCountText(alert))}</div>
                  <div class="detail">${htmlEscape(alert.detail || '')}</div>
                </div>
                <div>
                  ${alert.amount ? `<div class="amount">${money(alert.amount)}</div>` : ''}
                  <div class="open-link">Open</div>
                </div>
                ${attentionItemsHtml(alert)}
              </div>
            `).join('')
          : '<div class="muted">No dashboard alerts right now.</div>';
        target.querySelectorAll('[data-alert-tab]').forEach(card => {
          card.onclick = () => openAttentionTab(card.dataset.alertTab);
          card.onkeydown = event => {
            if (event.key === 'Enter' || event.key === ' ') {
              event.preventDefault();
              card.click();
            }
          };
        });
        target.querySelectorAll('[data-ack-job-invoice]').forEach(btn => {
          btn.onclick = async event => {
            event.stopPropagation();
            btn.disabled = true;
            btn.textContent = 'Saving...';
            try {
              await api('/api/home-alerts/job-invoice-ack', { method:'POST', body: JSON.stringify({ target_type: btn.dataset.ackJobType, target_id: btn.dataset.ackJobInvoice }) });
              await loadHomeAlerts();
            } catch (err) {
              btn.disabled = false;
              btn.textContent = 'Acknowledge';
              window.alert(err.message || 'Could not acknowledge this job.');
            }
          };
        });
      } catch (err) {
        target.innerHTML = `<div class="bad">Could not load dashboard alerts: ${htmlEscape(err.message)}</div>`;
      }
    }

    function applyPermissionVisibility() {
      document.querySelectorAll('[data-open-tab]').forEach(el => {
        const tab = el.dataset.openTab;
        el.classList.toggle('hidden', tab !== 'home' && !canViewTab(tab));
      });
      document.querySelectorAll('nav button[data-tab]').forEach(btn => {
        btn.classList.toggle('hidden', btn.dataset.tab !== 'home' && !canViewTab(btn.dataset.tab));
      });
      document.getElementById('systemAdminBtn')?.classList.toggle('hidden', !canViewTab('admin'));
      document.getElementById('systemRevisionBtn')?.classList.toggle('hidden', state.currentUser?.role !== 'Admin');
    }

    function updateNavForTab(tabName) {
      const activeGroup = navOptionsForTab(tabName);
      document.querySelector('.project-switcher')?.classList.toggle('hidden', activeGroup !== 'project');
      document.querySelectorAll('nav button').forEach(btn => {
        const tab = btn.dataset.tab;
        const isDisabledPo = !PO_FEATURE_ENABLED && ['fieldPo', 'officePo', 'projectPo'].includes(tab);
        btn.classList.toggle('hidden', !!isDisabledPo || !canViewTab(tab));
      });
      document.querySelectorAll('.nav-subgroup').forEach(group => {
        group.classList.toggle('hidden', group.dataset.navGroup !== activeGroup);
      });
      document.querySelector('nav button[data-tab="home"]')?.classList.toggle('active', activeGroup === 'home');
      document.querySelector('nav button[data-tab="dashboard"]')?.classList.toggle('active', activeGroup === 'project');
      document.querySelector('nav button[data-tab="fieldwiseAudit"]')?.classList.toggle('active', activeGroup === 'audit');
      document.querySelector('nav button[data-tab="fieldPo"][data-nav-level="main"]')?.classList.toggle('active', activeGroup === 'po');
      document.querySelector('nav button[data-tab="texasOps"]')?.classList.toggle('active', activeGroup === 'texas');
    }

    async function openTab(tabName) {
      if (tabName !== 'home' && !canViewTab(tabName)) tabName = 'home';
      if (isTexasReadOnly() && tabName !== 'texasOps') tabName = 'texasOps';
      if (isFieldPoOnly() && tabName !== 'fieldPo') tabName = 'fieldPo';
      if (!PO_FEATURE_ENABLED && ['fieldPo', 'officePo', 'projectPo'].includes(tabName)) tabName = 'home';
      markSaved();
      document.querySelectorAll('nav button').forEach(b => b.classList.remove('active'));
      updateNavForTab(tabName);
      document.querySelectorAll('.tab').forEach(t => t.classList.add('hidden'));
      document.querySelectorAll(`nav button[data-tab="${tabName}"]`).forEach(navButton => navButton.classList.add('active'));
      document.getElementById(tabName).classList.remove('hidden');
      if (tabName === 'home') loadHomeAlerts();
      if (tabName === 'dashboard') refreshOpenDetails();
      if (tabName === 'review') loadCosts();
      if (tabName === 'import') { loadImportHistory(); loadFieldTicketLines(); }
      if (tabName === 'fieldwiseAudit') loadFieldWiseAuditOmissions();
      if (tabName === 'invoices') { loadVendorInvoiceLines(); loadVendorAllocationHistory(); }
      if (tabName === 'billing') loadCustomerInvoices();
      if (tabName === 'projectPo') loadProjectPos();
      if (tabName === 'newMasterProject') loadNewMasterProjectForm();
      if (tabName === 'bids') loadBidDashboard();
      if (tabName === 'admin') loadUsers();
      if (tabName === 'activity') loadActivityLog();
      if (tabName === 'texasOps') loadTexasOpsDashboard();
      if (tabName === 'texasReminders') loadTexasReminderSetup();
      if (tabName === 'archivedProjects') loadArchivedProjects();
      if (tabName === 'jobOrderReport') loadJobOrderReport();
      if (tabName === 'fieldPo') loadFieldPo();
      if (tabName === 'officePo') loadOfficePos();
      if (tabName === 'cogSetup') loadCogSetup();
    }

    document.querySelectorAll('[data-open-tab]').forEach(btn => btn.onclick = async () => {
      if (!(await confirmDiscard())) return;
      openTab(btn.dataset.openTab);
    });
    document.querySelectorAll('nav button[data-tab]').forEach(btn => btn.onclick = async () => {
      if (!(await confirmDiscard())) return;
      openTab(btn.dataset.tab);
    });
    document.querySelectorAll('.home-card[data-open-tab]').forEach(card => card.onkeydown = async event => {
      if (!['Enter', ' '].includes(event.key)) return;
      event.preventDefault();
      if (!(await confirmDiscard())) return;
      openTab(card.dataset.openTab);
    });
    document.getElementById('refreshHomeAlerts').onclick = loadHomeAlerts;
    document.getElementById('homeLogo').onclick = async () => {
      if (!(await confirmDiscard())) return;
      openTab('home');
    };
    document.getElementById('systemMenuButton').onclick = event => {
      event.stopPropagation();
      document.getElementById('systemMenu').classList.toggle('hidden');
    };
    document.addEventListener('click', event => {
      if (!event.target.closest('.system-menu-wrap')) {
        document.getElementById('systemMenu').classList.add('hidden');
      }
    });
    document.getElementById('systemAdminBtn').onclick = async () => {
      document.getElementById('systemMenu').classList.add('hidden');
      if (!(await confirmDiscard())) return;
      openTab('admin');
    };
    document.getElementById('systemAccountBtn').onclick = () => {
      document.getElementById('systemMenu').classList.add('hidden');
      openAccountModal(false);
    };
    document.getElementById('systemRevisionBtn').onclick = async () => {
      document.getElementById('systemMenu').classList.add('hidden');
      if (!(await confirmDiscard())) return;
      window.location.href = '/server-health';
    };
    document.getElementById('systemLogoutBtn').onclick = async () => {
      await fetch('/api/logout', { method:'POST' });
      window.location.href = '/login';
    };
    document.getElementById('changePasswordForm').onsubmit = async event => {
      event.preventDefault();
      const form = event.target;
      const data = formDataObj(form);
      const error = document.getElementById('changePasswordError');
      error.textContent = '';
      if (data.new_password !== data.confirm_password) {
        error.textContent = 'New passwords do not match.';
        return;
      }
      if (String(data.new_password || '').length < 8) {
        error.textContent = 'Use at least 8 characters.';
        return;
      }
      if (data.new_password === 'TPE1776') {
        error.textContent = 'Choose a password different from the temporary password.';
        return;
      }
      try {
        await api('/api/change-password', { method:'POST', body: JSON.stringify(data) });
        document.getElementById('accountModal').dataset.force = '0';
        document.getElementById('accountModal').classList.add('hidden');
        await loadCurrentUser();
      } catch (err) {
        error.textContent = err.message;
      }
    };
    document.getElementById('vendorAllocationForm').onsubmit = async event => {
      event.preventDefault();
      const form = event.target;
      const groupId = form.dataset.groupId;
      const group = vendorAllocationGroups[groupId];
      const error = document.getElementById('vendorAllocationError');
      error.textContent = '';
      const targets = [...form.querySelectorAll('input[name="allocation_target"]:checked')].map(input => input.value);
      if (!group) {
        error.textContent = 'Invoice group is no longer available. Refresh and try again.';
        return;
      }
      if (targets.length < 2) {
        error.textContent = 'Choose at least two jobs to split across.';
        return;
      }
      await api('/api/vendor-invoice/allocate', {
        method: 'POST',
        body: JSON.stringify({
          project_id: state.projectId,
          source_file: group.source_file,
          ticket_or_invoice: group.ticket_or_invoice,
          vendor: group.vendor,
          targets
        })
      });
      closeVendorAllocationModal();
      markSaved();
      await refresh();
      await loadVendorInvoiceLines();
      await loadVendorAllocationHistory();
    };

    function metricAmount(summary, key) {
      return Number(summary.latest_metrics?.[key]?.amount || 0);
    }

    function trendBars(points, colorClass='') {
      if (!points || !points.length) return '<div class="muted">No history yet.</div>';
      const max = Math.max(...points.map(p => Math.abs(Number(p.amount || 0))), 1);
      return `<div>${points.slice(-8).map(p => {
        const width = Math.max(4, Math.min(100, Math.abs(Number(p.amount || 0)) / max * 100));
        const negative = Number(p.amount || 0) < 0;
        return `<div style="margin:8px 0"><div style="display:flex;justify-content:space-between;gap:10px"><span>${htmlEscape(p.report_date)}</span><strong class="${negative ? 'bad' : colorClass}">${money(p.amount)}</strong></div><div class="bar"><span style="width:${width}%;background:${negative ? 'var(--red)' : 'var(--blue)'}"></span></div></div>`;
      }).join('')}</div>`;
    }

    function defaultFinancialTrendRange() {
      const year = String(new Date().getFullYear());
      return { start: `${year}-01-01`, end: `${year}-12-31` };
    }

    function financialTrendRange() {
      const defaults = defaultFinancialTrendRange();
      return {
        start: localStorage.getItem('financialTrendStart') || defaults.start,
        end: localStorage.getItem('financialTrendEnd') || defaults.end,
      };
    }

    function rangeLabel(range) {
      return `${range.start || 'Start'} to ${range.end || 'End'}`;
    }

    function dateRangePoints(points, range=financialTrendRange()) {
      return (points || [])
        .filter(p => {
          const date = String(p.report_date || '');
          return (!range.start || date >= range.start) && (!range.end || date <= range.end);
        })
        .sort((a, b) => String(a.report_date || '').localeCompare(String(b.report_date || '')));
    }

    function linePath(points, xFor, yFor) {
      return points.map((p, i) => `${i ? 'L' : 'M'} ${xFor(p)} ${yFor(p)}`).join(' ');
    }

    function metricTrendValue(metricKey, value) {
      return ['current_ratio', 'future_current_ratio'].includes(metricKey) ? Number(value || 0).toFixed(2) : money(value);
    }

    function singleMetricTrendGraph(points, metricKey, label, range=financialTrendRange(), options={}) {
      const metricPoints = dateRangePoints(points || [], range);
      if (!metricPoints.length) return `<div class="muted">No ${htmlEscape(label)} history in ${htmlEscape(rangeLabel(range))}.</div>`;
      const width = options.width || 1120;
      const height = options.height || 560;
      const pad = options.pad || { left: metricKey === 'current_ratio' ? 70 : 92, right: 34, top: 30, bottom: 60 };
      const dates = [...new Set(metricPoints.map(p => p.report_date))].sort();
      const values = metricPoints.map(p => Number(p.amount || 0));
      let min = Math.min(0, ...values);
      let max = Math.max(0, ...values);
      if (min === max) { min -= 1; max += 1; }
      const plotW = width - pad.left - pad.right;
      const plotH = height - pad.top - pad.bottom;
      const xDate = d => pad.left + (dates.length === 1 ? plotW / 2 : dates.indexOf(d) / (dates.length - 1) * plotW);
      const yVal = v => pad.top + (max - Number(v || 0)) / (max - min) * plotH;
      const xFor = p => xDate(p.report_date);
      const yFor = p => yVal(p.amount);
      const zeroY = yVal(0);
      const yTicks = [max, (max + min) / 2, min];
      const dateLabels = dates.map(d => {
        const date = new Date(`${d}T00:00:00`);
        return Number.isNaN(date.getTime()) ? d : date.toLocaleDateString(undefined, { month: 'short', day: 'numeric' });
      });
      const circles = metricPoints.map(p => `<circle class="trend-point" cx="${xFor(p)}" cy="${yFor(p)}" r="5" fill="var(--blue)"><title>${p.report_date}: ${metricTrendValue(metricKey, p.amount)}</title></circle>`).join('');
      return `<svg class="trend-chart trend-chart-large" viewBox="0 0 ${width} ${height}" role="img" aria-label="${htmlEscape(label)} trend graph">
        ${yTicks.map(t => `<line class="trend-axis" x1="${pad.left}" x2="${width - pad.right}" y1="${yVal(t)}" y2="${yVal(t)}"></line><text class="trend-label" x="8" y="${yVal(t) + 4}">${htmlEscape(metricTrendValue(metricKey, t))}</text>`).join('')}
        <line class="trend-axis" x1="${pad.left}" x2="${width - pad.right}" y1="${zeroY}" y2="${zeroY}"></line>
        ${dates.map((d, i) => `<text class="trend-label" x="${xDate(d)}" y="${height - 18}" text-anchor="middle">${htmlEscape(dateLabels[i])}</text>`).join('')}
        <path class="trend-line-revenue" d="${linePath(metricPoints, xFor, yFor)}"></path>
        ${circles}
      </svg>
      <div class="trend-legend"><span><span class="trend-swatch" style="background:var(--blue)"></span>${htmlEscape(label)}</span><span>${htmlEscape(rangeLabel(range))}</span></div>`;
    }

    function trendLineGraph(summary, range=financialTrendRange(), options={}) {
      const netPoints = dateRangePoints(summary.history?.net_income || [], range);
      const revenuePoints = dateRangePoints(summary.history?.revenue || [], range);
      const allPoints = [...netPoints, ...revenuePoints];
      if (!allPoints.length) return `<div class="muted">No P&L history in ${htmlEscape(rangeLabel(range))}.</div>`;
      const width = options.width || 720;
      const height = options.height || 300;
      const pad = options.pad || { left: 78, right: 22, top: 24, bottom: 46 };
      const dates = [...new Set(allPoints.map(p => p.report_date))].sort();
      const values = allPoints.map(p => Number(p.amount || 0));
      let min = Math.min(0, ...values);
      let max = Math.max(0, ...values);
      if (min === max) { min -= 1; max += 1; }
      const plotW = width - pad.left - pad.right;
      const plotH = height - pad.top - pad.bottom;
      const xDate = d => pad.left + (dates.length === 1 ? plotW / 2 : dates.indexOf(d) / (dates.length - 1) * plotW);
      const yVal = v => pad.top + (max - Number(v || 0)) / (max - min) * plotH;
      const xFor = p => xDate(p.report_date);
      const yFor = p => yVal(p.amount);
      const zeroY = yVal(0);
      const yTicks = [max, (max + min) / 2, min];
      const dateLabels = dates.map(d => {
        const date = new Date(`${d}T00:00:00`);
        return Number.isNaN(date.getTime()) ? d : date.toLocaleDateString(undefined, { month: 'short', day: 'numeric' });
      });
      const circles = (points, color) => points.map(p => `<circle class="trend-point" cx="${xFor(p)}" cy="${yFor(p)}" r="4" fill="${color}"><title>${p.report_date}: ${money(p.amount)}</title></circle>`).join('');
      return `<svg class="trend-chart ${options.large ? 'trend-chart-large' : ''}" viewBox="0 0 ${width} ${height}" role="img" aria-label="Profitability trend graph">
        ${yTicks.map(t => `<line class="trend-axis" x1="${pad.left}" x2="${width - pad.right}" y1="${yVal(t)}" y2="${yVal(t)}"></line><text class="trend-label" x="8" y="${yVal(t) + 4}">${money(t)}</text>`).join('')}
        <line class="trend-axis" x1="${pad.left}" x2="${width - pad.right}" y1="${zeroY}" y2="${zeroY}"></line>
        ${dates.map((d, i) => `<text class="trend-label" x="${xDate(d)}" y="${height - 18}" text-anchor="middle">${htmlEscape(dateLabels[i])}</text>`).join('')}
        ${revenuePoints.length ? `<path class="trend-line-revenue" d="${linePath(revenuePoints, xFor, yFor)}"></path>${circles(revenuePoints, 'var(--blue)')}` : ''}
        ${netPoints.length ? `<path class="trend-line-profit" d="${linePath(netPoints, xFor, yFor)}"></path>${circles(netPoints, 'var(--green)')}` : ''}
      </svg>
      <div class="trend-legend"><span><span class="trend-swatch" style="background:var(--blue)"></span>Revenue</span><span><span class="trend-swatch" style="background:var(--green)"></span>Net Income</span><span>${htmlEscape(rangeLabel(range))}</span></div>`;
    }

    function renderProfitabilityTrend(summary, mode='graph', range=financialTrendRange()) {
      const controls = `<div style="display:flex;justify-content:space-between;align-items:center;gap:12px;margin-bottom:10px;flex-wrap:wrap">
        <div style="display:flex;align-items:end;gap:8px;flex-wrap:wrap">
          <div><label style="margin:0 0 4px">Start</label><input type="date" data-financial-trend-start value="${htmlEscape(range.start || '')}"></div>
          <div><label style="margin:0 0 4px">End</label><input type="date" data-financial-trend-end value="${htmlEscape(range.end || '')}"></div>
          <button class="btn" type="button" data-financial-trend-reset>Current Year</button>
          <button class="btn" type="button" data-financial-trend-expand>Expand Graph</button>
        </div>
        <div class="segmented" role="group" aria-label="Trend view">
          <button type="button" data-financial-trend-mode="graph" class="${mode === 'graph' ? 'active' : ''}">Graph</button>
          <button type="button" data-financial-trend-mode="bars" class="${mode === 'bars' ? 'active' : ''}">Bars</button>
        </div>
      </div>`;
      const filteredNet = dateRangePoints(summary.history?.net_income || [], range);
      const filteredRevenue = dateRangePoints(summary.history?.revenue || [], range);
      const body = mode === 'bars'
        ? `<h3>Net Income</h3>${trendBars(filteredNet)}<h3>Revenue</h3>${trendBars(filteredRevenue)}`
        : trendLineGraph(summary, range);
      const target = document.getElementById('financialProfitTrend');
      target.innerHTML = controls + body;
      target.querySelectorAll('[data-financial-trend-mode]').forEach(btn => btn.onclick = () => {
        localStorage.setItem('financialTrendMode', btn.dataset.financialTrendMode);
        renderProfitabilityTrend(summary, btn.dataset.financialTrendMode, financialTrendRange());
      });
      const startEl = target.querySelector('[data-financial-trend-start]');
      const endEl = target.querySelector('[data-financial-trend-end]');
      const updateRange = () => {
        localStorage.setItem('financialTrendStart', startEl.value || '');
        localStorage.setItem('financialTrendEnd', endEl.value || '');
        renderProfitabilityTrend(summary, mode, financialTrendRange());
      };
      if (startEl) startEl.onchange = updateRange;
      if (endEl) endEl.onchange = updateRange;
      const resetBtn = target.querySelector('[data-financial-trend-reset]');
      if (resetBtn) resetBtn.onclick = () => {
        const defaults = defaultFinancialTrendRange();
        localStorage.setItem('financialTrendStart', defaults.start);
        localStorage.setItem('financialTrendEnd', defaults.end);
        renderProfitabilityTrend(summary, mode, defaults);
      };
      const expandBtn = target.querySelector('[data-financial-trend-expand]');
      if (expandBtn) expandBtn.onclick = () => openTrendModal(summary, mode, financialTrendRange());
    }

    function openTrendModal(summary, mode='graph', range=financialTrendRange()) {
      const modal = document.getElementById('trendModal');
      document.getElementById('trendModalSubtitle').textContent = `${rangeLabel(range)} / ${mode === 'bars' ? 'bar view' : 'graph view'}`;
      const filteredNet = dateRangePoints(summary.history?.net_income || [], range);
      const filteredRevenue = dateRangePoints(summary.history?.revenue || [], range);
      document.getElementById('trendModalBody').innerHTML = mode === 'bars'
        ? `<h3>Net Income</h3>${trendBars(filteredNet)}<h3>Revenue</h3>${trendBars(filteredRevenue)}`
        : trendLineGraph(summary, range, { width: 1120, height: 560, large: true, pad: { left: 92, right: 34, top: 30, bottom: 60 } });
      modal.classList.remove('hidden');
    }

    function openMetricTrendModal(summary, metricKey, label) {
      const modal = document.getElementById('trendModal');
      const range = financialTrendRange();
      document.getElementById('trendModalSubtitle').textContent = `${label} / ${rangeLabel(range)}`;
      document.getElementById('trendModalBody').innerHTML = singleMetricTrendGraph(summary.history?.[metricKey] || [], metricKey, label, range);
      modal.classList.remove('hidden');
    }

    function showFinancialDuplicateModal(duplicates) {
      if (!duplicates || !duplicates.length) return;
      document.getElementById('financialDuplicateList').innerHTML = plainTable(
        ['Report Date','Type','File'],
        duplicates.map(d => [
          htmlEscape(d.report_date || ''),
          htmlEscape(d.report_type || ''),
          htmlEscape(d.file || '')
        ])
      );
      document.getElementById('financialDuplicateModal').classList.remove('hidden');
    }

    function financialReportActions(reports) {
      return (reports || []).map(r => `<div style="display:flex;align-items:center;justify-content:space-between;gap:8px;margin:4px 0">
        <span>${htmlEscape(r.source_file || '')}</span>
        <button class="btn danger" style="padding:5px 8px" type="button" data-delete-financial-report="${r.id}">Remove</button>
      </div>`).join('') || '<span class="muted">Missing</span>';
    }

    function renderFinancialReportHistory(summary) {
      const weeks = summary.report_weeks || [];
      if (!weeks.length) return '<div class="muted">No financial reports uploaded yet.</div>';
      return `<table><thead><tr><th>Week Ending</th><th>Status</th><th>P&L</th><th>Balance Sheet</th><th>Other</th><th>Reports</th></tr></thead><tbody>${weeks.map(w => {
        const statusClass = w.status === 'Complete' ? 'good' : 'warn';
        return `<tr>
          <td>${htmlEscape(w.report_date || '')}</td>
          <td><strong class="${statusClass}">${htmlEscape(w.status || '')}</strong></td>
          <td>${financialReportActions(w.pnl)}</td>
          <td>${financialReportActions(w.balance_sheet)}</td>
          <td>${financialReportActions(w.combined)}</td>
          <td>${Number(w.report_count || 0)}</td>
        </tr>`;
      }).join('')}</tbody></table>`;
    }

    function renderTexasApHistory(summary) {
      const entries = summary.ap_snapshots || [];
      const target = document.getElementById('texasApHistory');
      if (!target) return;
      target.innerHTML = entries.length
        ? `<table><thead><tr><th>Week Ending</th><th>AP Total</th><th>Entered By</th><th>Notes</th><th></th></tr></thead><tbody>${entries.map(entry => `
          <tr>
            <td>${htmlEscape(entry.report_date || '')}</td>
            <td><strong>${money(entry.ap_total)}</strong></td>
            <td>${htmlEscape(entry.created_by_username || '')}<div class="muted">${htmlEscape(entry.updated_at || entry.created_at || '')}</div></td>
            <td>${htmlEscape(entry.notes || '')}</td>
            <td><button class="btn danger" type="button" data-delete-texas-ap="${entry.id}">Remove</button></td>
          </tr>`).join('')}</tbody></table>`
        : '<div class="muted">No Texas AP totals entered yet.</div>';
      target.querySelectorAll('[data-delete-texas-ap]').forEach(btn => btn.onclick = async () => {
        if (!window.confirm('Remove this Texas AP total?')) return;
        await api('/api/texas-ap-delete', {
          method: 'POST',
          body: JSON.stringify({ id: btn.dataset.deleteTexasAp })
        });
        markSaved();
        await loadTexasOpsDashboard();
      });
    }

    function renderTexasReminderSetup(data) {
      const statusEl = document.getElementById('texasReminderStatus');
      const recipientsEl = document.getElementById('texasReminderRecipients');
      const logEl = document.getElementById('texasReminderLog');
      if (!statusEl || !recipientsEl || !logEl) return;
      const smtpText = data.smtp_configured
        ? '<span class="good">Email sending is configured.</span>'
        : '<span class="warn">Email sending is not configured on this server yet. Recipient selection can still be saved.</span>';
      const pnlText = data.pnl_uploaded
        ? '<span class="good">P&amp;L is uploaded for today.</span>'
        : '<span class="warn">No P&amp;L found for today.</span>';
      const settings = data.settings || {};
      document.getElementById('texasReminderEnabled').checked = Number(settings.enabled || 0) === 1;
      document.getElementById('texasReminderWeekday').value = String(settings.weekday ?? 4);
      document.getElementById('texasReminderTime').value = settings.reminder_time || '15:00';
      statusEl.innerHTML = `${smtpText}<br>${Number(settings.enabled || 0) === 1 ? htmlEscape(data.schedule || 'Weekly reminder') : '<span class="warn">Reminder is disabled.</span>'} / ${pnlText}`;
      const recipients = data.recipients || [];
      recipientsEl.innerHTML = recipients.length ? `<div class="grid cols-2">${recipients.map(user => {
        const isEmail = String(user.username || '').includes('@');
        const label = `${user.display_name || user.username}${user.display_name ? ' / ' + user.username : ''}`;
        return `<label style="display:flex;align-items:flex-start;gap:8px;border:1px solid var(--line);border-radius:6px;padding:9px;background:white">
          <input type="checkbox" style="width:auto;margin-top:3px" data-texas-reminder-user="${user.id}" ${Number(user.selected || 0) ? 'checked' : ''}>
          <span><strong>${htmlEscape(label)}</strong>${isEmail ? '' : '<div class="muted">Username is not an email address, so email cannot be sent to this user.</div>'}</span>
        </label>`;
      }).join('')}</div>` : '<div class="muted">No active users found.</div>';
      const logs = data.logs || [];
      logEl.innerHTML = logs.length
        ? `<h3>Recent Reminder Activity</h3><table><thead><tr><th>Date</th><th>Recipient</th><th>Status</th><th>Time</th><th>Note</th></tr></thead><tbody>${logs.map(log => `<tr>
          <td>${htmlEscape(log.reminder_date || '')}</td>
          <td>${htmlEscape(log.recipient || '')}</td>
          <td class="${log.status === 'sent' ? 'good' : log.status === 'failed' ? 'bad' : 'warn'}">${htmlEscape(log.status || '')}</td>
          <td>${htmlEscape(log.sent_at || '')}</td>
          <td>${htmlEscape(log.error || '')}</td>
        </tr>`).join('')}</tbody></table>`
        : '<div class="muted">No reminder activity yet.</div>';
    }

    async function loadTexasReminderSetup() {
      const data = await api('/api/texas-upload-reminders');
      renderTexasReminderSetup(data);
      await loadBillingReminderSetup();
      await loadPoReminderSetup();
    }

    function renderBillingReminderSetup(data) {
      document.getElementById('billingReminderEnabled').checked = Number(data.settings?.enabled || 0) === 1;
      document.getElementById('billingReminderWeekday').value = String(data.settings?.weekday ?? 4);
      document.getElementById('billingReminderTime').value = data.settings?.reminder_time || '15:00';
      document.getElementById('billingReminderLookback').value = data.settings?.lookback_days || 7;
      const smtpText = data.smtp_configured ? '<span class="good">Email sending is configured.</span>' : '<span class="warn">Email sending is not configured on this server yet. Recipient selection can still be saved.</span>';
      const enabledText = Number(data.settings?.enabled || 0) === 1 ? htmlEscape(data.schedule || 'Weekly reminder') : '<span class="warn">Reminder is disabled.</span>';
      document.getElementById('billingReminderStatus').innerHTML = `${smtpText}<br>${enabledText}<br><strong>${Number(data.missing_job_count || 0)}</strong> active job(s) currently have no customer invoice added in the last ${Number(data.settings?.lookback_days || 7)} day(s).`;
      const recipients = data.recipients || [];
      document.getElementById('billingReminderRecipients').innerHTML = recipients.length ? `<div class="grid cols-2">${recipients.map(user => {
        const isEmail = String(user.username || '').includes('@');
        const label = `${user.display_name || user.username}${user.display_name ? ' / ' + user.username : ''}`;
        return `<label style="display:flex;align-items:flex-start;gap:8px;border:1px solid var(--line);border-radius:6px;padding:9px;background:white">
          <input type="checkbox" style="width:auto;margin-top:3px" data-billing-reminder-user="${user.id}" ${Number(user.selected || 0) ? 'checked' : ''}>
          <span><strong>${htmlEscape(label)}</strong>${isEmail ? '' : '<div class="muted">Username is not an email address, so email cannot be sent to this user.</div>'}</span>
        </label>`;
      }).join('')}</div>` : '<div class="muted">No active users found.</div>';
      const jobs = data.missing_jobs || [];
      document.getElementById('billingReminderJobs').innerHTML = jobs.length ? `<h3>Current Jobs Missing Recent Customer Invoice</h3><table><thead><tr><th>Job</th><th>Customer</th><th>Project</th><th>Type</th><th>Last Added</th></tr></thead><tbody>${jobs.slice(0, 25).map(job => `<tr>
        <td><strong>${htmlEscape(job.job_number || '')}</strong><div class="muted">${htmlEscape(job.reference_code || '')}</div></td>
        <td>${htmlEscape(job.customer || '')}</td>
        <td>${htmlEscape(job.project_name || '')}</td>
        <td>${htmlEscape(job.job_type || '')}</td>
        <td>${htmlEscape(job.last_invoice_added_at || 'Never')}</td>
      </tr>`).join('')}</tbody></table>${jobs.length > 25 ? `<div class="muted">Showing 25 of ${jobs.length} jobs.</div>` : ''}` : '<div class="good">All active jobs have recent customer invoice activity.</div>';
      const logs = data.logs || [];
      document.getElementById('billingReminderLog').innerHTML = logs.length ? `<h3>Recent Billing Reminder Activity</h3><table><thead><tr><th>Date</th><th>Recipient</th><th>Status</th><th>Jobs</th><th>Time</th><th>Note</th></tr></thead><tbody>${logs.map(log => `<tr>
        <td>${htmlEscape(log.reminder_date || '')}</td>
        <td>${htmlEscape(log.recipient || '')}</td>
        <td class="${log.status === 'sent' ? 'good' : log.status === 'failed' ? 'bad' : 'warn'}">${htmlEscape(log.status || '')}</td>
        <td>${Number(log.job_count || 0)}</td>
        <td>${htmlEscape(log.sent_at || '')}</td>
        <td>${htmlEscape(log.error || '')}</td>
      </tr>`).join('')}</tbody></table>` : '<div class="muted">No billing reminder activity yet.</div>';
    }

    async function loadBillingReminderSetup() {
      const data = await api('/api/billing-invoice-reminders');
      renderBillingReminderSetup(data);
    }

    function renderPoReminderSetup(data) {
      document.getElementById('poReminderEnabled').checked = Number(data.settings?.enabled || 0) === 1;
      document.getElementById('poReminderWeekday').value = String(data.settings?.weekday ?? 4);
      document.getElementById('poReminderTime').value = data.settings?.reminder_time || '15:00';
      document.getElementById('poReminderUntouchedDays').value = data.settings?.untouched_days || 2;
      const smtpText = data.smtp_configured ? '<span class="good">Email sending is configured.</span>' : '<span class="warn">Email sending is not configured on this server yet. Recipient selection can still be saved.</span>';
      const enabledText = Number(data.settings?.enabled || 0) === 1 ? htmlEscape(data.schedule || 'Weekly reminder') : '<span class="warn">Reminder is disabled.</span>';
      document.getElementById('poReminderStatus').innerHTML = `${smtpText}<br>${enabledText}<br><strong>${Number(data.po_count || 0)}</strong> PO(s) currently untouched for ${Number(data.settings?.untouched_days || 2)} day(s).`;
      const recipients = data.recipients || [];
      document.getElementById('poReminderRecipients').innerHTML = recipients.length ? `<div class="grid cols-2">${recipients.map(user => {
        const isEmail = String(user.username || '').includes('@');
        const label = `${user.display_name || user.username}${user.display_name ? ' / ' + user.username : ''}`;
        return `<label style="display:flex;align-items:flex-start;gap:8px;border:1px solid var(--line);border-radius:6px;padding:9px;background:white">
          <input type="checkbox" style="width:auto;margin-top:3px" data-po-reminder-user="${user.id}" ${Number(user.selected || 0) ? 'checked' : ''}>
          <span><strong>${htmlEscape(label)}</strong>${isEmail ? '' : '<div class="muted">Username is not an email address, so email cannot be sent to this user.</div>'}</span>
        </label>`;
      }).join('')}</div>` : '<div class="muted">No active users found.</div>';
      const pos = data.purchase_orders || [];
      document.getElementById('poReminderList').innerHTML = pos.length ? `<h3>Current Untouched POs</h3><table><thead><tr><th>PO</th><th>Status</th><th>Vendor</th><th>Job / Order / COG</th><th>Created By</th><th>Last Touched</th></tr></thead><tbody>${pos.slice(0, 25).map(po => `<tr>
        <td><strong>${htmlEscape(po.po_number || '')}</strong></td>
        <td>${htmlEscape(po.status || '')}</td>
        <td>${htmlEscape(po.vendor || '')}</td>
        <td>${htmlEscape(po.job_number || '')}<div class="muted">${htmlEscape(po.job_label || '')}</div></td>
        <td>${htmlEscape(po.requested_by_username || '')}</td>
        <td>${htmlEscape(po.last_touched_at || po.created_at || '')}</td>
      </tr>`).join('')}</tbody></table>${pos.length > 25 ? `<div class="muted">Showing 25 of ${pos.length} POs.</div>` : ''}` : '<div class="good">No untouched POs are past the alert window.</div>';
      const logs = data.logs || [];
      document.getElementById('poReminderLog').innerHTML = logs.length ? `<h3>Recent PO Reminder Activity</h3><table><thead><tr><th>Date</th><th>Recipient</th><th>Status</th><th>POs</th><th>Time</th><th>Note</th></tr></thead><tbody>${logs.map(log => `<tr>
        <td>${htmlEscape(log.reminder_date || '')}</td>
        <td>${htmlEscape(log.recipient || '')}</td>
        <td class="${log.status === 'sent' ? 'good' : log.status === 'failed' ? 'bad' : 'warn'}">${htmlEscape(log.status || '')}</td>
        <td>${Number(log.po_count || 0)}</td>
        <td>${htmlEscape(log.sent_at || '')}</td>
        <td>${htmlEscape(log.error || '')}</td>
      </tr>`).join('')}</tbody></table>` : '<div class="muted">No PO reminder activity yet.</div>';
    }

    async function loadPoReminderSetup() {
      const data = await api('/api/po-untouched-reminders');
      renderPoReminderSetup(data);
    }

    async function loadTexasOpsDashboard() {
      const summary = await api('/api/texas-financial-summary');
      const revenue = metricAmount(summary, 'revenue');
      const expenses = metricAmount(summary, 'operating_expenses');
      const netIncome = metricAmount(summary, 'net_income');
      const cash = metricAmount(summary, 'cash');
      const workingCapital = metricAmount(summary, 'working_capital');
      const currentRatio = metricAmount(summary, 'current_ratio');
      const futureWorkingCapital = metricAmount(summary, 'future_working_capital');
      const futureCurrentRatio = metricAmount(summary, 'future_current_ratio');
      const balanceRows = [
        { key: 'cash', label: 'Cash', value: money(cash), helpText: 'Cash is pulled from the latest Balance Sheet cash or bank accounts line.' },
        { key: 'accounts_receivable', label: 'Accounts Receivable', value: money(metricAmount(summary, 'accounts_receivable')), helpText: 'Accounts Receivable comes from the latest Balance Sheet AR line.' },
        { key: 'accounts_payable', label: 'Accounts Payable', value: money(metricAmount(summary, 'accounts_payable')), helpText: 'Accounts Payable comes from the latest Balance Sheet AP line when present.' },
        { key: 'manual_accounts_payable', label: 'Provided Texas AP', value: money(metricAmount(summary, 'manual_accounts_payable')), helpText: 'Provided Texas AP is the weekly AP total entered below because accounting cannot provide AP by Texas class.' },
        { key: 'current_assets', label: 'Current Assets', value: money(metricAmount(summary, 'current_assets')), helpText: 'Current Assets comes from the latest Balance Sheet total current assets line.' },
        { key: 'current_liabilities', label: 'Current Liabilities', value: money(metricAmount(summary, 'current_liabilities')), helpText: 'Current Liabilities comes from the latest Balance Sheet total current liabilities line.' },
        { key: 'future_current_liabilities', label: 'Future Current Liabilities', value: money(metricAmount(summary, 'future_current_liabilities')), helpText: 'Future Current Liabilities equals imported Current Liabilities plus the latest provided Texas AP total.' },
        { key: 'working_capital', label: 'Working Capital', value: `<span class="${workingCapital >= 0 ? 'good' : 'bad'}">${money(workingCapital)}</span>`, helpText: 'Working Capital equals Current Assets minus Current Liabilities. It only calculates when both values come from the same report date.' },
        { key: 'future_working_capital', label: 'Future Working Capital', value: `<span class="${futureWorkingCapital >= 0 ? 'good' : 'bad'}">${money(futureWorkingCapital)}</span>`, helpText: 'Future Working Capital equals imported Current Assets minus Future Current Liabilities.' },
        { key: 'current_ratio', label: 'Current Ratio', value: currentRatio ? currentRatio.toFixed(2) : '', helpText: 'Current Ratio equals Current Assets divided by Current Liabilities. It only calculates when both values come from the same report date.' },
        { key: 'future_current_ratio', label: 'Future Current Ratio', value: futureCurrentRatio ? futureCurrentRatio.toFixed(2) : '', helpText: 'Future Current Ratio equals imported Current Assets divided by Future Current Liabilities.' },
        { key: 'total_assets', label: 'Total Assets', value: money(metricAmount(summary, 'total_assets')), helpText: 'Total Assets comes from the latest Balance Sheet total assets line.' },
        { key: 'total_liabilities', label: 'Total Liabilities', value: money(metricAmount(summary, 'total_liabilities')), helpText: 'Total Liabilities comes from the latest Balance Sheet total liabilities line.' },
        { key: 'equity', label: 'Equity', value: money(metricAmount(summary, 'equity')), helpText: 'Equity comes from the latest Balance Sheet equity or total equity line.' },
      ];
      document.getElementById('financialKpis').innerHTML = `
        <div class="panel kpi help-card">${help('Revenue comes from the latest uploaded P&L report. The importer looks for Total Income, Total Revenue, Revenue, or Sales.')}<div class="label">Revenue</div><div class="value">${money(revenue)}</div><div class="hint">Latest P&L</div></div>
        <div class="panel kpi help-card">${help('Operating Expenses comes from the latest uploaded P&L report. The importer looks for Total Expenses or Operating Expenses.')}<div class="label">Operating Expenses</div><div class="value">${money(expenses)}</div><div class="hint">Latest P&L</div></div>
        <div class="panel kpi help-card">${help('Net Income comes from the latest uploaded P&L report. It is the report line for Net Income, Net Profit, or Net Earnings.')}<div class="label">Net Income</div><div class="value ${netIncome >= 0 ? 'good' : 'bad'}">${money(netIncome)}</div><div class="hint">Latest P&L</div></div>
        <div class="panel kpi help-card">${help('Cash comes from the latest uploaded Balance Sheet. The importer looks for Bank Accounts, Cash, Cash in Bank, or Cash and Cash Equivalents.')}<div class="label">Cash</div><div class="value">${money(cash)}</div><div class="hint">Latest balance sheet</div></div>`;
      renderProfitabilityTrend(summary, localStorage.getItem('financialTrendMode') || 'graph');
      document.getElementById('financialBalanceSnapshot').innerHTML = `<table><thead><tr><th>Metric</th><th>Value</th></tr></thead><tbody>${balanceRows.map(row => `
        <tr class="trend-metric-row" data-financial-metric="${htmlEscape(row.key)}" data-financial-label="${htmlEscape(row.label)}" tabindex="0">
          <td>${htmlEscape(row.label)}</td>
          <td><span class="inline-help-cell">${row.value}${help(row.helpText)}</span></td>
        </tr>`).join('')}</tbody></table>`;
      document.querySelectorAll('[data-financial-metric]').forEach(row => {
        row.onclick = event => {
          if (event.target.closest('.help-marker')) return;
          openMetricTrendModal(summary, row.dataset.financialMetric, row.dataset.financialLabel);
        };
        row.onkeydown = event => {
          if (!['Enter', ' '].includes(event.key)) return;
          event.preventDefault();
          openMetricTrendModal(summary, row.dataset.financialMetric, row.dataset.financialLabel);
        };
      });
      renderTexasApHistory(summary);
      document.getElementById('financialReports').innerHTML = renderFinancialReportHistory(summary);
      document.querySelectorAll('[data-delete-financial-report]').forEach(btn => btn.onclick = async () => {
        const report = summary.reports.find(r => String(r.id) === String(btn.dataset.deleteFinancialReport));
        const label = [report?.report_date, report?.source_file].filter(Boolean).join(' / ') || 'this report';
        if (!window.confirm(`Remove ${label}? This will remove it from the Texas Ops dashboard.`)) return;
        await api('/api/texas-financial-delete', {
          method: 'POST',
          body: JSON.stringify({ report_id: btn.dataset.deleteFinancialReport })
        });
        markSaved();
        await loadTexasOpsDashboard();
      });
    }

    document.getElementById('financialUploadForm').onsubmit = async event => {
      event.preventDefault();
      const form = event.target;
      const resultEl = document.getElementById('financialUploadResult');
      resultEl.textContent = 'Uploading reports...';
      const data = new FormData(form);
      try {
        const res = await fetch('/api/texas-financial-import', { method:'POST', body:data });
        const payload = await res.json();
        if (!res.ok) throw new Error(payload.error || 'Upload failed');
        resultEl.textContent = `Imported ${payload.imported} report(s), ${payload.metric_count} metric(s). Skipped duplicates: ${payload.skipped_duplicates || 0}.`;
        form.reset();
        markSaved();
        await loadTexasOpsDashboard();
        showFinancialDuplicateModal(payload.duplicates || []);
      } catch (err) {
        resultEl.textContent = `Import failed: ${err.message}`;
      }
    };
    document.getElementById('texasApForm').onsubmit = async event => {
      event.preventDefault();
      const form = event.target;
      const resultEl = document.getElementById('texasApResult');
      resultEl.textContent = 'Saving AP total...';
      try {
        await api('/api/texas-ap', {
          method: 'POST',
          body: JSON.stringify(formDataObj(form))
        });
        resultEl.textContent = 'AP total saved.';
        form.reset();
        markSaved();
        await loadTexasOpsDashboard();
      } catch (err) {
        resultEl.textContent = `Save failed: ${err.message}`;
      }
    };
    document.getElementById('saveTexasReminderRecipients').onclick = async () => {
      const userIds = [...document.querySelectorAll('[data-texas-reminder-user]:checked')].map(el => Number(el.dataset.texasReminderUser));
      await api('/api/texas-upload-reminders', {
        method:'POST',
        body: JSON.stringify({
          user_ids: userIds,
          enabled: document.getElementById('texasReminderEnabled').checked ? 1 : 0,
          weekday: document.getElementById('texasReminderWeekday').value,
          reminder_time: document.getElementById('texasReminderTime').value
        })
      });
      document.getElementById('texasReminderResult').textContent = 'Reminder setup saved.';
      markSaved();
      await loadTexasReminderSetup();
    };
    document.getElementById('sendTexasReminderNow').onclick = async () => {
      const resultEl = document.getElementById('texasReminderResult');
      resultEl.textContent = 'Checking P&L status...';
      const result = await api('/api/texas-upload-reminders/send-now', { method:'POST', body: JSON.stringify({}) });
      resultEl.textContent = result.message || `Sent ${result.sent || 0}, failed ${result.failed || 0}, skipped ${result.skipped || 0}.`;
      await loadTexasReminderSetup();
    };
    document.getElementById('saveBillingReminder').onclick = async () => {
      const userIds = [...document.querySelectorAll('[data-billing-reminder-user]:checked')].map(el => Number(el.dataset.billingReminderUser));
      await api('/api/billing-invoice-reminders', {
        method:'POST',
        body: JSON.stringify({
          user_ids: userIds,
          enabled: document.getElementById('billingReminderEnabled').checked ? 1 : 0,
          weekday: document.getElementById('billingReminderWeekday').value,
          reminder_time: document.getElementById('billingReminderTime').value,
          lookback_days: document.getElementById('billingReminderLookback').value
        })
      });
      document.getElementById('billingReminderResult').textContent = 'Billing reminder saved.';
      markSaved();
      await loadBillingReminderSetup();
    };
    document.getElementById('sendBillingReminderNow').onclick = async () => {
      const resultEl = document.getElementById('billingReminderResult');
      resultEl.textContent = 'Checking customer invoice activity...';
      const result = await api('/api/billing-invoice-reminders/send-now', { method:'POST', body: JSON.stringify({}) });
      resultEl.textContent = result.message || `Found ${result.job_count || 0} job(s). Sent ${result.sent || 0}, failed ${result.failed || 0}, skipped ${result.skipped || 0}.`;
      await loadBillingReminderSetup();
    };
    document.getElementById('savePoReminder').onclick = async () => {
      const userIds = [...document.querySelectorAll('[data-po-reminder-user]:checked')].map(el => Number(el.dataset.poReminderUser));
      await api('/api/po-untouched-reminders', {
        method:'POST',
        body: JSON.stringify({
          user_ids: userIds,
          enabled: document.getElementById('poReminderEnabled').checked ? 1 : 0,
          weekday: document.getElementById('poReminderWeekday').value,
          reminder_time: document.getElementById('poReminderTime').value,
          untouched_days: document.getElementById('poReminderUntouchedDays').value
        })
      });
      document.getElementById('poReminderResult').textContent = 'PO reminder saved.';
      markSaved();
      await loadPoReminderSetup();
    };
    document.getElementById('sendPoReminderNow').onclick = async () => {
      const resultEl = document.getElementById('poReminderResult');
      resultEl.textContent = 'Checking untouched POs...';
      const result = await api('/api/po-untouched-reminders/send-now', { method:'POST', body: JSON.stringify({}) });
      resultEl.textContent = result.message || `Found ${result.po_count || 0} PO(s). Sent ${result.sent || 0}, failed ${result.failed || 0}, skipped ${result.skipped || 0}.`;
      await loadPoReminderSetup();
    };

    async function loadProjects() {
      state.projects = await api('/api/projects?status=all');
      const sel = document.getElementById('projectSelect');
      const activeProjects = state.projects.filter(p => (p.status || 'Active') !== 'Archived');
      const currentProject = state.projects.find(p => p.id === state.projectId);
      const selectorProjects = currentProject && (currentProject.status || 'Active') === 'Archived'
        ? [...activeProjects, currentProject]
        : activeProjects;
      sel.innerHTML = selectorProjects.map(p => `<option value="${p.id}">${htmlEscape(p.name)}${(p.status || 'Active') === 'Archived' ? ' (Archived)' : ''}</option>`).join('');
      const savedProjectId = Number(localStorage.getItem('selectedProjectId') || 0);
      if (!state.projectId && savedProjectId && activeProjects.some(p => p.id === savedProjectId)) state.projectId = savedProjectId;
      if (state.projectId && !selectorProjects.some(p => p.id === state.projectId)) state.projectId = null;
      if (!state.projectId && activeProjects[0]) state.projectId = activeProjects[0].id;
      if (state.projectId) sel.value = state.projectId;
      sel.onchange = async () => {
        if (!(await confirmDiscard())) {
          sel.value = state.projectId;
          return;
        }
        markSaved();
        selectedDashboardSubprojectId = null;
        state.projectId = Number(sel.value);
        localStorage.setItem('selectedProjectId', state.projectId);
        refresh();
      };
      await refresh();
      updateNavForTab(document.querySelector('.tab:not(.hidden)')?.id || 'home');
      if (!document.getElementById('home')?.classList.contains('hidden')) loadHomeAlerts();
    }

    async function loadArchivedProjects() {
      const archived = await api('/api/projects?status=archived');
      const wrap = document.getElementById('archivedProjectsTable');
      if (!archived.length) {
        wrap.innerHTML = '<p class="muted">No archived projects yet.</p>';
        return;
      }
      wrap.innerHTML = `<table><thead><tr><th>Project</th><th>Customer</th><th>Location</th><th>Closed</th><th></th></tr></thead><tbody>${archived.map(p => `
        <tr>
          <td><strong>${htmlEscape(p.name)}</strong><div class="muted">${htmlEscape(p.project_code || '')}</div></td>
          <td>${htmlEscape(p.customer || '')}</td>
          <td>${htmlEscape(p.location || '')}</td>
          <td>${htmlEscape((p.archived_at || p.closed_at || '').slice(0, 10))}</td>
          <td>
            <button class="btn" data-open-archived-project="${p.id}" type="button">Open</button>
            <button class="btn" data-restore-project="${p.id}" type="button">Restore</button>
          </td>
        </tr>`).join('')}</tbody></table>`;
      document.querySelectorAll('[data-open-archived-project]').forEach(btn => btn.onclick = async () => {
        if (!(await confirmDiscard())) return;
        state.projectId = Number(btn.dataset.openArchivedProject);
        localStorage.removeItem('selectedProjectId');
        await loadProjects();
        openTab('dashboard');
      });
      document.querySelectorAll('[data-restore-project]').forEach(btn => btn.onclick = async () => {
        await api(`/api/projects/${btn.dataset.restoreProject}/restore`, { method:'POST', body: JSON.stringify({}) });
        state.projectId = Number(btn.dataset.restoreProject);
        localStorage.setItem('selectedProjectId', state.projectId);
        await loadProjects();
        openTab('dashboard');
      });
    }

    function jobOrderFilterValue(id) {
      return String(document.getElementById(id)?.value || '').trim();
    }

    function jobOrderOptionValues(field) {
      return [...new Set(jobOrderReportRows.map(row => String(row[field] || '').trim()).filter(Boolean))]
        .sort((a, b) => a.localeCompare(b, undefined, { numeric: true, sensitivity: 'base' }));
    }

    function setJobOrderFilterOptions(id, field, allLabel) {
      const select = document.getElementById(id);
      if (!select) return;
      const current = select.value;
      const values = jobOrderOptionValues(field);
      select.innerHTML = `<option value="">${allLabel}</option>` + values.map(value => `<option value="${htmlEscape(value)}">${htmlEscape(value)}</option>`).join('');
      if (values.includes(current)) select.value = current;
    }

    function refreshJobOrderFilterOptions() {
      setJobOrderFilterOptions('jobOrderCustomerFilter', 'customer', 'All customers');
      setJobOrderFilterOptions('jobOrderProjectFilter', 'project_name', 'All master projects');
      setJobOrderFilterOptions('jobOrderTypeFilter', 'item_type', 'All types');
      setJobOrderFilterOptions('jobOrderStatusFilter', 'status', 'All statuses');
    }

    function sortJobOrderRows(rowsToSort) {
      const field = jobOrderSort.field || 'job_number';
      const direction = jobOrderSort.direction === 'desc' ? -1 : 1;
      return [...rowsToSort].sort((a, b) => {
        const left = String(a[field] || '').trim();
        const right = String(b[field] || '').trim();
        return left.localeCompare(right, undefined, { numeric: true, sensitivity: 'base' }) * direction;
      });
    }

    function jobOrderSortableHeader(field, label) {
      const active = jobOrderSort.field === field;
      const indicator = active ? `<span class="sort-indicator">${jobOrderSort.direction === 'asc' ? '▲' : '▼'}</span>` : '';
      return `<th><button class="sort-header" type="button" data-job-order-sort="${field}">${label}${indicator}</button></th>`;
    }

    function renderJobOrderReport() {
      const search = jobOrderFilterValue('jobOrderSearch').toLowerCase();
      const customer = jobOrderFilterValue('jobOrderCustomerFilter');
      const project = jobOrderFilterValue('jobOrderProjectFilter');
      const type = jobOrderFilterValue('jobOrderTypeFilter');
      const status = jobOrderFilterValue('jobOrderStatusFilter');
      const filtered = sortJobOrderRows(jobOrderReportRows.filter(row => {
        if (customer && String(row.customer || '') !== customer) return false;
        if (project && String(row.project_name || '') !== project) return false;
        if (type && String(row.item_type || '') !== type) return false;
        if (status && String(row.status || '') !== status) return false;
        if (!search) return true;
        return [
          row.job_number,
          row.item_type,
          row.customer,
          row.project_name,
          row.project_code,
          row.reference_code,
          row.description,
          row.project_description,
          row.location,
          row.status
        ].some(value => String(value || '').toLowerCase().includes(search));
      }));
      const headers = [
        ['job_number', 'Job / Order #'],
        ['item_type', 'Type'],
        ['customer', 'Customer'],
        ['project_name', 'Master Project'],
        ['reference_code', 'Ref'],
        ['description', 'Description'],
        ['project_description', 'Project Description'],
        ['status', 'Status']
      ].map(([field, label]) => jobOrderSortableHeader(field, label)).join('');
      document.getElementById('jobOrderReportCount').textContent = `${filtered.length} of ${jobOrderReportRows.length} active job/order reference(s) shown`;
      document.getElementById('jobOrderReportTable').innerHTML = filtered.length
        ? `<thead><tr>${headers}</tr></thead><tbody>${filtered.map(row => `
          <tr>
            <td><strong>${htmlEscape(row.job_number || '')}</strong></td>
            <td>${htmlEscape(row.item_type || '')}</td>
            <td>${htmlEscape(row.customer || '')}</td>
            <td>${htmlEscape(row.project_name || '')}<div class="muted">${htmlEscape([row.project_code, row.location].filter(Boolean).join(' / '))}</div></td>
            <td>${htmlEscape(row.reference_code || '')}</td>
            <td>${htmlEscape(row.description || '')}</td>
            <td>${htmlEscape(row.project_description || '')}</td>
            <td>${htmlEscape(row.status || '')}</td>
          </tr>`).join('')}</tbody>`
        : '<tbody><tr><td>No active job/order references match that search.</td></tr></tbody>';
      document.querySelectorAll('[data-job-order-sort]').forEach(btn => btn.onclick = () => {
        const field = btn.dataset.jobOrderSort;
        if (jobOrderSort.field === field) {
          jobOrderSort.direction = jobOrderSort.direction === 'asc' ? 'desc' : 'asc';
        } else {
          jobOrderSort = { field, direction: 'asc' };
        }
        renderJobOrderReport();
      });
    }

    async function loadJobOrderReport() {
      jobOrderReportRows = await api('/api/job-order-report');
      refreshJobOrderFilterOptions();
      renderJobOrderReport();
    }

    async function loadFieldPoJobs() {
      if (!jobOrderReportRows.length) jobOrderReportRows = await api('/api/job-order-report');
      const rowsWithKeys = jobOrderReportRows.filter(row => row.job_key);
      fieldPoJobChoices = rowsWithKeys.map(row => {
        const isCog = row.item_type === 'COG';
        const primary = isCog ? row.job_number : [row.job_number, row.reference_code].filter(Boolean).join(' / ');
        const secondary = isCog
          ? ['COG', row.reference_code ? `Internal COG ${row.reference_code}` : '', row.description].filter(Boolean).join(' - ')
          : [row.customer, row.project_name, row.item_type, row.description].filter(Boolean).join(' - ');
        return {
          key: row.job_key,
          primary,
          secondary,
          label: [primary, secondary].filter(Boolean).join(' - '),
          search: (isCog
            ? [row.job_number, row.description]
            : [row.job_number, row.reference_code, row.customer, row.project_name, row.item_type, row.description, row.project_description]
          ).filter(Boolean).join(' ').toLowerCase()
        };
      });
      renderFieldPoJobChoices('');
    }

    function selectFieldPoJob(choice) {
      document.getElementById('fieldPoJobSearch').value = choice.label;
      document.getElementById('fieldPoJobKey').value = choice.key;
      document.getElementById('fieldPoJobList').classList.add('hidden');
      updateFieldPoCustomerRequirement();
    }

    function updateFieldPoCustomerRequirement() {
      const jobKey = document.getElementById('fieldPoJobKey')?.value || '';
      const customerWrap = document.getElementById('fieldPoCustomerWrap');
      const customerInput = document.getElementById('fieldPoCustomer');
      const choice = fieldPoJobChoices.find(item => item.key === jobKey);
      const selectedCogName = String(choice?.primary || '').trim().toLowerCase().replace(/\s+/g, ' ');
      const requiresCustomer = jobKey.startsWith('cog:') && !COG_CUSTOMER_OPTIONAL_NAMES.has(selectedCogName);
      if (customerWrap) customerWrap.classList.toggle('hidden', !requiresCustomer);
      if (customerInput) {
        customerInput.required = requiresCustomer;
        customerInput.disabled = !requiresCustomer;
        customerInput.setAttribute('aria-required', requiresCustomer ? 'true' : 'false');
        if (!requiresCustomer) customerInput.value = '';
      }
    }

    function renderFieldPoJobChoices(query) {
      const list = document.getElementById('fieldPoJobList');
      if (!list) return;
      const needle = String(query || '').trim().toLowerCase();
      const matches = fieldPoJobChoices
        .filter(choice => !needle || choice.search.includes(needle) || choice.label.toLowerCase().includes(needle))
        .slice(0, 30);
      list.innerHTML = matches.length
        ? matches.map(choice => `<button class="combo-option" type="button" data-field-po-job="${htmlEscape(choice.key)}"><strong>${htmlEscape(choice.primary || choice.label)}</strong><span>${htmlEscape(choice.secondary || '')}</span></button>`).join('')
        : '<div class="combo-option"><strong>No matching job/order found</strong><span>Try another job number, COG, customer, or project.</span></div>';
      list.classList.toggle('hidden', !needle && !document.activeElement?.matches('#fieldPoJobSearch'));
      list.querySelectorAll('[data-field-po-job]').forEach(btn => {
        btn.onclick = () => {
          const choice = fieldPoJobChoices.find(item => item.key === btn.dataset.fieldPoJob);
          if (choice) selectFieldPoJob(choice);
        };
      });
    }

    function officePoJobOptions(po) {
      const currentValue = po.cog_category_id ? `cog:${po.cog_category_id}` : (po.change_order_id ? `change_order:${po.change_order_id}` : (po.subproject_id ? `subproject:${po.subproject_id}` : ''));
      return '<option value="">Choose job/order...</option>' + jobOrderReportRows.filter(row => row.job_key).map(row => {
        const label = [row.job_number, row.customer, row.project_name, row.reference_code].filter(Boolean).join(' - ');
        return `<option value="${htmlEscape(row.job_key)}" ${row.job_key === currentValue ? 'selected' : ''}>${htmlEscape(label)}</option>`;
      }).join('');
    }

    async function loadFieldPos() {
      const poRows = await api('/api/purchase-orders');
      const target = document.getElementById('fieldPoList');
      target.innerHTML = poRows.length
        ? poRows.map(po => {
          const statusClass = String(po.status || '').toLowerCase().replace(/\s+/g, '-');
          const pendingApproval = po.status === 'Pending Approval';
          const pickup = po.pickup_file
            ? `<a class="btn" href="/uploads/${encodeURIComponent(po.pickup_file)}" target="_blank" rel="noopener">Open Pickup Ticket</a>`
            : pendingApproval
            ? `<div class="pickup-upload"><strong>Pending PO admin approval.</strong><div class="field-po-card-meta">Wait for office to issue this PO before buying or uploading a pickup ticket.</div></div>`
            : `<form class="pickup-upload" data-pickup-form="${po.id}">
                <label style="margin-top:0">Upload pickup ticket photo</label>
                <input name="pickup_file" type="file" accept=".pdf,.png,.jpg,.jpeg,.webp" capture="environment" required>
                <button class="btn primary" type="submit">Upload Pickup Ticket</button>
              </form>`;
          return `<div class="field-po-card">
            <div class="field-po-card-header">
              <div>
                <div class="field-po-card-title">${htmlEscape(po.po_number || '')}</div>
                <div class="field-po-card-meta">${htmlEscape((po.created_at || '').slice(0, 16).replace('T', ' '))}</div>
              </div>
              <span class="field-po-status field-po-status-${htmlEscape(statusClass)}">${htmlEscape(po.status || '')}</span>
            </div>
            <div><strong>${htmlEscape(po.job_number || '')}</strong><div class="field-po-card-meta">${htmlEscape(po.job_label || '')}</div></div>
            ${po.customer_name ? `<div><strong>Customer:</strong> ${htmlEscape(po.customer_name || '')}</div>` : ''}
            <div><strong>Vendor:</strong> ${htmlEscape(po.vendor || '')}</div>
            <div><strong>Amount:</strong> ${money(po.estimated_amount || 0)}</div>
            <div class="field-po-card-meta">${htmlEscape(po.description || '')}</div>
            <div class="actions" style="margin-top:0">
              <a class="btn" href="/po/${po.id}" target="_blank" rel="noopener">Open PO</a>
            </div>
            ${pickup}
          </div>`;
        }).join('')
        : '<div class="field-po-card"><strong>No POs yet.</strong><div class="field-po-card-meta">Your created POs will show here.</div></div>';
      document.querySelectorAll('[data-pickup-form]').forEach(form => form.onsubmit = async event => {
        event.preventDefault();
        const poId = form.dataset.pickupForm;
        const data = new FormData(form);
        try {
          const button = form.querySelector('button');
          if (button) button.textContent = 'Uploading...';
          await api(`/api/purchase-orders/${poId}/pickup`, { method: 'POST', body: data });
          await loadFieldPos();
        } catch (err) {
          window.alert(err.message || 'Could not upload pickup ticket.');
        }
      });
    }

    async function loadFieldPo() {
      await loadFieldPoJobs();
      await loadFieldPos();
    }

    function renderPoReviewTable(rows, config) {
      const search = String(document.getElementById(config.searchId)?.value || '').trim().toLowerCase();
      const status = String(document.getElementById(config.statusId)?.value || '').trim();
      const filtered = rows.filter(po => {
        if (status && String(po.status || '') !== status) return false;
        if (!search) return true;
        return [
          po.po_number,
          po.job_number,
          po.job_label,
          po.customer_name,
          po.vendor,
          po.description,
          po.requested_by_username,
          po.status
        ].some(value => String(value || '').toLowerCase().includes(search));
      });
      document.getElementById(config.countId).textContent = `${filtered.length} of ${rows.length} PO(s) shown`;
      document.getElementById(config.tableId).innerHTML = filtered.length
        ? `<thead><tr><th>PO</th><th>Job / Order / COG</th><th>Customer</th><th>Vendor</th><th>Amount</th><th>Created By</th><th>Status</th><th>Close / Void Reason</th><th>Details</th><th></th></tr></thead><tbody>${filtered.map(po => {
          const statusOptions = ['Pending Approval','Issued','Picked Up','Closed','Void'].map(statusOption => `<option ${po.status === statusOption ? 'selected' : ''}>${statusOption}</option>`).join('');
          const eventLines = [
            po.pickup_uploaded_at ? `Pickup uploaded ${String(po.pickup_uploaded_at).slice(0, 16).replace('T', ' ')}${po.pickup_uploaded_by_username ? ' by ' + po.pickup_uploaded_by_username : ''}` : '',
            po.closed_at ? `Closed ${String(po.closed_at).slice(0, 16).replace('T', ' ')}${po.closed_by_username ? ' by ' + po.closed_by_username : ''}` : '',
            po.voided_at ? `Voided ${String(po.voided_at).slice(0, 16).replace('T', ' ')}${po.voided_by_username ? ' by ' + po.voided_by_username : ''}` : ''
          ].filter(Boolean).map(line => `<div class="muted">${htmlEscape(line)}</div>`).join('');
          return `
          <tr>
            <td><strong>${htmlEscape(po.po_number || '')}</strong><div class="muted">${htmlEscape((po.created_at || '').slice(0, 16).replace('T', ' '))}</div></td>
            <td><select data-office-po-field="${po.id}" data-field="job_key">${officePoJobOptions(po)}</select></td>
            <td><input data-office-po-field="${po.id}" data-field="customer_name" value="${htmlEscape(po.customer_name || '')}" placeholder="${po.cog_category_id ? 'Required' : 'Optional'}"></td>
            <td><input data-office-po-field="${po.id}" data-field="vendor" value="${htmlEscape(po.vendor || '')}"></td>
            <td><input data-office-po-field="${po.id}" data-field="estimated_amount" type="number" min="0" step="0.01" value="${Number(po.estimated_amount || 0)}"></td>
            <td>${htmlEscape(po.requested_by_username || '')}</td>
            <td><select data-office-po-field="${po.id}" data-field="status">${statusOptions}</select>${eventLines}</td>
            <td><textarea data-office-po-field="${po.id}" data-field="close_reason" placeholder="Required when closing">${htmlEscape(po.close_reason || '')}</textarea><textarea data-office-po-field="${po.id}" data-field="void_reason" placeholder="Required when voiding" style="margin-top:6px">${htmlEscape(po.void_reason || '')}</textarea></td>
            <td><textarea data-office-po-field="${po.id}" data-field="description">${htmlEscape(po.description || '')}</textarea>${po.attachment_file ? `<div><a class="pdf-link" href="/uploads/${encodeURIComponent(po.attachment_file)}" target="_blank" rel="noopener">Attachment</a></div>` : ''}${po.pickup_file ? `<div><a class="pdf-link" href="/uploads/${encodeURIComponent(po.pickup_file)}" target="_blank" rel="noopener">Pickup ticket</a></div>` : '<div class="muted">No pickup ticket</div>'}</td>
            <td class="actions-cell"><a class="btn" href="/po/${po.id}" target="_blank" rel="noopener">Open</a><button class="btn" data-save-office-po="${po.id}" type="button">Save</button>${state.currentUser?.role === 'Admin' ? `<button class="btn danger" data-delete-office-po="${po.id}" type="button">Delete</button>` : ''}</td>
          </tr>`;
        }).join('')}</tbody>`
        : '<tbody><tr><td>No POs match those filters.</td></tr></tbody>';
      document.querySelectorAll(`#${config.tableId} [data-save-office-po]`).forEach(btn => btn.onclick = async () => {
        const id = btn.dataset.saveOfficePo;
        const fields = {};
        document.querySelectorAll(`#${config.tableId} [data-office-po-field="${id}"]`).forEach(el => fields[el.dataset.field] = el.value);
        await api(`/api/purchase-orders/${id}`, {
          method: 'PUT',
          body: JSON.stringify(fields)
        });
        await config.reload();
      });
      document.querySelectorAll(`#${config.tableId} [data-delete-office-po]`).forEach(btn => btn.onclick = async () => {
        const id = btn.dataset.deleteOfficePo;
        const po = rows.find(item => String(item.id) === String(id));
        const label = po ? `${po.po_number || 'this PO'}${po.vendor ? ' / ' + po.vendor : ''}` : 'this PO';
        if (!window.confirm(`Delete ${label}?\n\nThis removes the PO from the tracker.`)) return;
        await api(`/api/purchase-orders/${id}`, { method: 'DELETE' });
        await config.reload();
      });
    }

    function renderOfficePos() {
      renderPoReviewTable(officePoRows, {
        searchId: 'officePoSearch',
        statusId: 'officePoStatusFilter',
        countId: 'officePoCount',
        tableId: 'officePoTable',
        reload: loadOfficePos
      });
    }

    async function loadOfficePos() {
      if (!jobOrderReportRows.length) jobOrderReportRows = await api('/api/job-order-report');
      officePoRows = await api('/api/purchase-orders');
      renderOfficePos();
    }

    async function loadCogSetup() {
      state.cogCategories = await api('/api/cog-categories');
      loadCogCategoryEditor();
    }

    async function loadProjectPos() {
      if (!jobOrderReportRows.length) jobOrderReportRows = await api('/api/job-order-report');
      projectPoRows = await api(`/api/purchase-orders?project_id=${state.projectId}`);
      renderProjectPos();
    }

    function renderProjectPos() {
      renderPoReviewTable(projectPoRows, {
        searchId: 'projectPoSearch',
        statusId: 'projectPoStatusFilter',
        countId: 'projectPoCount',
        tableId: 'projectPoTable',
        reload: loadProjectPos
      });
    }

    async function refresh() {
      if (!state.projectId) {
        document.getElementById('kpis').innerHTML = '<div class="panel">Create a master project to begin.</div>';
        return;
      }
      state.subprojects = await api(`/api/subprojects?project_id=${state.projectId}`);
      state.changeOrders = await api(`/api/change-orders?project_id=${state.projectId}`);
      state.cogCategories = await api('/api/cog-categories');
      state.rateSets = await api('/api/rate-sets');
      state.internalRates = await api('/api/internal-rates');
      fillSelects();
      await loadDashboard();
      loadSubprojectEditor();
      loadChangeOrderEditor();
      loadCogCategoryEditor();
      loadInternalRateEditor();
      loadImportHistory();
      loadFieldTicketLines();
      loadVendorInvoiceLines();
      loadVendorAllocationHistory();
      loadCustomerInvoices();
      refreshOpenDetails();
    }

    function fillSelects() {
      const spOpts = '<option value="">Unassigned</option>' + state.subprojects.map(s => `<option value="${s.id}">${s.job_number || ''} ${s.code} - ${s.name}</option>`).join('');
      ['coSubproject','invoiceSubproject','billingSubproject'].forEach(id => {
        const el = document.getElementById(id);
        if (el) el.innerHTML = spOpts;
      });
      const coOpts = '<option value="">Base Contract</option>' + state.changeOrders.map(c => `<option value="${c.id}">${[c.co_number, c.job_number].filter(Boolean).join(' / ')} - ${c.title || ''}</option>`).join('');
      ['invoiceCo','billingCo'].forEach(id => {
        const el = document.getElementById(id);
        if (el) el.innerHTML = coOpts;
      });
      document.querySelectorAll('[data-customer-invoice-allocation-row] select[data-allocation-field="target_key"]').forEach(el => {
        const current = el.value;
        el.innerHTML = customerInvoiceTargetOptions(current);
      });
      ensureCustomerInvoiceAllocationLine();
      const project = state.projects.find(p => p.id === state.projectId);
      const rateSetOptions = state.rateSets.map(r => `<option value="${r.id}">${r.name}${r.effective_date ? ' - ' + r.effective_date : ''}</option>`).join('');
      ['projectRateSet','newProjectRateSet','rateSetSelect'].forEach(id => {
        const el = document.getElementById(id);
        if (el) el.innerHTML = rateSetOptions;
      });
      if (project && document.getElementById('projectRateSet')) document.getElementById('projectRateSet').value = project.rate_set_id || '';
      if (project && document.getElementById('rateSetSelect')) document.getElementById('rateSetSelect').value = project.rate_set_id || '';
      if (project) {
        const projectForm = document.getElementById('projectForm');
        if (projectForm) {
          ['project_code','name','customer','location','customer_po','description'].forEach(field => {
            const input = projectForm.querySelector(`[name="${field}"]`);
            if (input) input.value = project[field] || '';
          });
        }
        const archiveBtn = document.getElementById('archiveProjectBtn');
        if (archiveBtn) {
          const isArchived = (project.status || 'Active') === 'Archived';
          archiveBtn.classList.toggle('hidden', isArchived);
        }
      }
    }

    function loadNewMasterProjectForm() {
      const form = document.getElementById('newProjectForm');
      if (!form) return;
      const rateSelect = document.getElementById('newProjectRateSet');
      const rateSetOptions = state.rateSets.map(r => `<option value="${r.id}">${r.name}${r.effective_date ? ' - ' + r.effective_date : ''}</option>`).join('');
      if (rateSelect) {
        rateSelect.innerHTML = rateSetOptions;
        const current = state.projects.find(p => p.id === state.projectId);
        if (current?.rate_set_id) rateSelect.value = current.rate_set_id;
      }
    }

    async function loadDashboard() {
      const s = await api(`/api/summary?project_id=${state.projectId}`);
      const marginClass = s.margin >= .2 ? 'good' : s.margin >= .1 ? 'warn' : 'bad';
      document.getElementById('projectBanner').innerHTML = `
        <div class="project-banner">
          <div class="project-title clickable help-card" id="masterProjectBannerCard" role="button" tabindex="0">
            ${help('Master project totals combine all subprojects and approved change orders. Click to open the full master detail view.')}
            <div class="eyebrow">Master Project</div>
            <div class="name">${s.project.name}</div>
            <div class="muted">${s.project.customer || ''} ${s.project.location ? ' / ' + s.project.location : ''}</div>
            <div class="muted">${s.project.customer_po ? 'PO # ' + s.project.customer_po : ''}</div>
            <div class="muted">${s.project.description || ''}</div>
          </div>
          ${s.subprojects.map(x => {
            const subprojectHelp = x.pricing_type === 'T&M'
              ? 'T&M subproject actual uses Field Wise labor, equipment, and Field Wise material costs. Vendor invoices are shown for reference but are not counted in T&M raw cost.'
              : 'Fixed subproject actual includes assigned cost records, including vendor invoices. Click to open this subproject detail.'
            return `<div class="job-chip clickable help-card" data-banner-subproject="${x.id}" role="button" tabindex="0">${help(subprojectHelp)}<div class="job">${[x.job_number, x.code].filter(Boolean).join(' ')}</div><div class="label">${x.name}</div><div class="label">${money(x.actual_cost)} actual</div></div>`;
          }).join('')}
        </div>`;
      const masterProjectBannerCard = document.getElementById('masterProjectBannerCard');
      if (masterProjectBannerCard) {
        masterProjectBannerCard.onclick = () => loadMasterDetail();
        masterProjectBannerCard.onkeydown = event => {
          if (event.key === 'Enter' || event.key === ' ') {
            event.preventDefault();
            masterProjectBannerCard.click();
          }
        };
      }
      document.querySelectorAll('[data-banner-subproject]').forEach(card => {
        card.onclick = () => {
          selectedDashboardSubprojectId = card.dataset.bannerSubproject;
          selectedDashboardChangeOrderId = null;
          renderSubprojectSummary(s);
          renderChangeOrders(s);
          renderBillingSummary(s);
          loadSubprojectDetail(card.dataset.bannerSubproject);
        };
        card.onkeydown = event => {
          if (event.key === 'Enter' || event.key === ' ') {
            event.preventDefault();
            card.click();
          }
        };
      });
      const baseContractInput = document.querySelector('#projectForm input[name="contract_value"]');
      if (baseContractInput) baseContractInput.value = Number(s.base_contract_value || 0).toFixed(2);
      document.getElementById('kpis').innerHTML = `
        <div class="panel kpi help-card">${help('Base value is the sum of fixed subproject contract values plus Field Wise sales for T&M subprojects. Approved COs are added on top.')}<div class="label">Contract + Approved COs</div><div class="value">${money(s.contract_value)}</div><div class="hint">Base: ${money(s.base_contract_value)} / Approved COs: ${money(s.approved_co_value)}</div></div>
        <div class="panel kpi help-card">${help('Raw actual cost comes from assigned cost records. For T&M work, vendor invoices are excluded from raw cost; Field Wise labor, equipment, and Field Wise material are used.')}<div class="label">Raw Actual Cost</div><div class="value">${money(s.actual_cost)}</div><div class="hint">${s.record_count} cost records</div></div>
        <div class="panel kpi help-card">${help('Projected profit equals Contract + Approved COs minus Raw Actual Cost. Margin equals Projected Profit divided by Contract + Approved COs.')}<div class="label">Projected Profit</div><div class="value ${s.profit >= 0 ? 'good' : 'bad'}">${money(s.profit)}</div><div class="hint ${marginClass}">Margin ${pct(s.margin)}</div></div>
        <div class="panel kpi clickable help-card" id="needsCodingKpi" role="button" tabindex="0">${help('Needs Coding counts cost records that are missing a subproject, cost type, or required internal/project rate. Click to review exceptions.')}<div class="label">Needs Coding</div><div class="value ${s.uncoded_count ? 'warn' : 'good'}">${s.uncoded_count}</div><div class="hint">${money(s.uncoded_cost)}</div></div>`;
      const needsCodingKpi = document.getElementById('needsCodingKpi');
      if (needsCodingKpi) {
        needsCodingKpi.onclick = async () => {
          if (!(await confirmDiscard())) return;
          openTab('review');
        };
        needsCodingKpi.onkeydown = event => {
          if (event.key === 'Enter' || event.key === ' ') {
            event.preventDefault();
            needsCodingKpi.click();
          }
        };
      }
      if (selectedDashboardSubprojectId && !s.subprojects.some(x => String(x.id) === String(selectedDashboardSubprojectId))) {
        selectedDashboardSubprojectId = null;
      }
      if (selectedDashboardChangeOrderId && !s.change_orders.some(x => String(x.id) === String(selectedDashboardChangeOrderId))) {
        selectedDashboardChangeOrderId = null;
      }
      renderSubprojectSummary(s);
      const showAllCos = document.getElementById('showAllCos');
      if (showAllCos) showAllCos.onclick = () => {
        selectedDashboardSubprojectId = null;
        selectedDashboardChangeOrderId = null;
        renderChangeOrders(s);
        renderSubprojectSummary(s);
        renderBillingSummary(s);
      };
      renderProjectHierarchy(s);
      renderChangeOrders(s);
      renderBillingSummary(s);
      const fieldMat = Number(s.material_compare?.field_ticket_material || 0);
      const vendorMat = Number(s.material_compare?.vendor_material || 0);
      document.getElementById('materialComparison').innerHTML = table(
        ['Metric','Amount'],
        [
          ['Field ticket material listed', money(fieldMat)],
          ['Vendor invoice material purchased', money(vendorMat)],
          ['Difference', money(vendorMat - fieldMat)]
        ]
      );
      document.getElementById('typeSummary').innerHTML = table(['Cost Type','Amount'], s.by_type.map(x => [x.label, money(x.amount)]));
    }

    function renderProjectHierarchy(summary) {
      const target = document.getElementById('projectHierarchy');
      if (!target) return;
      const changeOrdersBySubproject = {};
      (summary.change_orders || []).forEach(co => {
        const key = String(co.subproject_id || '');
        if (!changeOrdersBySubproject[key]) changeOrdersBySubproject[key] = [];
        changeOrdersBySubproject[key].push(co);
      });
      const isCollapsed = key => collapsedHierarchyNodes.has(key);
      const toggleGlyph = key => isCollapsed(key) ? '+' : '-';
      const coNode = co => `<div class="hierarchy-node change-order">
        <div class="node-title"><span>${htmlEscape([co.co_number, co.job_number].filter(Boolean).join(' / ') || co.order_type || 'Change Order')}</span><span>${money(co.sales_value || co.approved_value || 0)}</span></div>
        <div class="node-meta">${htmlEscape(co.title || '')}</div>
        <div class="node-values"><span>Type: ${htmlEscape(co.order_type || 'Change Order')}</span><span>Status: ${htmlEscape(co.status || '')}</span><span>Pricing: ${htmlEscape(co.pricing_type || 'Fixed')}</span><span>Actual: ${money(co.actual_cost || 0)}</span></div>
      </div>`;
      const subprojectNodes = (summary.subprojects || []).map(sp => {
        const nodeKey = `sp:${sp.id}`;
        const coNodes = (changeOrdersBySubproject[String(sp.id)] || []).map(coNode).join('');
        return `<div>
          <div class="hierarchy-node subproject collapsible" data-hierarchy-toggle="${nodeKey}" role="button" tabindex="0">
            <div class="node-title"><span class="hierarchy-title-left"><span class="hierarchy-toggle">${toggleGlyph(nodeKey)}</span><span>${htmlEscape([sp.job_number, sp.code].filter(Boolean).join(' ') || 'Subproject')} - ${htmlEscape(sp.name || '')}</span></span><span>${money(sp.sales_value || sp.contract_value || 0)}</span></div>
            <div class="node-meta">${htmlEscape(sp.pricing_type || 'Fixed')} pricing</div>
            <div class="node-values"><span>Actual: ${money(sp.actual_cost || 0)}</span><span>Profit: ${money(Number(sp.sales_value || sp.contract_value || 0) - Number(sp.actual_cost || 0))}</span></div>
          </div>
          <div class="hierarchy-children ${isCollapsed(nodeKey) ? 'collapsed' : ''}">${coNodes || '<div class="hierarchy-node change-order"><div class="node-meta">No change orders.</div></div>'}</div>
        </div>`;
      }).join('');
      const orphanCos = changeOrdersBySubproject[''] || [];
      const orphanKey = 'orphan-cos';
      const orphanSection = orphanCos.length ? `<div class="hierarchy-node subproject collapsible" data-hierarchy-toggle="${orphanKey}" role="button" tabindex="0"><div class="node-title"><span class="hierarchy-title-left"><span class="hierarchy-toggle">${toggleGlyph(orphanKey)}</span><span>Unassigned Change Orders</span></span><span>${orphanCos.length}</span></div><div class="node-meta">Change orders not tied to a subproject.</div></div><div class="hierarchy-children ${isCollapsed(orphanKey) ? 'collapsed' : ''}">${orphanCos.map(coNode).join('')}</div>` : '';
      const masterKey = `master:${summary.project?.id || state.projectId || 'current'}`;
      if (!initializedHierarchyNodes.has(masterKey)) {
        initializedHierarchyNodes.add(masterKey);
        collapsedHierarchyNodes.add(masterKey);
      }
      target.innerHTML = `<div class="hierarchy-wrap">
        <div class="hierarchy-node master collapsible" data-hierarchy-toggle="${masterKey}" role="button" tabindex="0">
          <div class="node-title"><span class="hierarchy-title-left"><span class="hierarchy-toggle">${toggleGlyph(masterKey)}</span><span>${htmlEscape(summary.project?.name || 'Master Project')}</span></span><span>${money(summary.contract_value || 0)}</span></div>
          <div class="node-meta">${htmlEscape(summary.project?.customer || '')}${summary.project?.location ? ' / ' + htmlEscape(summary.project.location) : ''}</div>
          <div class="node-values"><span>Subprojects: ${(summary.subprojects || []).length}</span><span>Change Orders / Child Projects: ${(summary.change_orders || []).length}</span><span>Actual: ${money(summary.actual_cost || 0)}</span></div>
        </div>
        <div class="hierarchy-children ${isCollapsed(masterKey) ? 'collapsed' : ''}">
          ${subprojectNodes || '<div class="muted">No subprojects yet.</div>'}
          ${orphanSection}
        </div>
      </div>`;
      target.querySelectorAll('[data-hierarchy-toggle]').forEach(node => {
        const toggle = () => {
          const key = node.dataset.hierarchyToggle;
          if (collapsedHierarchyNodes.has(key)) collapsedHierarchyNodes.delete(key);
          else collapsedHierarchyNodes.add(key);
          renderProjectHierarchy(summary);
        };
        node.onclick = event => {
          event.stopPropagation();
          toggle();
        };
        node.onkeydown = event => {
          if (event.key !== 'Enter' && event.key !== ' ') return;
          event.preventDefault();
          toggle();
        };
      });
    }

    function billingForDashboardScope(summary) {
      const selectedCo = selectedDashboardChangeOrderId ? summary.change_orders.find(x => String(x.id) === String(selectedDashboardChangeOrderId)) : null;
      const selectedSubproject = selectedDashboardSubprojectId ? summary.subprojects.find(x => String(x.id) === String(selectedDashboardSubprojectId)) : null;
      let label = 'Master Project';
      let contractValue = Number(summary.contract_value || 0);
      let invoices = summary.customer_invoices || [];
      if (selectedCo) {
        label = `${selectedCo.order_type || 'Change Order'} ${selectedCo.co_number || ''}${selectedCo.job_number ? ' / ' + selectedCo.job_number : ''}`;
        contractValue = Number(selectedCo.sales_value || 0);
      } else if (selectedSubproject) {
        label = `${selectedSubproject.job_number || ''} ${selectedSubproject.code || ''}`;
        contractValue = Number(selectedSubproject.sales_value || selectedSubproject.contract_value || 0);
      }
      const subprojectCoIds = selectedSubproject ? summary.change_orders
        .filter(co => String(co.subproject_id || '') === String(selectedSubproject.id))
        .map(co => String(co.id)) : [];
      const invoiceScopeAmount = invoice => {
        if (!selectedCo && !selectedSubproject) return Number(invoice.amount || 0);
        const allocations = invoice.allocations || [];
        if (selectedCo) {
          return allocations
            .filter(a => String(a.change_order_id || '') === String(selectedCo.id))
            .reduce((sum, a) => sum + Number(a.amount || 0), 0);
        }
        if (selectedSubproject) {
          return allocations
            .filter(a =>
              String(a.subproject_id || '') === String(selectedSubproject.id) ||
              subprojectCoIds.includes(String(a.change_order_id || ''))
            )
            .reduce((sum, a) => sum + Number(a.amount || 0), 0);
        }
        return Number(invoice.amount || 0);
      };
      if (selectedCo || selectedSubproject) {
        invoices = invoices
          .map(invoice => {
            const scopeAmount = invoiceScopeAmount(invoice);
            const allocatedTotal = Number(invoice.allocated_amount || invoice.amount || 0) || Number(invoice.amount || 0) || 1;
            return {
              ...invoice,
              scope_amount: scopeAmount,
              scope_paid_amount: Number(invoice.paid_amount || 0) * (scopeAmount / allocatedTotal),
            };
          })
          .filter(invoice => Math.abs(Number(invoice.scope_amount || 0)) >= 0.01);
      }
      const active = invoices.filter(i => i.status !== 'Void');
      const billed = active.reduce((sum, i) => sum + Number(i.scope_amount ?? i.amount ?? 0), 0);
      const paid = active.reduce((sum, i) => sum + Number(i.scope_paid_amount ?? i.paid_amount ?? 0), 0);
      const open = active
        .filter(i => !['Draft','Paid'].includes(i.status || ''))
        .reduce((sum, i) => sum + Math.max(0, Number(i.scope_amount ?? i.amount ?? 0) - Number(i.scope_paid_amount ?? i.paid_amount ?? 0)), 0);
      const remaining = Math.max(0, contractValue - billed);
      let stage = 'Not billed';
      if (contractValue && paid >= contractValue) stage = 'Paid in full';
      else if (contractValue && billed >= contractValue) stage = 'Fully billed';
      else if (billed) stage = 'Partially billed';
      return { label, contractValue, billed, paid, open, remaining, invoiceCount: active.length, stage };
    }

    function renderBillingSummary(summary) {
      const billing = billingForDashboardScope(summary);
      const billingSummaryEl = document.getElementById('billingSummary');
      if (!billingSummaryEl) return;
      const billedPct = billing.contractValue ? billing.billed / billing.contractValue : 0;
      const paidPct = billing.contractValue ? billing.paid / billing.contractValue : 0;
      billingSummaryEl.innerHTML = `
        <div class="muted" style="margin-bottom:8px">Showing invoicing for ${htmlEscape(billing.label)}</div>
        <div class="grid cols-4" style="margin-top:8px">
          <div class="kpi help-card">${help('Billing stage is based on customer invoices for the selected master project, subproject, or change order.')}<div class="label">Billing Stage</div><div class="value" style="font-size:22px">${billing.stage}</div><div class="hint">${billing.invoiceCount} customer invoice(s)</div></div>
          <div class="kpi help-card">${help('Invoiced amount is the total customer invoice amount for the selected scope. The percent compares invoiced amount to the selected sales or contract value.')}<div class="label">Invoiced Amount</div><div class="value">${money(billing.billed)}</div><div class="hint">${pct(billedPct)} of selected value</div><div class="bar"><span style="width:${Math.min(100, billedPct * 100)}%"></span></div></div>
          <div class="kpi help-card">${help('Paid amount is the total paid amount entered on customer invoices for the selected scope.')}<div class="label">Paid Amount</div><div class="value">${money(billing.paid)}</div><div class="hint">${pct(paidPct)} of selected value</div><div class="bar"><span style="width:${Math.min(100, paidPct * 100)}%"></span></div></div>
          <div class="kpi help-card">${help('Open AR equals invoiced amount minus paid amount. Remaining to bill equals selected sales or contract value minus invoiced amount.')}<div class="label">Open AR</div><div class="value ${Number(billing.open || 0) ? 'warn' : 'good'}">${money(billing.open)}</div><div class="hint">Remaining to bill ${money(billing.remaining)}</div></div>
        </div>`;
      const billingBtn = document.getElementById('openBillingFromDashboard');
      if (billingBtn) billingBtn.onclick = async () => {
        if (!(await confirmDiscard())) return;
        openTab('billing');
      };
    }

    function renderSubprojectSummary(summary) {
      document.getElementById('subprojectSummary').innerHTML = subprojectTable(summary.subprojects, selectedDashboardSubprojectId);
      document.querySelectorAll('#subprojectSummary [data-open-subproject]').forEach(btn => btn.onclick = () => {
        selectedDashboardSubprojectId = btn.dataset.openSubproject;
        selectedDashboardChangeOrderId = null;
        renderSubprojectSummary(summary);
        renderChangeOrders(summary);
        renderBillingSummary(summary);
        loadSubprojectDetail(btn.dataset.openSubproject);
      });
      document.querySelectorAll('#subprojectSummary [data-select-subproject]').forEach(row => row.onclick = event => {
        if (event.target.closest('button')) return;
        selectedDashboardSubprojectId = row.dataset.selectSubproject;
        selectedDashboardChangeOrderId = null;
        renderSubprojectSummary(summary);
        renderChangeOrders(summary);
        renderBillingSummary(summary);
      });
    }

    function renderChangeOrders(summary) {
      const selected = selectedDashboardSubprojectId ? summary.subprojects.find(x => String(x.id) === String(selectedDashboardSubprojectId)) : null;
      const orders = selected ? summary.change_orders.filter(x => String(x.subproject_id || '') === String(selected.id)) : summary.change_orders;
      const title = document.getElementById('coSummaryTitle');
      const showAll = document.getElementById('showAllCos');
      if (title) title.textContent = selected ? `Change Orders / Child Projects - ${selected.job_number || ''} ${selected.code}` : 'Change Orders / Child Projects';
      if (showAll) showAll.classList.toggle('hidden', !selected);
      document.getElementById('coSummary').innerHTML = orders.length
        ? `<table><thead><tr><th>Type</th><th>CO</th><th>Job / Order #</th><th>Pricing</th><th>Subproject</th><th>Status</th><th>Labor Hrs</th><th>Sales Value</th><th>Raw Actual</th><th>Profit</th></tr></thead><tbody>${orders.map(x => {
            const isSelected = String(selectedDashboardChangeOrderId || '') === String(x.id);
            return `<tr class="selectable-row ${isSelected ? 'selected' : ''}" data-select-co="${x.id}">
              <td>${htmlEscape(x.order_type || 'Change Order')}</td>
              <td>${x.co_number || ''}</td>
              <td>${x.job_number || ''}</td>
              <td>${x.pricing_type || 'Fixed'}</td>
              <td>${x.subproject_code || ''}</td>
              <td>${x.status || ''}</td>
              <td>${Number(x.labor_hours_used || 0).toFixed(2)}</td>
              <td>${money(x.sales_value)}</td>
              <td>${money(x.actual_cost)}</td>
              <td>${money(Number(x.sales_value || 0) - Number(x.actual_cost || 0))}</td>
            </tr>`;
          }).join('')}</tbody></table>`
        : '<table><tbody><tr><td>No change orders for this subproject.</td></tr></tbody></table>';
      document.querySelectorAll('#coSummary [data-select-co]').forEach(row => row.onclick = () => {
        selectedDashboardChangeOrderId = row.dataset.selectCo;
        const selectedCo = summary.change_orders.find(x => String(x.id) === String(selectedDashboardChangeOrderId));
        if (selectedCo?.subproject_id) selectedDashboardSubprojectId = String(selectedCo.subproject_id);
        renderSubprojectSummary(summary);
        renderChangeOrders(summary);
        renderBillingSummary(summary);
        if (selectedCo?.subproject_id) loadSubprojectDetail(selectedCo.subproject_id, true, selectedCo.id);
      });
    }

    function table(headers, data) {
      return `<table><thead><tr>${headers.map(h => `<th>${h}</th>`).join('')}</tr></thead><tbody>${data.map(r => `<tr>${r.map(c => `<td>${c}</td>`).join('')}</tr>`).join('')}</tbody></table>`;
    }

    function fieldWiseAuditRowsTable(title, rows, emptyText, options={}) {
      const shown = rows.slice(0, 250);
      const showOmitAction = Boolean(options.showOmitAction);
      const actionHeader = showOmitAction ? '<th>Action</th>' : '';
      return `<div class="panel" style="margin-top:14px">
        <h2>${htmlEscape(title)}</h2>
        <div class="muted">${rows.length} item(s)${rows.length > shown.length ? `, showing first ${shown.length}` : ''}</div>
        <div class="table-wrap" style="margin-top:10px"><table>${shown.length ? `<thead><tr><th>Ticket</th><th>Order #</th><th>Customer</th><th>Project</th><th>Type</th><th>Status</th><th>Date</th><th>Lines</th>${actionHeader}</tr></thead><tbody>${shown.map(r => `
          <tr>
            <td><strong>${htmlEscape(r.ticket_number || '')}</strong></td>
            <td>${htmlEscape(r.order_number || '')}</td>
            <td>${htmlEscape(r.customer || '')}</td>
            <td>${htmlEscape(r.project_name || '')}<div class="muted">${htmlEscape(r.reference_code || '')}</div></td>
            <td>${htmlEscape(r.item_type || '')}</td>
            <td>${htmlEscape(r.status || '')}</td>
            <td>${htmlEscape(r.ticket_date || '')}</td>
            <td>${Number(r.line_count || 0)}</td>
            ${showOmitAction ? `<td><button class="btn" type="button" data-omit-fieldwise="${htmlEscape(r.ticket_number || '')}" data-order-number="${htmlEscape(r.order_number || '')}" data-customer="${htmlEscape(r.customer || '')}" data-project-name="${htmlEscape(r.project_name || '')}">Mark OK</button></td>` : ''}
          </tr>`).join('')}</tbody>` : `<tbody><tr><td>${htmlEscape(emptyText)}</td></tr></tbody>`}</table></div>
      </div>`;
    }

    function renderFieldWiseAuditOmissions() {
      const el = document.getElementById('fieldWiseAuditOmissions');
      if (!el) return;
      const rows = fieldWiseAuditOmissions || [];
      el.innerHTML = `<div class="panel">
        <h2>Tickets Marked OK To Omit</h2>
        <p class="muted">These ticket/order combinations will not show as missing in future audits.</p>
        <div class="table-wrap" style="margin-top:10px"><table>${rows.length ? `<thead><tr><th>Ticket</th><th>Order #</th><th>Customer</th><th>Project</th><th>Reason</th><th>Marked By</th><th>Date</th><th></th></tr></thead><tbody>${rows.map(r => `
          <tr>
            <td><strong>${htmlEscape(r.ticket_number || '')}</strong></td>
            <td>${htmlEscape(r.order_number || '')}</td>
            <td>${htmlEscape(r.customer || '')}</td>
            <td>${htmlEscape(r.project_name || '')}</td>
            <td>${htmlEscape(r.reason || '')}</td>
            <td>${htmlEscape(r.omitted_by_username || '')}</td>
            <td>${htmlEscape((r.created_at || '').replace('T', ' '))}</td>
            <td><button class="btn danger" type="button" data-delete-audit-omission="${r.id}">Remove</button></td>
          </tr>`).join('')}</tbody>` : '<tbody><tr><td>No tickets have been marked OK to omit yet.</td></tr></tbody>'}</table></div>
      </div>`;
      document.querySelectorAll('[data-delete-audit-omission]').forEach(btn => btn.onclick = async () => {
        if (!window.confirm('Remove this OK-to-omit note? The ticket can show as missing again on future audits.')) return;
        await api(`/api/fieldwise-audit-omissions/${btn.dataset.deleteAuditOmission}`, { method: 'DELETE' });
        await loadFieldWiseAuditOmissions();
      });
    }

    async function loadFieldWiseAuditOmissions() {
      fieldWiseAuditOmissions = await api('/api/fieldwise-audit-omissions');
      renderFieldWiseAuditOmissions();
    }

    async function markFieldWiseTicketOmitted(row) {
      const reason = window.prompt(`Why is Field Wise ticket ${row.ticket_number} / order ${row.order_number} OK to omit?`, 'Valid Field Wise ticket, not required in this tracker');
      if (reason === null) return;
      await api('/api/fieldwise-audit-omissions', {
        method: 'POST',
        body: JSON.stringify({
          ticket_number: row.ticket_number,
          order_number: row.order_number,
          customer: row.customer,
          project_name: row.project_name,
          reason
        })
      });
      if (fieldWiseAuditData?.missing) {
        fieldWiseAuditData.missing = fieldWiseAuditData.missing.filter(r => !(String(r.ticket_number || '') === String(row.ticket_number || '') && String(r.order_number || '') === String(row.order_number || '')));
        if (fieldWiseAuditData.summary) {
          fieldWiseAuditData.summary.missing_count = Math.max(0, Number(fieldWiseAuditData.summary.missing_count || 0) - 1);
          fieldWiseAuditData.summary.omitted_count = Number(fieldWiseAuditData.summary.omitted_count || 0) + 1;
        }
        renderFieldWiseAudit();
      }
      await loadFieldWiseAuditOmissions();
    }

    function renderFieldWiseAudit() {
      const result = fieldWiseAuditData;
      const summaryEl = document.getElementById('fieldWiseAuditSummary');
      const tablesEl = document.getElementById('fieldWiseAuditTables');
      const exportBtn = document.getElementById('exportMissingTickets');
      if (!result) {
        summaryEl.innerHTML = '';
        tablesEl.innerHTML = '';
        exportBtn.disabled = true;
        return;
      }
      const s = result.summary || {};
      summaryEl.innerHTML = [
        ['Export Tickets', s.export_ticket_count || 0, `${s.export_line_count || 0} line(s)`],
        ['Missing Tracked', s.missing_count || 0, 'Ticket/order not found here'],
        ['Marked OK', s.omitted_count || 0, 'Hidden from missing list'],
        ['Untracked / No Order', Number(s.untracked_count || 0) + Number(s.no_order_count || 0), 'Can be omitted']
      ].map(([label, value, hint]) => `<div class="kpi"><div class="label">${htmlEscape(label)}</div><div class="value">${value}</div><div class="hint">${htmlEscape(hint)}</div></div>`).join('');
      const omitUntracked = document.getElementById('omitUntrackedAuditTickets').checked;
      const parts = [
        fieldWiseAuditRowsTable('Missing Tickets For Tracked Jobs', result.missing || [], 'No missing tracked tickets found.', { showOmitAction: true }),
        fieldWiseAuditRowsTable('Marked OK To Omit In This Export', result.omitted || [], 'No previously omitted tickets were found in this export.'),
        fieldWiseAuditRowsTable('Imported Here But Not In This Export', result.extra_imported || [], 'No extra imported tickets found.')
      ];
      if (!omitUntracked) {
        parts.push(fieldWiseAuditRowsTable('Untracked Order Numbers', result.untracked || [], 'No untracked order-number tickets found.'));
        parts.push(fieldWiseAuditRowsTable('Tickets With Blank Order Number', result.no_order || [], 'No blank order-number tickets found.'));
      }
      tablesEl.innerHTML = parts.join('');
      document.querySelectorAll('[data-omit-fieldwise]').forEach(btn => btn.onclick = async () => {
        const row = (fieldWiseAuditData?.missing || []).find(r => String(r.ticket_number || '') === String(btn.dataset.omitFieldwise || '') && String(r.order_number || '') === String(btn.dataset.orderNumber || ''));
        if (!row) return;
        await markFieldWiseTicketOmitted(row);
      });
      exportBtn.disabled = !(result.missing || []).length;
    }

    function csvCell(value) {
      const text = String(value ?? '');
      return /[",\r\n]/.test(text) ? `"${text.replace(/"/g, '""')}"` : text;
    }

    function downloadMissingTicketCsv() {
      if (!fieldWiseAuditData?.missing?.length) return;
      const headers = ['Ticket Number','Order Number','Customer','Project','Type','Reference','Status','Ticket Date','Line Count'];
      const lines = [headers.map(csvCell).join(',')];
      fieldWiseAuditData.missing.forEach(r => {
        lines.push([
          r.ticket_number,
          r.order_number,
          r.customer,
          r.project_name,
          r.item_type,
          r.reference_code,
          r.status,
          r.ticket_date,
          r.line_count || 0
        ].map(csvCell).join(','));
      });
      const blob = new Blob([lines.join('\r\n')], { type: 'text/csv' });
      const url = URL.createObjectURL(blob);
      const a = document.createElement('a');
      a.href = url;
      a.download = `missing-fieldwise-tickets-${new Date().toISOString().slice(0,10)}.csv`;
      document.body.appendChild(a);
      a.click();
      a.remove();
      URL.revokeObjectURL(url);
    }

    function invoiceKey(row) {
      return [row.source_file || '', row.ticket_or_invoice || '', row.vendor || '', row.record_date || ''].join('|');
    }

    function fieldWiseKey(row) {
      return [row.source_file || '', row.ticket_or_invoice || '', row.record_date || ''].join('|');
    }

    function costRecordGroups(rows) {
      const groups = [];
      const byKey = {};
      rows.forEach(row => {
        let key = '';
        let type = '';
        if (row.source === 'Vendor Invoice') {
          key = `vendor|${invoiceKey(row)}`;
          type = 'invoice';
        } else if (row.source === 'Field Wise' || row.source === 'Field Wise PDF') {
          key = `fieldwise|${fieldWiseKey(row)}`;
          type = 'fieldwise';
        } else {
          groups.push({ type: 'single', rows: [row] });
          return;
        }
        if (!byKey[key]) {
          byKey[key] = { type, key, id: `invoiceGroup${++invoiceGroupSeq}`, rows: [] };
          groups.push(byKey[key]);
        }
        byKey[key].rows.push(row);
      });
      return groups;
    }

    function vendorInvoiceGroups(rows) {
      const groups = [];
      const byKey = {};
      rows.forEach(row => {
        if (row.source !== 'Vendor Invoice') {
          groups.push({ type: 'single', rows: [row] });
          return;
        }
        const key = invoiceKey(row);
        if (!byKey[key]) {
          byKey[key] = { type: 'invoice', key, id: `invoiceGroup${++invoiceGroupSeq}`, rows: [] };
          groups.push(byKey[key]);
        }
        byKey[key].rows.push(row);
      });
      return groups;
    }

    function invoiceSubprojectLabel(rows) {
      const labels = [...new Set(rows.map(r => [r.job_number, r.subproject_code].filter(Boolean).join(' ')).filter(Boolean))];
      if (!labels.length) return '';
      return labels.length === 1 ? labels[0] : 'Multiple';
    }

    function invoiceSummaryButton(group, label) {
      return `<span class="invoice-toggle" data-toggle-icon="${group.id}">+</span>${label}`;
    }

    function invoiceSubprojectSelect(group) {
      const first = group.rows[0] || {};
      const options = '<option value="">Unassigned</option>' + state.subprojects.map(s => `<option value="${s.id}" ${String(first.subproject_id || '') === String(s.id) ? 'selected' : ''}>${s.job_number || ''} ${s.code} - ${s.name}</option>`).join('');
      return `<div style="display:flex;gap:6px;align-items:center;min-width:320px;flex-wrap:wrap" onclick="event.stopPropagation()">
        <select data-invoice-subproject="${group.id}" style="min-width:160px">${options}</select>
        <button class="btn" style="padding:7px 9px" type="button" data-save-invoice-subproject="${group.id}" data-source-file="${first.source_file || ''}" data-invoice="${first.ticket_or_invoice || ''}" data-vendor="${first.vendor || ''}">Save</button>
        <button class="btn" style="padding:7px 9px" type="button" data-allocate-invoice="${group.id}">Allocate</button>
      </div>`;
    }

    function openVendorAllocationModal(groupId) {
      const group = vendorAllocationGroups[groupId];
      if (!group) return;
      const total = group.rows.reduce((sum, r) => sum + Number(r.amount || 0), 0);
      const targets = [
        ...state.subprojects.map(sp => ({
          value: `sp:${sp.id}`,
          label: `${[sp.job_number, sp.code].filter(Boolean).join(' ')} - ${sp.name || 'Subproject'}`,
          meta: 'Subproject base'
        })),
        ...state.changeOrders.map(co => {
          const sp = state.subprojects.find(s => String(s.id) === String(co.subproject_id || ''));
          return {
            value: `co:${co.id}`,
            label: `${[co.co_number, co.job_number].filter(Boolean).join(' / ')} - ${co.title || 'Change Order'}`,
            meta: `Change order${sp ? ' under ' + [sp.job_number, sp.code].filter(Boolean).join(' ') : ''}`
          };
        })
      ];
      document.getElementById('vendorAllocationSubtitle').textContent = `${group.vendor || 'Vendor'} / ${group.ticket_or_invoice || 'Invoice'} / ${money(total)} total`;
      document.getElementById('vendorAllocationTargets').innerHTML = targets.length
        ? `<div class="table-wrap" style="max-height:420px"><table><thead><tr><th></th><th>Target</th><th>Type</th></tr></thead><tbody>${targets.map(t => `<tr>
            <td><input type="checkbox" name="allocation_target" value="${htmlEscape(t.value)}" style="width:auto"></td>
            <td>${htmlEscape(t.label)}</td>
            <td>${htmlEscape(t.meta)}</td>
          </tr>`).join('')}</tbody></table></div>`
        : '<p class="muted">Add subprojects or change orders before allocating invoices.</p>';
      document.getElementById('vendorAllocationError').textContent = '';
      const form = document.getElementById('vendorAllocationForm');
      form.dataset.groupId = groupId;
      document.getElementById('vendorAllocationModal').classList.remove('hidden');
    }

    function commonGroupValue(rows, field) {
      const values = [...new Set(rows.map(r => String(r[field] || '')))];
      return values.length === 1 ? values[0] : '';
    }

    function changeOrderOptionsForSubproject(subprojectId, selectedValue='') {
      const selected = String(selectedValue || '');
      const filtered = state.changeOrders.filter(c => !subprojectId || String(c.subproject_id || '') === String(subprojectId));
      return '<option value="">Base</option>' + filtered.map(c => `<option value="${c.id}" ${selected === String(c.id) ? 'selected' : ''}>${[c.co_number, c.job_number].filter(Boolean).join(' / ')}</option>`).join('');
    }

    function updateChangeOrderSelectForSubproject(coSelect, subprojectId, selectedValue=null) {
      if (!coSelect) return;
      const current = selectedValue === null ? coSelect.value : selectedValue;
      coSelect.innerHTML = changeOrderOptionsForSubproject(subprojectId, current);
      if (current && ![...coSelect.options].some(opt => opt.value === String(current))) {
        coSelect.value = '';
      }
    }

    function groupAssignmentSelect(group, field) {
      const selected = commonGroupValue(group.rows, field);
      if (field === 'subproject_id') {
        return `<select data-group-field="${field}" data-group="${group.id}" onclick="event.stopPropagation()">
          <option value="">Unassigned</option>
          ${state.subprojects.map(s => `<option value="${s.id}" ${selected === String(s.id) ? 'selected' : ''}>${s.job_number || ''} ${s.code}</option>`).join('')}
        </select>`;
      }
      const groupSubprojectId = commonGroupValue(group.rows, 'subproject_id');
      return `<select data-group-field="${field}" data-group="${group.id}" onclick="event.stopPropagation()">
        ${changeOrderOptionsForSubproject(groupSubprojectId, selected)}
      </select>`;
    }

    function groupSaveButton(group) {
      const ids = group.rows.map(r => r.id).join(',');
      return `<button class="btn" type="button" data-save-cost-group="${group.id}" data-record-ids="${ids}" onclick="event.stopPropagation()">Save All</button>`;
    }

    function pdfLink(row) {
      if (!row.source_file) return '';
      const viewerHref = `/pdf-viewer/${encodeURIComponent(row.source_file)}`;
      const rawHref = `/uploads/${encodeURIComponent(row.source_file)}`;
      return `<span onclick="event.stopPropagation();"><a class="pdf-link" href="${viewerHref}">View</a> <a class="pdf-link" href="${rawHref}" target="_blank" rel="noopener">Original</a></span>`;
    }

    function wireInvoiceToggles(scope=document) {
      scope.querySelectorAll('[data-toggle-invoice]').forEach(row => {
        row.onclick = event => {
          if (event.target.closest('a, button, select, input')) return;
          const groupId = row.dataset.toggleInvoice;
          const isOpen = row.dataset.open === 'true';
          row.dataset.open = isOpen ? 'false' : 'true';
          const icon = row.querySelector(`[data-toggle-icon="${groupId}"]`);
          if (icon) icon.textContent = isOpen ? '+' : '-';
          document.querySelectorAll(`[data-invoice-detail="${groupId}"]`).forEach(detail => detail.classList.toggle('hidden', isOpen));
        };
      });
    }

    function wireInvoiceSubprojectSaves(scope=document) {
      scope.querySelectorAll('[data-save-invoice-subproject]').forEach(btn => {
        btn.onclick = async event => {
          event.stopPropagation();
          const groupId = btn.dataset.saveInvoiceSubproject;
          const select = scope.querySelector(`[data-invoice-subproject="${groupId}"]`) || document.querySelector(`[data-invoice-subproject="${groupId}"]`);
          await api('/api/vendor-invoice/subproject', {
            method: 'POST',
            body: JSON.stringify({
              project_id: state.projectId,
              source_file: btn.dataset.sourceFile,
              ticket_or_invoice: btn.dataset.invoice,
              vendor: btn.dataset.vendor,
              subproject_id: select ? select.value : ''
            })
          });
          markSaved();
          await refresh();
        };
      });
      scope.querySelectorAll('[data-allocate-invoice]').forEach(btn => {
        btn.onclick = event => {
          event.stopPropagation();
          openVendorAllocationModal(btn.dataset.allocateInvoice);
        };
      });
    }

    function groupedReadonlyCostTable(headers, records, includeSubproject=false) {
      const groups = costRecordGroups(records);
      const body = groups.map(group => {
        if (group.type === 'single') {
          const r = group.rows[0];
          const cells = [
            r.record_date || '',
            ...(includeSubproject ? [[r.job_number, r.subproject_code].filter(Boolean).join(' ')] : []),
            r.source || '',
            r.ticket_or_invoice || '',
            pdfLink(r),
            r.cost_type || '',
            r.item || r.description || '',
            Number(r.qty || 0).toFixed(2),
            money(r.sales_rate || r.rate),
            money(r.sales_amount || r.amount),
            money(r.amount)
          ];
          return `<tr>${cells.map(c => `<td>${c}</td>`).join('')}</tr>`;
        }
        const first = group.rows[0];
        const total = group.rows.reduce((sum, r) => sum + Number(r.amount || 0), 0);
        const salesTotal = group.rows.reduce((sum, r) => sum + Number(r.sales_amount || r.amount || 0), 0);
        const qtyTotal = group.rows.reduce((sum, r) => sum + Number(r.qty || 0), 0);
        const groupLabel = group.type === 'fieldwise' ? 'Field Wise Ticket' : 'Vendor Invoice';
        const groupType = group.type === 'fieldwise' ? 'Imported ticket lines' : 'Material';
        const summary = [
          first.record_date || '',
          ...(includeSubproject ? [invoiceSubprojectLabel(group.rows)] : []),
          invoiceSummaryButton(group, groupLabel),
          first.ticket_or_invoice || '',
          pdfLink(first),
          groupType,
          `<span class="invoice-line-count">${group.rows.length} line item(s)</span>`,
          qtyTotal.toFixed(2),
          '',
          money(salesTotal),
          money(total)
        ];
        const details = group.rows.map(r => {
          const cells = [
            r.record_date || '',
            ...(includeSubproject ? [[r.job_number, r.subproject_code].filter(Boolean).join(' ')] : []),
            r.source || '',
            r.ticket_or_invoice || '',
            pdfLink(r),
            r.cost_type || '',
            r.item || r.description || '',
            Number(r.qty || 0).toFixed(2),
            money(r.sales_rate || r.rate),
            money(r.sales_amount || r.amount),
            money(r.amount)
          ];
          return `<tr class="invoice-detail hidden" data-invoice-detail="${group.id}">${cells.map(c => `<td>${c}</td>`).join('')}</tr>`;
        }).join('');
        return `<tr class="invoice-summary" data-toggle-invoice="${group.id}">${summary.map(c => `<td>${c}</td>`).join('')}</tr>${details}`;
      }).join('');
      return `<table><thead><tr>${headers.map(h => `<th>${h}</th>`).join('')}</tr></thead><tbody>${body}</tbody></table>`;
    }

    function uniqueOptions(records, getter) {
      return [...new Set(records.map(getter).filter(Boolean))].sort((a, b) => String(a).localeCompare(String(b)));
    }

    function optionList(values, selectedValue) {
      const selected = String(selectedValue || '');
      return values.map(v => `<option ${String(v) === selected ? 'selected' : ''}>${htmlEscape(v)}</option>`).join('');
    }

    function costRecordFilterMarkup(filterId, records, includeSubproject=false) {
      const sourceOptions = uniqueOptions(records, r => r.source || '');
      const typeOptions = uniqueOptions(records, r => r.cost_type || '');
      const subprojectOptions = uniqueOptions(records, r => [r.job_number, r.subproject_code].filter(Boolean).join(' '));
      return `<div class="cost-filter" data-cost-filter="${filterId}">
        <div class="cost-filter-bar">
          <div><label>Search</label><input data-cost-search="${filterId}" placeholder="Ticket, item, description, vendor, file"></div>
          ${includeSubproject ? `<div><label>Subproject</label><select data-cost-subproject="${filterId}"><option value="">All</option>${subprojectOptions.map(v => `<option value="${htmlEscape(v)}">${htmlEscape(v)}</option>`).join('')}</select></div>` : ''}
          <div><label>Source</label><select data-cost-source="${filterId}"><option value="">All</option>${sourceOptions.map(v => `<option value="${htmlEscape(v)}">${htmlEscape(v)}</option>`).join('')}</select></div>
          <div><label>Type</label><select data-cost-type="${filterId}"><option value="">All</option>${typeOptions.map(v => `<option value="${htmlEscape(v)}">${htmlEscape(v)}</option>`).join('')}</select></div>
          <button class="btn" type="button" data-cost-clear="${filterId}">Clear</button>
        </div>
        <div class="cost-filter-count" data-cost-count="${filterId}"></div>
        <div class="table-wrap" data-cost-table="${filterId}"></div>
      </div>`;
    }

    function initCostRecordFilter(scope, filterId, records, headers, includeSubproject=false) {
      const container = scope.querySelector(`[data-cost-filter="${filterId}"]`);
      if (!container) return;
      const searchEl = container.querySelector(`[data-cost-search="${filterId}"]`);
      const sourceEl = container.querySelector(`[data-cost-source="${filterId}"]`);
      const typeEl = container.querySelector(`[data-cost-type="${filterId}"]`);
      const subprojectEl = container.querySelector(`[data-cost-subproject="${filterId}"]`);
      const countEl = container.querySelector(`[data-cost-count="${filterId}"]`);
      const tableEl = container.querySelector(`[data-cost-table="${filterId}"]`);
      const clearEl = container.querySelector(`[data-cost-clear="${filterId}"]`);
      const draw = () => {
        const query = String(searchEl?.value || '').trim().toLowerCase();
        const source = sourceEl?.value || '';
        const type = typeEl?.value || '';
        const subproject = subprojectEl?.value || '';
        const filtered = records.filter(r => {
          const subLabel = [r.job_number, r.subproject_code].filter(Boolean).join(' ');
          const haystack = [
            r.record_date, r.source, r.source_file, r.ticket_or_invoice, r.vendor,
            r.cost_type, r.item, r.description, r.raw_cost_source, subLabel
          ].join(' ').toLowerCase();
          return (!query || haystack.includes(query)) &&
            (!source || r.source === source) &&
            (!type || r.cost_type === type) &&
            (!subproject || subLabel === subproject);
        });
        countEl.textContent = `Showing ${filtered.length} of ${records.length} cost record(s)`;
        tableEl.innerHTML = filtered.length
          ? groupedReadonlyCostTable(headers, filtered, includeSubproject)
          : '<table><tbody><tr><td>No cost records match these filters.</td></tr></tbody></table>';
        wireInvoiceToggles(tableEl);
      };
      [searchEl, sourceEl, typeEl, subprojectEl].filter(Boolean).forEach(el => {
        el.oninput = draw;
        el.onchange = draw;
      });
      if (clearEl) clearEl.onclick = () => {
        if (searchEl) searchEl.value = '';
        if (sourceEl) sourceEl.value = '';
        if (typeEl) typeEl.value = '';
        if (subprojectEl) subprojectEl.value = '';
        draw();
      };
      draw();
    }

    function vendorInvoiceLineFilterMarkup(filterId, lines) {
      const vendorOptions = uniqueOptions(lines, r => r.vendor || '');
      const subprojectOptions = uniqueOptions(lines, r => [r.job_number, r.subproject_code].filter(Boolean).join(' '));
      return `<div class="cost-filter" data-vendor-filter="${filterId}">
        <div class="cost-filter-bar">
          <div><label>Search</label><input data-vendor-search="${filterId}" placeholder="Invoice, item, description, vendor, file"></div>
          <div><label>Vendor</label><select data-vendor-name="${filterId}"><option value="">All</option>${vendorOptions.map(v => `<option value="${htmlEscape(v)}">${htmlEscape(v)}</option>`).join('')}</select></div>
          <div><label>Subproject</label><select data-vendor-subproject="${filterId}"><option value="">All</option>${subprojectOptions.map(v => `<option value="${htmlEscape(v)}">${htmlEscape(v)}</option>`).join('')}</select></div>
          <div><label>Date</label><input type="date" data-vendor-date="${filterId}"></div>
          <button class="btn" type="button" data-vendor-clear="${filterId}">Clear</button>
        </div>
        <div class="filter-summary">
          <div class="summary-pill"><div class="summary-label">Lines Shown</div><div class="summary-value" data-vendor-count="${filterId}">0 of ${lines.length}</div></div>
          <div class="summary-pill"><div class="summary-label">Amount Total</div><div class="summary-value amount" data-vendor-amount="${filterId}">$0.00</div></div>
        </div>
      </div>`;
    }

    function bidFilterMarkup(filterId, bids) {
      const customerOptions = uniqueOptions(bids, r => r.customer || '');
      const stageOptions = uniqueOptions(bids, r => r.stage || '');
      const estimatorOptions = uniqueOptions(bids, r => r.estimator || '');
      return `<div class="cost-filter" data-bid-filter="${filterId}">
        <div class="cost-filter-bar">
          <div><label>Search</label><input data-bid-search="${filterId}" placeholder="RFQ, customer, project, notes"></div>
          <div><label>Customer</label><select data-bid-customer="${filterId}"><option value="">All</option>${customerOptions.map(v => `<option value="${htmlEscape(v)}">${htmlEscape(v)}</option>`).join('')}</select></div>
          <div><label>Stage</label><select data-bid-stage="${filterId}"><option value="">All</option>${stageOptions.map(v => `<option value="${htmlEscape(v)}">${htmlEscape(v)}</option>`).join('')}</select></div>
          <div><label>Estimator</label><select data-bid-estimator="${filterId}"><option value="">All</option>${estimatorOptions.map(v => `<option value="${htmlEscape(v)}">${htmlEscape(v)}</option>`).join('')}</select></div>
          <button class="btn" type="button" data-bid-clear="${filterId}">Clear</button>
        </div>
        <div class="filter-summary">
          <div class="summary-pill"><div class="summary-label">Bids Shown</div><div class="summary-value" data-bid-count="${filterId}">0 of ${bids.length}</div></div>
          <div class="summary-pill"><div class="summary-label">Bid Price Total</div><div class="summary-value amount" data-bid-total="${filterId}">$0.00</div></div>
          <div class="summary-pill"><div class="summary-label">Weighted Total</div><div class="summary-value" data-bid-weighted="${filterId}">$0.00</div></div>
        </div>
      </div>`;
    }

    function isBidStale(bid) {
      if (['Won','Lost'].includes(String(bid.outcome || ''))) return false;
      const lastUpdate = Date.parse(bid.updated_at || bid.created_at || '');
      if (!lastUpdate) return false;
      return Date.now() - lastUpdate >= 5 * 24 * 60 * 60 * 1000;
    }

    function bidTableHtml(bids) {
      if (!bids.length) return '<tbody><tr><td>No bids match these filters.</td></tr></tbody>';
      const stageOptions = ['New RFQ','Go/No-Go','Estimating','Submitted','Award Pending','Closed'];
      const goNoGoOptions = ['Go','No Go','Review'];
      const outcomeOptions = ['Pending','Won','Lost'];
      return `<thead><tr><th>RFQ</th><th>Received</th><th>Customer</th><th>Project</th><th>Estimator</th><th>Stage</th><th>Due</th><th>Go / No-Go</th><th>Est. Cost</th><th>Margin</th><th>Bid Price</th><th>Prob.</th><th>Weighted</th><th>Outcome</th><th>Notes</th><th></th></tr></thead>
        <tbody>${bids.map(b => `<tr class="${isBidStale(b) ? 'bid-stale' : ''}" title="${isBidStale(b) ? 'No update in 5 or more days' : ''}">
          <td><input data-bid="${b.id}" data-field="rfq_no" value="${htmlEscape(b.rfq_no || '')}"></td>
          <td><input data-bid="${b.id}" data-field="date_received" type="date" value="${htmlEscape(b.date_received || '')}"></td>
          <td><input data-bid="${b.id}" data-field="customer" value="${htmlEscape(b.customer || '')}"></td>
          <td><input data-bid="${b.id}" data-field="project_name" value="${htmlEscape(b.project_name || '')}"></td>
          <td><input data-bid="${b.id}" data-field="estimator" value="${htmlEscape(b.estimator || '')}"></td>
          <td><select data-bid="${b.id}" data-field="stage">${optionList(stageOptions, b.stage)}</select></td>
          <td><input data-bid="${b.id}" data-field="bid_due_date" type="date" value="${htmlEscape(b.bid_due_date || '')}"></td>
          <td><select data-bid="${b.id}" data-field="go_no_go">${optionList(goNoGoOptions, b.go_no_go)}</select></td>
          <td><input data-bid="${b.id}" data-field="estimated_cost" type="number" step="0.01" value="${Number(b.estimated_cost || 0).toFixed(2)}"></td>
          <td><input data-bid="${b.id}" data-field="target_margin" type="number" step="0.01" value="${Number(b.target_margin || 0).toFixed(2)}"></td>
          <td><input data-bid="${b.id}" data-field="bid_price" type="number" step="0.01" value="${Number(b.bid_price || 0).toFixed(2)}"></td>
          <td><input data-bid="${b.id}" data-field="probability" type="number" step="0.01" value="${Number(b.probability || 0).toFixed(2)}"></td>
          <td>${money(b.weighted_value)}</td>
          <td><select data-bid="${b.id}" data-field="outcome">${optionList(outcomeOptions, b.outcome)}</select></td>
          <td><input data-bid="${b.id}" data-field="notes" value="${htmlEscape(b.notes || '')}"></td>
          <td><button class="btn" data-save-bid="${b.id}" type="button">Save</button></td>
        </tr>`).join('')}</tbody>`;
    }

    function fieldTicketLineFilterMarkup(filterId, lines) {
      const ticketOptions = uniqueOptions(lines, r => r.ticket_or_invoice || '');
      const subprojectOptions = uniqueOptions(lines, r => [r.job_number, r.subproject_code].filter(Boolean).join(' '));
      const typeOptions = uniqueOptions(lines, r => r.cost_type || '');
      return `<div class="cost-filter" data-field-ticket-filter="${filterId}">
        <div class="cost-filter-bar">
          <div><label>Search</label><input data-field-ticket-search="${filterId}" placeholder="Ticket, item, description, source file"></div>
          <div><label>Ticket</label><select data-field-ticket-number="${filterId}"><option value="">All</option>${ticketOptions.map(v => `<option value="${htmlEscape(v)}">${htmlEscape(v)}</option>`).join('')}</select></div>
          <div><label>Subproject</label><select data-field-ticket-subproject="${filterId}"><option value="">All</option>${subprojectOptions.map(v => `<option value="${htmlEscape(v)}">${htmlEscape(v)}</option>`).join('')}</select></div>
          <div><label>Type</label><select data-field-ticket-type="${filterId}"><option value="">All</option>${typeOptions.map(v => `<option value="${htmlEscape(v)}">${htmlEscape(v)}</option>`).join('')}</select></div>
          <button class="btn" type="button" data-field-ticket-clear="${filterId}">Clear</button>
        </div>
        <div class="filter-summary">
          <div class="summary-pill"><div class="summary-label">Lines Shown</div><div class="summary-value" data-field-ticket-count="${filterId}">0 of ${lines.length}</div></div>
          <div class="summary-pill"><div class="summary-label">Sales Total</div><div class="summary-value amount" data-field-ticket-sales="${filterId}">$0.00</div></div>
          <div class="summary-pill"><div class="summary-label">Raw Cost Total</div><div class="summary-value" data-field-ticket-raw="${filterId}">$0.00</div></div>
        </div>
      </div>`;
    }

    function invoiceDateInputValue(value) {
      const text = String(value || '').trim();
      const m = text.match(/^([0-9]{1,2})\/([0-9]{1,2})\/([0-9]{2,4})$/);
      if (!m) return text;
      const year = m[3].length === 2 ? `20${m[3]}` : m[3];
      return `${year}-${m[1].padStart(2, '0')}-${m[2].padStart(2, '0')}`;
    }

    function vendorInvoiceLinesTableHtml(lines) {
      vendorAllocationGroups = {};
      if (!lines.length) {
        return '<tbody><tr><td>No vendor invoice lines match these filters.</td></tr></tbody>';
      }
      const body = vendorInvoiceGroups(lines).map(group => {
        if (group.type === 'single') {
          const r = group.rows[0];
          return `<tr>
            <td>${r.ticket_or_invoice || ''}</td>
            <td>${pdfLink(r)}</td>
            <td>${r.record_date || ''}</td>
            <td>${r.vendor || ''}</td>
            <td>${[r.job_number, r.subproject_code].filter(Boolean).join(' ')}</td>
            <td>${r.item || ''}</td>
            <td>${r.description || ''}</td>
            <td>${Number(r.qty || 0).toFixed(2)}</td>
            <td>${money(r.rate)}</td>
            <td>${money(r.amount)}</td>
          </tr>`;
        }
        const first = group.rows[0];
        const total = group.rows.reduce((sum, r) => sum + Number(r.amount || 0), 0);
        const qtyTotal = group.rows.reduce((sum, r) => sum + Number(r.qty || 0), 0);
        vendorAllocationGroups[group.id] = {
          rows: group.rows,
          source_file: first.source_file || '',
          ticket_or_invoice: first.ticket_or_invoice || '',
          vendor: first.vendor || ''
        };
        const summary = `<tr class="invoice-summary" data-toggle-invoice="${group.id}">
          <td>${invoiceSummaryButton(group, first.ticket_or_invoice || 'Invoice')}</td>
          <td>${pdfLink(first)}</td>
          <td>${first.record_date || ''}</td>
          <td>${first.vendor || ''}</td>
          <td>${invoiceSubprojectSelect(group)}</td>
          <td><span class="invoice-line-count">${group.rows.length} line item(s)</span></td>
          <td>Click to view invoice items</td>
          <td>${qtyTotal.toFixed(2)}</td>
          <td></td>
          <td>${money(total)}</td>
        </tr>`;
        const details = group.rows.map(r => `<tr class="invoice-detail hidden" data-invoice-detail="${group.id}">
          <td>${r.ticket_or_invoice || ''}</td>
          <td>${pdfLink(r)}</td>
          <td>${r.record_date || ''}</td>
          <td>${r.vendor || ''}</td>
          <td>${[r.job_number, r.subproject_code].filter(Boolean).join(' ')}</td>
          <td>${r.item || ''}</td>
          <td>${r.description || ''}</td>
          <td>${Number(r.qty || 0).toFixed(2)}</td>
          <td>${money(r.rate)}</td>
          <td>${money(r.amount)}</td>
        </tr>`).join('');
        return summary + details;
      }).join('');
      return `
        <thead><tr><th>Invoice</th><th>PDF</th><th>Date</th><th>Vendor</th><th>Subproject</th><th>Item</th><th>Description</th><th>Qty</th><th>Unit Cost</th><th>Amount</th></tr></thead>
        <tbody>${body}</tbody>`;
    }

    function groupedEditableCostRows(records, spOpts, coOpts) {
      return costRecordGroups(records).map(group => {
        const rowHtml = c => `<tr>
          <td>${c.record_date || ''}</td><td>${c.source}</td><td>${c.ticket_or_invoice || ''}</td><td>${pdfLink(c)}</td>
          <td><select data-field="cost_type" data-id="${c.id}"><option>${c.cost_type || 'Uncoded'}</option><option>Labor</option><option>Material</option><option>Field Ticket Material</option><option>Equipment</option><option>Rental</option><option>Other</option></select></td>
          <td>${c.item || ''}</td><td>${c.description || ''}</td><td>${Number(c.qty || 0).toFixed(2)}</td><td>${money(c.sales_rate || c.rate)}</td><td>${money(c.sales_amount || c.amount)}</td><td>${money(c.amount)}</td><td class="${c.raw_cost_source === 'Missing project rate' || c.raw_cost_source === 'Missing internal rate' ? 'warn' : ''}">${c.raw_cost_source || ''}</td>
          <td><select data-field="subproject_id" data-id="${c.id}"><option value="">Unassigned</option>${spOpts}</select></td>
          <td><select data-field="change_order_id" data-id="${c.id}">${changeOrderOptionsForSubproject(c.subproject_id, c.change_order_id)}</select></td>
          <td><button class="btn" data-save="${c.id}">Save</button></td>
        </tr>`;
        if (group.type === 'single') return rowHtml(group.rows[0]);
        const first = group.rows[0];
        const total = group.rows.reduce((sum, r) => sum + Number(r.amount || 0), 0);
        const salesTotal = group.rows.reduce((sum, r) => sum + Number(r.sales_amount || r.amount || 0), 0);
        const qtyTotal = group.rows.reduce((sum, r) => sum + Number(r.qty || 0), 0);
        const subproject = invoiceSubprojectLabel(group.rows);
        const firstType = [...new Set(group.rows.map(r => r.cost_type || '').filter(Boolean))];
        const groupType = firstType.length === 1 ? firstType[0] : 'Mixed';
        const groupLabel = group.type === 'fieldwise' ? 'Field Wise Ticket' : 'Vendor Invoice';
        const rateSource = group.type === 'fieldwise' ? 'Field Wise import' : 'Vendor invoice';
        const summary = `<tr class="invoice-summary" data-toggle-invoice="${group.id}">
          <td>${first.record_date || ''}</td>
          <td>${invoiceSummaryButton(group, groupLabel)}</td>
          <td>${first.ticket_or_invoice || ''}</td>
          <td>${pdfLink(first)}</td>
          <td>${groupType}</td>
          <td><span class="invoice-line-count">${group.rows.length} line item(s)</span></td>
          <td>${group.type === 'fieldwise' ? first.source_file || '' : first.vendor || ''}</td>
          <td>${qtyTotal.toFixed(2)}</td>
          <td></td>
          <td>${money(salesTotal)}</td>
          <td>${money(total)}</td>
          <td>${rateSource}</td>
          <td>${groupAssignmentSelect(group, 'subproject_id')}</td>
          <td>${groupAssignmentSelect(group, 'change_order_id')}</td>
          <td>${groupSaveButton(group)}</td>
        </tr>`;
        const details = group.rows.map(c => rowHtml(c).replace('<tr>', `<tr class="invoice-detail hidden" data-invoice-detail="${group.id}">`)).join('');
        return summary + details;
      }).join('');
    }

    function subprojectTable(subprojects, selectedId=null) {
      return `<table><thead><tr><th>Job</th><th>Code</th><th>Name</th><th>Pricing</th><th>Sales / Contract</th><th>Labor Hrs</th><th>Budget</th><th>Raw Actual</th><th>Profit</th><th>Margin</th><th>Used</th></tr></thead><tbody>${subprojects.map(x => {
        const used = x.budget_total ? x.actual_cost / x.budget_total : 0;
        const laborUsed = Number(x.labor_hours_used || 0);
        const laborBudget = Number(x.budget_labor_hours || 0);
        const laborPct = laborBudget ? laborUsed / laborBudget : 0;
        const salesValue = x.pricing_type === 'T&M' ? x.sales_value : x.contract_value;
        const profit = Number(x.profit || 0);
        const margin = Number(x.margin || 0);
        const open = `<button class="btn" style="padding:4px 7px" data-open-subproject="${x.id}" type="button">Open</button>`;
        const selected = String(selectedId || '') === String(x.id);
        return `<tr class="selectable-row ${selected ? 'selected' : ''}" data-select-subproject="${x.id}">
          <td>${x.job_number || ''}</td>
          <td>${x.code}</td>
          <td>${open} ${x.name}</td>
          <td>${x.pricing_type || 'Fixed'}</td>
          <td>${money(salesValue)}</td>
          <td>${laborUsed.toFixed(2)} / ${laborBudget.toFixed(2)}<br><div class="bar"><span style="width:${Math.min(100, laborPct*100)}%"></span></div></td>
          <td>${money(x.budget_total)}</td>
          <td>${money(x.actual_cost)}</td>
          <td class="${profit >= 0 ? 'good' : 'bad'}">${money(profit)}</td>
          <td class="${margin >= 0 ? 'good' : 'bad'}">${pct(margin)}</td>
          <td><div class="bar"><span style="width:${Math.min(100, used*100)}%"></span></div> ${pct(used)}</td>
        </tr>`;
      }).join('')}</tbody></table>`;
    }

    async function copySubproject(id) {
      const original = state.subprojects.find(s => String(s.id) === String(id));
      const label = original ? [original.job_number, original.code, original.name].filter(Boolean).join(' - ') : 'this subproject';
      const form = document.getElementById('copySubprojectForm');
      form.reset();
      form.elements.subproject_id.value = id;
      if (original) {
        form.elements.code.value = original.code || '';
        form.elements.name.value = original.name || '';
        form.elements.pricing_type.value = original.pricing_type || 'Fixed';
        form.elements.contract_value.value = Number(original.contract_value || 0);
        form.elements.budget_labor_hours.value = Number(original.budget_labor_hours || 0);
        form.elements.budget_labor.value = Number(original.budget_labor || 0);
        form.elements.budget_material.value = Number(original.budget_material || 0);
        form.elements.budget_equipment.value = Number(original.budget_equipment || 0);
      }
      document.getElementById('copySubprojectMessage').textContent = `Copying ${label}. Enter the new job/order number, then adjust any copied setup information if needed.`;
      document.getElementById('copySubprojectError').textContent = '';
      document.getElementById('copySubprojectModal').classList.remove('hidden');
      setTimeout(() => form.elements.job_number?.focus(), 50);
    }

    window.copySubprojectById = copySubproject;

    document.getElementById('copySubprojectForm').onsubmit = async event => {
      event.preventDefault();
      const form = event.target;
      const subprojectId = form.elements.subproject_id.value;
      const jobNumber = form.elements.job_number.value.trim();
      const error = document.getElementById('copySubprojectError');
      error.textContent = '';
      if (!jobNumber) {
        error.textContent = 'Enter a job/order number before copying this subproject.';
        return;
      }
      try {
        await api(`/api/subprojects/${subprojectId}/copy`, {
          method:'POST',
          body: JSON.stringify({
            job_number: jobNumber,
            code: form.elements.code.value.trim(),
            name: form.elements.name.value.trim(),
            pricing_type: form.elements.pricing_type.value,
            contract_value: form.elements.contract_value.value,
            budget_labor_hours: form.elements.budget_labor_hours.value,
            budget_labor: form.elements.budget_labor.value,
            budget_material: form.elements.budget_material.value,
            budget_equipment: form.elements.budget_equipment.value
          })
        });
        closeCopySubprojectModal();
        markSaved();
        await refresh();
      } catch (err) {
        error.textContent = err.message || 'Could not copy subproject.';
      }
    };

    document.addEventListener('click', event => {
      const copyButton = event.target.closest('[data-copy-subproject], [data-copy-sp]');
      if (!copyButton) return;
      event.preventDefault();
      event.stopPropagation();
      copySubproject(copyButton.dataset.copySubproject || copyButton.dataset.copySp);
    }, true);

    function copyChangeOrder(id) {
      const original = state.changeOrders.find(c => String(c.id) === String(id));
      const label = original ? [original.co_number, original.job_number, original.title].filter(Boolean).join(' - ') : 'this change order';
      const form = document.getElementById('copyChangeOrderForm');
      form.reset();
      form.elements.change_order_id.value = id;
      form.elements.subproject_id.innerHTML = '<option value="">Unassigned</option>' + state.subprojects.map(s => `<option value="${s.id}">${htmlEscape([s.job_number, s.code, s.name].filter(Boolean).join(' - '))}</option>`).join('');
      if (original) {
        form.elements.subproject_id.value = original.subproject_id || '';
        form.elements.order_type.value = original.order_type || 'Change Order';
        form.elements.pricing_type.value = original.pricing_type || 'Fixed';
        form.elements.status.value = original.status || 'Pending';
        form.elements.title.value = original.title || '';
        form.elements.quoted_value.value = Number(original.quoted_value || 0);
        form.elements.approved_value.value = Number(original.approved_value || 0);
      }
      document.getElementById('copyChangeOrderMessage').textContent = `Copying ${label}. Enter the new CO number and job/order number, then adjust any copied setup information if needed.`;
      document.getElementById('copyChangeOrderError').textContent = '';
      document.getElementById('copyChangeOrderModal').classList.remove('hidden');
      setTimeout(() => form.elements.co_number?.focus(), 50);
    }

    window.copyChangeOrderById = copyChangeOrder;

    document.getElementById('copyChangeOrderForm').onsubmit = async event => {
      event.preventDefault();
      const form = event.target;
      const changeOrderId = form.elements.change_order_id.value;
      const coNumber = form.elements.co_number.value.trim();
      const jobNumber = form.elements.job_number.value.trim();
      const error = document.getElementById('copyChangeOrderError');
      error.textContent = '';
      if (!coNumber) {
        error.textContent = 'Enter a new CO number before copying this change order.';
        return;
      }
      if (!jobNumber) {
        error.textContent = 'Enter a new job/order number before copying this change order.';
        return;
      }
      try {
        await api(`/api/change-orders/${changeOrderId}/copy`, {
          method:'POST',
          body: JSON.stringify({
            subproject_id: form.elements.subproject_id.value,
            order_type: form.elements.order_type.value,
            co_number: coNumber,
            job_number: jobNumber,
            pricing_type: form.elements.pricing_type.value,
            status: form.elements.status.value,
            title: form.elements.title.value.trim(),
            quoted_value: form.elements.quoted_value.value,
            approved_value: form.elements.approved_value.value
          })
        });
        closeCopyChangeOrderModal();
        markSaved();
        await refresh();
      } catch (err) {
        error.textContent = err.message || 'Could not copy change order.';
      }
    };

    document.addEventListener('click', event => {
      const copyButton = event.target.closest('[data-copy-co]');
      if (!copyButton) return;
      event.preventDefault();
      event.stopPropagation();
      copyChangeOrder(copyButton.dataset.copyCo);
    }, true);

    async function loadSubprojectDetail(subprojectId, shouldScroll=true, changeOrderId=null) {
      openSubprojectDetailId = subprojectId;
      const coParam = changeOrderId ? `&change_order_id=${encodeURIComponent(changeOrderId)}` : '';
      const d = await api(`/api/subproject-detail?subproject_id=${subprojectId}${coParam}`);
      const panel = document.getElementById('subprojectDetailPanel');
      const costFilterId = `costFilter${++costFilterSeq}`;
      document.getElementById('subprojectDetailTitle').textContent = `${d.subproject.job_number || ''} ${d.subproject.code} - ${d.subproject.name} / ${d.scope_label || 'Base Contract'}`;
      panel.classList.remove('hidden');
      const laborPct = Number(d.labor_hours_used_pct || 0);
      const budgetPct = Number(d.budget_used || 0);
      document.getElementById('subprojectDetail').innerHTML = `
        <div class="grid cols-4">
          <div class="panel kpi help-card">${help(d.subproject.pricing_type === 'T&M' && !d.selected_change_order ? 'For T&M base work, sales value is the total Field Wise billable sales imported for this subproject.' : 'For fixed work, contract value is the entered subproject contract or selected change order value.')}<div class="label">${d.subproject.pricing_type === 'T&M' && !d.selected_change_order ? 'Field Wise Sales Value' : 'Contract Value'}</div><div class="value">${money(d.contract_value)}</div><div class="hint">${d.subproject.pricing_type || 'Fixed'} pricing</div></div>
          <div class="panel kpi help-card">${help(d.subproject.pricing_type === 'T&M' && !d.selected_change_order ? 'For T&M profit, raw actual cost uses Field Wise labor, equipment, and Field Wise material costs. Vendor invoices are visible below but not counted here.' : 'Raw actual cost is the assigned cost-record total for this scope.')}<div class="label">Raw Actual Cost</div><div class="value">${money(d.raw_actual)}</div><div class="hint">Budget used ${pct(budgetPct)}</div></div>
          <div class="panel kpi help-card">${help('Profit equals the sales or contract value minus raw actual cost. Margin equals profit divided by sales or contract value.')}<div class="label">Profit</div><div class="value ${d.profit >= 0 ? 'good' : 'bad'}">${money(d.profit)}</div><div class="hint">Margin ${pct(d.margin)}</div></div>
          <div class="panel kpi help-card">${help('Labor hours are summed from Labor cost records for the selected scope and compared against the subproject labor-hour budget.')}<div class="label">Labor Hours</div><div class="value">${Number(d.labor_hours_used || 0).toFixed(2)}</div><div class="hint">of ${Number(d.labor_hours_budget || 0).toFixed(2)} budgeted</div></div>
        </div>
        <div class="grid cols-2" style="margin-top:14px">
          <div>
            <h3>Budget And Labor</h3>
            ${table(['Metric','Value'], [
              ['Raw cost budget', money(d.budget_total)],
              ['Labor $ budget', money(d.subproject.budget_labor)],
              ['Material budget', money(d.subproject.budget_material)],
              ['Equipment budget', money(d.subproject.budget_equipment)],
              ['Raw actual cost', money(d.raw_actual)],
              ['Budget remaining', money(d.budget_total - d.raw_actual)],
              ['Labor hours used', `${Number(d.labor_hours_used || 0).toFixed(2)} / ${Number(d.labor_hours_budget || 0).toFixed(2)}`],
              ['Labor hours used %', pct(laborPct)]
            ])}
          </div>
          <div>
            <h3>Material Comparison</h3>
            ${table(['Metric','Amount'], [
              ['Field ticket material listed', money(d.field_ticket_material)],
              ['Vendor invoice material purchased', money(d.vendor_material)],
              ['Difference', money(d.vendor_material - d.field_ticket_material)]
            ])}
          </div>
        </div>
        <h3>Cost Type Breakdown</h3>
        ${table(['Type','Amount'], d.by_type.map(x => [x.label, money(x.amount)]))}
        <details>
          <summary>Cost Records (${d.records.length})</summary>
          <div class="detail-body">
            ${costRecordFilterMarkup(costFilterId, d.records)}
          </div>
        </details>
      `;
      initCostRecordFilter(panel, costFilterId, d.records, ['Date','Source','Ticket / Invoice','PDF','Type','Item','Qty','Unit Price','Sales','Raw Cost']);
      wireInvoiceToggles(panel);
      if (shouldScroll) panel.scrollIntoView({ behavior: 'smooth', block: 'start' });
    }

    document.getElementById('closeSubprojectDetail').onclick = () => {
      openSubprojectDetailId = null;
      document.getElementById('subprojectDetailPanel').classList.add('hidden');
    };

    async function loadMasterDetail(shouldScroll=true) {
      if (!state.projectId) return;
      masterDetailIsOpen = true;
      const d = await api(`/api/master-detail?project_id=${state.projectId}`);
      const panel = document.getElementById('masterDetailPanel');
      const costFilterId = `costFilter${++costFilterSeq}`;
      document.getElementById('masterDetailTitle').textContent = `${d.project.name} - Master Project Detail`;
      panel.classList.remove('hidden');
      const laborPct = Number(d.labor_hours_used_pct || 0);
      const budgetPct = Number(d.budget_used || 0);
      document.getElementById('masterDetail').innerHTML = `
        <div class="grid cols-4">
          <div class="panel kpi help-card">${help('Master value equals base subproject value plus approved change orders. T&M base value comes from Field Wise sales; fixed base value comes from entered contract values.')}<div class="label">Contract + Approved COs</div><div class="value">${money(d.contract_value)}</div><div class="hint">Base ${money(d.base_contract_value)} / COs ${money(d.approved_co_value)}</div></div>
          <div class="panel kpi help-card">${help('Master raw actual cost sums assigned cost records. For T&M scopes, vendor invoices are excluded from raw cost and Field Wise costs are used.')}<div class="label">Raw Actual Cost</div><div class="value">${money(d.raw_actual)}</div><div class="hint">Budget used ${pct(budgetPct)}</div></div>
          <div class="panel kpi help-card">${help('Master profit equals Contract + Approved COs minus Raw Actual Cost. Margin equals profit divided by Contract + Approved COs.')}<div class="label">Profit</div><div class="value ${d.profit >= 0 ? 'good' : 'bad'}">${money(d.profit)}</div><div class="hint">Margin ${pct(d.margin)}</div></div>
          <div class="panel kpi help-card">${help('Labor hours are summed from all Labor cost records in this master project and compared with the combined labor-hour budget.')}<div class="label">Labor Hours</div><div class="value">${Number(d.labor_hours_used || 0).toFixed(2)}</div><div class="hint">of ${Number(d.labor_hours_budget || 0).toFixed(2)} budgeted</div></div>
        </div>
        <div class="grid cols-2" style="margin-top:14px">
          <div>
            <h3>Budget And Labor</h3>
            ${table(['Metric','Value'], [
              ['Raw cost budget', money(d.budget_total)],
              ['Raw actual cost', money(d.raw_actual)],
              ['Budget remaining', money(d.budget_total - d.raw_actual)],
              ['Labor hours used', `${Number(d.labor_hours_used || 0).toFixed(2)} / ${Number(d.labor_hours_budget || 0).toFixed(2)}`],
              ['Labor hours used %', pct(laborPct)]
            ])}
          </div>
          <div>
            <h3>Material Comparison</h3>
            ${table(['Metric','Amount'], [
              ['Field ticket material listed', money(d.field_ticket_material)],
              ['Vendor invoice material purchased', money(d.vendor_material)],
              ['Difference', money(d.vendor_material - d.field_ticket_material)]
            ])}
          </div>
        </div>
        <h3>Subproject Rollup</h3>
        ${subprojectTable(d.subprojects)}
        <h3>Cost Type Breakdown</h3>
        ${table(['Type','Amount'], d.by_type.map(x => [x.label, money(x.amount)]))}
        <details>
          <summary>All Cost Records (${d.records.length})</summary>
          <div class="detail-body">
            ${costRecordFilterMarkup(costFilterId, d.records, true)}
          </div>
        </details>
      `;
      panel.querySelectorAll('[data-open-subproject]').forEach(btn => btn.onclick = () => loadSubprojectDetail(btn.dataset.openSubproject));
      initCostRecordFilter(panel, costFilterId, d.records, ['Date','Subproject','Source','Ticket / Invoice','PDF','Type','Item','Qty','Unit Price','Sales','Raw Cost'], true);
      wireInvoiceToggles(panel);
      if (shouldScroll) panel.scrollIntoView({ behavior: 'smooth', block: 'start' });
    }

    document.getElementById('openMasterDetail').onclick = loadMasterDetail;
    document.getElementById('closeMasterDetail').onclick = () => {
      masterDetailIsOpen = false;
      document.getElementById('masterDetailPanel').classList.add('hidden');
    };

    async function refreshOpenDetails() {
      if (masterDetailIsOpen && !document.getElementById('masterDetailPanel').classList.contains('hidden')) {
        await loadMasterDetail(false);
      }
      if (openSubprojectDetailId && !document.getElementById('subprojectDetailPanel').classList.contains('hidden')) {
        await loadSubprojectDetail(openSubprojectDetailId, false, selectedDashboardChangeOrderId);
      }
    }

    function loadSubprojectEditor() {
      const tableEl = document.getElementById('subprojectEditTable');
      if (!tableEl) return;
      tableEl.innerHTML = `
        <thead><tr><th>Job / Order #</th><th>Code</th><th>Name</th><th>Pricing</th><th>Contract Value</th><th>Labor Hours Budget</th><th>Labor $ Budget</th><th>Material Budget</th><th>Equipment Budget</th><th></th></tr></thead>
        <tbody>${state.subprojects.map(s => `<tr>
          <td><input data-sp="${s.id}" data-field="job_number" value="${s.job_number || ''}"></td>
          <td><input data-sp="${s.id}" data-field="code" value="${s.code || ''}"></td>
          <td><input data-sp="${s.id}" data-field="name" value="${s.name || ''}"></td>
          <td><select data-sp="${s.id}" data-field="pricing_type"><option ${s.pricing_type === 'Fixed' ? 'selected' : ''}>Fixed</option><option ${s.pricing_type === 'T&M' ? 'selected' : ''}>T&M</option></select></td>
          <td><input data-sp="${s.id}" data-field="contract_value" type="number" step="0.01" value="${s.contract_value || 0}"></td>
          <td><input data-sp="${s.id}" data-field="budget_labor_hours" type="number" step="0.01" value="${s.budget_labor_hours || 0}"></td>
          <td><input data-sp="${s.id}" data-field="budget_labor" type="number" step="0.01" value="${s.budget_labor || 0}"></td>
          <td><input data-sp="${s.id}" data-field="budget_material" type="number" step="0.01" value="${s.budget_material || 0}"></td>
          <td><input data-sp="${s.id}" data-field="budget_equipment" type="number" step="0.01" value="${s.budget_equipment || 0}"></td>
          <td class="actions-cell"><button class="btn" data-save-sp="${s.id}" type="button">Save</button><button class="btn" data-copy-sp="${s.id}" onclick="event.preventDefault(); event.stopPropagation(); window.copySubprojectById('${s.id}')" type="button">Copy</button><button class="btn danger" data-delete-sp="${s.id}" type="button">Delete</button></td>
        </tr>`).join('')}</tbody>`;
      tableEl.querySelectorAll('[data-sp]').forEach(el => {
        const markRowDirty = () => el.closest('tr')?.classList.add('setup-dirty');
        el.oninput = markRowDirty;
        el.onchange = markRowDirty;
      });
      document.querySelectorAll('[data-save-sp]').forEach(btn => btn.onclick = async () => {
        const id = btn.dataset.saveSp;
        const fields = {};
        document.querySelectorAll(`[data-sp="${id}"]`).forEach(el => fields[el.dataset.field] = el.value);
        await api(`/api/subprojects/${id}`, { method:'PUT', body: JSON.stringify(fields) });
        btn.closest('tr')?.classList.remove('setup-dirty');
        markSaved();
        await refresh();
      });
      document.querySelectorAll('[data-delete-sp]').forEach(btn => btn.onclick = async () => {
        const id = btn.dataset.deleteSp;
        const subproject = state.subprojects.find(s => String(s.id) === String(id));
        const label = subproject ? [subproject.job_number, subproject.code, subproject.name].filter(Boolean).join(' - ') : 'this subproject';
        if (!window.confirm(`Delete subproject ${label}?\n\nCost records, customer invoices, and change orders tied to it will be kept, but they will no longer be assigned to this subproject.`)) return;
        await api(`/api/subprojects/${id}`, { method:'DELETE' });
        if (String(selectedDashboardSubprojectId || '') === String(id)) selectedDashboardSubprojectId = null;
        if (String(openSubprojectDetailId || '') === String(id)) {
          openSubprojectDetailId = null;
          document.getElementById('subprojectDetailPanel').classList.add('hidden');
        }
        markSaved();
        await refresh();
      });
    }

    function loadChangeOrderEditor() {
      const tableEl = document.getElementById('changeOrderEditTable');
      if (!tableEl) return;
      if (!state.changeOrders.length) {
        tableEl.innerHTML = '<tbody><tr><td>No change orders have been added yet.</td></tr></tbody>';
        return;
      }
      const subprojectOptions = subprojectId => '<option value="">Unassigned</option>' + state.subprojects.map(s => `<option value="${s.id}" ${String(subprojectId || '') === String(s.id) ? 'selected' : ''}>${htmlEscape([s.job_number, s.code, s.name].filter(Boolean).join(' - '))}</option>`).join('');
      tableEl.innerHTML = `
        <thead><tr><th>Subproject</th><th>Type</th><th>CO Number</th><th>Job / Order #</th><th>Pricing</th><th>Status</th><th>Title</th><th>Quoted Value</th><th>Approved Value</th><th></th></tr></thead>
        <tbody>${state.changeOrders.map(c => `<tr>
          <td><select data-co-edit="${c.id}" data-field="subproject_id">${subprojectOptions(c.subproject_id)}</select></td>
          <td><select data-co-edit="${c.id}" data-field="order_type"><option ${c.order_type === 'Change Order' || !c.order_type ? 'selected' : ''}>Change Order</option><option ${c.order_type === 'Child Project' ? 'selected' : ''}>Child Project</option></select></td>
          <td><input data-co-edit="${c.id}" data-field="co_number" value="${htmlEscape(c.co_number || '')}"></td>
          <td><input data-co-edit="${c.id}" data-field="job_number" value="${htmlEscape(c.job_number || '')}"></td>
          <td><select data-co-edit="${c.id}" data-field="pricing_type"><option ${c.pricing_type === 'Fixed' ? 'selected' : ''}>Fixed</option><option ${c.pricing_type === 'T&M' ? 'selected' : ''}>T&M</option></select></td>
          <td><select data-co-edit="${c.id}" data-field="status"><option ${c.status === 'Pending' ? 'selected' : ''}>Pending</option><option ${c.status === 'Approved' ? 'selected' : ''}>Approved</option><option ${c.status === 'Rejected' ? 'selected' : ''}>Rejected</option><option ${c.status === 'Billed' ? 'selected' : ''}>Billed</option></select></td>
          <td><input data-co-edit="${c.id}" data-field="title" value="${htmlEscape(c.title || '')}"></td>
          <td><input data-co-edit="${c.id}" data-field="quoted_value" type="number" step="0.01" value="${c.quoted_value || 0}"></td>
          <td><input data-co-edit="${c.id}" data-field="approved_value" type="number" step="0.01" value="${c.approved_value || 0}"></td>
          <td class="actions-cell"><button class="btn" data-save-co="${c.id}" type="button">Save</button><button class="btn" data-copy-co="${c.id}" onclick="event.preventDefault(); event.stopPropagation(); window.copyChangeOrderById('${c.id}')" type="button">Copy</button><button class="btn danger" data-delete-co="${c.id}" type="button">Delete</button></td>
        </tr>`).join('')}</tbody>`;
      tableEl.querySelectorAll('[data-co-edit]').forEach(el => {
        const markRowDirty = () => el.closest('tr')?.classList.add('setup-dirty');
        el.oninput = markRowDirty;
        el.onchange = markRowDirty;
      });
      document.querySelectorAll('[data-save-co]').forEach(btn => btn.onclick = async () => {
        const id = btn.dataset.saveCo;
        const fields = {};
        document.querySelectorAll(`[data-co-edit="${id}"]`).forEach(el => fields[el.dataset.field] = el.value);
        try {
          await api(`/api/change-orders/${id}`, { method:'PUT', body: JSON.stringify(fields) });
          btn.closest('tr')?.classList.remove('setup-dirty');
          markSaved();
          await refresh();
        } catch (err) {
          window.alert(err.message || 'Could not save this change order / child project.');
        }
      });
      document.querySelectorAll('[data-delete-co]').forEach(btn => btn.onclick = async () => {
        const id = btn.dataset.deleteCo;
        const co = state.changeOrders.find(c => String(c.id) === String(id));
        const label = co ? [co.co_number, co.job_number, co.title].filter(Boolean).join(' - ') : 'this change order';
        if (!window.confirm(`Delete change order ${label}?\n\nCost records and customer invoices tied to it will be kept, but they will no longer be assigned to this change order.`)) return;
        await api(`/api/change-orders/${id}`, { method:'DELETE' });
        if (String(selectedDashboardChangeOrderId || '') === String(id)) selectedDashboardChangeOrderId = null;
        markSaved();
        await refresh();
      });
    }

    function loadInternalRateEditor() {
      const tableEl = document.getElementById('rateEditTable');
      if (!tableEl) return;
      const selectedRateSetId = Number(document.getElementById('rateSetSelect')?.value || 0);
      const shownRates = selectedRateSetId ? state.internalRates.filter(r => r.rate_set_id === selectedRateSetId) : state.internalRates;
      tableEl.innerHTML = `
        <thead><tr><th>Type</th><th>Category</th><th>Raw Rate</th><th></th></tr></thead>
        <tbody>${shownRates.map(r => `<tr>
          <td><select data-rate="${r.id}" data-field="category_type"><option ${r.category_type === 'Labor' ? 'selected' : ''}>Labor</option><option ${r.category_type === 'Equipment' ? 'selected' : ''}>Equipment</option></select></td>
          <td><input data-rate="${r.id}" data-field="category" value="${r.category || ''}"></td>
          <td><input data-rate="${r.id}" data-field="raw_rate" type="number" step="0.01" value="${r.raw_rate || 0}"></td>
          <td><button class="btn" data-save-rate="${r.id}">Save</button></td>
        </tr>`).join('')}</tbody>`;
      document.querySelectorAll('[data-save-rate]').forEach(btn => btn.onclick = async () => {
        const id = btn.dataset.saveRate;
        const fields = {};
        document.querySelectorAll(`[data-rate="${id}"]`).forEach(el => fields[el.dataset.field] = el.value);
        await api(`/api/internal-rates/${id}`, { method:'PUT', body: JSON.stringify(fields) });
        markSaved();
        await refresh();
      });
      const rateSetSelect = document.getElementById('rateSetSelect');
      if (rateSetSelect) rateSetSelect.onchange = loadInternalRateEditor;
    }

    function loadCogCategoryEditor() {
      const tableEl = document.getElementById('cogCategoryTable');
      if (!tableEl) return;
      tableEl.innerHTML = state.cogCategories.length
        ? `<thead><tr><th>Name</th><th>Internal COG Code</th><th>Description</th><th></th></tr></thead><tbody>${state.cogCategories.map(c => `<tr>
            <td><input data-cog="${c.id}" data-field="name" value="${htmlEscape(c.name || '')}"></td>
            <td><input data-cog="${c.id}" data-field="code" value="${htmlEscape(c.code || '')}"></td>
            <td><input data-cog="${c.id}" data-field="description" value="${htmlEscape(c.description || '')}"></td>
            <td class="actions-cell"><button class="btn" data-save-cog="${c.id}" type="button">Save</button><button class="btn danger" data-delete-cog="${c.id}" type="button">Delete</button></td>
          </tr>`).join('')}</tbody>`
        : '<tbody><tr><td>No COG categories yet.</td></tr></tbody>';
      tableEl.querySelectorAll('[data-cog]').forEach(el => {
        const markRowDirty = () => el.closest('tr')?.classList.add('setup-dirty');
        el.oninput = markRowDirty;
        el.onchange = markRowDirty;
      });
      tableEl.querySelectorAll('[data-save-cog]').forEach(btn => btn.onclick = async () => {
        const id = btn.dataset.saveCog;
        const fields = {};
        tableEl.querySelectorAll(`[data-cog="${id}"]`).forEach(el => fields[el.dataset.field] = el.value);
        await api(`/api/cog-categories/${id}`, { method:'PUT', body: JSON.stringify(fields) });
        btn.closest('tr')?.classList.remove('setup-dirty');
        markSaved();
        state.cogCategories = await api('/api/cog-categories');
        jobOrderReportRows = await api('/api/job-order-report');
        loadCogCategoryEditor();
        renderFieldPoJobChoices(document.getElementById('fieldPoJobSearch')?.value || '');
      });
      tableEl.querySelectorAll('[data-delete-cog]').forEach(btn => btn.onclick = async () => {
        const id = btn.dataset.deleteCog;
        const category = state.cogCategories.find(c => String(c.id) === String(id));
        const label = category ? [category.code, category.name].filter(Boolean).join(' - ') : 'this COG category';
        if (!window.confirm(`Delete COG category ${label}? Existing POs will keep their current label.`)) return;
        await api(`/api/cog-categories/${id}`, { method:'DELETE' });
        markSaved();
        state.cogCategories = await api('/api/cog-categories');
        jobOrderReportRows = await api('/api/job-order-report');
        loadCogCategoryEditor();
        renderFieldPoJobChoices(document.getElementById('fieldPoJobSearch')?.value || '');
      });
    }

    async function loadCosts() {
      if (!state.projectId) return;
      const costs = await api(`/api/cost-records?project_id=${state.projectId}`);
      const showAll = document.getElementById('showAllCosts')?.checked;
      const visibleCosts = showAll ? costs : costs.filter(c =>
        !c.subproject_id ||
        !c.cost_type ||
        c.cost_type === 'Uncoded' ||
        c.raw_cost_source === 'Missing project rate' ||
        c.raw_cost_source === 'Missing internal rate'
      );
      const spOpts = state.subprojects.map(s => `<option value="${s.id}">${s.job_number || ''} ${s.code}</option>`).join('');
      const coOpts = state.changeOrders.map(c => `<option value="${c.id}">${[c.co_number, c.job_number].filter(Boolean).join(' / ')}</option>`).join('');
      if (!visibleCosts.length) {
        document.getElementById('costTable').innerHTML = '<tbody><tr><td>No exceptions. Imported records are coded and flowing through automatically.</td></tr></tbody>';
        const showAllBox = document.getElementById('showAllCosts');
        if (showAllBox) showAllBox.onchange = loadCosts;
        return;
      }
      document.getElementById('costTable').innerHTML = `
        <thead><tr><th>Date</th><th>Source</th><th>Ticket / Invoice</th><th>PDF</th><th>Type</th><th>Item</th><th>Description</th><th>Qty</th><th>Unit Price</th><th>Sales</th><th>Raw Cost</th><th>Rate Source</th><th>Subproject</th><th>CO</th><th></th></tr></thead>
        <tbody>${groupedEditableCostRows(visibleCosts, spOpts, coOpts)}</tbody>`;
      wireEditableCostTable(document.getElementById('costTable'), costs, loadCosts);
      const showAllBox = document.getElementById('showAllCosts');
      if (showAllBox) showAllBox.onchange = loadCosts;
    }

    function wireEditableCostTable(scope, costs, reloadFn) {
      if (!scope) return;
      wireInvoiceToggles(scope);
      costs.forEach(c => {
        const sp = scope.querySelector(`select[data-id="${c.id}"][data-field="subproject_id"]`);
        const co = scope.querySelector(`select[data-id="${c.id}"][data-field="change_order_id"]`);
        if (sp) sp.value = c.subproject_id || '';
        if (co) updateChangeOrderSelectForSubproject(co, sp?.value || c.subproject_id || '', c.change_order_id || '');
        if (sp && co) {
          sp.onchange = () => updateChangeOrderSelectForSubproject(co, sp.value, co.value);
        }
      });
      scope.querySelectorAll('select[data-group-field="subproject_id"]').forEach(sp => {
        const groupId = sp.dataset.group;
        const co = scope.querySelector(`select[data-group="${groupId}"][data-group-field="change_order_id"]`);
        if (co) updateChangeOrderSelectForSubproject(co, sp.value, co.value);
        sp.onchange = event => {
          event.stopPropagation();
          updateChangeOrderSelectForSubproject(co, sp.value, co?.value || '');
        };
      });
      scope.querySelectorAll('[data-save]').forEach(btn => btn.onclick = async () => {
        const id = btn.dataset.save;
        const fields = {};
        scope.querySelectorAll(`select[data-id="${id}"]`).forEach(el => fields[el.dataset.field] = el.value);
        await api(`/api/cost-records/${id}`, { method:'PUT', body: JSON.stringify(fields) });
        markSaved();
        await refresh();
        if (reloadFn) await reloadFn();
      });
      scope.querySelectorAll('[data-save-cost-group]').forEach(btn => btn.onclick = async event => {
        event.stopPropagation();
        const groupId = btn.dataset.saveCostGroup;
        const fields = { ids: (btn.dataset.recordIds || '').split(',').filter(Boolean) };
        scope.querySelectorAll(`select[data-group="${groupId}"]`).forEach(el => fields[el.dataset.groupField] = el.value);
        await api('/api/cost-records/bulk-update', { method:'POST', body: JSON.stringify(fields) });
        markSaved();
        await refresh();
        if (reloadFn) await reloadFn();
      });
    }

    async function loadFieldTicketLines() {
      const tableEl = document.getElementById('fieldTicketLinesTable');
      const filterEl = document.getElementById('fieldTicketLineFilters');
      if (!tableEl || !state.projectId) return;
      const allCosts = await api(`/api/cost-records?project_id=${state.projectId}`);
      const tickets = allCosts.filter(c => c.source === 'Field Wise' || c.source === 'Field Wise PDF');
      if (!tickets.length) {
        if (filterEl) filterEl.innerHTML = '';
        tableEl.innerHTML = '<tbody><tr><td>No Field Wise ticket lines imported yet.</td></tr></tbody>';
        return;
      }
      const spOpts = state.subprojects.map(s => `<option value="${s.id}">${s.job_number || ''} ${s.code}</option>`).join('');
      const coOpts = state.changeOrders.map(c => `<option value="${c.id}">${[c.co_number, c.job_number].filter(Boolean).join(' / ')}</option>`).join('');
      const filterId = `fieldTicketFilter${++costFilterSeq}`;
      if (filterEl) filterEl.innerHTML = fieldTicketLineFilterMarkup(filterId, tickets);
      const container = filterEl?.querySelector(`[data-field-ticket-filter="${filterId}"]`);
      const searchEl = container?.querySelector(`[data-field-ticket-search="${filterId}"]`);
      const ticketEl = container?.querySelector(`[data-field-ticket-number="${filterId}"]`);
      const subprojectEl = container?.querySelector(`[data-field-ticket-subproject="${filterId}"]`);
      const typeEl = container?.querySelector(`[data-field-ticket-type="${filterId}"]`);
      const countEl = container?.querySelector(`[data-field-ticket-count="${filterId}"]`);
      const salesEl = container?.querySelector(`[data-field-ticket-sales="${filterId}"]`);
      const rawEl = container?.querySelector(`[data-field-ticket-raw="${filterId}"]`);
      const clearEl = container?.querySelector(`[data-field-ticket-clear="${filterId}"]`);
      const draw = () => {
        const query = String(searchEl?.value || '').trim().toLowerCase();
        const ticket = ticketEl?.value || '';
        const subproject = subprojectEl?.value || '';
        const type = typeEl?.value || '';
        const filtered = tickets.filter(r => {
          const subLabel = [r.job_number, r.subproject_code].filter(Boolean).join(' ');
          const haystack = [
            r.record_date, r.source, r.source_file, r.ticket_or_invoice,
            r.cost_type, r.item, r.description, r.raw_cost_source, subLabel
          ].join(' ').toLowerCase();
          return (!query || haystack.includes(query)) &&
            (!ticket || r.ticket_or_invoice === ticket) &&
            (!subproject || subLabel === subproject) &&
            (!type || r.cost_type === type);
        });
        const salesTotal = filtered.reduce((sum, r) => sum + Number(r.sales_amount || r.amount || 0), 0);
        const rawTotal = filtered.reduce((sum, r) => sum + Number(r.amount || 0), 0);
        if (countEl) countEl.textContent = `${filtered.length} of ${tickets.length}`;
        if (salesEl) salesEl.textContent = money(salesTotal);
        if (rawEl) rawEl.textContent = money(rawTotal);
        tableEl.innerHTML = filtered.length
          ? `<thead><tr><th>Date</th><th>Source</th><th>Ticket</th><th>PDF</th><th>Type</th><th>Item</th><th>Description</th><th>Qty</th><th>Unit Price</th><th>Sales</th><th>Raw Cost</th><th>Rate Source</th><th>Subproject</th><th>CO</th><th></th></tr></thead>
             <tbody>${groupedEditableCostRows(filtered, spOpts, coOpts)}</tbody>`
          : '<tbody><tr><td>No Field Wise ticket lines match these filters.</td></tr></tbody>';
        wireEditableCostTable(tableEl, filtered, loadFieldTicketLines);
      };
      [searchEl, ticketEl, subprojectEl, typeEl].filter(Boolean).forEach(el => {
        el.oninput = draw;
        el.onchange = draw;
      });
      if (clearEl) clearEl.onclick = () => {
        if (searchEl) searchEl.value = '';
        if (ticketEl) ticketEl.value = '';
        if (subprojectEl) subprojectEl.value = '';
        if (typeEl) typeEl.value = '';
        draw();
      };
      draw();
    }

    async function loadImportHistory() {
      const tableEl = document.getElementById('importHistoryTable');
      const filterEl = document.getElementById('importHistoryFilters');
      const countEl = document.getElementById('importHistoryCount');
      if (!tableEl || !state.projectId) return;
      const imports = await api(`/api/imports?project_id=${state.projectId}`);
      const sources = [...new Set(imports.map(row => row.source || '').filter(Boolean))].sort();
      if (filterEl) {
        filterEl.innerHTML = `
          <div><label>Search imports</label><input id="importHistorySearch" placeholder="Search file, source, date, amount, records"></div>
          <div><label>Source</label><select id="importHistorySource"><option value="">All sources</option>${sources.map(source => `<option>${htmlEscape(source)}</option>`).join('')}</select></div>
          <div style="align-self:end"><button class="btn" id="clearImportHistoryFilters" type="button">Clear</button></div>`;
      }
      const searchEl = document.getElementById('importHistorySearch');
      const sourceEl = document.getElementById('importHistorySource');
      const draw = () => {
        const query = String(searchEl?.value || '').trim().toLowerCase();
        const sourceFilter = String(sourceEl?.value || '');
        const filtered = imports.filter(row => {
          if (sourceFilter && row.source !== sourceFilter) return false;
          const haystack = [
            row.source,
            row.source_file,
            row.record_count,
            money(row.sales_amount),
            money(row.raw_amount),
            row.last_imported
          ].join(' ').toLowerCase();
          return !query || haystack.includes(query);
        });
        if (countEl) countEl.textContent = `${filtered.length} of ${imports.length} imported file group(s) shown`;
        tableEl.innerHTML = `
          <thead><tr><th>Source</th><th>File</th><th>Records</th><th>Sales</th><th>Raw Cost</th><th>Last Imported</th><th></th></tr></thead>
          <tbody>${filtered.length ? filtered.map(row => `<tr>
            <td>${htmlEscape(row.source || '')}</td>
            <td>${htmlEscape(row.source_file || '')}</td>
            <td>${row.record_count}</td>
            <td>${money(row.sales_amount)}</td>
            <td>${money(row.raw_amount)}</td>
            <td>${htmlEscape(row.last_imported || '')}</td>
            <td><button class="btn" data-delete-import="${htmlEscape(row.source_file || '')}" data-source="${htmlEscape(row.source || '')}">Remove Import</button></td>
          </tr>`).join('') : '<tr><td colspan="7">No imported files match the current filters.</td></tr>'}</tbody>`;
        document.querySelectorAll('[data-delete-import]').forEach(btn => btn.onclick = async () => {
          const file = btn.dataset.deleteImport;
          const source = btn.dataset.source;
          if (!window.confirm(`Remove imported records from ${file}?`)) return;
          await api('/api/imports/delete', {
            method: 'POST',
            body: JSON.stringify({ project_id: state.projectId, source_file: file, source })
          });
          await refresh();
        });
      };
      if (searchEl) searchEl.oninput = draw;
      if (sourceEl) sourceEl.onchange = draw;
      const clearBtn = document.getElementById('clearImportHistoryFilters');
      if (clearBtn) clearBtn.onclick = () => {
        if (searchEl) searchEl.value = '';
        if (sourceEl) sourceEl.value = '';
        draw();
      };
      draw();
    }

    async function loadVendorInvoiceLines() {
      const tableEl = document.getElementById('vendorInvoiceLinesTable');
      const filterEl = document.getElementById('vendorInvoiceLineFilters');
      if (!tableEl || !state.projectId) return;
      const lines = await api(`/api/vendor-invoice-lines?project_id=${state.projectId}`);
      if (!lines.length) {
        if (filterEl) filterEl.innerHTML = '';
        tableEl.innerHTML = '<tbody><tr><td>No vendor invoice lines imported yet.</td></tr></tbody>';
        return;
      }
      const filterId = `vendorFilter${++costFilterSeq}`;
      if (filterEl) filterEl.innerHTML = vendorInvoiceLineFilterMarkup(filterId, lines);
      const container = filterEl?.querySelector(`[data-vendor-filter="${filterId}"]`);
      const searchEl = container?.querySelector(`[data-vendor-search="${filterId}"]`);
      const vendorEl = container?.querySelector(`[data-vendor-name="${filterId}"]`);
      const subprojectEl = container?.querySelector(`[data-vendor-subproject="${filterId}"]`);
      const dateEl = container?.querySelector(`[data-vendor-date="${filterId}"]`);
      const countEl = container?.querySelector(`[data-vendor-count="${filterId}"]`);
      const amountEl = container?.querySelector(`[data-vendor-amount="${filterId}"]`);
      const clearEl = container?.querySelector(`[data-vendor-clear="${filterId}"]`);
      const draw = () => {
        const query = String(searchEl?.value || '').trim().toLowerCase();
        const vendor = vendorEl?.value || '';
        const subproject = subprojectEl?.value || '';
        const date = dateEl?.value || '';
        const filtered = lines.filter(r => {
          const subLabel = [r.job_number, r.subproject_code].filter(Boolean).join(' ');
          const haystack = [
            r.ticket_or_invoice, r.source_file, r.vendor, r.record_date, subLabel,
            r.item, r.description
          ].join(' ').toLowerCase();
          return (!query || haystack.includes(query)) &&
            (!vendor || r.vendor === vendor) &&
            (!subproject || subLabel === subproject) &&
            (!date || invoiceDateInputValue(r.record_date) === date);
        });
        const filteredAmount = filtered.reduce((sum, r) => sum + Number(r.amount || 0), 0);
        if (countEl) countEl.textContent = `${filtered.length} of ${lines.length}`;
        if (amountEl) amountEl.textContent = money(filteredAmount);
        tableEl.innerHTML = vendorInvoiceLinesTableHtml(filtered);
        wireInvoiceToggles(tableEl);
        wireInvoiceSubprojectSaves(tableEl);
      };
      [searchEl, vendorEl, subprojectEl, dateEl].filter(Boolean).forEach(el => {
        el.oninput = draw;
        el.onchange = draw;
      });
      if (clearEl) clearEl.onclick = () => {
        if (searchEl) searchEl.value = '';
        if (vendorEl) vendorEl.value = '';
        if (subprojectEl) subprojectEl.value = '';
        if (dateEl) dateEl.value = '';
        draw();
      };
      draw();
    }

    async function loadVendorAllocationHistory() {
      const tableEl = document.getElementById('vendorAllocationHistoryTable');
      if (!tableEl || !state.projectId) return;
      const allocations = await api(`/api/vendor-invoice-allocations?project_id=${state.projectId}`);
      if (!allocations.length) {
        tableEl.innerHTML = '<tbody><tr><td>No vendor invoice allocations have been recorded yet.</td></tr></tbody>';
        return;
      }
      tableEl.innerHTML = `<thead><tr><th>Allocated</th><th>Invoice</th><th>Vendor</th><th>Original Total</th><th>Split By</th><th>Targets</th></tr></thead>
        <tbody>${allocations.map(a => `<tr>
          <td>${htmlEscape(a.allocated_at || '')}</td>
          <td>${htmlEscape(a.ticket_or_invoice || '')}<div class="muted">${htmlEscape(a.source_file || '')}</div></td>
          <td>${htmlEscape(a.vendor || '')}</td>
          <td>${money(a.original_total)}</td>
          <td>${htmlEscape(a.allocated_by_username || '')}</td>
          <td>${(a.lines || []).map(line => `<div style="display:flex;justify-content:space-between;gap:12px;margin:3px 0"><span>${htmlEscape(line.target_label || '')}</span><strong>${money(line.amount)}</strong></div>`).join('')}</td>
        </tr>`).join('')}</tbody>`;
    }

    async function loadCustomerInvoices() {
      const tableEl = document.getElementById('customerInvoiceTable');
      const summaryEl = document.getElementById('customerInvoiceSummary');
      if (!tableEl || !state.projectId) return;
      const invoices = await api(`/api/customer-invoices?project_id=${state.projectId}`);
      const active = invoices.filter(i => i.status !== 'Void');
      const billed = active.reduce((sum, i) => sum + Number(i.amount || 0), 0);
      const paid = active.reduce((sum, i) => sum + Number(i.paid_amount || 0), 0);
      const open = active
        .filter(i => !['Draft','Paid'].includes(i.status || ''))
        .reduce((sum, i) => sum + Math.max(0, Number(i.amount || 0) - Number(i.paid_amount || 0)), 0);
      if (summaryEl) {
        summaryEl.innerHTML = table(['Invoices','Billed','Paid','Open AR'], [[active.length, money(billed), money(paid), money(open)]]);
      }
      if (!invoices.length) {
        tableEl.innerHTML = '<tbody><tr><td>No customer invoices entered yet.</td></tr></tbody>';
        return;
      }
      const spOpts = '<option value="">Unassigned</option>' + state.subprojects.map(s => `<option value="${s.id}">${s.job_number || ''} ${s.code}</option>`).join('');
      const coOpts = '<option value="">Base Contract</option>' + state.changeOrders.map(c => `<option value="${c.id}">${[c.co_number, c.job_number].filter(Boolean).join(' / ')}</option>`).join('');
      tableEl.innerHTML = `
        <thead><tr><th>Invoice #</th><th>File</th><th>Type</th><th>Subproject</th><th>CO</th><th>Allocation</th><th>Invoice Date</th><th>Due Date</th><th>Status</th><th>Amount</th><th>Paid</th><th>Open</th><th>Notes</th><th></th></tr></thead>
        <tbody>${invoices.map(i => {
          const openAmount = Math.max(0, Number(i.amount || 0) - Number(i.paid_amount || 0));
          const invoiceFile = i.invoice_file ? pdfLink({ source_file: i.invoice_file }) : '<span class="muted">Missing</span>';
          const allocations = i.allocations || [];
          const allocationSummary = allocations.length
            ? `<div class="muted">${allocations.length} allocation${allocations.length === 1 ? '' : 's'}</div>${allocations.map(a => `<div>${htmlEscape(a.target_label || 'Unassigned')}: <strong>${money(a.amount)}</strong></div>`).join('')}`
            : '<span class="muted">No allocations</span>';
          return `<tr>
            <td><input data-cinv="${i.id}" data-field="invoice_number" value="${htmlEscape(i.invoice_number || '')}"></td>
            <td>${invoiceFile}</td>
            <td><select data-cinv="${i.id}" data-field="billing_type">${['Progress','Base Contract','Change Order','T&M','Retainage','Final'].map(v => `<option ${i.billing_type === v ? 'selected' : ''}>${v}</option>`).join('')}</select></td>
            <td><select data-cinv="${i.id}" data-field="subproject_id">${spOpts}</select></td>
            <td><select data-cinv="${i.id}" data-field="change_order_id">${coOpts}</select></td>
            <td>${allocationSummary}</td>
            <td><input data-cinv="${i.id}" data-field="invoice_date" type="date" value="${htmlEscape(i.invoice_date || '')}"></td>
            <td><input data-cinv="${i.id}" data-field="due_date" type="date" value="${htmlEscape(i.due_date || '')}"></td>
            <td><select data-cinv="${i.id}" data-field="status">${['Draft','Sent','Partial','Paid','Overdue','Void'].map(v => `<option ${i.status === v ? 'selected' : ''}>${v}</option>`).join('')}</select></td>
            <td><input data-cinv="${i.id}" data-field="amount" type="number" step="0.01" value="${Number(i.amount || 0).toFixed(2)}"></td>
            <td><input data-cinv="${i.id}" data-field="paid_amount" type="number" step="0.01" value="${Number(i.paid_amount || 0).toFixed(2)}"></td>
            <td>${money(openAmount)}</td>
            <td><input data-cinv="${i.id}" data-field="notes" value="${htmlEscape(i.notes || '')}"></td>
            <td><button class="btn" data-save-customer-invoice="${i.id}" type="button">Save</button></td>
          </tr>`;
        }).join('')}</tbody>`;
      invoices.forEach(i => {
        const sp = tableEl.querySelector(`select[data-cinv="${i.id}"][data-field="subproject_id"]`);
        const co = tableEl.querySelector(`select[data-cinv="${i.id}"][data-field="change_order_id"]`);
        if (sp) sp.value = i.subproject_id || '';
        if (co) co.value = i.change_order_id || '';
      });
      tableEl.querySelectorAll('[data-save-customer-invoice]').forEach(btn => btn.onclick = async () => {
        const id = btn.dataset.saveCustomerInvoice;
        const fields = {};
        tableEl.querySelectorAll(`[data-cinv="${id}"]`).forEach(el => fields[el.dataset.field] = el.value);
        await api(`/api/customer-invoices/${id}`, { method:'PUT', body: JSON.stringify(fields) });
        markSaved();
        await refresh();
      });
    }

    async function loadBidDashboard() {
      const summary = await api('/api/bid-summary');
      document.getElementById('bidKpis').innerHTML = `
        <div class="panel kpi"><div class="label">Open Pipeline</div><div class="value">${money(summary.open_pipeline)}</div><div class="hint">${summary.open_count} open RFQ(s)</div></div>
        <div class="panel kpi"><div class="label">Weighted Forecast</div><div class="value">${money(summary.weighted_forecast)}</div></div>
        <div class="panel kpi"><div class="label">Win Rate</div><div class="value">${pct(summary.win_rate)}</div></div>
        <div class="panel kpi"><div class="label">Avg Target Margin</div><div class="value">${pct(summary.avg_target_margin)}</div></div>`;
      document.getElementById('bidStageSummary').innerHTML = table(['Stage','Count','Value'], summary.stage.map(x => [x.stage, x.count, money(x.value)]));
      document.getElementById('bidEstimatorSummary').innerHTML = table(['Estimator','Open RFQs','Open Value'], summary.estimator.map(x => [x.estimator, x.open_rfqs, money(x.open_value)]));
      const filterId = `bidFilter${++costFilterSeq}`;
      const filterEl = document.getElementById('bidFilters');
      const tableEl = document.getElementById('bidTable');
      filterEl.innerHTML = bidFilterMarkup(filterId, summary.bids);
      const container = filterEl.querySelector(`[data-bid-filter="${filterId}"]`);
      const searchEl = container.querySelector(`[data-bid-search="${filterId}"]`);
      const customerEl = container.querySelector(`[data-bid-customer="${filterId}"]`);
      const stageEl = container.querySelector(`[data-bid-stage="${filterId}"]`);
      const estimatorEl = container.querySelector(`[data-bid-estimator="${filterId}"]`);
      const countEl = container.querySelector(`[data-bid-count="${filterId}"]`);
      const totalEl = container.querySelector(`[data-bid-total="${filterId}"]`);
      const weightedEl = container.querySelector(`[data-bid-weighted="${filterId}"]`);
      const clearEl = container.querySelector(`[data-bid-clear="${filterId}"]`);
      const draw = () => {
        const query = String(searchEl.value || '').trim().toLowerCase();
        const customer = customerEl.value || '';
        const stage = stageEl.value || '';
        const estimator = estimatorEl.value || '';
        const filtered = summary.bids.filter(b => {
          const haystack = [b.rfq_no, b.customer, b.project_name, b.estimator, b.stage, b.go_no_go, b.outcome, b.notes].join(' ').toLowerCase();
          return (!query || haystack.includes(query)) &&
            (!customer || b.customer === customer) &&
            (!stage || b.stage === stage) &&
            (!estimator || b.estimator === estimator);
        });
        countEl.textContent = `${filtered.length} of ${summary.bids.length}`;
        totalEl.textContent = money(filtered.reduce((sum, b) => sum + Number(b.bid_price || 0), 0));
        weightedEl.textContent = money(filtered.reduce((sum, b) => sum + Number(b.weighted_value || 0), 0));
        tableEl.innerHTML = bidTableHtml(filtered);
        wireBidSaves(tableEl);
      };
      [searchEl, customerEl, stageEl, estimatorEl].forEach(el => {
        el.oninput = draw;
        el.onchange = draw;
      });
      clearEl.onclick = () => {
        searchEl.value = '';
        customerEl.value = '';
        stageEl.value = '';
        estimatorEl.value = '';
        draw();
      };
      draw();
    }

    function wireBidSaves(scope=document) {
      scope.querySelectorAll('[data-bid]').forEach(el => {
        el.oninput = () => el.closest('tr')?.classList.add('bid-dirty');
        el.onchange = () => el.closest('tr')?.classList.add('bid-dirty');
      });
      scope.querySelectorAll('[data-save-bid]').forEach(btn => {
        btn.onclick = async () => {
          const id = btn.dataset.saveBid;
          const fields = {};
          scope.querySelectorAll(`[data-bid="${id}"]`).forEach(el => fields[el.dataset.field] = el.value);
          await api(`/api/bids/${id}`, { method:'PUT', body: JSON.stringify(fields) });
          btn.closest('tr')?.classList.remove('bid-dirty');
          markSaved();
          await loadBidDashboard();
        };
      });
    }

    async function loadCurrentUser() {
      const me = await api('/api/me');
      state.currentUser = me;
      const hasAnyEdit = Object.values(me.permissions || {}).some(permission => Number(permission.can_edit || 0));
      const readOnly = !hasAnyEdit;
      document.body.classList.toggle('read-only', readOnly);
      document.body.classList.toggle('tx-read-only', me.role === 'TX/Read Only');
      document.body.classList.toggle('field-po-only', me.role === 'Field PO');
      document.getElementById('currentUser').textContent = `${me.display_name || me.username}${readOnly || me.role === 'Field PO' ? ' / ' + me.role : ''}`;
      applyPermissionVisibility();
      if (Number(me.must_change_password || 0)) openAccountModal(true);
    }

    function customerInvoiceTargetOptions(selectedKey = '') {
      const subprojectOptions = state.subprojects.map(s => {
        const key = `sp:${s.id}`;
        return `<option value="${key}" ${selectedKey === key ? 'selected' : ''}>${htmlEscape([s.job_number, s.code].filter(Boolean).join(' '))} - ${htmlEscape(s.name || '')}</option>`;
      }).join('');
      const changeOrderOptions = state.changeOrders.map(c => {
        const key = `co:${c.id}`;
        const type = c.order_type || 'Change Order';
        const label = [[c.co_number, c.job_number].filter(Boolean).join(' / '), c.title || type].filter(Boolean).join(' - ');
        return `<option value="${key}" ${selectedKey === key ? 'selected' : ''}>${htmlEscape(label)}</option>`;
      }).join('');
      return `<option value="">Choose allocation target...</option>${subprojectOptions}${changeOrderOptions}`;
    }

    function addCustomerInvoiceAllocationLine(values = {}) {
      const target = document.getElementById('customerInvoiceAllocations');
      if (!target) return;
      const row = document.createElement('div');
      row.className = 'grid cols-4 customer-invoice-allocation-row';
      row.dataset.customerInvoiceAllocationRow = '1';
      row.style.marginTop = '8px';
      row.innerHTML = `
        <div style="grid-column:span 2"><label>Job / Order</label><select data-allocation-field="target_key">${customerInvoiceTargetOptions(values.target_key || '')}</select></div>
        <div><label>Amount</label><input data-allocation-field="amount" type="number" step="0.01" value="${Number(values.amount || 0).toFixed(2)}"></div>
        <div><label>Note</label><input data-allocation-field="notes" value="${htmlEscape(values.notes || '')}" placeholder="Optional"></div>
        <div class="actions" style="grid-column:1 / -1"><button class="btn danger" data-remove-allocation-line type="button">Remove Line</button></div>`;
      row.querySelector('[data-remove-allocation-line]').onclick = () => {
        row.remove();
        updateCustomerInvoiceAllocationTotal();
      };
      row.querySelectorAll('input, select').forEach(el => el.oninput = updateCustomerInvoiceAllocationTotal);
      target.appendChild(row);
      updateCustomerInvoiceAllocationTotal();
    }

    function ensureCustomerInvoiceAllocationLine() {
      const target = document.getElementById('customerInvoiceAllocations');
      if (!target || target.querySelector('[data-customer-invoice-allocation-row]')) return;
      const sp = document.getElementById('billingSubproject')?.value || '';
      const co = document.getElementById('billingCo')?.value || '';
      const amount = document.querySelector('#customerInvoiceForm input[name="amount"]')?.value || 0;
      addCustomerInvoiceAllocationLine({ target_key: co ? `co:${co}` : sp ? `sp:${sp}` : '', amount });
    }

    function updateCustomerInvoiceAllocationTotal() {
      const totalEl = document.getElementById('customerInvoiceAllocationTotal');
      if (!totalEl) return;
      const total = [...document.querySelectorAll('[data-customer-invoice-allocation-row] [data-allocation-field="amount"]')]
        .reduce((sum, input) => sum + Number(input.value || 0), 0);
      const invoiceAmount = Number(document.querySelector('#customerInvoiceForm input[name="amount"]')?.value || 0);
      const diff = total - invoiceAmount;
      totalEl.innerHTML = `Allocated ${money(total)} of ${money(invoiceAmount)}${Math.abs(diff) >= 0.01 ? ` <span class="${diff > 0 ? 'bad' : 'warn'}">(${money(Math.abs(diff))} ${diff > 0 ? 'over' : 'remaining'})</span>` : ' <span class="good">(balanced)</span>'}`;
    }

    function collectCustomerInvoiceAllocations() {
      return [...document.querySelectorAll('[data-customer-invoice-allocation-row]')].map(row => {
        const targetKey = row.querySelector('[data-allocation-field="target_key"]')?.value || '';
        const amount = Number(row.querySelector('[data-allocation-field="amount"]')?.value || 0);
        const notes = row.querySelector('[data-allocation-field="notes"]')?.value || '';
        return { target_key: targetKey, amount, notes };
      }).filter(row => row.target_key && Math.abs(row.amount) >= 0.01);
    }

    async function loadUsers() {
      if (state.currentUser?.role !== 'Admin') return;
      if (!rolePermissionsData) rolePermissionsData = await api('/api/role-permissions');
      const users = await api('/api/users');
      const roleOptions = (rolePermissionsData?.roles || ['User','Read Only','TX/Read Only','Field PO','Admin']);
      document.getElementById('usersTable').innerHTML = `
        <thead><tr><th>Username</th><th>Name</th><th>Role</th><th>PO Trust</th><th>Status</th><th>Password</th><th></th></tr></thead>
        <tbody>${users.map(u => `<tr>
          <td>${u.username}</td>
          <td>${u.display_name || ''}</td>
          <td><select data-user-role="${u.id}">${roleOptions.map(role => `<option ${u.role === role ? 'selected' : ''}>${htmlEscape(role)}</option>`).join('')}</select></td>
          <td><label><input data-user-po-auto-issue="${u.id}" type="checkbox" ${Number(u.po_auto_issue || 0) ? 'checked' : ''}> Auto-issue</label></td>
          <td>${u.active ? 'Active' : 'Inactive'}${Number(u.must_change_password || 0) ? '<div class="muted">Must change password</div>' : ''}</td>
          <td><input data-user-password="${u.id}" type="password" placeholder="TPE1776"></td>
          <td>
            <button class="btn" data-reset-user="${u.id}" type="button">Reset Password</button>
            <button class="btn" data-toggle-user="${u.id}" data-active="${u.active}" type="button">${u.active ? 'Deactivate' : 'Activate'}</button>
          </td>
        </tr>`).join('')}</tbody>`;
      document.querySelectorAll('[data-reset-user]').forEach(btn => btn.onclick = async () => {
        const id = btn.dataset.resetUser;
        const password = document.querySelector(`[data-user-password="${id}"]`).value || 'TPE1776';
        await api(`/api/users/${id}`, { method:'PUT', body: JSON.stringify({ password }) });
        await loadUsers();
      });
      document.querySelectorAll('[data-toggle-user]').forEach(btn => btn.onclick = async () => {
        const id = btn.dataset.toggleUser;
        await api(`/api/users/${id}`, { method:'PUT', body: JSON.stringify({ active: btn.dataset.active === '1' ? 0 : 1 }) });
        await loadUsers();
      });
      document.querySelectorAll('[data-user-role]').forEach(sel => sel.onchange = async () => {
        await api(`/api/users/${sel.dataset.userRole}`, { method:'PUT', body: JSON.stringify({ role: sel.value }) });
        await loadUsers();
      });
      document.querySelectorAll('[data-user-po-auto-issue]').forEach(box => box.onchange = async () => {
        await api(`/api/users/${box.dataset.userPoAutoIssue}`, { method:'PUT', body: JSON.stringify({ po_auto_issue: box.checked ? 1 : 0 }) });
        await loadUsers();
      });
      renderRolePermissions();
    }

    function renderRolePermissions() {
      const wrap = document.getElementById('rolePermissions');
      if (!wrap || !rolePermissionsData) return;
      const permissions = rolePermissionsData.permissions || [];
      const matrix = rolePermissionsData.matrix || {};
      wrap.innerHTML = (rolePermissionsData.roles || []).map(role => {
        const locked = role === 'Admin';
        const rows = permissions.map(permission => {
          const value = matrix?.[role]?.[permission.key] || {};
          return `<tr>
            <td><strong>${htmlEscape(permission.label)}</strong><div class="muted">${htmlEscape(permission.group)}</div></td>
            <td><input type="checkbox" data-role-permission="${htmlEscape(role)}" data-permission-key="${htmlEscape(permission.key)}" data-permission-field="can_view" ${Number(value.can_view || 0) ? 'checked' : ''} ${locked ? 'disabled' : ''}></td>
            <td><input type="checkbox" data-role-permission="${htmlEscape(role)}" data-permission-key="${htmlEscape(permission.key)}" data-permission-field="can_edit" ${Number(value.can_edit || 0) ? 'checked' : ''} ${locked ? 'disabled' : ''}></td>
          </tr>`;
        }).join('');
        return `<div class="permission-role">
          <div class="permission-role-header">
            <span>${htmlEscape(role)}</span>
            ${locked ? '<span class="muted">Locked full access</span>' : `<button class="btn primary" data-save-role-permissions="${htmlEscape(role)}" type="button">Save Role</button>`}
          </div>
          <div class="table-wrap"><table><thead><tr><th>Area</th><th>View</th><th>Edit</th></tr></thead><tbody>${rows}</tbody></table></div>
        </div>`;
      }).join('');
      wrap.querySelectorAll('[data-permission-field="can_view"]').forEach(box => {
        box.onchange = () => {
          const editBox = wrap.querySelector(`[data-role-permission="${CSS.escape(box.dataset.rolePermission)}"][data-permission-key="${CSS.escape(box.dataset.permissionKey)}"][data-permission-field="can_edit"]`);
          if (!box.checked && editBox) editBox.checked = false;
        };
      });
      wrap.querySelectorAll('[data-permission-field="can_edit"]').forEach(box => {
        box.onchange = () => {
          const viewBox = wrap.querySelector(`[data-role-permission="${CSS.escape(box.dataset.rolePermission)}"][data-permission-key="${CSS.escape(box.dataset.permissionKey)}"][data-permission-field="can_view"]`);
          if (box.checked && viewBox) viewBox.checked = true;
        };
      });
      wrap.querySelectorAll('[data-save-role-permissions]').forEach(btn => {
        btn.onclick = async () => {
          const role = btn.dataset.saveRolePermissions;
          const payload = {};
          permissions.forEach(permission => {
            const viewBox = wrap.querySelector(`[data-role-permission="${CSS.escape(role)}"][data-permission-key="${CSS.escape(permission.key)}"][data-permission-field="can_view"]`);
            const editBox = wrap.querySelector(`[data-role-permission="${CSS.escape(role)}"][data-permission-key="${CSS.escape(permission.key)}"][data-permission-field="can_edit"]`);
            payload[permission.key] = { can_view: !!viewBox?.checked, can_edit: !!editBox?.checked };
          });
          await api('/api/role-permissions', { method:'POST', body: JSON.stringify({ role, permissions: payload }) });
          rolePermissionsData = await api('/api/role-permissions');
          renderRolePermissions();
        };
      });
    }

    async function loadRolePermissions() {
      if (state.currentUser?.role !== 'Admin') return;
      rolePermissionsData = await api('/api/role-permissions');
      renderRolePermissions();
    }

    let activityRows = [];
    function renderActivityLog() {
      const search = String(document.getElementById('activitySearch')?.value || '').trim().toLowerCase();
      const action = document.getElementById('activityAction')?.value || '';
      const filtered = activityRows.filter(row => {
        if (action && row.action !== action) return false;
        if (!search) return true;
        return [
          row.actor_username,
          row.action,
          row.entity_type,
          row.entity_label,
          row.details,
          row.created_at
        ].some(value => String(value || '').toLowerCase().includes(search));
      });
      document.getElementById('activityCount').textContent = `${filtered.length} of ${activityRows.length} activity record(s) shown`;
      document.getElementById('activityTable').innerHTML = filtered.length ? `
        <thead><tr><th>When</th><th>User</th><th>Action</th><th>Target</th><th>Details</th></tr></thead>
        <tbody>${filtered.map(row => `<tr>
          <td>${htmlEscape(row.created_at || '')}</td>
          <td>${htmlEscape(row.actor_username || 'system')}</td>
          <td><strong>${htmlEscape(row.action || '')}</strong></td>
          <td>${htmlEscape([row.entity_type, row.entity_label].filter(Boolean).join(' / '))}</td>
          <td>${htmlEscape(row.details || '')}</td>
        </tr>`).join('')}</tbody>`
        : '<tbody><tr><td>No activity matches the current filters.</td></tr></tbody>';
    }

    async function loadActivityLog() {
      if (state.currentUser?.role !== 'Admin') return;
      const limit = document.getElementById('activityLimit')?.value || '100';
      activityRows = await api(`/api/activity?limit=${encodeURIComponent(limit)}`);
      const actionEl = document.getElementById('activityAction');
      const currentAction = actionEl.value || '';
      const actions = [...new Set(activityRows.map(row => row.action).filter(Boolean))].sort();
      actionEl.innerHTML = '<option value="">All actions</option>' + actions.map(item => `<option ${item === currentAction ? 'selected' : ''}>${htmlEscape(item)}</option>`).join('');
      renderActivityLog();
    }

    document.getElementById('projectForm').onsubmit = async e => {
      e.preventDefault();
      const data = formDataObj(e.target);
      if (!state.projectId) return;
      await api(`/api/projects/${state.projectId}`, { method:'PUT', body: JSON.stringify(data) });
      markSaved();
      await loadProjects();
    };
    document.getElementById('newProjectForm').onsubmit = async e => {
      e.preventDefault();
      const saved = await api('/api/projects', { method:'POST', body: JSON.stringify(formDataObj(e.target)) });
      state.projectId = saved.id;
      localStorage.setItem('selectedProjectId', state.projectId);
      e.target.reset();
      markSaved();
      await loadProjects();
      openTab('setup');
    };
    document.getElementById('archiveProjectBtn').onclick = async () => {
      if (!state.projectId) return;
      const project = state.projects.find(p => p.id === state.projectId);
      const ok = window.confirm(`Close and archive ${project?.name || 'this project'}? It will move out of the active project list.`);
      if (!ok) return;
      await api(`/api/projects/${state.projectId}/archive`, { method:'POST', body: JSON.stringify({}) });
      localStorage.removeItem('selectedProjectId');
      state.projectId = null;
      markSaved();
      await loadProjects();
      openTab('archivedProjects');
    };
    document.getElementById('refreshArchivedProjects').onclick = () => loadArchivedProjects();
    document.getElementById('refreshJobOrderReport').onclick = () => loadJobOrderReport();
    document.getElementById('refreshFieldPos').onclick = () => loadFieldPos();
    document.getElementById('refreshActivity').onclick = () => loadActivityLog();
    document.getElementById('activitySearch').oninput = () => renderActivityLog();
    document.getElementById('activityAction').onchange = () => renderActivityLog();
    document.getElementById('activityLimit').onchange = () => loadActivityLog();
    document.getElementById('refreshOfficePos').onclick = () => loadOfficePos();
    document.getElementById('officePoSearch').oninput = () => renderOfficePos();
    document.getElementById('officePoStatusFilter').onchange = () => renderOfficePos();
    document.getElementById('refreshProjectPos').onclick = () => loadProjectPos();
    document.getElementById('projectPoSearch').oninput = () => renderProjectPos();
    document.getElementById('projectPoStatusFilter').onchange = () => renderProjectPos();
    document.getElementById('jobOrderSearch').oninput = () => renderJobOrderReport();
    ['jobOrderCustomerFilter','jobOrderProjectFilter','jobOrderTypeFilter','jobOrderStatusFilter'].forEach(id => {
      document.getElementById(id).onchange = () => renderJobOrderReport();
    });
    document.getElementById('clearJobOrderFilters').onclick = () => {
      ['jobOrderSearch','jobOrderCustomerFilter','jobOrderProjectFilter','jobOrderTypeFilter','jobOrderStatusFilter'].forEach(id => {
        const el = document.getElementById(id);
        if (el) el.value = '';
      });
      renderJobOrderReport();
    };
    document.getElementById('fieldPoJobSearch').oninput = event => {
      document.getElementById('fieldPoJobKey').value = '';
      updateFieldPoCustomerRequirement();
      renderFieldPoJobChoices(event.target.value);
      document.getElementById('fieldPoJobList').classList.remove('hidden');
    };
    document.getElementById('fieldPoJobSearch').onfocus = event => {
      renderFieldPoJobChoices(event.target.value);
      document.getElementById('fieldPoJobList').classList.remove('hidden');
    };
    document.addEventListener('click', event => {
      if (!event.target.closest('.combo-wrap')) {
        document.getElementById('fieldPoJobList')?.classList.add('hidden');
      }
    });
    document.getElementById('fieldPoForm').onsubmit = async e => {
      e.preventDefault();
      const resultEl = document.getElementById('fieldPoResult');
      if (!document.getElementById('fieldPoJobKey').value) {
        resultEl.innerHTML = '<span class="bad">Choose a matching Job / Order # / COG from the list.</span>';
        document.getElementById('fieldPoJobSearch').focus();
        document.getElementById('fieldPoJobList').classList.remove('hidden');
        return;
      }
      const selectedPoChoice = fieldPoJobChoices.find(item => item.key === document.getElementById('fieldPoJobKey').value);
      const selectedCogName = String(selectedPoChoice?.primary || '').trim().toLowerCase().replace(/\s+/g, ' ');
      const selectedRequiresCustomer = document.getElementById('fieldPoJobKey').value.startsWith('cog:') && !COG_CUSTOMER_OPTIONAL_NAMES.has(selectedCogName);
      if (selectedRequiresCustomer && !String(document.getElementById('fieldPoCustomer').value || '').trim()) {
        resultEl.innerHTML = '<span class="bad">Enter the customer for this shop / COG PO.</span>';
        document.getElementById('fieldPoCustomer').focus();
        return;
      }
      resultEl.textContent = 'Creating PO...';
      const form = e.target;
      const data = new FormData(form);
      try {
        const saved = await api('/api/purchase-orders', { method:'POST', body: data });
        resultEl.innerHTML = `<strong class="good">PO Created: ${htmlEscape(saved.po_number)}</strong><div class="muted">Status: ${htmlEscape(saved.status || '')}</div>`;
        form.reset();
        document.getElementById('fieldPoJobKey').value = '';
        updateFieldPoCustomerRequirement();
        await loadFieldPoJobs();
        await loadFieldPos();
      } catch (err) {
        resultEl.innerHTML = `<span class="bad">${htmlEscape(err.message || 'Could not create PO.')}</span>`;
      }
    };
    document.getElementById('subprojectForm').onsubmit = async e => {
      e.preventDefault();
      await api('/api/subprojects', { method:'POST', body: JSON.stringify({ ...formDataObj(e.target), project_id: state.projectId }) });
      e.target.reset(); markSaved(); await refresh();
    };
    document.getElementById('coForm').onsubmit = async e => {
      e.preventDefault();
      try {
        await api('/api/change-orders', { method:'POST', body: JSON.stringify({ ...formDataObj(e.target), project_id: state.projectId }) });
        e.target.reset(); markSaved(); await refresh();
        updateCoPricingFields();
      } catch (err) {
        window.alert(err.message || 'Could not add this change order / child project.');
      }
    };
    function updateCoPricingFields() {
      const pricing = document.getElementById('coPricingType')?.value || 'Fixed';
      const approved = document.getElementById('coApprovedValue');
      if (!approved) return;
      approved.disabled = pricing === 'T&M';
      approved.placeholder = pricing === 'T&M' ? 'Calculated from Field Wise tickets' : '';
      if (pricing === 'T&M') approved.value = '0';
    }
    document.getElementById('coPricingType').onchange = updateCoPricingFields;
    document.getElementById('invoiceForm').onsubmit = async e => {
      e.preventDefault();
      await api('/api/invoices', { method:'POST', body: JSON.stringify({ ...formDataObj(e.target), project_id: state.projectId }) });
      e.target.reset(); markSaved(); await refresh();
    };
    document.getElementById('addCustomerInvoiceAllocation').onclick = () => addCustomerInvoiceAllocationLine();
    document.getElementById('splitCustomerInvoiceEvenly').onclick = () => {
      const rows = [...document.querySelectorAll('[data-customer-invoice-allocation-row]')];
      const amount = Number(document.querySelector('#customerInvoiceForm input[name="amount"]')?.value || 0);
      if (!rows.length || !amount) {
        updateCustomerInvoiceAllocationTotal();
        return;
      }
      const share = Math.floor((amount / rows.length) * 100) / 100;
      let allocated = 0;
      rows.forEach((row, idx) => {
        const input = row.querySelector('[data-allocation-field="amount"]');
        const value = idx === rows.length - 1 ? Number((amount - allocated).toFixed(2)) : share;
        allocated += value;
        if (input) input.value = value.toFixed(2);
      });
      updateCustomerInvoiceAllocationTotal();
    };
    document.querySelector('#customerInvoiceForm input[name="amount"]').oninput = updateCustomerInvoiceAllocationTotal;
    document.getElementById('billingSubproject').onchange = () => {
      const first = document.querySelector('[data-customer-invoice-allocation-row] select[data-allocation-field="target_key"]');
      if (first && !first.value) first.value = document.getElementById('billingSubproject').value ? `sp:${document.getElementById('billingSubproject').value}` : '';
      updateCustomerInvoiceAllocationTotal();
    };
    document.getElementById('billingCo').onchange = () => {
      const first = document.querySelector('[data-customer-invoice-allocation-row] select[data-allocation-field="target_key"]');
      if (first && document.getElementById('billingCo').value) first.value = `co:${document.getElementById('billingCo').value}`;
      updateCustomerInvoiceAllocationTotal();
    };
    document.getElementById('customerInvoiceForm').onsubmit = async e => {
      e.preventDefault();
      const data = new FormData(e.target);
      data.append('project_id', state.projectId);
      const allocations = collectCustomerInvoiceAllocations();
      const invoiceAmount = Number(data.get('amount') || 0);
      const allocationTotal = allocations.reduce((sum, row) => sum + Number(row.amount || 0), 0);
      if (!allocations.length) {
        alert('Add at least one allocation line for this customer invoice.');
        return;
      }
      if (Math.abs(allocationTotal - invoiceAmount) >= 0.01) {
        alert(`Allocation total must equal the invoice amount. Allocated ${money(allocationTotal)} of ${money(invoiceAmount)}.`);
        return;
      }
      data.append('allocations', JSON.stringify(allocations));
      await api('/api/customer-invoices', { method:'POST', body: data });
      e.target.reset();
      document.getElementById('customerInvoiceAllocations').innerHTML = '';
      ensureCustomerInvoiceAllocationLine();
      markSaved();
      fillSelects();
      await refresh();
      openTab('billing');
    };
    document.getElementById('bidForm').onsubmit = async e => {
      e.preventDefault();
      await api('/api/bids', { method:'POST', body: JSON.stringify(formDataObj(e.target)) });
      e.target.reset();
      markSaved();
      await loadBidDashboard();
    };
    document.getElementById('userForm').onsubmit = async e => {
      e.preventDefault();
      await api('/api/users', { method:'POST', body: JSON.stringify(formDataObj(e.target)) });
      e.target.reset();
      markSaved();
      await loadUsers();
    };
    document.getElementById('importForm').onsubmit = async e => {
      e.preventDefault();
      const fileInput = e.target.querySelector('input[type="file"]');
      const resultEl = document.getElementById('importResult');
      const selectedFiles = Array.from(fileInput.files || []);
      if (!selectedFiles.length) {
        resultEl.textContent = 'Choose one or more Field Wise files first.';
        return;
      }
      resultEl.textContent = `Importing ${selectedFiles.length} Field Wise file(s)...`;
      const results = [];
      for (const file of selectedFiles) {
        const fd = new FormData();
        fd.append('file', file);
        fd.append('project_id', state.projectId);
        try {
          const data = await api('/api/import-fieldwise', { method:'POST', body: fd });
          results.push({ file: file.name, ok: true, data });
        } catch (err) {
          results.push({ file: file.name, ok: false, error: err.message });
        }
      }
      const importedTotal = results.reduce((sum, r) => sum + (r.ok ? Number(r.data.imported || 0) : 0), 0);
      const skippedTotal = results.reduce((sum, r) => sum + (r.ok ? Number(r.data.skipped || 0) : 0), 0);
      const failedTotal = results.filter(r => !r.ok).length;
      const detailLines = results.map(r => {
        if (!r.ok) return `${r.file}: failed - ${r.error}`;
        const d = r.data;
        const matched = d.matched_change_order_id ? `auto-coded to CO for job/order ${d.order_number}` : d.matched_subproject_id ? `auto-coded to order ${d.order_number}` : `needs review for order ${d.order_number || 'blank'}`;
        const skipped = d.skipped ? `, skipped ${d.skipped} duplicate line item(s)` : '';
        return `${r.file}: imported ${d.imported} record(s)${skipped}, ${matched}`;
      });
      resultEl.innerHTML = `Imported ${importedTotal} total Field Wise record(s) from ${selectedFiles.length} file(s). Skipped duplicates: ${skippedTotal}. Failed: ${failedTotal}.<br>${detailLines.map(x => `<span>${htmlEscape(x)}</span>`).join('<br>')}`;
      fileInput.value = '';
      markSaved();
      await refresh();
      await loadImportHistory();
      await loadFieldTicketLines();
    };
    document.getElementById('fieldWiseAuditForm').onsubmit = async e => {
      e.preventDefault();
      const fileInput = e.target.querySelector('input[type="file"]');
      const resultEl = document.getElementById('fieldWiseAuditResult');
      if (!fileInput.files.length) {
        resultEl.textContent = 'Choose the Field Wise ticket export first.';
        return;
      }
      resultEl.textContent = 'Running Field Wise audit...';
      const data = new FormData();
      data.append('file', fileInput.files[0]);
      try {
        fieldWiseAuditData = await api('/api/fieldwise-audit', { method: 'POST', body: data });
        const s = fieldWiseAuditData.summary || {};
        resultEl.textContent = `Audit complete. Missing tracked tickets: ${s.missing_count || 0}. Untracked/blank order tickets: ${Number(s.untracked_count || 0) + Number(s.no_order_count || 0)}.`;
        renderFieldWiseAudit();
      } catch (err) {
        fieldWiseAuditData = null;
        renderFieldWiseAudit();
        resultEl.textContent = `Audit failed: ${err.message}`;
      }
    };
    document.getElementById('omitUntrackedAuditTickets').onchange = () => renderFieldWiseAudit();
    document.getElementById('exportMissingTickets').onclick = () => downloadMissingTicketCsv();
    document.getElementById('addRate').onclick = async () => {
      await api('/api/internal-rates', {
        method: 'POST',
        body: JSON.stringify({
          category_type: document.getElementById('rateType').value,
          rate_set_id: document.getElementById('rateSetSelect').value,
          category: document.getElementById('rateCategory').value,
          raw_rate: document.getElementById('rateRaw').value
        })
      });
      document.getElementById('rateCategory').value = '';
      document.getElementById('rateRaw').value = '0';
      markSaved();
      await refresh();
    };
    document.getElementById('addCogCategory').onclick = async () => {
      await api('/api/cog-categories', {
        method: 'POST',
        body: JSON.stringify({
          code: document.getElementById('cogCode').value,
          name: document.getElementById('cogName').value,
          description: document.getElementById('cogDescription').value
        })
      });
      ['cogCode','cogName','cogDescription'].forEach(id => document.getElementById(id).value = '');
      markSaved();
      state.cogCategories = await api('/api/cog-categories');
      jobOrderReportRows = await api('/api/job-order-report');
      loadCogCategoryEditor();
      renderFieldPoJobChoices(document.getElementById('fieldPoJobSearch')?.value || '');
    };
    document.getElementById('importVendorInvoice').onclick = async () => {
      const fileInput = document.getElementById('vendorInvoiceFile');
      const resultEl = document.getElementById('vendorImportResult');
      if (!fileInput.files.length) {
        resultEl.textContent = 'Choose one or more vendor invoice PDFs first.';
        return;
      }
      const selectedFiles = Array.from(fileInput.files);
      resultEl.textContent = `Importing ${selectedFiles.length} vendor invoice PDF(s)...`;
      try {
        const results = [];
        for (const file of selectedFiles) {
          const fd = new FormData();
          fd.append('file', file);
          fd.append('project_id', state.projectId);
          try {
            const data = await api('/api/import-vendor-invoice', { method: 'POST', body: fd });
            results.push({ file: file.name, ok: true, data });
          } catch (err) {
            results.push({ file: file.name, ok: false, error: err.message });
          }
        }
        const importedLines = results.filter(r => r.ok && !r.data.duplicate).reduce((sum, r) => sum + Number(r.data.imported || 0), 0);
        const duplicateCount = results.filter(r => r.ok && r.data.duplicate).length;
        const failedCount = results.filter(r => !r.ok).length;
        const detailLines = results.map(r => {
          if (!r.ok) return `${r.file}: failed - ${r.error}`;
          const d = r.data;
          if (d.duplicate) return `${r.file}: duplicate ${d.vendor || 'Vendor'} invoice ${d.invoice_number || ''}, already has ${d.existing_line_count || 0} line item(s) totaling ${money(d.existing_total)}`;
          const matched = d.matched_subproject_id ? `auto-coded to PO/order ${d.order_number}` : `needs subproject review`;
          return `${r.file}: imported ${d.imported} line item(s) from ${d.vendor || 'vendor'} invoice ${d.invoice_number || ''}, ${matched}`;
        });
        resultEl.innerHTML = `Imported ${importedLines} total line item(s) from ${selectedFiles.length} PDF(s). Duplicates: ${duplicateCount}. Failed: ${failedCount}.<br>${detailLines.map(x => `<span>${htmlEscape(x)}</span>`).join('<br>')}`;
        fileInput.value = '';
        markSaved();
        await refresh();
        await loadVendorInvoiceLines();
      } catch (err) {
        resultEl.textContent = `Import failed: ${err.message}`;
      }
    };

    updateCoPricingFields();
    if ('serviceWorker' in navigator) {
      window.addEventListener('load', async () => {
        const registrations = await navigator.serviceWorker.getRegistrations();
        await Promise.all(registrations.map(registration => registration.unregister()));
        if ('caches' in window) {
          const keys = await caches.keys();
          await Promise.all(keys.map(key => caches.delete(key)));
        }
      });
    }
    (async () => {
      await loadCurrentUser();
      if (isTexasReadOnly()) {
        openTab('texasOps');
      } else if (state.currentUser?.role === 'Field PO') {
        openTab('fieldPo');
      } else {
        await loadProjects();
      }
    })();
  </script>
</body>
</html>
"""


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        print("%s - %s" % (self.address_string(), fmt % args))

    def do_GET(self):
        try:
            parsed = urlparse(self.path)
            qs = parse_qs(parsed.query)
            if parsed.path == "/manifest.json":
                return text_response(self, pwa_manifest_json(), "application/manifest+json")
            if parsed.path == "/service-worker.js":
                return text_response(self, service_worker_js(), "application/javascript; charset=utf-8")
            if parsed.path == "/offline":
                return text_response(self, offline_html())
            if parsed.path == "/login":
                if current_user(self):
                    return redirect_response(self, "/")
                return text_response(self, LOGIN_HTML)
            public_paths = ("/brand/", "/manifest.json", "/service-worker.js", "/offline")
            user = current_user(self)
            if not user and not parsed.path.startswith(public_paths):
                if parsed.path.startswith("/api/"):
                    return json_response(self, {"error": "Login required"}, 401)
                return redirect_response(self, "/login")
            if is_texas_read_only(user):
                allowed_paths = ("/", "/api/me", "/api/texas-financial-summary")
                if parsed.path not in allowed_paths and not parsed.path.startswith(public_paths):
                    if parsed.path.startswith("/api/"):
                        return json_response(self, {"error": "Texas Operations access only."}, 403)
                    return redirect_response(self, "/")
            if is_field_po_only(user):
                allowed_paths = ("/", "/api/me", "/api/job-order-report", "/api/purchase-orders")
                if parsed.path not in allowed_paths and not parsed.path.startswith(public_paths) and not parsed.path.startswith("/po/") and not parsed.path.startswith("/uploads/"):
                    if parsed.path.startswith("/api/"):
                        return json_response(self, {"error": "PO access only."}, 403)
                    return redirect_response(self, "/")
            if parsed.path in ("/developer-revision", "/server-health"):
                if user.get("role") != "Admin":
                    return text_response(self, "Admin required", "text/plain", 403)
                return text_response(self, developer_revision_html(user))
            if parsed.path in ("/api/developer-revision", "/api/server-health"):
                if user.get("role") != "Admin":
                    return json_response(self, {"error": "Admin required"}, 403)
                return json_response(self, app_revision_info())
            if parsed.path == "/api/admin/database-backup":
                if user.get("role") != "Admin":
                    return text_response(self, "Admin required", "text/plain", 403)
                filename = f"pm-tracker-backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}.sqlite3"
                log_activity(user, "downloaded backup", "Database", filename)
                return download_response(self, sqlite_backup_bytes(), filename, "application/vnd.sqlite3")
            if parsed.path == "/":
                return text_response(self, HTML)
            if parsed.path.startswith("/po/"):
                po_id = parsed.path.rsplit("/", 1)[-1]
                po = one("SELECT * FROM purchase_orders WHERE id = ?", (po_id,))
                if not po:
                    return text_response(self, "PO not found", "text/plain", 404)
                if is_field_po_only(user) and po["requested_by_user_id"] != user["id"]:
                    return text_response(self, "Not found", "text/plain", 404)
                return text_response(self, purchase_order_html(po))
            if parsed.path.startswith("/uploads/"):
                requested_upload = Path(unquote(parsed.path.removeprefix("/uploads/"))).name
                if is_field_po_only(user):
                    allowed_attachment = one(
                        """
                        SELECT id
                        FROM purchase_orders
                        WHERE requested_by_user_id = ?
                          AND (attachment_file = ? OR pickup_file = ?)
                        """,
                        (user["id"], requested_upload, requested_upload),
                    )
                    if not allowed_attachment:
                        return text_response(self, "Not found", "text/plain", 404)
                path = upload_attachment_path(requested_upload)
                if not path:
                    return text_response(self, "Not found", "text/plain", 404)
                content_types = {
                    ".pdf": "application/pdf",
                    ".png": "image/png",
                    ".jpg": "image/jpeg",
                    ".jpeg": "image/jpeg",
                    ".webp": "image/webp",
                }
                return file_response(self, path, content_types.get(path.suffix.lower(), "application/octet-stream"))
            if parsed.path.startswith("/pdf-viewer/"):
                html = pdf_viewer_html(parsed.path.removeprefix("/pdf-viewer/"))
                if not html:
                    return text_response(self, "Not found", "text/plain", 404)
                return text_response(self, html)
            if parsed.path.startswith("/pdf-page/"):
                parts = parsed.path.removeprefix("/pdf-page/").rsplit("/", 1)
                if len(parts) != 2 or not parts[1].endswith(".png"):
                    return text_response(self, "Not found", "text/plain", 404)
                page_index = parts[1].removesuffix(".png")
                try:
                    image_data = render_pdf_page_png(parts[0], page_index)
                except Exception:
                    image_data = None
                if not image_data:
                    return text_response(self, "Not found", "text/plain", 404)
                return bytes_response(self, image_data, "image/png")
            if parsed.path.startswith("/brand/"):
                file_name = Path(unquote(parsed.path.removeprefix("/brand/"))).name
                path = (BRAND_DIR / file_name).resolve()
                brand_root = BRAND_DIR.resolve()
                if brand_root not in path.parents or not path.exists() or path.suffix.lower() not in (".png", ".jpg", ".jpeg", ".webp"):
                    return text_response(self, "Not found", "text/plain", 404)
                content_types = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".webp": "image/webp"}
                return file_response(self, path, content_types.get(path.suffix.lower(), "application/octet-stream"))
            if parsed.path == "/api/job-order-report":
                return json_response(
                    self,
                    rows(
                        """
                        SELECT
                          'subproject:' || sp.id AS job_key,
                          sp.job_number,
                          'Subproject' AS item_type,
                          p.customer,
                          p.name AS project_name,
                          p.project_code,
                          sp.code AS reference_code,
                          sp.name AS description,
                          p.description AS project_description,
                          p.location,
                          COALESCE(p.status, 'Active') AS status
                        FROM subprojects sp
                        JOIN projects p ON p.id = sp.project_id
                        WHERE COALESCE(p.status, 'Active') <> 'Archived'
                          AND TRIM(COALESCE(sp.job_number, '')) <> ''
                        UNION ALL
                        SELECT
                          'change_order:' || co.id AS job_key,
                          co.job_number,
                          COALESCE(co.order_type, 'Change Order') AS item_type,
                          p.customer,
                          p.name AS project_name,
                          p.project_code,
                          CASE
                            WHEN sp.code IS NOT NULL THEN co.co_number || ' / ' || sp.code
                            ELSE co.co_number
                          END AS reference_code,
                          co.title AS description,
                          p.description AS project_description,
                          p.location,
                          COALESCE(co.status, 'Pending') AS status
                        FROM change_orders co
                        JOIN projects p ON p.id = co.project_id
                        LEFT JOIN subprojects sp ON sp.id = co.subproject_id
                        WHERE COALESCE(p.status, 'Active') <> 'Archived'
                          AND TRIM(COALESCE(co.job_number, '')) <> ''
                        UNION ALL
                        SELECT
                          'cog:' || cg.id AS job_key,
                          cg.name AS job_number,
                          'COG' AS item_type,
                          '' AS customer,
                          'COG Category' AS project_name,
                          cg.code AS project_code,
                          cg.code AS reference_code,
                          cg.description,
                          '' AS project_description,
                          '' AS location,
                          'Active' AS status
                        FROM cog_categories cg
                        WHERE COALESCE(cg.active, 1) = 1
                        ORDER BY customer, project_name, job_number, item_type
                        """
                    ),
                )
            if parsed.path == "/api/fieldwise-audit-omissions":
                return json_response(
                    self,
                    rows(
                        """
                        SELECT *
                        FROM fieldwise_audit_omissions
                        ORDER BY created_at DESC, ticket_number, order_number
                        """
                    ),
                )
            if parsed.path == "/api/projects":
                status_filter = (qs.get("status", ["active"])[0] or "active").lower()
                if status_filter == "all":
                    return json_response(self, rows("SELECT * FROM projects ORDER BY CASE WHEN status = 'Archived' THEN 1 ELSE 0 END, project_code"))
                if status_filter == "archived":
                    return json_response(self, rows("SELECT * FROM projects WHERE status = 'Archived' ORDER BY COALESCE(archived_at, closed_at, created_at) DESC, project_code"))
                return json_response(self, rows("SELECT * FROM projects WHERE COALESCE(status, 'Active') <> 'Archived' ORDER BY project_code"))
            if parsed.path == "/api/me":
                return json_response(self, current_user_with_permissions(self))
            if parsed.path == "/api/users":
                if not require_admin(self):
                    return json_response(self, {"error": "Admin required"}, 403)
                return json_response(self, rows("SELECT id, username, display_name, role, active, COALESCE(po_auto_issue, 0) AS po_auto_issue, COALESCE(must_change_password, 0) AS must_change_password, created_at FROM users ORDER BY username"))
            if parsed.path == "/api/role-permissions":
                if not require_admin(self):
                    return json_response(self, {"error": "Admin required"}, 403)
                return json_response(
                    self,
                    {
                        "roles": list(ROLE_NAMES),
                        "permissions": PERMISSION_DEFINITIONS,
                        "matrix": {role: role_permissions(role) for role in ROLE_NAMES},
                    },
                )
            if parsed.path == "/api/activity":
                if not require_admin(self):
                    return json_response(self, {"error": "Admin required"}, 403)
                try:
                    limit = max(1, min(500, int(qs.get("limit", ["100"])[0] or 100)))
                except Exception:
                    limit = 100
                return json_response(
                    self,
                    rows(
                        """
                        SELECT *
                        FROM activity_log
                        ORDER BY created_at DESC, id DESC
                        LIMIT ?
                        """,
                        (limit,),
                    ),
                )
            if parsed.path == "/api/purchase-orders":
                if not can_use_field_po(user):
                    return json_response(self, {"error": "PO access required."}, 403)
                if is_field_po_only(user):
                    return json_response(
                        self,
                        rows("SELECT * FROM purchase_orders WHERE requested_by_user_id = ? ORDER BY created_at DESC, id DESC", (user["id"],)),
                    )
                project_filter = qs.get("project_id", [""])[0]
                if project_filter:
                    return json_response(
                        self,
                        rows("SELECT * FROM purchase_orders WHERE project_id = ? ORDER BY created_at DESC, id DESC", (project_filter,)),
                    )
                return json_response(self, rows("SELECT * FROM purchase_orders ORDER BY created_at DESC, id DESC"))
            if parsed.path == "/api/home-alerts":
                return json_response(self, home_alerts())
            if parsed.path == "/api/bid-summary":
                return json_response(self, bid_summary())
            if parsed.path == "/api/texas-financial-summary":
                return json_response(self, texas_financial_summary())
            if parsed.path == "/api/texas-upload-reminders":
                if not require_editor(self):
                    return json_response(self, {"error": "Admin or standard user access required."}, 403)
                return json_response(self, texas_upload_reminder_status())
            if parsed.path == "/api/billing-invoice-reminders":
                if not require_editor(self):
                    return json_response(self, {"error": "Admin or standard user access required."}, 403)
                return json_response(self, billing_invoice_reminder_status())
            if parsed.path == "/api/po-untouched-reminders":
                if not require_editor(self):
                    return json_response(self, {"error": "Admin or standard user access required."}, 403)
                return json_response(self, po_untouched_reminder_status())
            if parsed.path == "/api/bids":
                return json_response(self, rows("SELECT * FROM bid_requests ORDER BY bid_due_date, rfq_no"))
            if parsed.path == "/api/rate-sets":
                return json_response(self, rows("SELECT * FROM rate_sets WHERE active = 1 ORDER BY name"))
            if parsed.path == "/api/subprojects":
                return json_response(self, rows("SELECT * FROM subprojects WHERE project_id = ? ORDER BY code", (qs.get("project_id", [""])[0],)))
            if parsed.path == "/api/change-orders":
                return json_response(self, rows("SELECT * FROM change_orders WHERE project_id = ? ORDER BY co_number", (qs.get("project_id", [""])[0],)))
            if parsed.path == "/api/cog-categories":
                return json_response(self, rows("SELECT * FROM cog_categories WHERE COALESCE(active, 1) = 1 ORDER BY name, code"))
            if parsed.path == "/api/customer-invoices":
                project_id = qs.get("project_id", [""])[0]
                invoice_rows = rows(
                    """
                    SELECT ci.*, COALESCE(SUM(cia.amount), ci.amount, 0) AS allocated_amount,
                           sp.job_number, sp.code AS subproject_code, co.co_number, co.job_number AS co_job_number
                    FROM customer_invoices ci
                    LEFT JOIN customer_invoice_allocations cia ON cia.customer_invoice_id = ci.id
                    LEFT JOIN subprojects sp ON sp.id = ci.subproject_id
                    LEFT JOIN change_orders co ON co.id = ci.change_order_id
                    WHERE ci.project_id = ?
                    GROUP BY ci.id
                    ORDER BY ci.invoice_date DESC, ci.id DESC
                    """,
                    (project_id,),
                )
                allocation_rows = rows(
                    """
                    SELECT
                      cia.*,
                      sp.job_number,
                      sp.code AS subproject_code,
                      sp.name AS subproject_name,
                      co.co_number,
                      co.job_number AS co_job_number,
                      co.title AS co_title,
                      COALESCE(co.order_type, 'Change Order') AS order_type
                    FROM customer_invoice_allocations cia
                    LEFT JOIN subprojects sp ON sp.id = cia.subproject_id
                    LEFT JOIN change_orders co ON co.id = cia.change_order_id
                    WHERE cia.project_id = ?
                    ORDER BY cia.id
                    """,
                    (project_id,),
                )
                by_invoice = {}
                for invoice in invoice_rows:
                    invoice["allocations"] = []
                    by_invoice[invoice["id"]] = invoice
                for allocation in allocation_rows:
                    if allocation.get("change_order_id"):
                        pieces = [allocation.get("co_number"), allocation.get("co_job_number")]
                        label = " / ".join(str(x) for x in pieces if x) or allocation.get("order_type") or "Change Order"
                        if allocation.get("co_title"):
                            label = f"{label} - {allocation.get('co_title')}"
                    else:
                        pieces = [allocation.get("job_number"), allocation.get("subproject_code")]
                        label = " ".join(str(x) for x in pieces if x) or "Unassigned"
                        if allocation.get("subproject_name"):
                            label = f"{label} - {allocation.get('subproject_name')}"
                    allocation["target_label"] = label
                    by_invoice.get(allocation["customer_invoice_id"], {}).get("allocations", []).append(allocation)
                return json_response(self, invoice_rows)
            if parsed.path == "/api/internal-rates":
                return json_response(self, rows("SELECT * FROM internal_rates WHERE active = 1 ORDER BY rate_set_id, category_type, category"))
            if parsed.path == "/api/cost-records":
                return json_response(self, rows("SELECT * FROM cost_records WHERE project_id = ? ORDER BY CASE WHEN subproject_id IS NULL THEN 0 ELSE 1 END, record_date DESC, id DESC", (qs.get("project_id", [""])[0],)))
            if parsed.path == "/api/imports":
                return json_response(
                    self,
                    rows(
                        """
                        SELECT
                          source,
                          source_file,
                          COUNT(*) AS record_count,
                          COALESCE(SUM(sales_amount), 0) AS sales_amount,
                          COALESCE(SUM(amount), 0) AS raw_amount,
                          MAX(created_at) AS last_imported
                        FROM cost_records
                        WHERE project_id = ?
                          AND source IN ('Field Wise', 'Field Wise PDF', 'Vendor Invoice')
                          AND source_file IS NOT NULL
                        GROUP BY source, source_file
                        ORDER BY MAX(created_at) DESC
                        """,
                        (qs.get("project_id", [""])[0],),
                    ),
                )
            if parsed.path == "/api/vendor-invoice-lines":
                return json_response(
                    self,
                    rows(
                        """
                        SELECT cr.*, sp.job_number, sp.code AS subproject_code
                        FROM cost_records cr
                        LEFT JOIN subprojects sp ON sp.id = cr.subproject_id
                        WHERE cr.project_id = ?
                          AND cr.source = 'Vendor Invoice'
                        ORDER BY cr.record_date DESC, cr.ticket_or_invoice, cr.id
                        """,
                        (qs.get("project_id", [""])[0],),
                    ),
                )
            if parsed.path == "/api/vendor-invoice-allocations":
                allocation_rows = rows(
                    """
                    SELECT *
                    FROM vendor_invoice_allocations
                    WHERE project_id = ?
                    ORDER BY allocated_at DESC, id DESC
                    """,
                    (qs.get("project_id", [""])[0],),
                )
                line_rows = rows(
                    """
                    SELECT
                      vl.*,
                      sp.job_number,
                      sp.code AS subproject_code,
                      sp.name AS subproject_name,
                      co.co_number,
                      co.job_number AS co_job_number,
                      co.title AS co_title
                    FROM vendor_invoice_allocation_lines vl
                    LEFT JOIN subprojects sp ON sp.id = vl.subproject_id
                    LEFT JOIN change_orders co ON co.id = vl.change_order_id
                    WHERE vl.allocation_id IN (
                      SELECT id FROM vendor_invoice_allocations WHERE project_id = ?
                    )
                    ORDER BY vl.id
                    """,
                    (qs.get("project_id", [""])[0],),
                )
                by_allocation = {}
                for allocation in allocation_rows:
                    item = dict(allocation)
                    item["lines"] = []
                    by_allocation[item["id"]] = item
                for line in line_rows:
                    label = [line.get("job_number"), line.get("subproject_code")].filter(Boolean)
                    if line.get("change_order_id"):
                        co_label = [line.get("co_number"), line.get("co_job_number")].filter(Boolean)
                        label = [" / ".join(co_label) or "Change Order", line.get("co_title") or ""]
                    target_label = " - ".join(str(x) for x in label if x) or "Unassigned"
                    by_allocation.get(line["allocation_id"], {}).get("lines", []).append({**dict(line), "target_label": target_label})
                return json_response(self, list(by_allocation.values()))
            if parsed.path == "/api/summary":
                summary = project_summary(qs.get("project_id", [""])[0])
                return json_response(self, summary or {"error": "Project not found"}, 200 if summary else 404)
            if parsed.path == "/api/subproject-detail":
                detail = subproject_detail(qs.get("subproject_id", [""])[0], qs.get("change_order_id", [""])[0] or None)
                return json_response(self, detail or {"error": "Subproject not found"}, 200 if detail else 404)
            if parsed.path == "/api/master-detail":
                detail = master_project_detail(qs.get("project_id", [""])[0])
                return json_response(self, detail or {"error": "Project not found"}, 200 if detail else 404)
            return text_response(self, "Not found", "text/plain", 404)
        except Exception as e:
            traceback.print_exc()
            return json_response(self, {"error": str(e)}, 500)

    def do_POST(self):
        try:
            parsed = urlparse(self.path)
            if parsed.path == "/api/login":
                data = parse_json(self)
                user = one("SELECT * FROM users WHERE username = ? AND active = 1", (data.get("username"),))
                if not user or not verify_password(data.get("password"), user["password_hash"]):
                    return json_response(self, {"error": "Invalid login"}, 401)
                token = create_session(user["id"])
                return login_success_response(self, token)
            if parsed.path == "/api/logout":
                token = parse_cookie_header(self.headers.get("Cookie")).get("pm_session")
                if token:
                    execute("DELETE FROM user_sessions WHERE session_token = ?", (token,))
                return logout_response(self)
            if not current_user(self):
                return json_response(self, {"error": "Login required"}, 401)
            if parsed.path == "/api/change-password":
                data = parse_json(self)
                user = current_user(self)
                account = one("SELECT id, password_hash FROM users WHERE id = ?", (user["id"],))
                if not account or not verify_password(data.get("current_password"), account["password_hash"]):
                    return json_response(self, {"error": "Current password is incorrect."}, 400)
                new_password = str(data.get("new_password") or "")
                if len(new_password) < 8:
                    return json_response(self, {"error": "Use at least 8 characters."}, 400)
                if new_password != data.get("confirm_password"):
                    return json_response(self, {"error": "New passwords do not match."}, 400)
                if new_password == DEFAULT_NEW_USER_PASSWORD:
                    return json_response(self, {"error": "Choose a password different from the temporary password."}, 400)
                execute(
                    "UPDATE users SET password_hash = ?, must_change_password = 0 WHERE id = ?",
                    (hash_password(new_password), user["id"]),
                )
                token = parse_cookie_header(self.headers.get("Cookie")).get("pm_session")
                with db() as con:
                    con.execute("DELETE FROM user_sessions WHERE user_id = ? AND session_token <> ?", (user["id"], token or ""))
                return json_response(self, {"ok": True})
            if parsed.path == "/api/purchase-orders":
                user = current_user(self)
                if not can_use_field_po(user):
                    return json_response(self, {"error": "PO access required."}, 403)
                form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": self.headers.get("Content-Type")})
                job_key = form.getvalue("job_key")
                customer_name = str(form.getvalue("customer_name") or "").strip()
                vendor = str(form.getvalue("vendor") or "").strip()
                description = str(form.getvalue("description") or "").strip()
                estimated_amount = money(form.getvalue("estimated_amount"))
                if not job_key:
                    return json_response(self, {"error": "Choose a job/order number."}, 400)
                if not vendor:
                    return json_response(self, {"error": "Enter the vendor."}, 400)
                if not description:
                    return json_response(self, {"error": "Enter what you are buying."}, 400)
                attachment_file = None
                file_item = form["attachment"] if "attachment" in form else None
                if file_item is not None and getattr(file_item, "filename", ""):
                    safe_name = Path(file_item.filename).name
                    suffix = Path(safe_name).suffix.lower()
                    if suffix not in (".pdf", ".png", ".jpg", ".jpeg", ".webp"):
                        return json_response(self, {"error": "Attachment must be a PDF or image."}, 400)
                    attachment_file = f"{datetime.now().strftime('%Y%m%d%H%M%S')}-po-{safe_name}"
                    path = UPLOAD_DIR / attachment_file
                    with open(path, "wb") as f:
                        f.write(file_item.file.read())
                now = datetime.now().isoformat(timespec="seconds")
                with db() as con:
                    job_ref = job_reference_for_po(con, job_key)
                    if not job_ref:
                        if attachment_file:
                            try:
                                (UPLOAD_DIR / attachment_file).unlink()
                            except Exception:
                                pass
                        return json_response(self, {"error": "That job/order is no longer available."}, 400)
                    if po_customer_required_for_ref(job_ref) and not customer_name:
                        if attachment_file:
                            try:
                                (UPLOAD_DIR / attachment_file).unlink()
                            except Exception:
                                pass
                        return json_response(self, {"error": "Enter the customer for this shop / COG PO."}, 400)
                    po_number = next_po_number(con)
                    initial_status = "Issued" if can_auto_issue_po(user) else "Pending Approval"
                    cur = con.execute(
                        """
                        INSERT INTO purchase_orders (
                          po_number, project_id, subproject_id, change_order_id, cog_category_id, job_number, job_label, customer_name,
                          vendor, description, estimated_amount, attachment_file, status,
                          requested_by_user_id, requested_by_username, created_at, updated_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            po_number,
                            job_ref["project_id"],
                            job_ref["subproject_id"],
                            job_ref["change_order_id"],
                            job_ref["cog_category_id"],
                            job_ref["job_number"],
                            job_ref["job_label"],
                            customer_name,
                            vendor,
                            description,
                            estimated_amount,
                            attachment_file,
                            initial_status,
                            user["id"],
                            user["username"],
                            now,
                            now,
                        ),
                    )
                    log_activity(user, "created PO", "Purchase Order", po_number, {"status": initial_status, "vendor": vendor, "estimated_amount": estimated_amount})
                    return json_response(self, {"id": cur.lastrowid, "po_number": po_number, "status": initial_status})
            if parsed.path.startswith("/api/purchase-orders/") and parsed.path.endswith("/pickup"):
                user = current_user(self)
                if not can_use_field_po(user):
                    return json_response(self, {"error": "PO access required."}, 403)
                po_id = parsed.path.split("/")[-2]
                with db() as con:
                    po = con.execute("SELECT * FROM purchase_orders WHERE id = ?", (po_id,)).fetchone()
                    if not po:
                        return json_response(self, {"error": "PO not found."}, 404)
                    if is_field_po_only(user) and po["requested_by_user_id"] != user["id"]:
                        return json_response(self, {"error": "PO not found."}, 404)
                    if po["status"] in ("Closed", "Void"):
                        return json_response(self, {"error": "Cannot upload to a closed or void PO."}, 400)
                    if po["status"] == "Pending Approval":
                        return json_response(self, {"error": "This PO is pending office approval. Upload the pickup ticket after it is issued."}, 400)
                    form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": self.headers.get("Content-Type")})
                    file_item = form["pickup_file"] if "pickup_file" in form else None
                    if file_item is None or not getattr(file_item, "filename", ""):
                        return json_response(self, {"error": "Choose a pickup ticket photo or PDF."}, 400)
                    safe_name = Path(file_item.filename).name
                    suffix = Path(safe_name).suffix.lower()
                    if suffix not in (".pdf", ".png", ".jpg", ".jpeg", ".webp"):
                        return json_response(self, {"error": "Pickup ticket must be a PDF or image."}, 400)
                    saved_name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}-pickup-{safe_name}"
                    path = UPLOAD_DIR / saved_name
                    with open(path, "wb") as f:
                        f.write(file_item.file.read())
                    now = datetime.now().isoformat(timespec="seconds")
                    con.execute(
                        """
                        UPDATE purchase_orders
                        SET pickup_file = ?,
                            pickup_uploaded_at = ?,
                            pickup_uploaded_by_user_id = ?,
                            pickup_uploaded_by_username = ?,
                            status = 'Picked Up',
                            updated_at = ?
                        WHERE id = ?
                        """,
                        (saved_name, now, user["id"], user["username"], now, po_id),
                    )
                    log_activity(user, "uploaded pickup ticket", "Purchase Order", po["po_number"] if po else po_id, {"pickup_file": saved_name, "uploaded_at": now})
                return json_response(self, {"ok": True, "pickup_file": saved_name, "status": "Picked Up"})
            if parsed.path == "/api/role-permissions":
                if not require_admin(self):
                    return json_response(self, {"error": "Admin required"}, 403)
                data = parse_json(self)
                role = clean_role(data.get("role"))
                if role == "Admin":
                    return json_response(self, {"error": "Admin permissions are always full access."}, 400)
                updates = data.get("permissions") or {}
                allowed_keys = {p["key"] for p in PERMISSION_DEFINITIONS}
                now = datetime.now().isoformat(timespec="seconds")
                with db() as con:
                    for key, values in updates.items():
                        if key not in allowed_keys:
                            continue
                        can_view = 1 if values.get("can_view") else 0
                        can_edit = 1 if values.get("can_edit") and can_view else 0
                        con.execute(
                            """
                            INSERT INTO role_permissions (role, permission_key, can_view, can_edit, updated_at)
                            VALUES (?, ?, ?, ?, ?)
                            ON CONFLICT(role, permission_key) DO UPDATE SET
                              can_view = excluded.can_view,
                              can_edit = excluded.can_edit,
                              updated_at = excluded.updated_at
                            """,
                            (role, key, can_view, can_edit, now),
                        )
                log_activity(current_user(self), "updated role permissions", "Role", role, {"permissions": updates})
                return json_response(self, {"ok": True})
            if parsed.path == "/api/home-alerts/job-invoice-ack":
                user = current_user(self)
                if not require_editor(self):
                    return json_response(self, {"error": "Read-only users cannot acknowledge alerts."}, 403)
                data = parse_json(self)
                target_type = str(data.get("target_type") or "").strip()
                target_id = str(data.get("target_id") or "").strip()
                if target_type == "subproject":
                    target = one(
                        """
                        SELECT sp.id, sp.job_number, sp.code AS reference_code, sp.name AS job_name,
                               p.project_code, p.name AS project_name
                        FROM subprojects sp
                        JOIN projects p ON p.id = sp.project_id
                        WHERE sp.id = ?
                        """,
                        (target_id,),
                    )
                elif target_type == "change_order":
                    target = one(
                        """
                        SELECT co.id, co.job_number, co.co_number AS reference_code, co.title AS job_name,
                               p.project_code, p.name AS project_name
                        FROM change_orders co
                        JOIN projects p ON p.id = co.project_id
                        WHERE co.id = ?
                        """,
                        (target_id,),
                    )
                else:
                    return json_response(self, {"error": "Choose a valid job target."}, 400)
                if not target:
                    return json_response(self, {"error": "Job target not found."}, 404)
                now = datetime.now().isoformat(timespec="seconds")
                execute(
                    """
                    INSERT INTO job_invoice_alert_acknowledgements (
                      target_type, target_id, acknowledged_at, acknowledged_by_user_id, acknowledged_by_username, notes
                    )
                    VALUES (?, ?, ?, ?, ?, ?)
                    ON CONFLICT(target_type, target_id) DO UPDATE SET
                      acknowledged_at = excluded.acknowledged_at,
                      acknowledged_by_user_id = excluded.acknowledged_by_user_id,
                      acknowledged_by_username = excluded.acknowledged_by_username,
                      notes = excluded.notes
                    """,
                    (
                        target_type,
                        target_id,
                        now,
                        user["id"],
                        user["username"],
                        str(data.get("notes") or "Acknowledged from Needs Attention").strip(),
                    ),
                )
                label = " / ".join(str(x or "").strip() for x in (target.get("job_number"), target.get("reference_code"), target.get("job_name")) if str(x or "").strip())
                log_activity(user, "acknowledged job invoice alert", "Customer Billing", label, {"target_type": target_type, "target_id": target_id, "next_check_days": 5})
                return json_response(self, {"ok": True, "acknowledged_at": now})
            if not require_editor(self):
                return json_response(self, {"error": "Read-only users cannot make changes."}, 403)
            if parsed.path == "/api/texas-upload-reminders":
                actor = current_user(self)
                data = parse_json(self)
                selected_ids = {int(x) for x in data.get("user_ids", []) if str(x).isdigit()}
                enabled = 1 if str(data.get("enabled", "1")).lower() not in ("0", "false", "no") else 0
                try:
                    weekday = max(0, min(6, int(data.get("weekday", 4))))
                except Exception:
                    weekday = 4
                reminder_time = str(data.get("reminder_time") or "15:00").strip()
                if not re.match(r"^\d{2}:\d{2}$", reminder_time):
                    reminder_time = "15:00"
                now = datetime.now().isoformat(timespec="seconds")
                with db() as con:
                    con.execute(
                        """
                        INSERT INTO texas_upload_reminder_settings (id, enabled, weekday, reminder_time, updated_at)
                        VALUES (1, ?, ?, ?, ?)
                        ON CONFLICT(id) DO UPDATE SET
                          enabled = excluded.enabled,
                          weekday = excluded.weekday,
                          reminder_time = excluded.reminder_time,
                          updated_at = excluded.updated_at
                        """,
                        (enabled, weekday, reminder_time, now),
                    )
                    con.execute("UPDATE texas_upload_reminder_recipients SET active = 0")
                    for user_id in selected_ids:
                        con.execute(
                            """
                            INSERT INTO texas_upload_reminder_recipients (user_id, active, created_at)
                            VALUES (?, 1, ?)
                            ON CONFLICT(user_id) DO UPDATE SET active = 1
                            """,
                            (user_id, now),
                        )
                log_activity(actor, "updated reminder setup", "Texas Ops", "P&L upload reminder", {"enabled": enabled, "weekday": weekday, "reminder_time": reminder_time, "recipient_count": len(selected_ids)})
                return json_response(self, texas_upload_reminder_status())
            if parsed.path == "/api/texas-upload-reminders/send-now":
                actor = current_user(self)
                parse_json(self)
                result = run_texas_upload_reminder(force=True)
                log_activity(actor, "ran reminder check", "Texas Ops", "P&L upload reminder", result)
                result["status"] = texas_upload_reminder_status()
                return json_response(self, result)
            if parsed.path == "/api/billing-invoice-reminders":
                actor = current_user(self)
                data = parse_json(self)
                selected_ids = {int(x) for x in data.get("user_ids", []) if str(x).isdigit()}
                enabled = 1 if str(data.get("enabled", "1")).lower() not in ("0", "false", "no") else 0
                try:
                    weekday = max(0, min(6, int(data.get("weekday", 4))))
                except Exception:
                    weekday = 4
                reminder_time = str(data.get("reminder_time") or "15:00").strip()
                if not re.match(r"^\d{2}:\d{2}$", reminder_time):
                    reminder_time = "15:00"
                try:
                    lookback_days = max(1, min(60, int(data.get("lookback_days", 7))))
                except Exception:
                    lookback_days = 7
                now = datetime.now().isoformat(timespec="seconds")
                with db() as con:
                    con.execute(
                        """
                        INSERT INTO billing_invoice_reminder_settings (id, enabled, weekday, reminder_time, lookback_days, updated_at)
                        VALUES (1, ?, ?, ?, ?, ?)
                        ON CONFLICT(id) DO UPDATE SET
                          enabled = excluded.enabled,
                          weekday = excluded.weekday,
                          reminder_time = excluded.reminder_time,
                          lookback_days = excluded.lookback_days,
                          updated_at = excluded.updated_at
                        """,
                        (enabled, weekday, reminder_time, lookback_days, now),
                    )
                    con.execute("UPDATE billing_invoice_reminder_recipients SET active = 0")
                    for user_id in selected_ids:
                        con.execute(
                            """
                            INSERT INTO billing_invoice_reminder_recipients (user_id, active, created_at)
                            VALUES (?, 1, ?)
                            ON CONFLICT(user_id) DO UPDATE SET active = 1
                            """,
                            (user_id, now),
                        )
                log_activity(actor, "updated reminder setup", "Customer Billing", "Missing customer invoice reminder", {"enabled": enabled, "weekday": weekday, "reminder_time": reminder_time, "lookback_days": lookback_days, "recipient_count": len(selected_ids)})
                return json_response(self, billing_invoice_reminder_status())
            if parsed.path == "/api/billing-invoice-reminders/send-now":
                actor = current_user(self)
                parse_json(self)
                result = run_billing_invoice_reminder(force=True)
                log_activity(actor, "ran reminder check", "Customer Billing", "Missing customer invoice reminder", result)
                result["status"] = billing_invoice_reminder_status()
                return json_response(self, result)
            if parsed.path == "/api/po-untouched-reminders":
                actor = current_user(self)
                data = parse_json(self)
                selected_ids = {int(x) for x in data.get("user_ids", []) if str(x).isdigit()}
                enabled = 1 if str(data.get("enabled", "1")).lower() not in ("0", "false", "no") else 0
                try:
                    weekday = max(0, min(6, int(data.get("weekday", 4))))
                except Exception:
                    weekday = 4
                reminder_time = str(data.get("reminder_time") or "15:00").strip()
                if not re.match(r"^\d{2}:\d{2}$", reminder_time):
                    reminder_time = "15:00"
                try:
                    untouched_days = max(1, min(30, int(data.get("untouched_days", 2))))
                except Exception:
                    untouched_days = 2
                now = datetime.now().isoformat(timespec="seconds")
                with db() as con:
                    con.execute(
                        """
                        INSERT INTO po_untouched_reminder_settings (id, enabled, weekday, reminder_time, untouched_days, updated_at)
                        VALUES (1, ?, ?, ?, ?, ?)
                        ON CONFLICT(id) DO UPDATE SET
                          enabled = excluded.enabled,
                          weekday = excluded.weekday,
                          reminder_time = excluded.reminder_time,
                          untouched_days = excluded.untouched_days,
                          updated_at = excluded.updated_at
                        """,
                        (enabled, weekday, reminder_time, untouched_days, now),
                    )
                    con.execute("UPDATE po_untouched_reminder_recipients SET active = 0")
                    for user_id in selected_ids:
                        con.execute(
                            """
                            INSERT INTO po_untouched_reminder_recipients (user_id, active, created_at)
                            VALUES (?, 1, ?)
                            ON CONFLICT(user_id) DO UPDATE SET active = 1
                            """,
                            (user_id, now),
                        )
                log_activity(actor, "updated reminder setup", "Purchase Orders", "Untouched PO reminder", {"enabled": enabled, "weekday": weekday, "reminder_time": reminder_time, "untouched_days": untouched_days, "recipient_count": len(selected_ids)})
                return json_response(self, po_untouched_reminder_status())
            if parsed.path == "/api/po-untouched-reminders/send-now":
                actor = current_user(self)
                parse_json(self)
                result = run_po_untouched_reminder(force=True)
                log_activity(actor, "ran reminder check", "Purchase Orders", "Untouched PO reminder", result)
                result["status"] = po_untouched_reminder_status()
                return json_response(self, result)
            if parsed.path == "/api/fieldwise-audit-omissions":
                user = current_user(self)
                data = parse_json(self)
                ticket_number = str(data.get("ticket_number") or "").strip()
                order_number = str(data.get("order_number") or "").strip()
                if not ticket_number:
                    return json_response(self, {"error": "Ticket number is required."}, 400)
                if not order_number:
                    return json_response(self, {"error": "Order number is required."}, 400)
                with db() as con:
                    con.execute(
                        """
                        INSERT INTO fieldwise_audit_omissions (
                          ticket_number, order_number, customer, project_name, reason,
                          omitted_by_user_id, omitted_by_username, created_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(ticket_number, order_number) DO UPDATE SET
                          customer = excluded.customer,
                          project_name = excluded.project_name,
                          reason = excluded.reason,
                          omitted_by_user_id = excluded.omitted_by_user_id,
                          omitted_by_username = excluded.omitted_by_username,
                          created_at = excluded.created_at
                        """,
                        (
                            ticket_number,
                            order_number,
                            str(data.get("customer") or "").strip(),
                            str(data.get("project_name") or "").strip(),
                            str(data.get("reason") or "").strip(),
                            user["id"],
                            user["username"],
                            datetime.now().isoformat(timespec="seconds"),
                        ),
                    )
                log_activity(user, "marked ticket OK to omit", "Field Wise Audit", f"{ticket_number} / {order_number}", {"customer": data.get("customer"), "project_name": data.get("project_name"), "reason": data.get("reason")})
                return json_response(self, {"ok": True})
            if parsed.path == "/api/fieldwise-audit":
                form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": self.headers.get("Content-Type")})
                file_item = form["file"] if "file" in form else None
                if file_item is None or not getattr(file_item, "filename", ""):
                    return json_response(self, {"error": "Choose a Field Wise ticket export."}, 400)
                safe_name = Path(file_item.filename).name
                if Path(safe_name).suffix.lower() not in (".xlsx", ".xlsm"):
                    return json_response(self, {"error": "Field Wise audit export must be an Excel file."}, 400)
                path = UPLOAD_DIR / f"{datetime.now().strftime('%Y%m%d%H%M%S')}-audit-{safe_name}"
                with open(path, "wb") as f:
                    f.write(file_item.file.read())
                try:
                    result = fieldwise_audit_result(path)
                finally:
                    try:
                        path.unlink()
                    except Exception:
                        pass
                return json_response(self, result)
            if parsed.path == "/api/import-fieldwise":
                form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": self.headers.get("Content-Type")})
                project_id = form.getvalue("project_id")
                file_item = form["file"]
                safe_name = Path(file_item.filename).name
                path = UPLOAD_DIR / f"{datetime.now().strftime('%Y%m%d%H%M%S')}-{safe_name}"
                with open(path, "wb") as f:
                    f.write(file_item.file.read())
                if safe_name.lower().endswith(".pdf"):
                    result = import_fieldwise_pdf(path, project_id)
                else:
                    result = import_fieldwise_xlsx(path, project_id)
                return json_response(self, {
                    "imported": result["count"],
                    "skipped": result["skipped"],
                    "order_number": result["order_number"],
                    "matched_subproject_id": result["matched_subproject_id"],
                    "matched_change_order_id": result.get("matched_change_order_id"),
                })
            if parsed.path == "/api/import-vendor-invoice":
                form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": self.headers.get("Content-Type")})
                project_id = form.getvalue("project_id")
                file_item = form["file"]
                safe_name = Path(file_item.filename).name
                path = UPLOAD_DIR / f"{datetime.now().strftime('%Y%m%d%H%M%S')}-{safe_name}"
                with open(path, "wb") as f:
                    f.write(file_item.file.read())
                result = import_vendor_invoice_pdf(path, project_id)
                return json_response(
                    self,
                    {
                        "imported": result["count"],
                        "skipped": result["skipped"],
                        "order_number": result["order_number"],
                        "matched_subproject_id": result["matched_subproject_id"],
                        "vendor": result["vendor"],
                        "invoice_number": result["invoice_number"],
                        "duplicate": result.get("duplicate", False),
                        "existing_line_count": result.get("existing_line_count", 0),
                        "existing_total": result.get("existing_total", 0),
                        "existing_source_file": result.get("existing_source_file", ""),
                    },
                )
            if parsed.path == "/api/texas-financial-import":
                form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": self.headers.get("Content-Type")})
                report_date = form.getvalue("report_date") or None
                report_type = form.getvalue("report_type") or "combined"
                file_items = form["files"] if "files" in form else []
                if not isinstance(file_items, list):
                    file_items = [file_items]
                imported = 0
                metric_count = 0
                details = []
                duplicates = []
                for file_item in file_items:
                    if not getattr(file_item, "filename", ""):
                        continue
                    safe_name = Path(file_item.filename).name
                    path = UPLOAD_DIR / f"{datetime.now().strftime('%Y%m%d%H%M%S')}-{safe_name}"
                    with open(path, "wb") as f:
                        f.write(file_item.file.read())
                    result = import_financial_report(path, report_date, report_type)
                    if result.get("duplicate"):
                        try:
                            path.unlink()
                        except Exception:
                            pass
                        duplicates.append({
                            "file": result.get("original_source_file") or safe_name,
                            "report_date": result.get("report_date"),
                            "report_type": result.get("report_type"),
                            "existing_report_id": result.get("report_id"),
                        })
                        details.append({"file": safe_name, "duplicate": True, "metrics": 0})
                        continue
                    imported += 1
                    metric_count += result["count"]
                    details.append({"file": safe_name, "duplicate": False, "metrics": result["count"], "report_date": result.get("report_date")})
                return json_response(self, {"imported": imported, "metric_count": metric_count, "details": details, "duplicates": duplicates, "skipped_duplicates": len(duplicates)})

            if parsed.path == "/api/customer-invoices":
                actor = current_user(self)
                now = datetime.now().isoformat(timespec="seconds")
                form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": self.headers.get("Content-Type")})
                file_item = form["invoice_file"] if "invoice_file" in form else None
                if file_item is None or not getattr(file_item, "filename", ""):
                    return json_response(self, {"error": "Attach our invoice PDF before adding the customer invoice."}, 400)
                safe_name = Path(file_item.filename).name
                if not safe_name.lower().endswith(".pdf"):
                    return json_response(self, {"error": "Our invoice attachment must be a PDF."}, 400)
                saved_name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}-{safe_name}"
                path = UPLOAD_DIR / saved_name
                with open(path, "wb") as f:
                    f.write(file_item.file.read())
                project_id = form.getvalue("project_id")
                invoice_amount = money(form.getvalue("amount"))
                try:
                    allocation_payload = json.loads(form.getvalue("allocations") or "[]")
                except json.JSONDecodeError:
                    allocation_payload = []
                if not allocation_payload:
                    target_key = f"co:{form.getvalue('change_order_id')}" if form.getvalue("change_order_id") else f"sp:{form.getvalue('subproject_id')}" if form.getvalue("subproject_id") else ""
                    allocation_payload = [{"target_key": target_key, "amount": invoice_amount, "notes": ""}]
                allocation_total = sum(money(item.get("amount")) for item in allocation_payload)
                if not allocation_payload or abs(allocation_total - invoice_amount) >= 0.01:
                    return json_response(self, {"error": "Allocation total must match the invoice amount."}, 400)
                with db() as con:
                    target_rows = []
                    for item in allocation_payload:
                        target_key = str(item.get("target_key") or "")
                        if ":" not in target_key:
                            return json_response(self, {"error": "Each allocation line needs a job/order target."}, 400)
                        target_type, target_id = target_key.split(":", 1)
                        if target_type == "sp":
                            sp = con.execute("SELECT id FROM subprojects WHERE id = ? AND project_id = ?", (target_id, project_id)).fetchone()
                            if not sp:
                                return json_response(self, {"error": "One allocation target is not part of this master project."}, 400)
                            target_rows.append({"subproject_id": sp["id"], "change_order_id": None, "amount": money(item.get("amount")), "notes": str(item.get("notes") or "").strip()})
                        elif target_type == "co":
                            co = con.execute("SELECT id, subproject_id FROM change_orders WHERE id = ? AND project_id = ?", (target_id, project_id)).fetchone()
                            if not co:
                                return json_response(self, {"error": "One allocation target is not part of this master project."}, 400)
                            target_rows.append({"subproject_id": co["subproject_id"], "change_order_id": co["id"], "amount": money(item.get("amount")), "notes": str(item.get("notes") or "").strip()})
                        else:
                            return json_response(self, {"error": "Each allocation target must be a subproject or change order."}, 400)
                    primary = target_rows[0] if target_rows else {"subproject_id": None, "change_order_id": None}
                    cursor = con.execute(
                        """
                        INSERT INTO customer_invoices (
                          project_id, subproject_id, change_order_id, invoice_number, billing_type,
                          invoice_date, due_date, status, amount, paid_amount, invoice_file, notes, created_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            project_id,
                            primary["subproject_id"],
                            primary["change_order_id"],
                            form.getvalue("invoice_number"),
                            form.getvalue("billing_type") or "Progress",
                            form.getvalue("invoice_date"),
                            form.getvalue("due_date"),
                            form.getvalue("status") or "Draft",
                            invoice_amount,
                            money(form.getvalue("paid_amount")),
                            saved_name,
                            form.getvalue("notes"),
                            now,
                        ),
                    )
                    new_id = cursor.lastrowid
                    for target in target_rows:
                        con.execute(
                            """
                            INSERT INTO customer_invoice_allocations (
                              customer_invoice_id, project_id, subproject_id, change_order_id, amount,
                              notes, created_by_user_id, created_by_username, created_at
                            )
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                new_id,
                                project_id,
                                target["subproject_id"],
                                target["change_order_id"],
                                target["amount"],
                                target["notes"],
                                actor["id"] if actor else None,
                                actor["username"] if actor else "",
                                now,
                            ),
                        )
                return json_response(self, {"id": new_id})

            data = parse_json(self)
            now = datetime.now().isoformat(timespec="seconds")
            if parsed.path == "/api/texas-ap":
                actor = current_user(self)
                if not require_editor(self):
                    return json_response(self, {"error": "Admin or standard user access required."}, 403)
                report_date = date_text(data.get("report_date"))
                if not report_date:
                    return json_response(self, {"error": "Choose a week ending date."}, 400)
                ap_total = money(data.get("ap_total"))
                notes = str(data.get("notes") or "").strip()
                with db() as con:
                    con.execute(
                        """
                        INSERT INTO texas_ap_snapshots (
                          report_date, ap_total, notes, created_by_user_id, created_by_username, created_at, updated_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(report_date) DO UPDATE SET
                          ap_total = excluded.ap_total,
                          notes = excluded.notes,
                          created_by_user_id = excluded.created_by_user_id,
                          created_by_username = excluded.created_by_username,
                          updated_at = excluded.updated_at
                        """,
                        (report_date, ap_total, notes, actor["id"], actor.get("username"), now, now),
                    )
                log_activity(actor, "saved Texas AP total", "Texas Ops", report_date, {"ap_total": ap_total, "notes": notes})
                return json_response(self, {"ok": True, "report_date": report_date, "ap_total": ap_total})
            if parsed.path == "/api/texas-ap-delete":
                actor = current_user(self)
                if not require_editor(self):
                    return json_response(self, {"error": "Admin or standard user access required."}, 403)
                ap_id = data.get("id")
                with db() as con:
                    entry = con.execute("SELECT * FROM texas_ap_snapshots WHERE id = ?", (ap_id,)).fetchone()
                    deleted = con.execute("DELETE FROM texas_ap_snapshots WHERE id = ?", (ap_id,)).rowcount
                if deleted and entry:
                    log_activity(actor, "deleted Texas AP total", "Texas Ops", entry["report_date"], {"ap_total": entry["ap_total"], "notes": entry["notes"]})
                return json_response(self, {"deleted": deleted})
            if parsed.path == "/api/texas-financial-delete":
                report_id = data.get("report_id")
                with db() as con:
                    report = con.execute("SELECT * FROM financial_reports WHERE id = ?", (report_id,)).fetchone()
                    metric_count = con.execute("DELETE FROM financial_metrics WHERE report_id = ?", (report_id,)).rowcount
                    report_count = con.execute("DELETE FROM financial_reports WHERE id = ?", (report_id,)).rowcount
                if report_count:
                    log_activity(current_user(self), "deleted financial report", "Texas Ops", f"{report['report_date']} / {report['source_file']}", {"report_id": report_id, "report_type": report["report_type"], "deleted_metrics": metric_count})
                return json_response(self, {"deleted": report_count, "deleted_metrics": metric_count})
            if parsed.path == "/api/users":
                if not require_admin(self):
                    return json_response(self, {"error": "Admin required"}, 403)
                po_auto_issue = 1 if str(data.get("po_auto_issue") or "") in ("1", "true", "on") else 0
                new_id = execute(
                    "INSERT INTO users (username, display_name, password_hash, role, active, po_auto_issue, must_change_password, created_at) VALUES (?, ?, ?, ?, 1, ?, 1, ?)",
                    (data.get("username"), data.get("display_name"), hash_password(DEFAULT_NEW_USER_PASSWORD), clean_role(data.get("role")), po_auto_issue, now),
                )
                log_activity(current_user(self), "created user", "User", data.get("username"), {"display_name": data.get("display_name"), "role": clean_role(data.get("role")), "po_auto_issue": po_auto_issue})
                return json_response(self, {"id": new_id})
            if parsed.path.startswith("/api/projects/") and parsed.path.endswith("/archive"):
                project_id = parsed.path.split("/")[-2]
                execute(
                    "UPDATE projects SET status = 'Archived', closed_at = COALESCE(closed_at, ?), archived_at = ? WHERE id = ?",
                    (now, now, project_id),
                )
                return json_response(self, {"ok": True})
            if parsed.path.startswith("/api/projects/") and parsed.path.endswith("/restore"):
                project_id = parsed.path.split("/")[-2]
                execute(
                    "UPDATE projects SET status = 'Active', closed_at = NULL, archived_at = NULL WHERE id = ?",
                    (project_id,),
                )
                return json_response(self, {"ok": True})
            if parsed.path == "/api/projects":
                new_id = execute(
                    "INSERT INTO projects (project_code, name, customer, location, customer_po, description, rate_set_id, contract_value, status, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'Active', ?)",
                    (data.get("project_code"), data.get("name"), data.get("customer"), data.get("location"), data.get("customer_po"), data.get("description"), data.get("rate_set_id") or None, money(data.get("contract_value")), now),
                )
                return json_response(self, {"id": new_id})
            if parsed.path == "/api/bids":
                estimated_cost = money(data.get("estimated_cost"))
                target_margin = money(data.get("target_margin"))
                probability = money(data.get("probability"))
                bid_price = bid_price_value(estimated_cost, target_margin)
                outcome = data.get("outcome") or "Pending"
                weighted = weighted_bid_value(outcome, bid_price, probability)
                new_id = execute(
                    """
                    INSERT INTO bid_requests (
                      rfq_no, date_received, customer, project_name, estimator, stage, bid_due_date,
                      go_no_go, estimated_cost, target_margin, bid_price, probability, weighted_value,
                      submission_status, outcome, notes, created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        data.get("rfq_no"),
                        data.get("date_received"),
                        data.get("customer"),
                        data.get("project_name"),
                        data.get("estimator"),
                        data.get("stage"),
                        data.get("bid_due_date"),
                        data.get("go_no_go"),
                        estimated_cost,
                        target_margin,
                        bid_price,
                        probability,
                        weighted,
                        data.get("stage"),
                        outcome,
                        data.get("notes"),
                        now,
                        now,
                    ),
                )
                return json_response(self, {"id": new_id})
            if parsed.path == "/api/subprojects":
                new_id = execute(
                    "INSERT INTO subprojects (project_id, job_number, code, name, pricing_type, contract_value, budget_labor_hours, budget_labor, budget_material, budget_equipment, budget_vendor, budget_other) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 0)",
                    (
                        data.get("project_id"),
                        data.get("job_number"),
                        data.get("code"),
                        data.get("name"),
                        data.get("pricing_type") or "Fixed",
                        money(data.get("contract_value")),
                        money(data.get("budget_labor_hours")),
                        money(data.get("budget_labor")),
                        money(data.get("budget_material")),
                        money(data.get("budget_equipment")),
                    ),
                )
                execute(
                    "UPDATE projects SET contract_value = (SELECT COALESCE(SUM(contract_value), 0) FROM subprojects WHERE project_id = ?) WHERE id = ?",
                    (data.get("project_id"), data.get("project_id")),
                )
                return json_response(self, {"id": new_id})
            if parsed.path.startswith("/api/subprojects/") and parsed.path.endswith("/copy"):
                subproject_id = parsed.path.split("/")[-2]
                new_job_number = str(data.get("job_number") or "").strip()
                if not new_job_number:
                    return json_response(self, {"error": "New job/order number is required."}, 400)
                with db() as con:
                    original = con.execute("SELECT * FROM subprojects WHERE id = ?", (subproject_id,)).fetchone()
                    if not original:
                        return json_response(self, {"error": "Subproject not found."}, 404)
                    copied_name = str(data.get("name") or original["name"] or "").strip()
                    copied_pricing_type = data.get("pricing_type") or original["pricing_type"] or "Fixed"
                    if not copied_name:
                        return json_response(self, {"error": "Name is required."}, 400)
                    existing_codes = {
                        str(r["code"] or "").strip()
                        for r in con.execute("SELECT code FROM subprojects WHERE project_id = ?", (original["project_id"],)).fetchall()
                    }
                    new_code = str(data.get("code") or original["code"] or "").strip()
                    if not new_code or new_code in existing_codes:
                        new_code = new_job_number
                    base_code = new_code
                    suffix = 2
                    while new_code in existing_codes:
                        new_code = f"{base_code}-{suffix}"
                        suffix += 1
                    cur = con.execute(
                        """
                        INSERT INTO subprojects (
                          project_id, job_number, code, name, pricing_type, contract_value,
                          budget_labor_hours, budget_labor, budget_material, budget_equipment,
                          budget_vendor, budget_other
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            original["project_id"],
                            new_job_number,
                            new_code,
                            copied_name,
                            copied_pricing_type,
                            money(data.get("contract_value") if "contract_value" in data else original["contract_value"]),
                            money(data.get("budget_labor_hours") if "budget_labor_hours" in data else original["budget_labor_hours"]),
                            money(data.get("budget_labor") if "budget_labor" in data else original["budget_labor"]),
                            money(data.get("budget_material") if "budget_material" in data else original["budget_material"]),
                            money(data.get("budget_equipment") if "budget_equipment" in data else original["budget_equipment"]),
                            money(original["budget_vendor"]),
                            money(original["budget_other"]),
                        ),
                    )
                    con.execute(
                        "UPDATE projects SET contract_value = (SELECT COALESCE(SUM(contract_value), 0) FROM subprojects WHERE project_id = ?) WHERE id = ?",
                        (original["project_id"], original["project_id"]),
                    )
                    return json_response(self, {"id": cur.lastrowid, "code": new_code})
            if parsed.path == "/api/change-orders":
                with db() as con:
                    duplicate_message = duplicate_job_order_message(con, data.get("project_id"), data.get("job_number"))
                    if duplicate_message:
                        return json_response(self, {"error": duplicate_message}, 400)
                    cur = con.execute(
                        "INSERT INTO change_orders (project_id, subproject_id, co_number, job_number, order_type, pricing_type, title, status, quoted_value, approved_value) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (
                            data.get("project_id"),
                            data.get("subproject_id") or None,
                            data.get("co_number"),
                            data.get("job_number"),
                            clean_order_type(data.get("order_type")),
                            data.get("pricing_type") or "Fixed",
                            data.get("title"),
                            data.get("status"),
                            money(data.get("quoted_value")),
                            0 if data.get("pricing_type") == "T&M" else money(data.get("approved_value")),
                        ),
                    )
                    new_id = cur.lastrowid
                return json_response(self, {"id": new_id})
            if parsed.path == "/api/cog-categories":
                code = str(data.get("code") or "").strip().upper()
                name = str(data.get("name") or "").strip()
                description = str(data.get("description") or "").strip()
                if not code:
                    return json_response(self, {"error": "Internal COG code is required."}, 400)
                if not name:
                    return json_response(self, {"error": "COG name is required."}, 400)
                with db() as con:
                    existing = con.execute("SELECT id FROM cog_categories WHERE lower(name) = lower(?)", (name,)).fetchone()
                    if existing:
                        con.execute(
                            "UPDATE cog_categories SET code = ?, description = ?, active = 1 WHERE id = ?",
                            (code, description, existing["id"]),
                        )
                        new_id = existing["id"]
                    else:
                        cur = con.execute(
                            "INSERT INTO cog_categories (code, name, description, active, created_at) VALUES (?, ?, ?, 1, ?)",
                            (code, name, description, datetime.now().isoformat(timespec="seconds")),
                        )
                        new_id = cur.lastrowid
                return json_response(self, {"id": new_id})
            if parsed.path.startswith("/api/change-orders/") and parsed.path.endswith("/copy"):
                change_order_id = parsed.path.split("/")[-2]
                new_co_number = str(data.get("co_number") or "").strip()
                new_job_number = str(data.get("job_number") or "").strip()
                if not new_co_number:
                    return json_response(self, {"error": "New CO number is required."}, 400)
                if not new_job_number:
                    return json_response(self, {"error": "New job/order number is required."}, 400)
                with db() as con:
                    original = con.execute("SELECT * FROM change_orders WHERE id = ?", (change_order_id,)).fetchone()
                    if not original:
                        return json_response(self, {"error": "Change order not found."}, 404)
                    subproject_id = data.get("subproject_id") or original["subproject_id"]
                    if subproject_id:
                        subproject = con.execute("SELECT id FROM subprojects WHERE id = ? AND project_id = ?", (subproject_id, original["project_id"])).fetchone()
                        if not subproject:
                            return json_response(self, {"error": "Selected subproject does not belong to this master project."}, 400)
                    duplicate = con.execute(
                        "SELECT id FROM change_orders WHERE project_id = ? AND COALESCE(subproject_id, '') = COALESCE(?, '') AND co_number = ?",
                        (original["project_id"], subproject_id, new_co_number),
                    ).fetchone()
                    if duplicate:
                        return json_response(self, {"error": "A change order with that CO number already exists for the selected subproject."}, 400)
                    duplicate_job = duplicate_job_order_message(con, original["project_id"], new_job_number)
                    if duplicate_job:
                        return json_response(self, {"error": duplicate_job}, 400)
                    pricing_type = data.get("pricing_type") or original["pricing_type"] or "Fixed"
                    cur = con.execute(
                        """
                        INSERT INTO change_orders (
                          project_id, subproject_id, co_number, job_number, order_type, pricing_type,
                          title, status, quoted_value, approved_value
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            original["project_id"],
                            subproject_id or None,
                            new_co_number,
                            new_job_number,
                            clean_order_type(data.get("order_type") if "order_type" in data else original["order_type"]),
                            pricing_type,
                            str(data.get("title") if "title" in data else original["title"] or "").strip(),
                            data.get("status") or original["status"] or "Pending",
                            money(data.get("quoted_value") if "quoted_value" in data else original["quoted_value"]),
                            0 if pricing_type == "T&M" else money(data.get("approved_value") if "approved_value" in data else original["approved_value"]),
                        ),
                    )
                    return json_response(self, {"id": cur.lastrowid})
            if parsed.path == "/api/internal-rates":
                existing_rate = one(
                    "SELECT id FROM internal_rates WHERE rate_set_id = ? AND category_type = ? AND category = ?",
                    (data.get("rate_set_id") or None, data.get("category_type"), data.get("category")),
                )
                if existing_rate:
                    execute(
                        "UPDATE internal_rates SET raw_rate = ?, active = 1 WHERE id = ?",
                        (money(data.get("raw_rate")), existing_rate["id"]),
                    )
                    new_id = existing_rate["id"]
                else:
                    new_id = execute(
                        "INSERT INTO internal_rates (rate_set_id, category_type, category, raw_rate, active) VALUES (?, ?, ?, ?, 1)",
                        (data.get("rate_set_id") or None, data.get("category_type"), data.get("category"), money(data.get("raw_rate"))),
                    )
                updated = apply_internal_rate(data.get("category_type"), data.get("category"), money(data.get("raw_rate")), data.get("rate_set_id"))
                return json_response(self, {"id": new_id, "updated_cost_records": updated})
            if parsed.path == "/api/imports/delete":
                with db() as con:
                    deleted = con.execute(
                        """
                        DELETE FROM cost_records
                        WHERE project_id = ?
                          AND source = ?
                          AND source_file = ?
                        """,
                        (data.get("project_id"), data.get("source"), data.get("source_file")),
                    ).rowcount
                if deleted:
                    log_activity(current_user(self), "deleted import", str(data.get("source") or "Import"), str(data.get("source_file") or ""), {"project_id": data.get("project_id"), "deleted_records": deleted})
                return json_response(self, {"deleted": deleted})
            if parsed.path == "/api/cost-records/bulk-update":
                ids = [str(x) for x in data.get("ids", []) if str(x).isdigit()]
                if not ids:
                    return json_response(self, {"updated": 0})
                change_order_id = data.get("change_order_id") or None
                subproject_id = data.get("subproject_id") or None
                placeholders = ",".join(["?"] * len(ids))
                with db() as con:
                    updated = con.execute(
                        f"""
                        UPDATE cost_records
                        SET subproject_id = ?,
                            change_order_id = ?,
                            amount = CASE
                              WHEN cost_type = 'Field Ticket Material' THEN
                                CASE WHEN ? IS NOT NULL OR COALESCE((SELECT pricing_type FROM subprojects WHERE id = ?), 'Fixed') = 'T&M'
                                  THEN COALESCE(sales_amount, 0) * ? ELSE 0 END
                              ELSE amount
                            END,
                            raw_rate = CASE
                              WHEN cost_type = 'Field Ticket Material' THEN
                                CASE WHEN ? IS NOT NULL OR COALESCE((SELECT pricing_type FROM subprojects WHERE id = ?), 'Fixed') = 'T&M'
                                  THEN COALESCE(sales_rate, rate, 0) * ? ELSE 0 END
                              ELSE raw_rate
                            END,
                            raw_cost_source = CASE
                              WHEN cost_type = 'Field Ticket Material' THEN
                                CASE
                                  WHEN ? IS NOT NULL THEN 'CO T&M material estimate at 35% margin'
                                  WHEN COALESCE((SELECT pricing_type FROM subprojects WHERE id = ?), 'Fixed') = 'T&M' THEN 'Subproject T&M material estimate at 35% margin'
                                  ELSE 'Usage only - not budget cost'
                                END
                              ELSE raw_cost_source
                            END
                        WHERE id IN ({placeholders})
                        """,
                        (
                            subproject_id,
                            change_order_id,
                            change_order_id,
                            subproject_id,
                            CO_MATERIAL_COST_FACTOR,
                            change_order_id,
                            subproject_id,
                            CO_MATERIAL_COST_FACTOR,
                            change_order_id,
                            subproject_id,
                            *ids,
                        ),
                    ).rowcount
                return json_response(self, {"updated": updated})
            if parsed.path == "/api/vendor-invoice/subproject":
                with db() as con:
                    updated = con.execute(
                        """
                        UPDATE cost_records
                        SET subproject_id = ?
                        WHERE project_id = ?
                          AND source = 'Vendor Invoice'
                          AND source_file = ?
                          AND ticket_or_invoice = ?
                          AND COALESCE(vendor, '') = COALESCE(?, '')
                        """,
                        (
                            data.get("subproject_id") or None,
                            data.get("project_id"),
                            data.get("source_file"),
                            data.get("ticket_or_invoice"),
                            data.get("vendor") or "",
                        ),
                    ).rowcount
                return json_response(self, {"updated": updated})
            if parsed.path == "/api/vendor-invoice/allocate":
                actor = current_user(self)
                project_id = data.get("project_id")
                targets = [str(t) for t in data.get("targets", []) if str(t).startswith(("sp:", "co:"))]
                targets = list(dict.fromkeys(targets))
                if len(targets) < 2:
                    return json_response(self, {"error": "Choose at least two allocation targets."}, 400)
                with db() as con:
                    invoice_rows = con.execute(
                        """
                        SELECT *
                        FROM cost_records
                        WHERE project_id = ?
                          AND source = 'Vendor Invoice'
                          AND source_file = ?
                          AND ticket_or_invoice = ?
                          AND COALESCE(vendor, '') = COALESCE(?, '')
                        ORDER BY id
                        """,
                        (project_id, data.get("source_file"), data.get("ticket_or_invoice"), data.get("vendor") or ""),
                    ).fetchall()
                    if not invoice_rows:
                        return json_response(self, {"error": "Invoice was not found."}, 404)
                    total_amount = sum(money(r["amount"]) for r in invoice_rows)
                    if not total_amount:
                        return json_response(self, {"error": "Invoice amount is zero."}, 400)
                    target_rows = []
                    for target in targets:
                        target_type, target_id = target.split(":", 1)
                        if target_type == "sp":
                            sp = con.execute("SELECT id FROM subprojects WHERE id = ? AND project_id = ?", (target_id, project_id)).fetchone()
                            if sp:
                                target_rows.append({"subproject_id": sp["id"], "change_order_id": None})
                        elif target_type == "co":
                            co = con.execute("SELECT id, subproject_id FROM change_orders WHERE id = ? AND project_id = ?", (target_id, project_id)).fetchone()
                            if co:
                                target_rows.append({"subproject_id": co["subproject_id"], "change_order_id": co["id"]})
                    if len(target_rows) < 2:
                        return json_response(self, {"error": "At least two selected targets must belong to this master project."}, 400)
                    first = invoice_rows[0]
                    now_alloc = datetime.now().isoformat(timespec="seconds")
                    allocation_cursor = con.execute(
                        """
                        INSERT INTO vendor_invoice_allocations (
                          project_id, source_file, ticket_or_invoice, vendor, original_total,
                          allocation_count, allocated_by_user_id, allocated_by_username, allocated_at, notes
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            project_id,
                            first["source_file"],
                            first["ticket_or_invoice"],
                            first["vendor"],
                            total_amount,
                            len(target_rows),
                            actor["id"] if actor else None,
                            actor["username"] if actor else "",
                            now_alloc,
                            json.dumps({"source_line_count": len(invoice_rows)}, default=str),
                        ),
                    )
                    allocation_id = allocation_cursor.lastrowid
                    con.execute(
                        """
                        DELETE FROM cost_records
                        WHERE project_id = ?
                          AND source = 'Vendor Invoice'
                          AND source_file = ?
                          AND ticket_or_invoice = ?
                          AND COALESCE(vendor, '') = COALESCE(?, '')
                        """,
                        (project_id, data.get("source_file"), data.get("ticket_or_invoice"), data.get("vendor") or ""),
                    )
                    base_share = round(total_amount / len(target_rows), 2)
                    allocated_total = 0
                    for idx, target in enumerate(target_rows):
                        amount = round(total_amount - allocated_total, 2) if idx == len(target_rows) - 1 else base_share
                        allocated_total += amount
                        con.execute(
                            """
                            INSERT INTO vendor_invoice_allocation_lines (allocation_id, subproject_id, change_order_id, amount)
                            VALUES (?, ?, ?, ?)
                            """,
                            (allocation_id, target["subproject_id"], target["change_order_id"], amount),
                        )
                        con.execute(
                            """
                            INSERT INTO cost_records (
                              project_id, subproject_id, change_order_id, source, source_file, ticket_or_invoice,
                              record_date, status, cost_type, item, description, qty, rate, amount,
                              sales_rate, sales_amount, raw_rate, raw_cost_source, vendor, notes, created_at
                            )
                            VALUES (?, ?, ?, 'Vendor Invoice', ?, ?, ?, 'Allocated', 'Material', ?, ?, 1, ?, ?, 0, 0, ?, 'Vendor invoice even allocation', ?, ?, ?)
                            """,
                            (
                                project_id,
                                target["subproject_id"],
                                target["change_order_id"],
                                first["source_file"],
                                first["ticket_or_invoice"],
                                first["record_date"],
                                "Allocated Material",
                                f"Even allocation of invoice {first['ticket_or_invoice'] or ''} across {len(target_rows)} jobs",
                                amount,
                                amount,
                                amount,
                                first["vendor"],
                                json.dumps({"allocation_id": allocation_id, "allocated_from_line_count": len(invoice_rows), "allocated_total": total_amount, "allocation_count": len(target_rows), "allocated_by": actor["username"] if actor else ""}, default=str),
                                now_alloc,
                            ),
                        )
                return json_response(self, {"allocated": len(target_rows), "total_amount": total_amount, "share": round(total_amount / len(target_rows), 2)})
            if parsed.path == "/api/invoices":
                new_id = execute(
                    """
                    INSERT INTO cost_records (
                      project_id, subproject_id, change_order_id, source, ticket_or_invoice, record_date, status,
                      cost_type, description, amount, vendor, created_at
                    ) VALUES (?, ?, ?, 'Vendor Invoice', ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (data.get("project_id"), data.get("subproject_id") or None, data.get("change_order_id") or None, data.get("ticket_or_invoice"), data.get("record_date"), data.get("status"), data.get("cost_type"), data.get("description"), money(data.get("amount")), data.get("vendor"), now),
                )
                return json_response(self, {"id": new_id})
            return json_response(self, {"error": "Not found"}, 404)
        except Exception as e:
            traceback.print_exc()
            return json_response(self, {"error": str(e)}, 500)

    def do_PUT(self):
        try:
            parsed = urlparse(self.path)
            data = parse_json(self)
            if not current_user(self):
                return json_response(self, {"error": "Login required"}, 401)
            if not require_editor(self):
                return json_response(self, {"error": "Read-only users cannot make changes."}, 403)
            if parsed.path.startswith("/api/users/"):
                if not require_admin(self):
                    return json_response(self, {"error": "Admin required"}, 403)
                user_id = parsed.path.rsplit("/", 1)[-1]
                target_user = one("SELECT username, display_name, role, active FROM users WHERE id = ?", (user_id,))
                if "password" in data and data.get("password"):
                    execute("UPDATE users SET password_hash = ?, must_change_password = 1 WHERE id = ?", (hash_password(data.get("password")), user_id))
                if "role" in data:
                    execute("UPDATE users SET role = ? WHERE id = ?", (clean_role(data.get("role")), user_id))
                if "active" in data:
                    execute("UPDATE users SET active = ? WHERE id = ?", (1 if str(data.get("active")) == "1" else 0, user_id))
                if "po_auto_issue" in data:
                    execute("UPDATE users SET po_auto_issue = ? WHERE id = ?", (1 if str(data.get("po_auto_issue")) == "1" else 0, user_id))
                log_activity(current_user(self), "updated user", "User", target_user["username"] if target_user else user_id, {"changed_fields": list(data.keys()), "new_role": clean_role(data.get("role")) if "role" in data else None, "active": data.get("active") if "active" in data else None, "po_auto_issue": data.get("po_auto_issue") if "po_auto_issue" in data else None})
                return json_response(self, {"ok": True})
            if parsed.path.startswith("/api/purchase-orders/"):
                po_id = parsed.path.rsplit("/", 1)[-1]
                actor = current_user(self)
                status = str(data.get("status") or "Pending Approval").strip()
                if status not in ("Pending Approval", "Issued", "Picked Up", "Closed", "Void"):
                    return json_response(self, {"error": "Choose a valid PO status."}, 400)
                vendor = str(data.get("vendor") or "").strip()
                customer_name = str(data.get("customer_name") or "").strip()
                description = str(data.get("description") or "").strip()
                if not vendor:
                    return json_response(self, {"error": "Vendor is required."}, 400)
                if not description:
                    return json_response(self, {"error": "Details are required."}, 400)
                with db() as con:
                    current_po = con.execute("SELECT * FROM purchase_orders WHERE id = ?", (po_id,)).fetchone()
                    if not current_po:
                        return json_response(self, {"error": "PO not found."}, 404)
                    job_ref = job_reference_for_po(con, data.get("job_key"))
                    if not job_ref:
                        return json_response(self, {"error": "Choose a valid job/order number."}, 400)
                    if po_customer_required_for_ref(job_ref) and not customer_name:
                        return json_response(self, {"error": "Customer is required for shop / COG POs."}, 400)
                    now = datetime.now().isoformat(timespec="seconds")
                    close_reason = str(data.get("close_reason") or current_po["close_reason"] or "").strip()
                    void_reason = str(data.get("void_reason") or current_po["void_reason"] or "").strip()
                    closed_at = current_po["closed_at"]
                    closed_by_user_id = current_po["closed_by_user_id"]
                    closed_by_username = current_po["closed_by_username"]
                    voided_at = current_po["voided_at"]
                    voided_by_user_id = current_po["voided_by_user_id"]
                    voided_by_username = current_po["voided_by_username"]
                    action_detail = None
                    if status == "Closed":
                        if not close_reason:
                            return json_response(self, {"error": "Close reason is required to close a PO."}, 400)
                        if current_po["status"] != "Closed" or not closed_at:
                            closed_at = now
                            closed_by_user_id = actor["id"]
                            closed_by_username = actor["username"]
                            action_detail = ("closed PO", {"close_reason": close_reason, "closed_at": now})
                        voided_at = None
                        voided_by_user_id = None
                        voided_by_username = None
                        void_reason = ""
                    elif status == "Void":
                        if not void_reason:
                            return json_response(self, {"error": "Void reason is required to void a PO."}, 400)
                        if current_po["status"] != "Void" or not voided_at:
                            voided_at = now
                            voided_by_user_id = actor["id"]
                            voided_by_username = actor["username"]
                            action_detail = ("voided PO", {"void_reason": void_reason, "voided_at": now})
                        closed_at = None
                        closed_by_user_id = None
                        closed_by_username = None
                        close_reason = ""
                    else:
                        closed_at = None
                        closed_by_user_id = None
                        closed_by_username = None
                        close_reason = ""
                        voided_at = None
                        voided_by_user_id = None
                        voided_by_username = None
                        void_reason = ""
                    updated = con.execute(
                        """
                        UPDATE purchase_orders
                        SET project_id = ?,
                            subproject_id = ?,
                            change_order_id = ?,
                            cog_category_id = ?,
                            job_number = ?,
                            job_label = ?,
                            customer_name = ?,
                            vendor = ?,
                            description = ?,
                            estimated_amount = ?,
                            status = ?,
                            closed_at = ?,
                            closed_by_user_id = ?,
                            closed_by_username = ?,
                            close_reason = ?,
                            voided_at = ?,
                            voided_by_user_id = ?,
                            voided_by_username = ?,
                            void_reason = ?,
                            updated_at = ?
                        WHERE id = ?
                        """,
                        (
                            job_ref["project_id"],
                            job_ref["subproject_id"],
                            job_ref["change_order_id"],
                            job_ref["cog_category_id"],
                            job_ref["job_number"],
                            job_ref["job_label"],
                            customer_name,
                            vendor,
                            description,
                            money(data.get("estimated_amount")),
                            status,
                            closed_at,
                            closed_by_user_id,
                            closed_by_username,
                            close_reason,
                            voided_at,
                            voided_by_user_id,
                            voided_by_username,
                            void_reason,
                            now,
                            po_id,
                        ),
                    ).rowcount
                    if not updated:
                        return json_response(self, {"error": "PO not found."}, 404)
                if action_detail:
                    log_activity(actor, action_detail[0], "Purchase Order", current_po["po_number"], action_detail[1])
                log_activity(actor, "updated PO", "Purchase Order", current_po["po_number"], {"status": status, "vendor": vendor, "estimated_amount": money(data.get("estimated_amount"))})
                return json_response(self, {"ok": True})
            if parsed.path.startswith("/api/cost-records/"):
                record_id = parsed.path.rsplit("/", 1)[-1]
                cost_type = data.get("cost_type") or None
                if cost_type == "Field Ticket Material":
                    change_order_id = data.get("change_order_id") or None
                    subproject_id = data.get("subproject_id") or None
                    execute(
                        """
                        UPDATE cost_records
                        SET subproject_id = ?,
                            change_order_id = ?,
                            cost_type = ?,
                            amount = CASE
                              WHEN ? IS NOT NULL OR COALESCE((SELECT pricing_type FROM subprojects WHERE id = ?), 'Fixed') = 'T&M'
                                THEN COALESCE(sales_amount, 0) * ?
                              ELSE 0
                            END,
                            raw_rate = CASE
                              WHEN ? IS NOT NULL OR COALESCE((SELECT pricing_type FROM subprojects WHERE id = ?), 'Fixed') = 'T&M'
                                THEN COALESCE(sales_rate, rate, 0) * ?
                              ELSE 0
                            END,
                            raw_cost_source = CASE
                              WHEN ? IS NOT NULL THEN 'CO T&M material estimate at 35% margin'
                              WHEN COALESCE((SELECT pricing_type FROM subprojects WHERE id = ?), 'Fixed') = 'T&M' THEN 'Subproject T&M material estimate at 35% margin'
                              ELSE 'Usage only - not budget cost'
                            END
                        WHERE id = ?
                        """,
                        (
                            subproject_id,
                            change_order_id,
                            cost_type,
                            change_order_id,
                            subproject_id,
                            CO_MATERIAL_COST_FACTOR,
                            change_order_id,
                            subproject_id,
                            CO_MATERIAL_COST_FACTOR,
                            change_order_id,
                            subproject_id,
                            record_id,
                        ),
                    )
                else:
                    execute(
                        "UPDATE cost_records SET subproject_id = ?, change_order_id = ?, cost_type = ? WHERE id = ?",
                        (data.get("subproject_id") or None, data.get("change_order_id") or None, cost_type, record_id),
                    )
                return json_response(self, {"ok": True})
            if parsed.path.startswith("/api/projects/"):
                project_id = parsed.path.rsplit("/", 1)[-1]
                execute(
                    """
                    UPDATE projects
                    SET project_code = ?, name = ?, customer = ?, location = ?, customer_po = ?, description = ?, rate_set_id = ?
                    WHERE id = ?
                    """,
                    (
                        data.get("project_code"),
                        data.get("name"),
                        data.get("customer"),
                        data.get("location"),
                        data.get("customer_po"),
                        data.get("description"),
                        data.get("rate_set_id") or None,
                        project_id,
                    ),
                )
                return json_response(self, {"ok": True})
            if parsed.path.startswith("/api/bids/"):
                bid_id = parsed.path.rsplit("/", 1)[-1]
                estimated_cost = money(data.get("estimated_cost"))
                target_margin = money(data.get("target_margin"))
                probability = money(data.get("probability"))
                bid_price = bid_price_value(estimated_cost, target_margin, data.get("bid_price"))
                outcome = data.get("outcome") or "Pending"
                weighted = weighted_bid_value(outcome, bid_price, probability)
                execute(
                    """
                    UPDATE bid_requests
                    SET rfq_no = ?, date_received = ?, customer = ?, project_name = ?, estimator = ?,
                        stage = ?, bid_due_date = ?, go_no_go = ?, estimated_cost = ?, target_margin = ?,
                        bid_price = ?, probability = ?, weighted_value = ?, submission_status = ?,
                        outcome = ?, notes = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        data.get("rfq_no"),
                        data.get("date_received"),
                        data.get("customer"),
                        data.get("project_name"),
                        data.get("estimator"),
                        data.get("stage"),
                        data.get("bid_due_date"),
                        data.get("go_no_go"),
                        estimated_cost,
                        target_margin,
                        bid_price,
                        probability,
                        weighted,
                        data.get("stage"),
                        outcome,
                        data.get("notes"),
                        datetime.now().isoformat(timespec="seconds"),
                        bid_id,
                    ),
                )
                return json_response(self, {"ok": True})
            if parsed.path.startswith("/api/subprojects/"):
                subproject_id = parsed.path.rsplit("/", 1)[-1]
                execute(
                    """
                    UPDATE subprojects
                    SET job_number = ?, code = ?, name = ?, pricing_type = ?, contract_value = ?, budget_labor_hours = ?, budget_labor = ?, budget_material = ?, budget_equipment = ?, budget_vendor = 0, budget_other = 0
                    WHERE id = ?
                    """,
                    (
                        data.get("job_number"),
                        data.get("code"),
                        data.get("name"),
                        data.get("pricing_type") or "Fixed",
                        money(data.get("contract_value")),
                        money(data.get("budget_labor_hours")),
                        money(data.get("budget_labor")),
                        money(data.get("budget_material")),
                        money(data.get("budget_equipment")),
                        subproject_id,
                    ),
                )
                subproject = one("SELECT project_id FROM subprojects WHERE id = ?", (subproject_id,))
                if subproject:
                    execute(
                        """
                        UPDATE cost_records
                        SET amount = CASE
                              WHEN change_order_id IS NOT NULL OR COALESCE((SELECT pricing_type FROM subprojects WHERE id = cost_records.subproject_id), 'Fixed') = 'T&M'
                                THEN COALESCE(sales_amount, 0) * ?
                              ELSE 0
                            END,
                            raw_rate = CASE
                              WHEN change_order_id IS NOT NULL OR COALESCE((SELECT pricing_type FROM subprojects WHERE id = cost_records.subproject_id), 'Fixed') = 'T&M'
                                THEN COALESCE(sales_rate, rate, 0) * ?
                              ELSE 0
                            END,
                            raw_cost_source = CASE
                              WHEN change_order_id IS NOT NULL THEN 'CO T&M material estimate at 35% margin'
                              WHEN COALESCE((SELECT pricing_type FROM subprojects WHERE id = cost_records.subproject_id), 'Fixed') = 'T&M' THEN 'Subproject T&M material estimate at 35% margin'
                              ELSE 'Usage only - not budget cost'
                            END
                        WHERE subproject_id = ?
                          AND cost_type = 'Field Ticket Material'
                        """,
                        (CO_MATERIAL_COST_FACTOR, CO_MATERIAL_COST_FACTOR, subproject_id),
                    )
                    execute(
                        "UPDATE projects SET contract_value = (SELECT COALESCE(SUM(contract_value), 0) FROM subprojects WHERE project_id = ?) WHERE id = ?",
                        (subproject["project_id"], subproject["project_id"]),
                    )
                return json_response(self, {"ok": True})
            if parsed.path.startswith("/api/change-orders/"):
                change_order_id = parsed.path.rsplit("/", 1)[-1]
                pricing_type = data.get("pricing_type") or "Fixed"
                with db() as con:
                    change_order = con.execute("SELECT project_id FROM change_orders WHERE id = ?", (change_order_id,)).fetchone()
                    if not change_order:
                        return json_response(self, {"error": "Change order not found."}, 404)
                    duplicate_message = duplicate_job_order_message(con, change_order["project_id"], data.get("job_number"), change_order_id)
                    if duplicate_message:
                        return json_response(self, {"error": duplicate_message}, 400)
                    con.execute(
                        """
                        UPDATE change_orders
                        SET subproject_id = ?, order_type = ?, co_number = ?, job_number = ?, pricing_type = ?, title = ?, status = ?, quoted_value = ?, approved_value = ?
                        WHERE id = ?
                        """,
                        (
                            data.get("subproject_id") or None,
                            clean_order_type(data.get("order_type")),
                            data.get("co_number"),
                            data.get("job_number"),
                            pricing_type,
                            data.get("title"),
                            data.get("status") or "Pending",
                            money(data.get("quoted_value")),
                            0 if pricing_type == "T&M" else money(data.get("approved_value")),
                            change_order_id,
                        ),
                    )
                    con.execute(
                        """
                        UPDATE cost_records
                        SET subproject_id = ?,
                            amount = CASE
                              WHEN cost_type = 'Field Ticket Material' THEN COALESCE(sales_amount, 0) * ?
                              ELSE amount
                            END,
                            raw_rate = CASE
                              WHEN cost_type = 'Field Ticket Material' THEN COALESCE(sales_rate, rate, 0) * ?
                              ELSE raw_rate
                            END,
                            raw_cost_source = CASE
                              WHEN cost_type = 'Field Ticket Material' THEN 'CO T&M material estimate at 35% margin'
                              ELSE raw_cost_source
                            END
                        WHERE change_order_id = ?
                        """,
                        (data.get("subproject_id") or None, CO_MATERIAL_COST_FACTOR, CO_MATERIAL_COST_FACTOR, change_order_id),
                    )
                return json_response(self, {"ok": True})
            if parsed.path.startswith("/api/customer-invoices/"):
                invoice_id = parsed.path.rsplit("/", 1)[-1]
                with db() as con:
                    invoice = con.execute("SELECT * FROM customer_invoices WHERE id = ?", (invoice_id,)).fetchone()
                    con.execute(
                        """
                        UPDATE customer_invoices
                        SET subproject_id = ?, change_order_id = ?, invoice_number = ?, billing_type = ?,
                            invoice_date = ?, due_date = ?, status = ?, amount = ?, paid_amount = ?, notes = ?
                        WHERE id = ?
                        """,
                        (
                            data.get("subproject_id") or None,
                            data.get("change_order_id") or None,
                            data.get("invoice_number"),
                            data.get("billing_type") or "Progress",
                            data.get("invoice_date"),
                            data.get("due_date"),
                            data.get("status") or "Draft",
                            money(data.get("amount")),
                            money(data.get("paid_amount")),
                            data.get("notes"),
                            invoice_id,
                        ),
                    )
                    if invoice:
                        existing_count = con.execute("SELECT COUNT(*) AS c FROM customer_invoice_allocations WHERE customer_invoice_id = ?", (invoice_id,)).fetchone()["c"]
                        if existing_count <= 1:
                            con.execute("DELETE FROM customer_invoice_allocations WHERE customer_invoice_id = ?", (invoice_id,))
                            con.execute(
                                """
                                INSERT INTO customer_invoice_allocations (
                                  customer_invoice_id, project_id, subproject_id, change_order_id, amount, notes, created_at
                                )
                                VALUES (?, ?, ?, ?, ?, ?, ?)
                                """,
                                (
                                    invoice_id,
                                    invoice["project_id"],
                                    data.get("subproject_id") or None,
                                    data.get("change_order_id") or None,
                                    money(data.get("amount")),
                                    "Updated single-invoice allocation",
                                    datetime.now().isoformat(timespec="seconds"),
                                ),
                            )
                return json_response(self, {"ok": True})
            if parsed.path.startswith("/api/internal-rates/"):
                rate_id = parsed.path.rsplit("/", 1)[-1]
                execute(
                    "UPDATE internal_rates SET category_type = ?, category = ?, raw_rate = ? WHERE id = ?",
                    (data.get("category_type"), data.get("category"), money(data.get("raw_rate")), rate_id),
                )
                rate = one("SELECT rate_set_id FROM internal_rates WHERE id = ?", (rate_id,))
                updated = apply_internal_rate(data.get("category_type"), data.get("category"), money(data.get("raw_rate")), rate["rate_set_id"] if rate else None)
                return json_response(self, {"ok": True, "updated_cost_records": updated})
            if parsed.path.startswith("/api/cog-categories/"):
                category_id = parsed.path.rsplit("/", 1)[-1]
                code = str(data.get("code") or "").strip().upper()
                name = str(data.get("name") or "").strip()
                description = str(data.get("description") or "").strip()
                if not code:
                    return json_response(self, {"error": "Internal COG code is required."}, 400)
                if not name:
                    return json_response(self, {"error": "COG name is required."}, 400)
                with db() as con:
                    duplicate = con.execute("SELECT id FROM cog_categories WHERE lower(name) = lower(?) AND id != ?", (name, category_id)).fetchone()
                    if duplicate:
                        return json_response(self, {"error": "That COG name already exists."}, 400)
                    updated = con.execute(
                        "UPDATE cog_categories SET code = ?, name = ?, description = ?, active = 1 WHERE id = ?",
                        (code, name, description, category_id),
                    ).rowcount
                if not updated:
                    return json_response(self, {"error": "COG category not found."}, 404)
                return json_response(self, {"ok": True})
            return json_response(self, {"error": "Not found"}, 404)
        except Exception as e:
            traceback.print_exc()
            return json_response(self, {"error": str(e)}, 500)

    def do_DELETE(self):
        try:
            parsed = urlparse(self.path)
            if not current_user(self):
                return json_response(self, {"error": "Login required"}, 401)
            if not require_editor(self):
                return json_response(self, {"error": "Read-only users cannot make changes."}, 403)
            if parsed.path.startswith("/api/fieldwise-audit-omissions/"):
                omission_id = parsed.path.rsplit("/", 1)[-1]
                with db() as con:
                    omission = con.execute("SELECT * FROM fieldwise_audit_omissions WHERE id = ?", (omission_id,)).fetchone()
                    deleted = con.execute("DELETE FROM fieldwise_audit_omissions WHERE id = ?", (omission_id,)).rowcount
                if not deleted:
                    return json_response(self, {"error": "Omission note not found."}, 404)
                log_activity(current_user(self), "removed OK-to-omit note", "Field Wise Audit", f"{omission['ticket_number']} / {omission['order_number']}" if omission else omission_id)
                return json_response(self, {"ok": True})
            if parsed.path.startswith("/api/cog-categories/"):
                category_id = parsed.path.rsplit("/", 1)[-1]
                with db() as con:
                    updated = con.execute("UPDATE cog_categories SET active = 0 WHERE id = ?", (category_id,)).rowcount
                if not updated:
                    return json_response(self, {"error": "COG category not found."}, 404)
                return json_response(self, {"ok": True})
            if parsed.path.startswith("/api/purchase-orders/"):
                if not require_admin(self):
                    return json_response(self, {"error": "Admin required"}, 403)
                po_id = parsed.path.rsplit("/", 1)[-1]
                with db() as con:
                    po = con.execute("SELECT * FROM purchase_orders WHERE id = ?", (po_id,)).fetchone()
                    deleted = con.execute("DELETE FROM purchase_orders WHERE id = ?", (po_id,)).rowcount
                if not deleted:
                    return json_response(self, {"error": "PO not found."}, 404)
                log_activity(current_user(self), "deleted PO", "Purchase Order", po["po_number"] if po else po_id, {"vendor": po["vendor"] if po else "", "amount": po["estimated_amount"] if po else 0})
                return json_response(self, {"ok": True})
            if parsed.path.startswith("/api/subprojects/"):
                subproject_id = parsed.path.rsplit("/", 1)[-1]
                with db() as con:
                    subproject = con.execute("SELECT project_id FROM subprojects WHERE id = ?", (subproject_id,)).fetchone()
                    if not subproject:
                        return json_response(self, {"error": "Subproject not found."}, 404)
                    con.execute("UPDATE change_orders SET subproject_id = NULL WHERE subproject_id = ?", (subproject_id,))
                    con.execute("UPDATE customer_invoices SET subproject_id = NULL WHERE subproject_id = ?", (subproject_id,))
                    con.execute("UPDATE customer_invoice_allocations SET subproject_id = NULL WHERE subproject_id = ?", (subproject_id,))
                    con.execute(
                        """
                        UPDATE cost_records
                        SET subproject_id = NULL,
                            amount = CASE
                              WHEN change_order_id IS NULL AND cost_type = 'Field Ticket Material' THEN 0
                              ELSE amount
                            END,
                            raw_rate = CASE
                              WHEN change_order_id IS NULL AND cost_type = 'Field Ticket Material' THEN 0
                              ELSE raw_rate
                            END,
                            raw_cost_source = CASE
                              WHEN change_order_id IS NULL AND cost_type = 'Field Ticket Material' THEN 'Usage only - not budget cost'
                              ELSE raw_cost_source
                            END
                        WHERE subproject_id = ?
                        """,
                        (subproject_id,),
                    )
                    con.execute("DELETE FROM subprojects WHERE id = ?", (subproject_id,))
                    con.execute(
                        "UPDATE projects SET contract_value = (SELECT COALESCE(SUM(contract_value), 0) FROM subprojects WHERE project_id = ?) WHERE id = ?",
                        (subproject["project_id"], subproject["project_id"]),
                    )
                return json_response(self, {"ok": True})
            if parsed.path.startswith("/api/change-orders/"):
                change_order_id = parsed.path.rsplit("/", 1)[-1]
                with db() as con:
                    change_order = con.execute("SELECT id FROM change_orders WHERE id = ?", (change_order_id,)).fetchone()
                    if not change_order:
                        return json_response(self, {"error": "Change order not found."}, 404)
                    con.execute("UPDATE customer_invoices SET change_order_id = NULL WHERE change_order_id = ?", (change_order_id,))
                    con.execute("UPDATE customer_invoice_allocations SET change_order_id = NULL WHERE change_order_id = ?", (change_order_id,))
                    con.execute(
                        """
                        UPDATE cost_records
                        SET change_order_id = NULL,
                            amount = CASE
                              WHEN cost_type = 'Field Ticket Material'
                                THEN CASE WHEN COALESCE((SELECT pricing_type FROM subprojects WHERE id = cost_records.subproject_id), 'Fixed') = 'T&M'
                                  THEN COALESCE(sales_amount, 0) * ? ELSE 0 END
                              ELSE amount
                            END,
                            raw_rate = CASE
                              WHEN cost_type = 'Field Ticket Material'
                                THEN CASE WHEN COALESCE((SELECT pricing_type FROM subprojects WHERE id = cost_records.subproject_id), 'Fixed') = 'T&M'
                                  THEN COALESCE(sales_rate, rate, 0) * ? ELSE 0 END
                              ELSE raw_rate
                            END,
                            raw_cost_source = CASE
                              WHEN cost_type = 'Field Ticket Material'
                                THEN CASE WHEN COALESCE((SELECT pricing_type FROM subprojects WHERE id = cost_records.subproject_id), 'Fixed') = 'T&M'
                                  THEN 'Subproject T&M material estimate at 35% margin' ELSE 'Usage only - not budget cost' END
                              ELSE raw_cost_source
                            END
                        WHERE change_order_id = ?
                        """,
                        (CO_MATERIAL_COST_FACTOR, CO_MATERIAL_COST_FACTOR, change_order_id),
                    )
                    con.execute("DELETE FROM change_orders WHERE id = ?", (change_order_id,))
                return json_response(self, {"ok": True})
            return json_response(self, {"error": "Not found"}, 404)
        except Exception as e:
            traceback.print_exc()
            return json_response(self, {"error": str(e)}, 500)


if __name__ == "__main__":
    init_db()
    start_background_jobs()
    print(f"Company Dashboard running at http://{HOST}:{PORT}")
    print("Press Ctrl+C to stop.")
    ThreadingHTTPServer((HOST, PORT), Handler).serve_forever()
